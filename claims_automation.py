"""Claims automation utilities.

This module builds an Excel output for credit lines (A–AB) from Source File 1,
optionally enriching Detail/Header Narration using Master File 1 and Source File 2.

Key behaviors:
- Fixed and mapped columns per business rules
- Detail/Header Narration concatenation per employee with Excel line breaks
- Robust header detection for Source File 2 (headers may start below row 1)

The functions are written to be composable and testable.
"""

import argparse
import logging
from datetime import date, datetime, timedelta
from typing import List, Dict, Any, Optional, Union, IO

from dateutil import parser as date_parser
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, PatternFill

# ---------------------------------------------------------------------------
# Configuration & constants
# ---------------------------------------------------------------------------

# Excel uses CRLF for line breaks inside cells; we also add a blank line between entries
EXCEL_LF = "\r\n"
EXCEL_DOUBLE_LF = EXCEL_LF + EXCEL_LF

logger = logging.getLogger(__name__)


OUTPUT_HEADERS: List[str] = [
    "Doc No", "Doc Dt", "Seq No", "Ref Seq No", "Manual Entry Y/N",
    "Main A/C", "Sub A/C", "Div", "Dept", "Anly1", "Anly2", "Acty1", "Acty2",
    "Currency", "FC Amt", "LC Amt", "Dr/Cr", "Detail Narration", "Header Narration",
    "Paym Mode", "Chq Book Id", "Chq No", "Chq Dt", "Payee Name", "Val Date",
    "Doc Ref", "TH Doc ref", "Due Dt",
] + [f"FLEX_{i:02d}" for i in range(1, 51)] + [
    "Party Code", "NOP/NOR", "Tax Code", "Expense Code", "DISC Code",
] + [f"TH_FLEX_{i:02d}" for i in range(1, 51)]


SOURCE1_EXPECTED_COLS: List[str] = [
    "Document Number",
    "Posting Date",
    "Sequence Number",
    "Main Acct",
    "Sub Acct",
    "Division",
    "MIDIS Department",
    "Anly2",
    "Currency",
    "Amount",
    "LC Amt",
    "Dr/Cr",
    "Detail Narration",
    "Doc Ref.",
    "Due Dt.",
]

MASTER1_EXPECTED_COLS: List[str] = [
    "Employee Name",
    "Orion ID",
]

SOURCE2_REQUIRED_FOR_NARRATION: List[str] = [
    "Employee",  # Column C
    "Purpose/Description",  # Column F
    "Benefit Item",  # Column K
    "Benefit Amount",  # Column L
]

# Master File 2 supported headers
MASTER2_HEADERS_OPTIONS = {
    "main_account": ["Main Account", "Main A/C", "Main Acc", "Account"],
    "benefit_type": ["Benefit Type"],
    "benefit_item": ["Benefit Item", "Item"],
    "description_contains": ["Description contains", "Description", "Contains"],
}


COL_INDEX = {name: idx for idx, name in enumerate(SOURCE1_EXPECTED_COLS, start=1)}


# ---------------------------------------------------------------------------
# Utilities
# ---------------------------------------------------------------------------

def last_day_of_previous_month(today: Optional[date] = None) -> date:
    """Return the last calendar day of the previous month relative to `today`."""
    if today is None:
        today = date.today()
    first_of_month = today.replace(day=1)
    last_prev = first_of_month - timedelta(days=1)
    return last_prev


def to_excel_date_string(d: date | None) -> str:
    """Format a date as dd/mm/YYYY for Excel cells."""
    if d is None:
        return ""
    return d.strftime("%d/%m/%Y")


def parse_date_maybe(value: Any) -> str:
    """Best-effort parse of a date-like value into an Excel-friendly string.

    Returns the original string if parsing fails.
    """
    if value is None or value == "":
        return ""
    if isinstance(value, (datetime, date)):
        return to_excel_date_string(value if isinstance(value, date) else value.date())
    try:
        dt = date_parser.parse(str(value), dayfirst=True, fuzzy=True)
        return to_excel_date_string(dt.date())
    except Exception:
        return str(value)


def normalize_header(s: Any) -> str:
    """Normalize a header cell value for case-insensitive comparisons."""
    if s is None:
        return ""
    # Lowercase, trim, and strip common trailing punctuation to be tolerant of variants like
    # "Description contains:" vs "Description contains"
    return str(s).strip().lower().strip(" .:-_/")


def _normalize_category(value: Any) -> str:
    """Normalize labels like 'Travel Expense(s)'/'Business Expense(s)' to a comparable key.

    - lowercases and trims
    - strips a single trailing 's' to be tolerant of singular/plural
    """
    v = normalize_header(value)
    if v.endswith("s"):
        v = v[:-1]
    return v


def _is_any_category(value: Any, labels: List[str]) -> bool:
    v_norm = _normalize_category(value)
    return any(_normalize_category(lbl) == v_norm for lbl in labels)


def find_header_index(headers: List[Any], options: List[str]) -> Optional[int]:
    """Find the index of the first header cell matching any of the `options` (case-insensitive)."""
    norm = [normalize_header(h) for h in headers]
    option_norm = [o.strip().lower() for o in options]
    for i, h in enumerate(norm):
        if h in option_norm:
            return i
    return None


def blank_if_zero(value: Any) -> Any:
    """Return an empty string if the numeric value equals zero; otherwise return the original value.

    Handles strings like "0" or "0.00" and ignores commas.
    """
    if value is None:
        return ""
    s = str(value).strip()
    if s == "":
        return ""
    try:
        numeric = float(s.replace(",", ""))
        if numeric == 0.0:
            return ""
    except Exception:
        # Non-numeric; keep as-is
        return value
    return value


def to_float_or_none(value: Any) -> Optional[float]:
    """Convert a value to float if possible; otherwise return None.

    Handles strings with commas and parentheses for negatives, e.g., "(1,234.56)" → -1234.56.
    Returns None for blanks or unparseable values.
    """
    if value is None:
        return None
    s = str(value).strip()
    if s == "":
        return None
    try:
        negative = s.startswith("(") and s.endswith(")")
        s_clean = s.replace(",", "").replace("$", "").replace("€", "")
        if negative:
            s_clean = s_clean.strip("()")
        f = float(s_clean)
        return -f if negative else f
    except Exception:
        return None


# ---------------------------------------------------------------------------
# Readers
# ---------------------------------------------------------------------------

def read_source1_rows(path_or_stream: Union[str, IO[bytes]]) -> List[Dict[str, Any]]:
    """Read Source File 1, ensuring the expected header row is the first row.

    Returns a list of dicts keyed by `SOURCE1_EXPECTED_COLS`.
    """
    wb = load_workbook(filename=path_or_stream, data_only=True)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    missing = [h for h in SOURCE1_EXPECTED_COLS if h not in headers]
    if missing:
        raise ValueError("Missing required columns in Source File 1: " + ", ".join(missing))

    header_to_index = {h: headers.index(h) for h in headers}

    rows: List[Dict[str, Any]] = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if all((c is None or str(c).strip() == "") for c in r):
            continue
        row_map = {h: r[header_to_index[h]] if h in header_to_index and header_to_index[h] < len(r) else "" for h in SOURCE1_EXPECTED_COLS}
        rows.append(row_map)
    logger.info("Read %d rows from Source File 1", len(rows))
    return rows


def read_master1_map(path_or_stream: Union[str, IO[bytes]]) -> Dict[str, str]:
    """Read Master File 1 and return a map of Orion ID -> Employee Name.

    Accepts multiple variants of the employee name header.
    """
    wb = load_workbook(filename=path_or_stream, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]

    # Accept alternative header labels
    idx_name = find_header_index(headers, [
        "Employee Name from SAP Employee Benefit report",
        "Employee Name",
        "Employee",
    ])
    idx_orion = find_header_index(headers, [
        "Orion ID",
        "OrionID",
        "Orion Id",
    ])

    if idx_name is None or idx_orion is None:
        raise ValueError(
            "Missing required columns in Master File 1. Expected one of Employee Name headers: "
            "['Employee Name from SAP Employee Benefit report','Employee Name','Employee'] and Orion ID headers: "
            "['Orion ID','OrionID','Orion Id']"
        )

    mapping: Dict[str, str] = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        orion = r[idx_orion] if idx_orion < len(r) else None
        name = r[idx_name] if idx_name < len(r) else None
        if orion is None or str(orion).strip() == "":
            continue
        mapping[str(orion).strip()] = str(name).strip() if name is not None else ""
    logger.info("Loaded %d Orion ID -> Employee Name mappings", len(mapping))
    return mapping


def find_orion_id_for_employee(master1_map: Optional[Dict[str, str]], employee_name: str) -> str:
    """Lookup Orion ID by employee name using Master File 1 mapping.

    `master1_map` is Orion ID -> Employee Name. We invert it on the fly using a
    normalized, whitespace-trimmed comparison to tolerate minor variations.
    Returns an empty string when no match is found or inputs are missing.
    """
    if not master1_map or not employee_name:
        return ""
    target = str(employee_name).strip()
    if target == "":
        return ""
    target_norm = normalize_header(target)
    logger.debug("Acty1: looking up Orion ID for employee='%s' (norm='%s')", target, target_norm)
    # First, try exact normalized match
    for orion_id, name in master1_map.items():
        name_norm = normalize_header(name)
        if name_norm == target_norm:
            logger.debug("Acty1: found Orion ID '%s' for employee '%s'", orion_id, target)
            return str(orion_id).strip()
    # Next, try token-set match to handle formats like "Doe, John" vs "John Doe"
    def tokenize(n: str) -> set[str]:
        import re
        tokens = [t for t in re.split(r"[^a-z0-9]+", normalize_header(n)) if len(t) >= 2]
        return set(tokens)
    target_tokens = tokenize(target)
    if target_tokens:
        for orion_id, name in master1_map.items():
            if tokenize(name) == target_tokens:
                logger.debug("Acty1: token-set match Orion ID '%s' for employee '%s'", orion_id, target)
                return str(orion_id).strip()
    logger.debug("Acty1: no Orion ID found for employee '%s'", target)
    return ""


def find_source2_header_row(ws, required_cols: List[str], max_search_rows: int = 25) -> Optional[Dict[str, int]]:
    """Scan the first `max_search_rows` to locate a header row that contains all `required_cols`."""
    required_norm = [normalize_header(c) for c in required_cols]
    for row_idx in range(1, max_search_rows + 1):
        row_vals = [cell.value for cell in ws[row_idx]]
        norm_vals = [normalize_header(v) for v in row_vals]
        if all(req in norm_vals for req in required_norm):
            mapping: Dict[str, int] = {}
            for c in required_cols:
                mapping[c] = norm_vals.index(normalize_header(c))
            mapping["__header_row_index__"] = row_idx  # sentinel for the found row
            return mapping
    return None


def read_source2_rows(path_or_stream: Union[str, IO[bytes]]) -> List[Dict[str, Any]]:
    """Read Source File 2 with a fixed header row at row 3 and data starting at row 4.

    Duplicate headers in row 3 are preserved as `Header`, `Header__2`, ... to keep order.
    """
    wb = load_workbook(filename=path_or_stream, data_only=True)
    ws = wb.active

    header_row_index = 3

    # Build a complete header list from row 3, preserving order and duplicates
    header_cells = [cell.value for cell in ws[header_row_index]]
    normalized_seen: Dict[str, int] = {}
    full_headers: List[str] = []
    for h in header_cells:
        base = str(h) if h is not None else ""
        norm = normalize_header(base)
        if norm in normalized_seen:
            normalized_seen[norm] += 1
            full_headers.append(f"{base}__{normalized_seen[norm]}")
        else:
            normalized_seen[norm] = 1
            full_headers.append(base)

    rows: List[Dict[str, Any]] = []
    for r in ws.iter_rows(min_row=header_row_index + 1, values_only=True):
        if all((c is None or str(c).strip() == "") for c in r):
            continue
        row_obj: Dict[str, Any] = {}
        for idx, key in enumerate(full_headers):
            row_obj[str(key)] = r[idx] if idx < len(r) else ""
        rows.append(row_obj)
    logger.info("Read %d rows from Source File 2 using fixed header row 3", len(rows))
    return rows


def read_master2_entries(path_or_stream: Union[str, IO[bytes]]) -> List[Dict[str, Any]]:
    """Read Master File 2 entries used to map debit Main A/C.

    Supports either exact Benefit Item matching or substring matching via Description contains.
    Returns a list of dicts with keys: main_account, benefit_type, benefit_item, description_contains.
    """
    wb = load_workbook(filename=path_or_stream, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]

    def idx_for(options: List[str]) -> Optional[int]:
        return find_header_index(headers, options)

    idx_main = idx_for(MASTER2_HEADERS_OPTIONS["main_account"])  # required
    idx_bt = idx_for(MASTER2_HEADERS_OPTIONS["benefit_type"])  # required
    idx_bi = idx_for(MASTER2_HEADERS_OPTIONS["benefit_item"])  # optional
    idx_desc = idx_for(MASTER2_HEADERS_OPTIONS["description_contains"])  # optional

    if idx_main is None or idx_bt is None:
        raise ValueError("Missing required columns in Master File 2: Main Account and Benefit Type are required.")

    entries: List[Dict[str, Any]] = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if all((c is None or str(c).strip() == "") for c in r):
            continue
        main_account = r[idx_main] if idx_main < len(r) else ""
        benefit_type = r[idx_bt] if idx_bt < len(r) else ""
        benefit_item = r[idx_bi] if (idx_bi is not None and idx_bi < len(r)) else ""
        desc_contains = r[idx_desc] if (idx_desc is not None and idx_desc < len(r)) else ""
        entries.append({
            "main_account": str(main_account).strip() if main_account is not None else "",
            "benefit_type": str(benefit_type).strip() if benefit_type is not None else "",
            "benefit_item": str(benefit_item).strip() if benefit_item is not None else "",
            "description_contains": str(desc_contains).strip() if desc_contains is not None else "",
        })
    logger.info("Loaded %d entries from Master File 2", len(entries))
    return entries


def _categories_equivalent(a: str, b: str) -> bool:
    """Loose comparison for categories like Business/Travel Expenses/Expense/Other ..."""
    def simplify(x: str) -> str:
        v = _normalize_category(x)
        v = v.replace("other ", "").strip()
        return v
    return simplify(a) == simplify(b)


def _match_main_account_from_master2(entries: List[Dict[str, Any]], benefit_type: str, benefit_item: str) -> str:
    bt_norm = normalize_header(benefit_type)
    item_norm = normalize_header(benefit_item)
    haystack = f"{benefit_type} {benefit_item}".lower()

    # First try exact (benefit_item) match when benefit_item present in entry and category is compatible
    for e in entries:
        e_bt = e.get("benefit_type", "")
        e_item = e.get("benefit_item", "")
        if e_item:
            if (_categories_equivalent(e_bt, bt_norm) or _is_any_category(e_bt, ["Business Expense", "Business Expenses"])) and normalize_header(e_item) == item_norm:
                return e.get("main_account", "")

    # Fallback to token/keyword match using description_contains across Benefit Type + Item text
    import re
    for e in entries:
        e_bt = e.get("benefit_type", "")
        e_desc = e.get("description_contains", "")
        if not e_desc:
            continue
        # Only consider entries for Business Expenses (or broadly compatible with provided category)
        if not (_categories_equivalent(e_bt, bt_norm) or _is_any_category(e_bt, ["Business Expense", "Business Expenses"])):
            continue
        tokens = [t.strip() for t in re.split(r"[\+\-_/;,|]+", e_desc.lower()) if t.strip()]
        for token in tokens:
            if len(token) >= 3 and token in haystack:
                return e.get("main_account", "")

    return ""


# ---------------------------------------------------------------------------
# Business logic
# ---------------------------------------------------------------------------

def _extract_duplicate_header_values(row: Dict[str, Any], header_label: str) -> List[Any]:
    """Return values for all occurrences of a header that might appear multiple times.

    When reading Source File 2, duplicate headers are stored as `Header`, `Header__2`, `Header__3`, ...
    This helper retrieves them in left-to-right order without relying on the exact suffix number.
    """
    target = normalize_header(header_label)
    values: List[Any] = []
    for key in row.keys():  # preserves original column order
        key_norm = normalize_header(key)
        # Normalize away the duplicate suffix by splitting on the sentinel "__"
        base = key_norm.split("__", 1)[0]
        if base == target:
            values.append(row.get(key, ""))
    return values


def build_detail_narration_for_credit(
    src_row: Dict[str, Any],
    user_id_col: Optional[str],
    master1_map: Optional[Dict[str, str]],
    source2_rows: Optional[List[Dict[str, Any]]],
    fallback_detail: str,
) -> str:
    """Build the Detail Narration for a credit line.

    If inputs are incomplete, fall back to the source detail narration.
    """
    if not user_id_col or not master1_map or not source2_rows:
        return fallback_detail

    user_id_val = src_row.get(user_id_col, "")
    user_id = str(user_id_val).strip() if user_id_val is not None else ""
    if user_id == "":
        return fallback_detail

    employee_name = master1_map.get(user_id, "").strip()
    if employee_name == "":
        return fallback_detail

    matches = [r for r in source2_rows if str(r.get("Employee", "")).strip() == employee_name]
    if not matches:
        return fallback_detail

    def snippet(r: Dict[str, Any]) -> str:
        f = str(r.get("Purpose/Description", "")).strip()
        k = str(r.get("Benefit Item", "")).strip()
        l = str(r.get("Benefit Amount", "")).strip()
        parts = [p for p in [f, k, l] if p]
        return "-".join(parts)

    body = " ".join(snippet(r) for r in matches)
    result = f"{body} {employee_name}".strip()

    # Ensure the final result does not exceed 240 characters
    if len(result) > 240:
        result = result[:237] + "..."

    return result


def build_output_rows_from_source1(
    src_rows: List[Dict[str, Any]],
    master1_map: Optional[Dict[str, str]] = None,
    source2_rows: Optional[List[Dict[str, Any]]] = None,
    user_id_col: Optional[str] = None,
) -> List[List[Any]]:
    """Build output rows (list of lists) according to the specified column mapping rules."""
    rows_out: List[List[Any]] = []

    doc_no_fixed = 1
    seq_no_fixed = 1
    doc_dt_val = last_day_of_previous_month()
    doc_dt_str = to_excel_date_string(doc_dt_val)

    for r in src_rows:
        sub_acct = r.get("Sub Acct", "")
        division = r.get("Division", "")
        currency = r.get("Currency", "")
        fc_amt = to_float_or_none(r.get("Amount", ""))
        lc_amt = to_float_or_none(r.get("LC Amt", ""))
        detail_narr = r.get("Detail Narration", "")
        doc_ref_src = r.get("Doc Ref.", "")
        due_dt_src = r.get("Due Dt.", "")
        dept = r.get("MIDIS Department", "")
        anly1 = ""
        anly2 = r.get("Anly2", "")
       

        row = [""] * len(OUTPUT_HEADERS)

        def set_col(name: str, value: Any) -> None:
            idx = OUTPUT_HEADERS.index(name)
            row[idx] = value

        set_col("Doc No", doc_no_fixed)
        set_col("Doc Dt", doc_dt_str)
        set_col("Seq No", seq_no_fixed)
        # Use string type for Main A/C to keep consistent with debit rows and Streamlit preview
        set_col("Main A/C", "12501")
        set_col("Sub A/C", sub_acct)
        set_col("Div", division)
        #logger.debug("Division value = '%s'", division)
        set_col("Dept", "GEN")
        set_col("Anly1", anly1)
        set_col("Anly2", anly2)
        
        set_col("Currency", currency)
        set_col("FC Amt", fc_amt)
        set_col("LC Amt", None)
        set_col("Dr/Cr", "C")

        narr = build_detail_narration_for_credit(
            r, user_id_col=user_id_col, master1_map=master1_map, source2_rows=source2_rows, fallback_detail=str(detail_narr)
        )
        set_col("Detail Narration", narr)
        set_col("Header Narration", narr)

        set_col("Val Date", doc_dt_str)
        set_col("Doc Ref", doc_ref_src)
        set_col("TH Doc ref", doc_ref_src)
        # Per requirement: Due Dt should be the same as Doc Dt for all rows
        set_col("Due Dt", doc_dt_str)

        rows_out.append(row)

    logger.info("Built %d output rows", len(rows_out))
    return rows_out


def build_debit_rows_from_source2(
    source2_rows: Optional[List[Dict[str, Any]]],
    master2_entries: Optional[List[Dict[str, Any]]] = None,
    master1_map: Optional[Dict[str, str]] = None,
    default_div: Optional[str] = None,
    default_dept: Optional[str] = None,
    default_anly1: Optional[str] = None,
    default_anly2: Optional[str] = None,
    default_currency: Optional[str] = None,
    doc_ref: Optional[str] = None,
    division_map: Optional[Dict[str, str]] = None
) -> List[List[Any]]:
    """Build debit rows where the number of rows equals the number of Source File 2 rows.

    Implements custom mapping for Main A/C (column D) based on Benefit Type and Purpose/Description.
    """
    if not source2_rows:
        return []

    rows_out: List[List[Any]] = []
    doc_no_fixed = 1
    seq_no_fixed = 1
    doc_dt_val = last_day_of_previous_month()
    doc_dt_str = to_excel_date_string(doc_dt_val)

    
    abu_dhabi_variants = [
        "ABU DHABI", "ABUDHABI", "ABUDABI", "AUH", "ABUDHABAI", "ABUDHBAI"
    ]

    for src in source2_rows:
        row = [""] * len(OUTPUT_HEADERS)

        def set_col(name: str, value: Any) -> None:
            idx = OUTPUT_HEADERS.index(name)
            row[idx] = value

        set_col("Doc No", doc_no_fixed)
        set_col("Doc Dt", doc_dt_str)
        set_col("Seq No", seq_no_fixed)

        bt_values = _extract_duplicate_header_values(src, "Benefit Type")
        bt_primary = str(bt_values[0]).strip() if bt_values else ""
        purpose_desc = str(src.get("Purpose/Description", "") or "").upper()
        is_abu_dhabi = any(v in purpose_desc for v in abu_dhabi_variants)

        debit_main_ac = ""
        if _is_any_category(bt_primary, ["Travel Expense", "Travel Expenses"]):
            if not is_abu_dhabi:
                debit_main_ac = "54901"
            else:
                bt_secondary = str(bt_values[1]).strip() if len(bt_values) > 1 else bt_primary
                item_values = _extract_duplicate_header_values(src, "Benefit Item")
                benefit_item = str(item_values[0]).strip() if item_values else ""
                if master2_entries:
                    debit_main_ac = _match_main_account_from_master2(master2_entries, bt_secondary, benefit_item)
        elif _is_any_category(bt_primary, ["Business Expenses", "Business Expense"]):
            bt_secondary = str(bt_values[1]).strip() if len(bt_values) > 1 else bt_primary
            item_values = _extract_duplicate_header_values(src, "Benefit Item")
            benefit_item = str(item_values[0]).strip() if item_values else ""
            if master2_entries:
                debit_main_ac = _match_main_account_from_master2(master2_entries, bt_secondary, benefit_item)
        # If no match found, Main A/C remains blank

        if debit_main_ac:
            set_col("Main A/C", debit_main_ac)

        
        employee_name = str(src.get("Employee", "")).strip()
        orion_id = find_orion_id_for_employee(master1_map, employee_name)
        acty1_value = ""

        
        main_ac_val = row[OUTPUT_HEADERS.index("Main A/C")]

     
        bt_values = _extract_duplicate_header_values(src, "Benefit Type")
        benefit_type = str(bt_values[0]).strip() if bt_values else ""
        item_values = _extract_duplicate_header_values(src, "Benefit Item")
        benefit_item = str(item_values[0]).strip() if item_values else ""

        keywords = ["telephone", "phone", "communication"]

        if main_ac_val in ["54901", "54902"]:
            acty1_value = orion_id
        else:
            haystack = f"{benefit_type} {benefit_item}".lower()
            if any(kw in haystack for kw in keywords):
                acty1_value = orion_id

        if acty1_value:
            set_col("Acty1", acty1_value)
            
        division = division_map.get(orion_id, "").strip().upper() if division_map else ""
        acty2_value = "OMOBIL" if division in ["POMN", "PKWT"] else "NET ETSL"
        set_col("Acty2", acty2_value)

        if default_div is not None and str(default_div).strip() != "":
            set_col("Div", str(default_div).strip())
        set_col("Dept", "GEN")
        if default_anly1 is not None and str(default_anly1).strip() != "":
            set_col("Anly1", str(default_anly1).strip())
        if default_anly2 is not None and str(default_anly2).strip() != "":
            set_col("Anly2", str(default_anly2).strip())
        if default_currency is not None and str(default_currency).strip() != "":
            set_col("Currency", str(default_currency).strip())

        benefit_amount_num = to_float_or_none(src.get("Benefit Amount", ""))
        set_col("FC Amt", benefit_amount_num)
        set_col("LC Amt", None)
        set_col("Dr/Cr", "D")
        set_col("Val Date", doc_dt_str)
        set_col("Due Dt", doc_dt_str)
        if doc_ref and str(doc_ref).strip() != "":
            set_col("Doc Ref", str(doc_ref).strip())
            set_col("TH Doc ref", str(doc_ref).strip()) 

        purpose = str(src.get("Purpose/Description", "") or "").strip()
        benefit_item_text = str(src.get("Benefit Item", "") or "").strip()
        benefit_amount_text = str(src.get("Benefit Amount", "") or "").strip()
        employee_text = str(src.get("Employee", "") or "").strip()
        parts = [p for p in [purpose, benefit_item_text, benefit_amount_text, employee_text] if p]
        debit_detail = " - ".join(parts)
        if debit_detail:
            set_col("Detail Narration", debit_detail)
            set_col("Header Narration", debit_detail)

        rows_out.append(row)

    #logger.info("Built %d debit rows from Source File 2", len(rows_out))
    return rows_out


def diagnose_debit_mapping(
    source2_rows: Optional[List[Dict[str, Any]]],
    master2_entries: Optional[List[Dict[str, Any]]],
) -> List[Dict[str, Any]]:
    """Return a row-by-row explanation of how debit Main A/C is determined.

    Each dict contains: primary_benefit_type, secondary_benefit_type, benefit_item,
    decision, matched_main_account.
    """
    diagnostics: List[Dict[str, Any]] = []
    if not source2_rows:
        return diagnostics
    for src in source2_rows:
        bt_values = _extract_duplicate_header_values(src, "Benefit Type")
        bt_primary = str(bt_values[0]).strip() if bt_values else ""
        bt_secondary = str(bt_values[1]).strip() if len(bt_values) > 1 else bt_primary
        item_values = _extract_duplicate_header_values(src, "Benefit Item")
        benefit_item = str(item_values[0]).strip() if item_values else ""

        matched = ""
        decision = ""
        if _is_any_category(bt_primary, ["Travel Expense", "Travel Expenses"]):
            matched = "54901"
            decision = "travel_expense"
        elif _is_any_category(bt_primary, ["Business Expenses", "Business Expense"]) and master2_entries:
            matched = _match_main_account_from_master2(master2_entries, bt_secondary, benefit_item)
            decision = "business_match" if matched else "business_no_match"
        else:
            decision = "no_rule"

        diagnostics.append({
            "primary_benefit_type": bt_primary,
            "secondary_benefit_type": bt_secondary,
            "benefit_item": benefit_item,
            "matched_main_account": matched,
            "decision": decision,
        })
    return diagnostics


def derive_defaults_from_source1(src_rows: List[Dict[str, Any]]) -> Dict[str, str]:
    """Derive default Div/Dept/Anly1/Anly2 values from Source File 1.

    Strategy: take the first non-empty occurrence for each field across the file.
    """
    defaults = {
        "Div": "",
        "Dept": "",
        "Anly1": "",
        "Anly2": "",
    }
    for r in src_rows:
        if not defaults["Div"]:
            defaults["Div"] = str(r.get("Division", "") or "").strip()
        if not defaults["Dept"]:
            defaults["Dept"] = str(r.get("MIDIS Department", "") or "").strip()
        if not defaults["Anly2"]:
            defaults["Anly2"] = str(r.get("Anly2", "") or "").strip()
        # Anly1 has no direct source; keep empty unless specified later
        if all(defaults.values()):
            break
    return defaults


# ---------------------------------------------------------------------------
# Writers
# ---------------------------------------------------------------------------

def _apply_multiline_alignment(ws, column_names: List[str]) -> None:
    """Enable wrap text and top vertical alignment for given column names."""
    for name in column_names:
        if name not in OUTPUT_HEADERS:
            continue
        col_idx = OUTPUT_HEADERS.index(name) + 1  # 1-based index
        for r in range(2, ws.max_row + 1):
            ws.cell(row=r, column=col_idx).alignment = Alignment(wrap_text=True, vertical="top")


def write_output_excel(path_or_stream: Union[str, IO[bytes]], rows: List[List[Any]]) -> None:
    """Write the output workbook and apply formatting to narration columns."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Output"
    ws.append(OUTPUT_HEADERS)
    for row in rows:
        ws.append(row)

    # Improve readability
    _apply_multiline_alignment(ws, ["Detail Narration", "Header Narration"]) 

    # Highlight rows where Main A/C = 54909 in purple
    purple_fill = PatternFill(start_color="800080", end_color="800080", fill_type="solid")
    main_ac_col_idx = OUTPUT_HEADERS.index("Main A/C") + 1  # 1-based index
    
    for row_idx in range(2, ws.max_row + 1):  # Start from row 2 (after header)
        cell_value = ws.cell(row=row_idx, column=main_ac_col_idx).value
        if str(cell_value).strip() == "54909":
            # Highlight the entire row in purple
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col_idx).fill = purple_fill

    wb.save(path_or_stream)


def write_source1_template(path: str) -> None:
    """Write a blank Source File 1 template with expected headers."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Source1 Template"
    ws.append(SOURCE1_EXPECTED_COLS)
    wb.save(path)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def run(
    source1_path: str,
    output_path: str,
    master1_path: Optional[str] = None,
    source2_path: Optional[str] = None,
    user_id_col: Optional[str] = None,
    master2_path: Optional[str] = None,
    division_map: Optional[Dict[str, str]] = None,  # ✅ Add this
) -> None:
    """CLI runner for building the Excel output from files on disk."""
    logger.info("Starting build: source1=%s, master1=%s, source2=%s", source1_path, master1_path, source2_path)
    src_rows = read_source1_rows(source1_path)
    master1_map = read_master1_map(master1_path) if master1_path else None
    source2_rows = read_source2_rows(source2_path) if source2_path else None
    master2_entries = read_master2_entries(master2_path) if master2_path else None

    credit_rows = build_output_rows_from_source1(
        src_rows, master1_map=master1_map, source2_rows=source2_rows, user_id_col=user_id_col
    )
    debit_rows = build_debit_rows_from_source2(
    source2_rows,
    master2_entries=master2_entries,
    master1_map=master1_map,
    division_map=division_map  # ✅ Pass it here
)

    out_rows = credit_rows + debit_rows
    write_output_excel(output_path, out_rows)
    logger.info("Output written to %s with %d rows", output_path, len(out_rows))
    print(f"Output written to {output_path} with {len(out_rows)} rows.")


def main() -> None:
    """CLI entrypoint."""
    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(name)s: %(message)s")

    parser = argparse.ArgumentParser(description="Claims Automation (Credits A–AB)")
    parser.add_argument("--source1", default="source1.xlsx", help="Path to Source File 1 Excel")
    parser.add_argument("--output", default="claims_output.xlsx", help="Path to output Excel")
    parser.add_argument("--make-template", action="store_true", help="Create a Source1 Excel template and exit")
    parser.add_argument("--template-out", default="source1_template.xlsx", help="Path to write the Source1 template")
    parser.add_argument("--master1", help="Path to Master File 1 (Employee Name, Orion ID)")
    parser.add_argument("--source2", help="Path to Source File 2 (with Employee, Purpose/Description, Benefit Item, Benefit Amount, Benefit Type)")
    parser.add_argument("--master2", help="Path to Master File 2 (mapping Benefit Type/Item or Description contains to Main Account)")
    parser.add_argument("--user-id-col", default="Sub Acct", help="Column name in Source File 1 that contains the user/Orion ID")
    parser.add_argument("--debug", action="store_true", help="Enable verbose debug logging")
    args = parser.parse_args()

    if args.debug:
        logging.getLogger().setLevel(logging.DEBUG)
        logger.debug("Debug logging enabled")

    if args.make_template:
        write_source1_template(args.template_out)
        logger.info("Template written to %s", args.template_out)
        print(f"Template written to {args.template_out}.")
        return

    run(
    args.source1,
    args.output,
    master1_path=args.master1,
    source2_path=args.source2,
    user_id_col=args.user_id_col,
    master2_path=args.master2
)

if __name__ == "__main__":
    main()


