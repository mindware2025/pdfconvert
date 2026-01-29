"""
MIBB Quotation processing module.
- Uses same header extraction as IBM quotations
- Custom table structure: Part Number, Description, Start Date, End Date, QTY, Price USD
- MIBB-specific terms and conditions
"""

from datetime import datetime
from io import BytesIO
import os
import re
import fitz  # PyMuPDF
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
import logging
from pathlib import Path

# Configure MIBB-specific logging
MIBB_LOG_DIR = Path("mibb_logs")
MIBB_LOG_DIR.mkdir(exist_ok=True)

# Create logger for MIBB extraction
mibb_logger = logging.getLogger('mibb_extraction')
mibb_logger.setLevel(logging.DEBUG)

# Remove existing handlers to avoid duplicates
for handler in mibb_logger.handlers[:]:
    mibb_logger.removeHandler(handler)

# Create file handler with timestamp
log_timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
log_file_path = MIBB_LOG_DIR / f'mibb_extraction_{log_timestamp}.log'
file_handler = logging.FileHandler(log_file_path, mode='w', encoding='utf-8')
file_handler.setLevel(logging.DEBUG)

# Create formatter
formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
file_handler.setFormatter(formatter)

# Add handler to logger
mibb_logger.addHandler(file_handler)
mibb_logger.propagate = False

def log_debug(message):
    """Helper function to log debug messages"""
    mibb_logger.debug(message)
    print(f"[MIBB DEBUG] {message}")  # Also print to console for immediate feedback


def parse_euro_number(value: str):
    """Parse EU-formatted numbers like '733,00' -> 733.00"""
    try:
        if value is None:
            return None
        s = str(value).strip().replace(" ", "")
        if "." in s and "," in s:
            if s.rfind(",") > s.rfind("."):
                s = s.replace(".", "").replace(",", ".")
            else:
                s = s.replace(",", "")
        else:
            s = s.replace(",", ".")
        return float(s)
    except Exception:
        return None


def extract_mibb_header_from_pdf(file_like) -> dict:
    """
    Extract header information from MIBB quotation PDF.
    Uses same logic as IBM header extraction.
    Returns: dict with header fields
    """
    log_debug("="*80)
    log_debug("MIBB HEADER EXTRACTION STARTED")
    log_debug("="*80)
    
    try:
        doc = fitz.open(stream=file_like.read(), filetype="pdf")
        log_debug(f"PDF opened successfully: {len(doc)} pages")
    except Exception as e:
        log_debug(f"ERROR opening PDF: {e}")
        return {}
    
    # Collect lines
    lines = []
    for page_num, page in enumerate(doc):
        page_text = page.get_text("text") or page.get_text()
        page_lines = []
        for l in page_text.splitlines():
            if l and l.strip():
                lines.append(l.rstrip())
                page_lines.append(l.rstrip())
        log_debug(f"Page {page_num + 1}: Extracted {len(page_lines)} lines")
    
    log_debug(f"Total lines extracted: {len(lines)}")
    log_debug("\nFirst 50 lines of PDF:")
    for idx, line in enumerate(lines[:50]):
        log_debug(f"  Line {idx:3d}: {line}")
    
    # Header fields (same as IBM)
    header_info = {
        "Customer Name": "",
        "Bid Number": "",
        
        "Select Territory": "",
        "Government Entity (GOE)": "",
        "Reseller Name": "",
        "City": "",
        "Country": "",
        "Maximum End User Price (MEP)": "",
        "Total Value Seller Revenue Opportunity": "",
        "Bid Expiration Date": ""
    }
    
    # Parse header info (same logic as IBM)
    log_debug("\nParsing header information...")
    fields_found = 0
    for i, line in enumerate(lines):
        if "Customer Name:" in line:
            value = lines[i].strip() if i + 1 < len(lines) else ""
            header_info["Customer Name"] = value
            log_debug(f"  [Line {i}] Customer Name: '{value}'")
            fields_found += 1
        if "Reseller Name:" in line:
            value = lines[i].strip() if i + 1 < len(lines) else ""
            header_info["Reseller Name"] = value
            log_debug(f"  [Line {i}] Reseller Name: '{value}'")
            fields_found += 1
        if "Bid Number:" in line or "Quote Number:" in line:
            value = lines[i].strip() if i + 1 < len(lines) else ""
            header_info["Bid Number"] = value
            fields_found += 1
        if "Business Partner of Record:" in line :
            value = lines[i].strip() if i + 1 < len(lines) else ""
            header_info["Business Partner of Record"] = value
            fields_found += 1
       
        if "Select Territory:" in line:
            value = lines[i + 1].strip() if i + 1 < len(lines) else ""
            header_info["Select Territory"] = value
            log_debug(f"  [Line {i}] Select Territory: '{value}'")
            fields_found += 1
        if "Government Entity" in line:
            value = lines[i + 1].strip() if i + 1 < len(lines) else ""
            header_info["Government Entity (GOE)"] = value
            log_debug(f"  [Line {i}] Government Entity (GOE): '{value}'")
            fields_found += 1
        
        if "Bid Expiration Date:" in line or "Quote Expiration Date:" in line:
            value = lines[i].strip() if i + 1 < len(lines) else ""
            header_info["Bid Expiration Date"] = value
            log_debug(f"  [Line {i}] Bid Expiration Date: '{value}'")
            fields_found += 1
        if (
            "Maximum End User Price" in line
            or "Total Value Seller Revenue Opportunity" in line
            or "MEP" in line
        ):
            # Look for MEP value in same line or next line
            if ":" in line:
                mep_part = line.split(":", 1)[1].strip()
                if mep_part:
                    mep_clean = re.sub(r'\s*(USD).*$', '', mep_part).strip()
                    mep_value = parse_euro_number(mep_clean)
                    if mep_value:
                        header_info["Maximum End User Price (MEP)"] = f"{mep_value:,.2f}"
                        log_debug(f"  [Line {i}] MEP found in same line: '{mep_part}' -> {mep_value:,.2f}")
                        fields_found += 1
            elif i + 1 < len(lines):
                next_line = lines[i + 1].strip()
                if "USD" in next_line or "," in next_line:
                    mep_clean = re.sub(r'\s*(USD).*$', '', next_line).strip()
                    mep_value = parse_euro_number(mep_clean)
                    if mep_value:
                        header_info["Maximum End User Price (MEP)"] = f"{mep_value:,.2f}"
                        log_debug(f"  [Line {i}] MEP found in next line: '{next_line}' -> {mep_value:,.2f}")
                        fields_found += 1
    
    log_debug(f"\nHeader extraction complete: {fields_found} fields found")
    log_debug("\nExtracted header information:")
    for key, value in header_info.items():
        log_debug(f"  {key:35s}: {value}")
    log_debug("="*80 + "\n")
    
    return header_info


def extract_mibb_table_from_pdf(file_like) -> list:
    """
    Extract table data from page 2 of MIBB quotation PDF.
    Looks for "Subscription Quotation - Parts Information" table.
    Returns: list of rows, each row is [Part Number, Description, Start Date, End Date, QTY, Price USD]
    """
    log_debug("="*80)
    log_debug("MIBB TABLE EXTRACTION STARTED")
    log_debug("="*80)
    
    try:
        doc = fitz.open(stream=file_like.read(), filetype="pdf")
        log_debug(f"PDF opened for table extraction: {len(doc)} pages")
    except Exception as e:
        log_debug(f"ERROR opening PDF for table extraction: {e}")
        return []

    if len(doc) == 0:
        log_debug("ERROR: PDF has 0 pages")
        return []

    # ------------------------------------------------------------------
    # Page selection: don't assume the table is on page 2.
    # We scan all pages for strong signals, then fall back gracefully.
    # ------------------------------------------------------------------
    marker_patterns = [
        "parts information",
        "subscription quotation",
        "quotation - parts information",
    ]
    header_signals = [
        "part number",
        "coverage start",
        "coverage end",
        "quantity",
        "qty",
        "bid ext",
        "bid extended",
    ]

    candidate_pages: list[tuple[int, int, int]] = []
    # (page_index, marker_score, header_score)

    for page_idx in range(len(doc)):
        try:
            page_text = (doc[page_idx].get_text("text") or doc[page_idx].get_text() or "")
        except Exception as e:
            log_debug(f"[PAGE SCAN] Could not read text for page {page_idx+1}: {e}")
            continue

        text_lower = page_text.lower()
        marker_score = sum(1 for p in marker_patterns if p in text_lower)
        header_score = sum(1 for s in header_signals if s in text_lower)

        # Only consider pages with at least some evidence.
        if marker_score > 0 or header_score >= 2:
            candidate_pages.append((page_idx, marker_score, header_score))
            log_debug(
                f"[PAGE SCAN] Candidate page {page_idx+1}: marker_score={marker_score}, header_score={header_score}"
            )

    if candidate_pages:
        # Prefer higher marker_score, then header_score, then earlier page
        candidate_pages.sort(key=lambda t: (-t[1], -t[2], t[0]))
        page_index, marker_score, header_score = candidate_pages[0]
        log_debug(
            f"[PAGE SELECT] Using page {page_index+1} (marker_score={marker_score}, header_score={header_score})"
        )
        page = doc[page_index]
    else:
        # Fallback to page 2 if present, else page 1
        page_index = 1 if len(doc) >= 2 else 0
        log_debug(
            f"[PAGE SELECT] No candidates found; falling back to page {page_index+1}"
        )
        page = doc[page_index]
    
    # Try to extract tables using PyMuPDF's table detection
    log_debug(f"\n[STRATEGY 1] Attempting PyMuPDF table detection on page {page_index+1}...")
    try:
        tables = page.find_tables()
        log_debug(f"Found {len(tables)} table(s) using PyMuPDF")
        
        if tables:
            # Use the first table found (should be Parts Information table)
            table = tables[0]
            rows = table.extract()
            log_debug(f"Table extracted: {len(rows)} rows found")
            
            if len(rows) < 2:  # Need at least header + 1 data row
                log_debug("WARNING: Table has less than 2 rows, falling back to text extraction")
                raise Exception("Insufficient rows")
            
            log_debug(f"\nHeader row: {rows[0]}")
            
            # Find column indices
            header_row = rows[0]
            part_num_col = None
            desc_col = None
            start_date_col = None
            end_date_col = None
            qty_col = None
            bid_ext_svp_col = None
            
            log_debug("\nIdentifying column indices:")
            for idx, header in enumerate(header_row):
                header_upper = str(header).upper() if header else ""
                log_debug(f"  Column {idx}: '{header}'")
                if "PART NUMBER" in header_upper or "PART#" in header_upper:
                    part_num_col = idx
                    log_debug(f"    -> Part Number column: {idx}")
                elif "DESCRIPTION" in header_upper:
                    desc_col = idx
                    log_debug(f"    -> Description column: {idx}")
                elif "COVERAGE START" in header_upper or ("START" in header_upper and "DATE" in header_upper):
                    start_date_col = idx
                    log_debug(f"    -> Start Date column: {idx}")
                elif "COVERAGE END" in header_upper or ("END" in header_upper and "DATE" in header_upper):
                    end_date_col = idx
                    log_debug(f"    -> End Date column: {idx}")
                elif "QUANTITY" in header_upper or "QTY" in header_upper:
                    qty_col = idx
                    log_debug(f"    -> Quantity column: {idx}")
                elif "BID EXT SVP" in header_upper or "BID EXTENDED" in header_upper:
                    bid_ext_svp_col = idx
                    log_debug(f"    -> Bid Ext SVP column: {idx}")
            
            extracted_data = []
            
            # Process data rows (skip header)
            log_debug(f"\nProcessing {len(rows)-1} data rows:")
            for row_idx, row in enumerate(rows[1:], start=1):
                if not row or len(row) == 0:
                    log_debug(f"  Row {row_idx}: Empty row, skipping")
                    continue
                
                part_number = str(row[part_num_col]).strip() if part_num_col is not None and part_num_col < len(row) else ""
                description = str(row[desc_col]).strip() if desc_col is not None and desc_col < len(row) else ""
                start_date = str(row[start_date_col]).strip() if start_date_col is not None and start_date_col < len(row) else ""
                end_date = str(row[end_date_col]).strip() if end_date_col is not None and end_date_col < len(row) else ""
                qty_str = str(row[qty_col]).strip() if qty_col is not None and qty_col < len(row) else "1"
                price_str = str(row[bid_ext_svp_col]).strip() if bid_ext_svp_col is not None and bid_ext_svp_col < len(row) else "0"
                
                log_debug(f"\n  Row {row_idx}:")
                log_debug(f"    Part Number: '{part_number}'")
                log_debug(f"    Description: '{description[:50]}...'")
                log_debug(f"    Start Date: '{start_date}'")
                log_debug(f"    End Date: '{end_date}'")
                log_debug(f"    Qty (raw): '{qty_str}'")
                log_debug(f"    Price (raw): '{price_str}'")
                
                # Validate part number (should start with D and have alphanumeric)
                if not part_number or not re.match(r'^D[A-Z0-9]{5,7}', part_number):
                    log_debug(f"    SKIPPED: Invalid part number format")
                    continue
                
                # Parse quantity
                try:
                    qty = int(float(qty_str.replace(',', ''))) if qty_str else 1
                except:
                    qty = 1
                    log_debug(f"    WARNING: Could not parse quantity, using default: 1")
                
                # Parse price (Bid Ext SVP)
                try:
                    price_usd = parse_euro_number(price_str)
                    if price_usd is None:
                        price_usd = 0.0
                        log_debug(f"    WARNING: Could not parse price, using 0.0")
                except Exception as e:
                    price_usd = 0.0
                    log_debug(f"    ERROR parsing price: {e}")
                
                # Clean dates
                start_date = start_date.replace(' ', '')
                end_date = end_date.replace(' ', '')
                
                log_debug(f"    FINAL: Part={part_number}, Qty={qty}, Price={price_usd}")
                
                extracted_data.append([
                    part_number,
                    description,
                    start_date,
                    end_date,
                    qty,
                    price_usd
                ])
            
            log_debug(f"\n[STRATEGY 1 SUCCESS] Extracted {len(extracted_data)} rows using PyMuPDF table detection")
            log_debug("="*80 + "\n")
            return extracted_data
    except Exception as e:
        log_debug(f"[STRATEGY 1 FAILED] PyMuPDF table detection error: {e}")
        log_debug("Falling back to text-based extraction...")
    
    # Fallback: Text-based extraction
    log_debug(f"\n[STRATEGY 2] Text-based extraction (fallback) on page {page_index+1}...")
    page_text = page.get_text("text") or page.get_text()
    lines = []
    for l in page_text.splitlines():
        if l and l.strip():
            lines.append(l.rstrip())
    
    log_debug(f"Extracted {len(lines)} lines from page 2 text")
    log_debug("\nFirst 100 lines of page 2:")
    for idx, line in enumerate(lines[:100]):
        log_debug(f"  Line {idx:3d}: {line}")
    
    extracted_data = []
    
    # Look for "Parts Information" or "Subscription Quotation" table
    table_start_idx = None
    log_debug("\nSearching for table start marker...")
    for i, line in enumerate(lines):
        if "Parts Information" in line or ("Subscription" in line and "Quotation" in line):
            table_start_idx = i
            log_debug(f"  Found table start at line {i}: '{line}'")
            break
    
    if table_start_idx is None:
        log_debug("  ERROR: Could not find 'Parts Information' or 'Subscription Quotation' marker")
        log_debug("="*80 + "\n")
        return []
    
    # Look for table headers
    #
    # NOTE: In these MIBB PDFs, the header is spread across multiple lines:
    #   "Part Number"
    #   "Description"
    #   ...
    #   "Coverage Start"
    #   "Coverage End"
    #   "Quantity"
    # so we only require a line with "Part Number" and then treat that as
    # the anchor; data rows start after that group of header lines.
    header_line_idx = None
    log_debug("\nSearching for header line...")
    for i in range(table_start_idx, min(table_start_idx + 15, len(lines))):
        line = lines[i].upper()
        if "PART NUMBER" in line:
            header_line_idx = i
            log_debug(f"  Found header anchor at line {i}: '{lines[i]}'")
            break
    
    if header_line_idx is None:
        log_debug("  ERROR: Could not find header line containing 'Part Number'")
        log_debug("="*80 + "\n")
        return []
    
    # Extract data rows after header
    
    part_number_pattern = re.compile(r'\b[A-Z][A-Z0-9]{5,7}\b')


    date_pattern = re.compile(r'\b\d{2}/\d{2}/\d{4}\b')  # Format: DD/MM/YYYY
    
    log_debug(f"\nExtracting data rows starting from line {header_line_idx + 1}:")
    i = header_line_idx + 1
    row_count = 0
    while i < len(lines):
        line = lines[i]
        
        # Look for part number
        part_match = part_number_pattern.search(line)
        if not part_match:
            i += 1
            continue
        
        row_count += 1
        part_number = part_match.group()
        log_debug(f"\n  Row {row_count} (line {i}):")
        log_debug(f"    Raw line: '{line}'")
        log_debug(f"    Part Number found: '{part_number}'")
        
        # Extract data from vertical table structure
        # Structure: Part Number (i), Description (i+1), Transaction Type (i+2), Billing (i+3),
        #            Coverage Start (i+4), Coverage End (i+5), Quantity (i+6), ... Bid Ext SVP (i+15 or so)
        
        # Extract description (next line after part number)
        description = ""
        if i + 1 < len(lines):
            description = lines[i + 1].strip()
            # Clean up description - remove if it's just "Planning Analytics" or similar
            if len(description) > 100:
                description = description[:100]
        
        log_debug(f"    Description: '{description[:50]}...'")
        
        # Extract dates - Coverage Start is ~4 lines after part number, Coverage End is ~5 lines
        start_date = ""
        end_date = ""
        dates_found = []
        
        # Search up to 10 lines ahead for dates
        for j in range(i + 1, min(i + 11, len(lines))):
            line_text = lines[j]
            found_dates = date_pattern.findall(line_text)
            if found_dates:
                dates_found.extend(found_dates)
                log_debug(f"    Found dates at line {j}: {found_dates}")
        
        # Remove duplicates while preserving order
        seen = set()
        unique_dates = []
        for d in dates_found:
            if d not in seen:
                seen.add(d)
                unique_dates.append(d)
        
        if len(unique_dates) >= 1:
            start_date = unique_dates[0]
        if len(unique_dates) >= 2:
            end_date = unique_dates[1]
        
        log_debug(f"    Dates found: Start='{start_date}', End='{end_date}'")
        qty = 1
        qty_extracted = False
        
        # QTY is ALWAYS the first number after Coverage End
        # Coverage End is at: i+5 (based on the IBM vertical layout)
        start_search = i + 6
        
        for j in range(start_search, min(i + 15, len(lines))):
            line_j = lines[j].strip()
        
            # Stop if we hit non-numeric placeholder
            if line_j == "-":
                break
        
            # Match integer or decimal
            qty_match = re.match(r'^(\d+(?:\.\d+)?)$', line_j)
            if qty_match:
                qty_val = float(qty_match.group(1))
        
                # QTY in IBM MIBB PDFs is always integer (85, 30, 838)
                if qty_val.is_integer():
                    qty = int(qty_val)
                    log_debug(f"    Extracted QTY at line {j}: {qty}")
                    qty_extracted = True
                    break
        
        if not qty_extracted:
            log_debug("    QTY not found, defaulting to 1")
        
        # Extract Bid Ext SVP (Price USD) - this appears after Bid Unit SVP
        # Structure: Entitled Unit -> Entitled Extended -> Discount% -> Bid Unit SVP -> Bid Ext SVP
        price_usd = 0.0
        price_pattern = re.compile(r'\b(\d{1,3}(?:[.,]\d{3})*[.,]\d{2})\b')
        
        seen_discount = False
        price_after_discount_count = 0
        
        for j in range(i + 1, min(i + 30, len(lines))):
            line_j = lines[j].strip()
        
            # Stop if next part number → next item block
            if part_number_pattern.search(line_j) and j != i:
                break
        
            # Step 1: detect discount line (xx.xx)
            if not seen_discount:
                m = re.match(r'^(\d+\.\d{2})$', line_j)
                if m:
                    seen_discount = True
                    log_debug(f"    Found Discount% at line {j}: {m.group(1)}")
                    continue
        
            # Step 2: after discount, collect prices
            if seen_discount:
                matches = price_pattern.findall(line_j)
                if not matches:
                    continue
        
                raw = matches[0]
        
                # Clean "15,857.60" or "125.750,28"
                if "," in raw and "." in raw:
                    cleaned = raw.replace(",", "")
                else:
                    cleaned = raw.replace(",", ".")
        
                try:
                    val = float(cleaned)
                except:
                    continue
        
                price_after_discount_count += 1
        
                if price_after_discount_count == 1:
                    log_debug(f"    Unit SVP found at line {j}: {val}")
                    continue
        
                if price_after_discount_count == 2:
                    price_usd = val
                    log_debug(f"    Bid Ext SVP found at line {j}: {price_usd}")
                    break
        
        log_debug(f"    Final: Part={part_number}, Desc={description[:30]}, Start={start_date}, End={end_date}, Qty={qty}, Price={price_usd}")
        
        # Accept row if we have part number and at least one date
        if part_number and (start_date or end_date):
            extracted_data.append([
                part_number,
                description,
                start_date,
                end_date,
                qty,
                price_usd
            ])
            log_debug(f"    ✓ Row added successfully")
        else:
            log_debug(f"    ✗ Row skipped: missing part_number or dates (part_number={bool(part_number)}, start_date={bool(start_date)}, end_date={bool(end_date)})")
        
        i += 1
        
        if i < len(lines):
            next_line = lines[i].upper()
            if ("PAGE" in next_line and "OF" in next_line) or \
               ("TOTAL" in next_line and "PRICE" in next_line) or \
               ("AUTORENEWAL" in next_line):
                log_debug(f"\n  Stopping at line {i}: End marker found ('{lines[i]}')")
                break
    
    log_debug(f"\n[STRATEGY 2 COMPLETE] Extracted {len(extracted_data)} rows using text-based extraction")
    log_debug("\nFinal extracted data summary:")
    for idx, row in enumerate(extracted_data, 1):
        log_debug(f"  Row {idx}: Part={row[0]}, Qty={row[4]}, Price={row[5]}")
    log_debug("="*80 + "\n")
    
    return extracted_data


def get_mibb_terms_section(header_info,data):
    """
    Generate MIBB-specific terms and conditions section.
    Returns list of (cell_address, text, style_dict) tuples.
    """
    quote_validity = header_info.get("Bid Expiration Date", "XXXX")
    totalprice = sum(float(row[5]) for row in data if len(row) > 5 and row[5])
    terms = [
        ("B29", "Terms and Conditions:", {"bold": True, "size": 11, "color": "1F497D"}),
        ("C30", f"""• 30 Days from POE Date.
• Quote Validity: {quote_validity} as per the quote
• Mindware requires full payment of this invoice (BP Price USD {totalprice:,.2f})if WHT is applicable on offshore payment
• Pricing valid for this transaction only."""),
        ("C31", "1. Compliance Review", {"bold": True}),
        ("C32", """Transaction Agreement Reseller ("Reseller") shall keep and maintain all records necessary to establish its compliance with the Agreement for at least three years after the Agreement end date. IBM and/or VAD or their auditors may periodically review Reseller's compliance with the Agreement, and may do so either remotely, on Reseller's premises during normal business hours, or a combination thereof. In connection with any such review, Reseller's agrees to provide IBM and/or VAD, or their auditor, with relevant records and system tools output on request. IBM and/or VAD may reproduce and retain copies of such records and output.
If, during any such review, it is determined that Reseller has failed to comply with any material term of this Agreement, in addition to IBM's and or VAD's rights under law and the terms of this Agreement, for transactions that are the subject of the breach, Reseller agrees to refund the amount equal to the discount or fees, if any, that IBM gave Reseller through VAD for the applicable Products or Services, or IBM and or VAD may offset any amounts due to Reseller from VAD.
IBM's audit rights with respect to special bids are set forth further in Section 6."""),
        ("C33", "2. Compliance with Laws", {"bold": True}),
        ("C34", """Each party will comply with all laws and regulations applicable to its business and content, including, without limitation, those prohibiting corruption and bribery, such as the U.S. Foreign Corrupt Practices Act and those governing transactions with government and public entities, antitrust and competition, taxes and export insider trading, securities, and financial reporting, consumer transactions, and regarding data privacy. Each party will procure all licenses and pay all fees and other charges required for such compliance."""),
        ("C35", "3. Prohibition of Inappropriate Conduct", {"bold": True}),
        ("C36", """Reseller will not directly or indirectly make or give, offer or promise to make or give, or authorize the making or giving of any payment, gift, or other thing of value or advantage (including, for example, accommodations, air fare, entertainment or meals) to any person or entity for (a) the purpose of (i) wrongfully influencing any act or decision, (ii) inducing any act or omission to act in violation of a lawful duty; (iii) inducing a misuse of influence or (iv) securing any improper advantage, or (b) any purpose that is otherwise unlawful under any applicable anti-corruption or anti-bribery law, including the U.S. Foreign Corrupt Practices Act. VAD may terminate this Agreement immediately if Reseller breaches this Section or if VAD reasonably believes such a breach has occurred or is likely to occur."""),
        ("C37", "4. Code of Conduct", {"bold": True}),
        ("C38", """Reseller agrees to comply with the IBM Code of Conduct, a current version of which is available on the following IBM Internet website: 
https://www.ibm.com/investor/att/pdf/IBM_Business_Conduct_Guidelines .pdf
Reseller agrees to comply with the Midis Group Code of Conduct, a current version of which is available on the Midis Group Website: 
https://www.midisgroup.com/wp-content/uploads/2024/08/Code-of-Conduct-2023-English.pdf """),
        ("C39", """Definitions:
"Company" refers to the MIBB entity identified at the top of the first page of this Legal Quotation.
"Partner" refers to the distributor entity identified in the "Distributor Name" section on the first page of this Legal Quotation.
"End User" refers to the end-user customer entity identified in the "Customer Name" section on the first page of this Legal Quotation, which is purchasing from or through Partner for its own internal use only.
"T&M Services" refers to time-based engagements sold by half or full-day SKUs with corresponding Company SOWs.
"Packaged Services" refers to standardized offerings tied to IBM part codes and IBM service descriptions.
"Bespoke Services" refers to tailored solutions governed by SOWs through unique Company SKUs and supporting SOWs.
"SOW" refers to the applicable statement of work accompanying this Legal Quotation."""),
        ("C40", """Acceptance of this Legal Quotation requires Partner to issue a valid Purchase Order ("PO") as indicated in this Legal Quotation or, where available, to select and complete the e-sign option.
The PO must (i) reference this Legal Quotation number, (ii) include the email address of the End User contact, and (iii) include the Partner email address to which the invoice(s) will be sent (or a physical address if a physical invoice is required by applicable law).
This Legal Quotation includes (i) the applicable contractual discount, if any, or (ii) the special pricing, if any, for this particular transaction, as agreed by Company and Partner, which special pricing will take precedence over the otherwise applicable contractual discount. Prices are exclusive of use, sales, value added, and other applicable taxes, which will be paid or reimbursed by Partner.
Invoices will be sent by email except where otherwise required by applicable law, and shall be paid to Company by Partner within 30 days of the invoice date or as otherwise specified elsewhere in this Legal Quotation or Partner Agreement."""),
        ("C41", """Unless otherwise specified, all software products will be delivered electronically and deemed accepted upon delivery of access to such software products (i.e. making such software products available for download).
The software licenses within this Legal Quotation shall be for End User's internal use only, even if the installation location in the quote detail for a license specifies an entity that is different than the End User, except as may be otherwise set forth in a separate signed written agreement between End User and Company.
The governing terms for this Legal Quotation consist of Company's standard distributor or partner contract terms and conditions (as applicable), unless superseded by a separate signed agreement ("Governing Terms"). The software, services and support hereunder are sold to Partner strictly for the purpose of resale and not for any internal or other use by Partner.
	Unless otherwise agreed in writing, the products and services are purchased solely under the terms and conditions of the IBM License Terms including but not limited to IBM Passport Advantage and IBM Cloud Offerings available at https://www.ibm.com. No other terms apply. In the event of any inconsistencies between the existing agreement and License Terms, the terms of the License Terms prevail."""),
        ("C42", """For all professional services, the scope, deliverables, and timelines shall be defined in the applicable SOW or service description accompanying this Legal Quotation. Changes to the scope of services after acceptance of this Legal Quotation must be agreed in writing by Company and may result in additional charges or revised delivery timelines.
Scheduling of professional services is subject to resource availability. Company will make reasonable efforts to accommodate requested dates but reserves the right to propose alternatives.
Any expected expenses for the delivery of professional services, including but not limited to travel, accommodation and subsistence, are defined as an explicit cost on the above quote, or incorporated in the agreed fee for the Statement of Work."""),
        ("C43", """T&M Services are offered on a half-day or full-day basis under predefined SKUs. Each engagement is supported by a corresponding SOW, which outlines the scope, estimated effort, and associated deliverables. Time-based billing shall apply, and services will be invoiced according to actual time spent.
For Packaged Services, the standard service description shall apply unless otherwise agreed in writing.
Bespoke service deliverables shall be owned by the End User, unless otherwise specified in the SOW. IBM proprietary materials and intellectual property remains the property of IBM."""),
        ("C44", """Commodities included on this quotation are subject to shipping restrictions under applicable laws, including but not limited to United States and/or European Union export laws, and are authorized for delivery only to the destination shown. Diversion contrary to such applicable laws is prohibited.
Subscription licenses and software maintenance are not perpetual and begin with delivery of license keys. SaaS and education subscriptions and managed services begin when the service is provisioned. Support subscription rates and SaaS subscription rates are subject to change upon renewal."""),
        ("C45", """If you purchase a multi-year subscription license, SaaS or education subscription, managed service or software maintenance, or multi-year renewal, your purchase is for the full value of all years of the offering, even if required payments are annual. Partner irrevocably commits to pay such fees to Company for the entirety of the Term. In the event you fail to pay any annual payment, and such default shall continue for a period of thirty (30) days, then any and all remaining amounts for the relevant offering shall become immediately due and payable. All Orders, including renewals, are subject to acceptance by Company in its discretion. All purchases are final, with no right to a refund, except as expressly provided under the applicable license or service terms."""),
        ("C46", """By accepting this Legal Quotation, Partner agrees that no other terms and conditions apply to this transaction, including, without limitation, those on a PO or other document issued by Partner, End User or any other third party.
Each party shall keep all confidential information it receives using the same protections that it applies to its own information of like importance, but in no event less than reasonable care, and may use such information solely for the purposes contemplated by this transaction or as otherwise agreed mutually in writing by both parties."""),
        ("C47", """Under no circumstances shall either party's liability arising out of or in connection with the Products or a party's performance with this Agreement exceed the aggregate amount of the fees paid by Partner and all orders regardless of whether such claim for liability is alleged to arise in contract, tort (including negligence) or otherwise. In no event shall either party be liable for indirect, special, incidental, or punitive damages including, without limitation, damages resulting from loss of use, loss of data, loss of profits, or loss of business arising out of, or in connection with, the products, services and/or solutions or Partner's performance of any of its obligations under this Agreement. Limitation of liability in this clause does not apply to intellectual property, confidentiality, compliance breaches by Partner and any other liability which cannot be excluded or limited under applicable law.
These General Terms and Conditions are governed by and construed according to the laws of England and Wales and each party irrevocably and unconditionally submits to the non-exclusive jurisdiction of the courts of Dubai International Financial Centre. The 1980 U.N. Convention on Contracts for the International Sale of Goods shall not apply."""),
    ]
    
    return terms


def estimate_line_count(text, max_chars_per_line=80):
    """Estimate number of lines needed for wrapped text"""
    lines = text.split('\n')
    total_lines = 0
    for line in lines:
        if not line:
            total_lines += 1
        else:
            wrapped = len(line) // max_chars_per_line + (1 if (len(line) % max_chars_per_line) else 0)
            total_lines += max(1, wrapped)
    return total_lines


def create_mibb_excel(
    data: list,
    header_info: dict,
    logo_path: str,
    output: BytesIO
):
    """
    Create MIBB Quotation Excel file.
    
    Args:
        data: list of rows, each row is [Part Number, Description, Start Date, End Date, QTY, Price USD]
        header_info: dict with header fields (same as IBM)
        logo_path: path to logo image
        output: BytesIO object to write Excel to
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Quotation"
    ws.sheet_view.showGridLines = False

    # --- Header / Branding ---
    ws.merge_cells("B1:C2")
    if logo_path and os.path.exists(logo_path):
        img = Image(logo_path)
        img.width = 1.87 * 96
        img.height = 0.56 * 96
        ws.add_image(img, "B1")
        ws.row_dimensions[1].height = 25
        ws.row_dimensions[2].height = 25

    ws.merge_cells("D3:G3")
    ws["D3"] = "Quotation"
    ws["D3"].font = Font(size=20, color="1F497D")
    ws["D3"].alignment = Alignment(horizontal="center", vertical="center")

    # Column widths
    ws.column_dimensions[get_column_letter(2)].width = 8
    ws.column_dimensions[get_column_letter(3)].width = 15
    ws.column_dimensions[get_column_letter(4)].width = 50
    ws.column_dimensions[get_column_letter(5)].width = 10
    ws.column_dimensions[get_column_letter(6)].width = 14
    ws.column_dimensions[get_column_letter(7)].width = 14
    ws.column_dimensions[get_column_letter(8)].width = 15
    ws.column_dimensions[get_column_letter(9)].width = 15
    ws.column_dimensions[get_column_letter(10)].width = 18
    ws.column_dimensions[get_column_letter(11)].width = 15
    ws.column_dimensions[get_column_letter(12)].width = 18

    # Left side labels and values
    left_labels = ["Date:", "From:", "Email:", "Contact:", "", "Company:", "Attn:", "Email:"]
    left_values = [
        datetime.today().strftime('%d/%m/%Y'),
        "Eliana Youssef",
        "E.youssef@mindware.net",
        "+961 123 456 758",
        "",
        header_info.get('Reseller Name', 'empty'),
        "empty",
        "empty"
    ]
    row_positions = [5, 6, 7, 8, 9, 10, 11, 12]
    for row, label, value in zip(row_positions, left_labels, left_values):
        if label:
            ws[f"C{row}"] = label
            ws[f"C{row}"].font = Font(bold=True, color="1F497D")
        if value:
            ws[f"D{row}"] = value
            ws[f"D{row}"].font = Font(color="1F497D")

    # Right side labels and values
    right_labels = [
        "", " ",  " ","Payment Terms:"
    ]
    right_values = [
        header_info.get('Customer Name', ''),
        header_info.get('Bid Number', ''),
        
        header_info.get('Business Partner of Record', ''),
       
        "As aligned with Mindware"
    ]
    for row, label, value in zip(row_positions, right_labels, right_values):
        ws.merge_cells(f"H{row}:L{row}")
        ws[f"H{row}"] = f"{label} {value}"
        ws[f"H{row}"].font = Font(bold=True, color="1F497D")
        ws[f"H{row}"].alignment = Alignment(horizontal="left", vertical="center")

    # --- Table Headers ---
    headers = [
        "Sl", "Part Number", "Description", "Start Date", "End Date", "QTY","Partner Price USD", "Price USD"
    ]
    
    header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    for col, header in enumerate(headers, start=2):
        ws.merge_cells(start_row=16, start_column=col, end_row=17, end_column=col)
        cell = ws.cell(row=16, column=col, value=header)
        cell.font = Font(bold=True, size=13, color="1F497D")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = header_fill

    row_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    start_row = 18

    # --- Data Rows ---
    for idx, row in enumerate(data, start=1):
        excel_row = start_row + idx - 1
        
        # Extract data: [Part Number, Description, Start Date, End Date, QTY, Price USD]
        part_number = row[0] if len(row) > 0 else ""
        description = row[1] if len(row) > 1 else ""
        start_date = row[2] if len(row) > 2 else ""
        end_date = row[3] if len(row) > 3 else ""
        qty = row[4] if len(row) > 4 else 0
        price_usd = row[5] if len(row) > 5 else 0

        ws.cell(row=excel_row, column=2, value=idx).font = Font(size=11, color="1F497D")
        ws.cell(row=excel_row, column=2).alignment = Alignment(horizontal="center", vertical="center")
        
        ws.cell(row=excel_row, column=3, value=part_number).font = Font(size=11, color="1F497D")
        ws.cell(row=excel_row, column=4, value=description).font = Font(size=11, color="1F497D")
        ws.cell(row=excel_row, column=5, value=start_date).font = Font(size=11, color="1F497D")
        ws.cell(row=excel_row, column=6, value=end_date).font = Font(size=11, color="1F497D")
        ws.cell(row=excel_row, column=7, value=qty).font = Font(size=11, color="1F497D")
        partner_formula = f"=ROUNDUP(I{excel_row}*0.99, 2)"   # H = Price USD column
        ws.cell(row=excel_row, column=8, value=partner_formula)
        ws.cell(row=excel_row, column=8).font = Font(size=11, color="1F497D")
        ws.cell(row=excel_row, column=8).number_format = '"USD"#,##0.00'
        ws.cell(row=excel_row, column=8).alignment = Alignment(horizontal="center", vertical="center")
        
        ws.cell(row=excel_row, column=9, value=price_usd).font = Font(size=11, color="1F497D")
        
        # Format price as USD currency
        ws.cell(row=excel_row, column=9).number_format = '"USD"#,##0.00'
        
        
        for col in range(2, 9):
            ws.cell(row=excel_row, column=col).fill = row_fill
        
        ws.cell(row=excel_row, column=4).alignment = Alignment(wrap_text=True, horizontal="left", vertical="center")
        for col in [2, 3, 5, 6, 7, 8]:
            ws.cell(row=excel_row, column=col).alignment = Alignment(horizontal="center", vertical="center")

    # --- Summary row (if data exists) ---
    if data:
        data_end_row = start_row + len(data) - 1
        summary_row = data_end_row + 2
        
        ws.merge_cells(f"C{summary_row}:G{summary_row}")
        ws[f"C{summary_row}"] = "Total Price USD"
        ws[f"C{summary_row}"].font = Font(bold=True, color="1F497D")
        ws[f"C{summary_row}"].alignment = Alignment(horizontal="right")
        
        # Calculate total
        total_sum = sum(float(row[5]) if len(row) > 5 and row[5] else 0 for row in data)
        total_formula = f"=SUM(H{start_row}:H{data_end_row})"
        ws[f"H{summary_row}"] = total_formula
        ws[f"H{summary_row}"].number_format = '"USD"#,##0.00'
        ws[f"H{summary_row}"].font = Font(bold=True, color="1F497D")
        ws[f"H{summary_row}"].fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    else:
        summary_row = start_row + 1

    # --- Terms and Conditions Section ---
    terms = get_mibb_terms_section(header_info, data)
    terms_start_row = summary_row + 3
    
    # Adjust terms cell addresses based on where table ends
    adjusted_terms = []
    row_offset = terms_start_row - 29  # 29 is the anchor in terms_template.py
    for cell_addr, text, *style in terms:
        try:
            if len(cell_addr) >= 2 and cell_addr[1:].isdigit():
                col_letter = cell_addr[0]
                original_row = int(cell_addr[1:])
                new_row = original_row + row_offset
                new_cell_addr = f"{col_letter}{new_row}"
                adjusted_terms.append((new_cell_addr, text, *style))
            else:
                adjusted_terms.append((cell_addr, text, *style))
        except Exception:
            adjusted_terms.append((cell_addr, text, *style))

    for cell_addr, text, *style in adjusted_terms:
        try:
            if len(cell_addr) >= 2 and cell_addr[1:].isdigit():
                row_num = int(cell_addr[1:])
                col_letter = cell_addr[0]
                merge_rows = style[0].get("merge_rows") if style else None
                end_row = row_num + (merge_rows - 1 if merge_rows else 0)
                is_bold_title = style and "bold" in style[0] and style[0].get("bold") is True
                if is_bold_title:
                    ws.merge_cells(f"{col_letter}{row_num}:E{end_row}")
                    ws.row_dimensions[row_num].height = 24
                else:
                    ws.merge_cells(f"{col_letter}{row_num}:H{end_row}")
                    line_count = estimate_line_count(str(text), max_chars_per_line=80)
                    total_height = max(18, line_count * 16)
                    ws.row_dimensions[row_num].height = total_height
                ws[cell_addr] = text
                ws[cell_addr].alignment = Alignment(wrap_text=True, vertical="top")
                if style and "bold" in style[0]:
                    ws[cell_addr].font = Font(**style[0])
        except Exception:
            pass

    # --- Page Setup ---
    last_row = ws.max_row
    ws.print_area = f"A1:L{last_row}"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.15
    ws.page_margins.right = 0.15
    ws.page_margins.top = 0.25
    ws.page_margins.bottom = 0.25
    ws.page_margins.header = 0.15
    ws.page_margins.footer = 0.15
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.draft = False
    ws.page_setup.blackAndWhite = False
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    wb.calculation.fullCalcOnLoad = True
    wb.save(output)
    output.seek(0)
