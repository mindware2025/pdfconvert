"""
MIBB Quotation processing module.
- Uses same header extraction as IBM quotations
- Custom table structure: Part Number, Description, Start Date, End Date, QTY, Price USD
- MIBB-specific terms and conditions
"""

from datetime import datetime
from io import BytesIO
import os
import re
import fitz  # PyMuPDF
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter


def parse_euro_number(value: str):
    """Parse EU-formatted numbers like '733,00' -> 733.00"""
    try:
        if value is None:
            return None
        s = str(value).strip().replace(" ", "")
        if "." in s and "," in s:
            if s.rfind(",") > s.rfind("."):
                s = s.replace(".", "").replace(",", ".")
            else:
                s = s.replace(",", "")
        else:
            s = s.replace(",", ".")
        return float(s)
    except Exception:
        return None


def extract_mibb_header_from_pdf(file_like) -> dict:
    """
    Extract header information from MIBB quotation PDF.
    Uses same logic as IBM header extraction.
    Returns: dict with header fields
    """
    doc = fitz.open(stream=file_like.read(), filetype="pdf")
    
    # Collect lines
    lines = []
    for page_num, page in enumerate(doc):
        page_text = page.get_text("text") or page.get_text()
        for l in page_text.splitlines():
            if l and l.strip():
                lines.append(l.rstrip())
    
    # Header fields (same as IBM)
    header_info = {
        "Customer Name": "",
        "Bid Number": "",
        "PA Agreement Number": "",
        "PA Site Number": "",
        "Select Territory": "",
        "Government Entity (GOE)": "",
        "Reseller Name": "",
        "City": "",
        "Country": "",
        "Maximum End User Price (MEP)": "",
        "Total Value Seller Revenue Opportunity": "",
        "Bid Expiration Date": ""
    }
    
    # Parse header info (same logic as IBM)
    for i, line in enumerate(lines):
        if "Customer Name:" in line:
            header_info["Customer Name"] = lines[i + 1].strip() if i + 1 < len(lines) else ""
        if "Reseller Name:" in line:
            header_info["Reseller Name"] = lines[i + 1].strip() if i + 1 < len(lines) else ""
        if "Bid Number:" in line or "Quote Number:" in line:
            header_info["Bid Number"] = lines[i + 1].strip() if i + 1 < len(lines) else ""
        if "PA Agreement Number:" in line:
            header_info["PA Agreement Number"] = lines[i + 1].strip() if i + 1 < len(lines) else ""
        if "PA Site Number:" in line:
            header_info["PA Site Number"] = lines[i + 1].strip() if i + 1 < len(lines) else ""
        if "Select Territory:" in line:
            header_info["Select Territory"] = lines[i + 1].strip() if i + 1 < len(lines) else ""
        if "Government Entity" in line:
            header_info["Government Entity (GOE)"] = lines[i + 1].strip() if i + 1 < len(lines) else ""
        if "City:" in line:
            header_info["City"] = lines[i + 1].strip() if i + 1 < len(lines) else ""
        if "Country:" in line:
            header_info["Country"] = lines[i + 1].strip() if i + 1 < len(lines) else ""
        if "Bid Expiration Date:" in line or "Quote Expiration Date:" in line:
            header_info["Bid Expiration Date"] = lines[i + 1].strip() if i + 1 < len(lines) else ""
        if (
            "Maximum End User Price" in line
            or "Total Value Seller Revenue Opportunity" in line
            or "MEP" in line
        ):
            # Look for MEP value in same line or next line
            if ":" in line:
                mep_part = line.split(":", 1)[1].strip()
                if mep_part:
                    mep_clean = re.sub(r'\s*(USD).*$', '', mep_part).strip()
                    mep_value = parse_euro_number(mep_clean)
                    if mep_value:
                        header_info["Maximum End User Price (MEP)"] = f"{mep_value:,.2f}"
            elif i + 1 < len(lines):
                next_line = lines[i + 1].strip()
                if "USD" in next_line or "," in next_line:
                    mep_clean = re.sub(r'\s*(USD).*$', '', next_line).strip()
                    mep_value = parse_euro_number(mep_clean)
                    if mep_value:
                        header_info["Maximum End User Price (MEP)"] = f"{mep_value:,.2f}"
    
    return header_info


def get_mibb_terms_section(header_info):
    """
    Generate MIBB-specific terms and conditions section.
    Returns list of (cell_address, text, style_dict) tuples.
    """
    quote_validity = header_info.get("Bid Expiration Date", "XXXX")
    
    terms = [
        ("B29", "Terms and Conditions:", {"bold": True, "size": 11, "color": "1F497D"}),
        ("C30", f"""• 30 Days from POE Date.
• Quote Validity: {quote_validity} as per the quote
• Pricing valid for this transaction only."""),
        ("C31", "1. Compliance Review", {"bold": True}),
        ("C32", """Transaction Agreement Reseller ("Reseller") shall keep and maintain all records necessary to establish its compliance with the Agreement for at least three years after the Agreement end date. IBM and/or VAD or their auditors may periodically review Reseller's compliance with the Agreement, and may do so either remotely, on Reseller's premises during normal business hours, or a combination thereof. In connection with any such review, Reseller's agrees to provide IBM and/or VAD, or their auditor, with relevant records and system tools output on request. IBM and/or VAD may reproduce and retain copies of such records and output.
If, during any such review, it is determined that Reseller has failed to comply with any material term of this Agreement, in addition to IBM's and or VAD's rights under law and the terms of this Agreement, for transactions that are the subject of the breach, Reseller agrees to refund the amount equal to the discount or fees, if any, that IBM gave Reseller through VAD for the applicable Products or Services, or IBM and or VAD may offset any amounts due to Reseller from VAD.
IBM's audit rights with respect to special bids are set forth further in Section 6."""),
        ("C33", "2. Compliance with Laws", {"bold": True}),
        ("C34", """Each party will comply with all laws and regulations applicable to its business and content, including, without limitation, those prohibiting corruption and bribery, such as the U.S. Foreign Corrupt Practices Act and those governing transactions with government and public entities, antitrust and competition, taxes and export insider trading, securities, and financial reporting, consumer transactions, and regarding data privacy. Each party will procure all licenses and pay all fees and other charges required for such compliance."""),
        ("C35", "3. Prohibition of Inappropriate Conduct", {"bold": True}),
        ("C36", """Reseller will not directly or indirectly make or give, offer or promise to make or give, or authorize the making or giving of any payment, gift, or other thing of value or advantage (including, for example, accommodations, air fare, entertainment or meals) to any person or entity for (a) the purpose of (i) wrongfully influencing any act or decision, (ii) inducing any act or omission to act in violation of a lawful duty; (iii) inducing a misuse of influence or (iv) securing any improper advantage, or (b) any purpose that is otherwise unlawful under any applicable anti-corruption or anti-bribery law, including the U.S. Foreign Corrupt Practices Act. VAD may terminate this Agreement immediately if Reseller breaches this Section or if VAD reasonably believes such a breach has occurred or is likely to occur."""),
        ("C37", "4. Code of Conduct", {"bold": True}),
        ("C38", """Reseller agrees to comply with the IBM Code of Conduct, a current version of which is available on the following IBM Internet website: 
https://www.ibm.com/investor/att/pdf/IBM_Business_Conduct_Guidelines .pdf
Reseller agrees to comply with the Midis Group Code of Conduct, a current version of which is available on the Midis Group Website: 
https://www.midisgroup.com/wp-content/uploads/2024/08/Code-of-Conduct-2023-English.pdf """),
        ("C39", "5. Special Bids", {"bold": True}),
        ("C40", """Reseller may request a Special Bid (a special discount or price) on a specific End User transaction. VAD may, at its sole discretion, approve or reject a Special Bid based on the information provided by Reseller in its Special Bid request. If VAD approves a Special Bid, then the price provided by VAD shall only be valid for the applicable Special Bid, and its validity shall be subject to all the terms and conditions set out in this Agreement, including IBM's Special Bid authorization notice ("Special Bid Addendum"). Further, IBM provides Special Bids through VAD to Reseller on the basis that the information Reseller provided in its Special Bid request is truthful and accurate. If the information provided in the Special Bid request changes, Reseller must immediately notify VAD. In such event, VAD reserves the right to modify the terms of, or cancel any Special Bid authorization it may have provided. If Reseller fails to provide truthful and accurate information on Special Bid requests, then VAD shall be entitled to recover from Reseller (and Reseller is obligated to repay) the amount of any discounts IBM provided through VAD in the Special Bid and take any other actions authorized under this Agreement or applicable law. Special Bid authorizations and the terms applicable to Special Bids are IBM's confidential information, which is subject to the applicable confidentiality agreement.
Reseller accepts the terms of the Special Bid by:
a. submitting an order under the Special Bid authorization;
b. accepting the Products or Services for which Reseller is receiving a Special Bid;
c. providing the Products or Services to its Customer; or
d. paying for the Products or Services."""),
        ("C41", """The Special Bid discount or price for eligible Products or Services is subject to the following:
a. no other discounts, incentive offerings, rebates, or promotions apply, unless VAD specifies otherwise in writing;
b. availability of the Products or Services;
c. Reseller's acceptance of the additional terms contained in the Special Bid Addendum (which occurs upon Reseller's acceptance of the Special Bid, as set forth above)
d. Reseller's advising the local VAD financing entity/organization of any Special Bid pricing for any Products or Services for which Reseller arranges financing; and
e. Resale of the Products or Services by Reseller to the End User associated with the Special Bid by the date indicated in the Special Bid request.
If reseller is a Distributor, Reseller may only market the Products and Services to the Resellers that Reseller has identified in the Special Bid request as bidding to the End User.
Reseller is responsible to require Reseller's Resellers who do not have a contract with IBM to market such Products and Services to comply with the Special Bid terms contained in this Agreement and in the applicable Special Bid Addendum that IBM provides for the Special Bid through VAD.
If Reseller is requesting a specific End User price or discount in the Special Bid, Reseller shall ensure, and shall require any Resellers to also ensure, that the intended End User receives the financial benefit of such price or discount."""),
        ("C42", "6. IBM's Audit of Special Bid Transactions", {"bold": True}),
        ("C43", """IBM may audit directly or through VAD any Special Bid transactions in accordance with the terms of this Section 
a. Upon VAD's request, Reseller will promptly provide VAD or its auditors with all relevant Documentation to enable VAD and/or IBM to verify that all information provided in support of a Special Bid request was truthful and accurate and that IBM Products and Services have been or will be supplied to the End User in accordance with the terms of the Special Bid, including, but not limited to, i) documentation that identifies the dates of sale and delivery and End User prices for IBM Products and Services, such as invoices, delivery orders, contracts and purchase orders by and between Reseller and any Reseller and by and between Reseller or any Reseller and an End User and ii) documentation that demonstrates that Reseller or Reseller's Reseller, as applicable, own and use the Special Bid Products for at least the Service Period to provide the service offering described in the terms of the Special Bid to End Users (collectively, items i) and ii) being the "Documentation").
	b. In any case where reseller is unable to provide the Documentation because of confidentiality obligations owed to an End User, whether arising by written contract or applicable law, Reseller will promptly provide VAD with written evidence of, and any Documentation not subject to, those obligations. In addition, Reseller will promptly and in writing seek the End User's consent to waive confidentiality restrictions to permit VAD and IBM to conduct their audit as intended. Should the End User refuse to grant that consent, Reseller will i) provide VAD with a copy of the waiver request and written proof of that refusal and ii) identify appropriate contacts at the End User with whom VAD may elect to discuss the refusal.
c. Reseller hereby waives any objection to i) VAD and/or IBM sharing Special Bid information directly with the End User, notwithstanding the terms of any agreement that would prohibit VAD from doing so, and otherwise communicating (both orally and in writing) with the End User, as VAD deems necessary and appropriate to complete its desired compliance review and ii) the End User sharing Special Bid information directly with IBM/VAD. In this subsection (c), "Special Bid information" includes, but is not limited to, the types and quantity of Products and anticipated End User prices and delivery dates set forth in a Special Bid. VAD may invalidate a Special Bid if in respect of such Special Bid, Reseller fails to comply with this Section 4.9.1 or the applicable Special Bid terms. In that event, IBM/VAD shall be entitled to recover from Reseller (and Reseller is obligated to repay) the amount of any discounts IBM provided in the Special Bid. IBM may also take any other actions authorized under the Agreement or applicable law."""),
        ("C44", """Definitions:
"Company" refers to the MIBB entity identified at the top of the first page of this Legal Quotation.
"Partner" refers to the distributor entity identified in the "Distributor Name" section on the first page of this Legal Quotation.
"End User" refers to the end-user customer entity identified in the "Customer Name" section on the first page of this Legal Quotation, which is purchasing from or through Partner for its own internal use only.
"T&M Services" refers to time-based engagements sold by half or full-day SKUs with corresponding Company SOWs.
"Packaged Services" refers to standardized offerings tied to IBM part codes and IBM service descriptions.
"Bespoke Services" refers to tailored solutions governed by SOWs through unique Company SKUs and supporting SOWs.
"SOW" refers to the applicable statement of work accompanying this Legal Quotation."""),
        ("C45", """Acceptance of this Legal Quotation requires Partner to issue a valid Purchase Order ("PO") as indicated in this Legal Quotation or, where available, to select and complete the e-sign option.
The PO must (i) reference this Legal Quotation number, (ii) include the email address of the End User contact, and (iii) include the Partner email address to which the invoice(s) will be sent (or a physical address if a physical invoice is required by applicable law).
This Legal Quotation includes (i) the applicable contractual discount, if any, or (ii) the special pricing, if any, for this particular transaction, as agreed by Company and Partner, which special pricing will take precedence over the otherwise applicable contractual discount. Prices are exclusive of use, sales, value added, and other applicable taxes, which will be paid or reimbursed by Partner.
Invoices will be sent by email except where otherwise required by applicable law, and shall be paid to Company by Partner within 30 days of the invoice date or as otherwise specified elsewhere in this Legal Quotation or Partner Agreement."""),
        ("C46", """Unless otherwise specified, all software products will be delivered electronically and deemed accepted upon delivery of access to such software products (i.e. making such software products available for download).
The software licenses within this Legal Quotation shall be for End User's internal use only, even if the installation location in the quote detail for a license specifies an entity that is different than the End User, except as may be otherwise set forth in a separate signed written agreement between End User and Company.
The governing terms for this Legal Quotation consist of Company's standard distributor or partner contract terms and conditions (as applicable), unless superseded by a separate signed agreement ("Governing Terms"). The software, services and support hereunder are sold to Partner strictly for the purpose of resale and not for any internal or other use by Partner.
	Unless otherwise agreed in writing, the products and services are purchased solely under the terms and conditions of the IBM License Terms including but not limited to IBM Passport Advantage and IBM Cloud Offerings available at https://www.ibm.com. No other terms apply. In the event of any inconsistencies between the existing agreement and License Terms, the terms of the License Terms prevail."""),
        ("C47", """For all professional services, the scope, deliverables, and timelines shall be defined in the applicable SOW or service description accompanying this Legal Quotation. Changes to the scope of services after acceptance of this Legal Quotation must be agreed in writing by Company and may result in additional charges or revised delivery timelines.
Scheduling of professional services is subject to resource availability. Company will make reasonable efforts to accommodate requested dates but reserves the right to propose alternatives.
Any expected expenses for the delivery of professional services, including but not limited to travel, accommodation and subsistence, are defined as an explicit cost on the above quote, or incorporated in the agreed fee for the Statement of Work."""),
        ("C48", """T&M Services are offered on a half-day or full-day basis under predefined SKUs. Each engagement is supported by a corresponding SOW, which outlines the scope, estimated effort, and associated deliverables. Time-based billing shall apply, and services will be invoiced according to actual time spent.
For Packaged Services, the standard service description shall apply unless otherwise agreed in writing.
Bespoke service deliverables shall be owned by the End User, unless otherwise specified in the SOW. IBM proprietary materials and intellectual property remains the property of IBM."""),
        ("C49", """Commodities included on this quotation are subject to shipping restrictions under applicable laws, including but not limited to United States and/or European Union export laws, and are authorized for delivery only to the destination shown. Diversion contrary to such applicable laws is prohibited.
Subscription licenses and software maintenance are not perpetual and begin with delivery of license keys. SaaS and education subscriptions and managed services begin when the service is provisioned. Support subscription rates and SaaS subscription rates are subject to change upon renewal."""),
        ("C50", """If you purchase a multi-year subscription license, SaaS or education subscription, managed service or software maintenance, or multi-year renewal, your purchase is for the full value of all years of the offering, even if required payments are annual. Partner irrevocably commits to pay such fees to Company for the entirety of the Term. In the event you fail to pay any annual payment, and such default shall continue for a period of thirty (30) days, then any and all remaining amounts for the relevant offering shall become immediately due and payable. All Orders, including renewals, are subject to acceptance by Company in its discretion. All purchases are final, with no right to a refund, except as expressly provided under the applicable license or service terms."""),
        ("C51", """By accepting this Legal Quotation, Partner agrees that no other terms and conditions apply to this transaction, including, without limitation, those on a PO or other document issued by Partner, End User or any other third party.
Each party shall keep all confidential information it receives using the same protections that it applies to its own information of like importance, but in no event less than reasonable care, and may use such information solely for the purposes contemplated by this transaction or as otherwise agreed mutually in writing by both parties."""),
        ("C52", """Under no circumstances shall either party's liability arising out of or in connection with the Products or a party's performance with this Agreement exceed the aggregate amount of the fees paid by Partner and all orders regardless of whether such claim for liability is alleged to arise in contract, tort (including negligence) or otherwise. In no event shall either party be liable for indirect, special, incidental, or punitive damages including, without limitation, damages resulting from loss of use, loss of data, loss of profits, or loss of business arising out of, or in connection with, the products, services and/or solutions or Partner's performance of any of its obligations under this Agreement. Limitation of liability in this clause does not apply to intellectual property, confidentiality, compliance breaches by Partner and any other liability which cannot be excluded or limited under applicable law.
These General Terms and Conditions are governed by and construed according to the laws of England and Wales and each party irrevocably and unconditionally submits to the non-exclusive jurisdiction of the courts of Dubai International Financial Centre. The 1980 U.N. Convention on Contracts for the International Sale of Goods shall not apply."""),
    ]
    
    return terms


def estimate_line_count(text, max_chars_per_line=80):
    """Estimate number of lines needed for wrapped text"""
    lines = text.split('\n')
    total_lines = 0
    for line in lines:
        if not line:
            total_lines += 1
        else:
            wrapped = len(line) // max_chars_per_line + (1 if (len(line) % max_chars_per_line) else 0)
            total_lines += max(1, wrapped)
    return total_lines


def create_mibb_excel(
    data: list,
    header_info: dict,
    logo_path: str,
    output: BytesIO
):
    """
    Create MIBB Quotation Excel file.
    
    Args:
        data: list of rows, each row is [Part Number, Description, Start Date, End Date, QTY, Price USD]
        header_info: dict with header fields (same as IBM)
        logo_path: path to logo image
        output: BytesIO object to write Excel to
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Quotation"
    ws.sheet_view.showGridLines = False

    # --- Header / Branding ---
    ws.merge_cells("B1:C2")
    if logo_path and os.path.exists(logo_path):
        img = Image(logo_path)
        img.width = 1.87 * 96
        img.height = 0.56 * 96
        ws.add_image(img, "B1")
        ws.row_dimensions[1].height = 25
        ws.row_dimensions[2].height = 25

    ws.merge_cells("D3:G3")
    ws["D3"] = "Quotation"
    ws["D3"].font = Font(size=20, color="1F497D")
    ws["D3"].alignment = Alignment(horizontal="center", vertical="center")

    # Column widths
    ws.column_dimensions[get_column_letter(2)].width = 8
    ws.column_dimensions[get_column_letter(3)].width = 15
    ws.column_dimensions[get_column_letter(4)].width = 50
    ws.column_dimensions[get_column_letter(5)].width = 10
    ws.column_dimensions[get_column_letter(6)].width = 14
    ws.column_dimensions[get_column_letter(7)].width = 14
    ws.column_dimensions[get_column_letter(8)].width = 15
    ws.column_dimensions[get_column_letter(9)].width = 15
    ws.column_dimensions[get_column_letter(10)].width = 18
    ws.column_dimensions[get_column_letter(11)].width = 15
    ws.column_dimensions[get_column_letter(12)].width = 18

    # Left side labels and values
    left_labels = ["Date:", "From:", "Email:", "Contact:", "", "Company:", "Attn:", "Email:"]
    left_values = [
        datetime.today().strftime('%d/%m/%Y'),
        "Priyansha Kapoor",
        "p.Kapoor@mindware.net",
        "+971 55 456 6650",
        "",
        header_info.get('Reseller Name', 'empty'),
        "empty",
        "empty"
    ]
    row_positions = [5, 6, 7, 8, 9, 10, 11, 12]
    for row, label, value in zip(row_positions, left_labels, left_values):
        if label:
            ws[f"C{row}"] = label
            ws[f"C{row}"].font = Font(bold=True, color="1F497D")
        if value:
            ws[f"D{row}"] = value
            ws[f"D{row}"].font = Font(color="1F497D")

    # Right side labels and values
    right_labels = [
        "End User:", "Bid Number:", "Agreement Number:", "PA Site Number:", "",
        "Select Territory:", "Government Entity (GOE):", "Payment Terms:"
    ]
    right_values = [
        header_info.get('Customer Name', ''),
        header_info.get('Bid Number', ''),
        header_info.get('PA Agreement Number', ''),
        header_info.get('PA Site Number', ''),
        "",
        header_info.get('Select Territory', ''),
        header_info.get('Government Entity (GOE)', ''),
        "As aligned with Mindware"
    ]
    for row, label, value in zip(row_positions, right_labels, right_values):
        ws.merge_cells(f"H{row}:L{row}")
        ws[f"H{row}"] = f"{label} {value}"
        ws[f"H{row}"].font = Font(bold=True, color="1F497D")
        ws[f"H{row}"].alignment = Alignment(horizontal="left", vertical="center")

    # --- Table Headers ---
    headers = [
        "Sl", "Part Number", "Description", "Start Date", "End Date", "QTY", "Price USD"
    ]
    
    header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    for col, header in enumerate(headers, start=2):
        ws.merge_cells(start_row=16, start_column=col, end_row=17, end_column=col)
        cell = ws.cell(row=16, column=col, value=header)
        cell.font = Font(bold=True, size=13, color="1F497D")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = header_fill

    row_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    start_row = 18

    # --- Data Rows ---
    for idx, row in enumerate(data, start=1):
        excel_row = start_row + idx - 1
        
        # Extract data: [Part Number, Description, Start Date, End Date, QTY, Price USD]
        part_number = row[0] if len(row) > 0 else ""
        description = row[1] if len(row) > 1 else ""
        start_date = row[2] if len(row) > 2 else ""
        end_date = row[3] if len(row) > 3 else ""
        qty = row[4] if len(row) > 4 else 0
        price_usd = row[5] if len(row) > 5 else 0

        ws.cell(row=excel_row, column=2, value=idx).font = Font(size=11, color="1F497D")
        ws.cell(row=excel_row, column=2).alignment = Alignment(horizontal="center", vertical="center")
        
        ws.cell(row=excel_row, column=3, value=part_number).font = Font(size=11, color="1F497D")
        ws.cell(row=excel_row, column=4, value=description).font = Font(size=11, color="1F497D")
        ws.cell(row=excel_row, column=5, value=start_date).font = Font(size=11, color="1F497D")
        ws.cell(row=excel_row, column=6, value=end_date).font = Font(size=11, color="1F497D")
        ws.cell(row=excel_row, column=7, value=qty).font = Font(size=11, color="1F497D")
        ws.cell(row=excel_row, column=8, value=price_usd).font = Font(size=11, color="1F497D")
        
        # Format price as USD currency
        ws.cell(row=excel_row, column=8).number_format = '"USD"#,##0.00'
        
        for col in range(2, 9):
            ws.cell(row=excel_row, column=col).fill = row_fill
        
        ws.cell(row=excel_row, column=4).alignment = Alignment(wrap_text=True, horizontal="left", vertical="center")
        for col in [2, 3, 5, 6, 7, 8]:
            ws.cell(row=excel_row, column=col).alignment = Alignment(horizontal="center", vertical="center")

    # --- Summary row (if data exists) ---
    if data:
        data_end_row = start_row + len(data) - 1
        summary_row = data_end_row + 2
        
        ws.merge_cells(f"C{summary_row}:G{summary_row}")
        ws[f"C{summary_row}"] = "Total Price USD"
        ws[f"C{summary_row}"].font = Font(bold=True, color="1F497D")
        ws[f"C{summary_row}"].alignment = Alignment(horizontal="right")
        
        # Calculate total
        total_sum = sum(float(row[5]) if len(row) > 5 and row[5] else 0 for row in data)
        total_formula = f"=SUM(H{start_row}:H{data_end_row})"
        ws[f"H{summary_row}"] = total_formula
        ws[f"H{summary_row}"].number_format = '"USD"#,##0.00'
        ws[f"H{summary_row}"].font = Font(bold=True, color="1F497D")
        ws[f"H{summary_row}"].fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    else:
        summary_row = start_row + 1

    # --- Terms and Conditions Section ---
    terms = get_mibb_terms_section(header_info)
    terms_start_row = summary_row + 3
    
    # Adjust terms cell addresses based on where table ends
    adjusted_terms = []
    row_offset = terms_start_row - 29  # 29 is the anchor in terms_template.py
    for cell_addr, text, *style in terms:
        try:
            if len(cell_addr) >= 2 and cell_addr[1:].isdigit():
                col_letter = cell_addr[0]
                original_row = int(cell_addr[1:])
                new_row = original_row + row_offset
                new_cell_addr = f"{col_letter}{new_row}"
                adjusted_terms.append((new_cell_addr, text, *style))
            else:
                adjusted_terms.append((cell_addr, text, *style))
        except Exception:
            adjusted_terms.append((cell_addr, text, *style))

    for cell_addr, text, *style in adjusted_terms:
        try:
            if len(cell_addr) >= 2 and cell_addr[1:].isdigit():
                row_num = int(cell_addr[1:])
                col_letter = cell_addr[0]
                merge_rows = style[0].get("merge_rows") if style else None
                end_row = row_num + (merge_rows - 1 if merge_rows else 0)
                is_bold_title = style and "bold" in style[0] and style[0].get("bold") is True
                if is_bold_title:
                    ws.merge_cells(f"{col_letter}{row_num}:E{end_row}")
                    ws.row_dimensions[row_num].height = 24
                else:
                    ws.merge_cells(f"{col_letter}{row_num}:H{end_row}")
                    line_count = estimate_line_count(str(text), max_chars_per_line=80)
                    total_height = max(18, line_count * 16)
                    ws.row_dimensions[row_num].height = total_height
                ws[cell_addr] = text
                ws[cell_addr].alignment = Alignment(wrap_text=True, vertical="top")
                if style and "bold" in style[0]:
                    ws[cell_addr].font = Font(**style[0])
        except Exception:
            pass

    # --- Page Setup ---
    last_row = ws.max_row
    ws.print_area = f"A1:L{last_row}"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.15
    ws.page_margins.right = 0.15
    ws.page_margins.top = 0.25
    ws.page_margins.bottom = 0.25
    ws.page_margins.header = 0.15
    ws.page_margins.footer = 0.15
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.draft = False
    ws.page_setup.blackAndWhite = False
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    wb.calculation.fullCalcOnLoad = True
    wb.save(output)
    output.seek(0)
