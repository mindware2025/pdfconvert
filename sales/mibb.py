"""
MIBB Quotation processing module.
- Uses same header extraction as IBM quotations
- Custom table structure: Part Number, Description, Start Date, End Date, QTY, Price USD
- MIBB-specific terms and conditions
"""

from datetime import datetime, timedelta
from io import BytesIO
import os
import re
import fitz  # PyMuPDF
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
import logging
from pathlib import Path
from xml.etree import ElementTree as ET

import xlrd


try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - defensive import
    load_workbook = None

# Configure MIBB-specific logging
# MIBB_LOG_DIR = Path("mibb_logs")
# MIBB_LOG_DIR.mkdir(exist_ok=True)

# # Create logger for MIBB extraction
# mibb_logger = logging.getLogger('mibb_extraction')
# mibb_logger.setLevel(logging.DEBUG)

# # Remove existing handlers to avoid duplicates
# for handler in mibb_logger.handlers[:]:
#     mibb_logger.removeHandler(handler)

# # Create file handler with timestamp
# log_timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
# log_file_path = MIBB_LOG_DIR / f'mibb_extraction_{log_timestamp}.log'
# file_handler = logging.FileHandler(log_file_path, mode='w', encoding='utf-8')
# file_handler.setLevel(logging.DEBUG)

# # Create formatter
# formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
# file_handler.setFormatter(formatter)

# # Add handler to logger
# mibb_logger.addHandler(file_handler)
# mibb_logger.propagate = False

# def log_debug(message):
#     """Helper function to log debug messages"""
#     mibb_logger.debug(message)
#     print(f"[MIBB DEBUG] {message}")  # Also print to console for immediate feedback

def log_debug(message):
    # logging disabled
    return

def correct_mibb_descriptions(extracted_data, master_map=None):
    """
    MIBB rows: [part_number, description, start_date, end_date, qty, price_usd]

    Policy:
    - If master_map uploaded:
        - If SKU found: use master description
        - If SKU not found: blank description
    - If master_map NOT uploaded:
        - blank description for all rows
    """
    corrected = []
    use_master = bool(master_map)

    for row in extracted_data:
        # make sure description column exists
        if len(row) < 2:
            row = row + [""] * (2 - len(row))

        part = str(row[0]).strip().upper()

        if use_master:
            row[1] = master_map.get(part, "")   # ✅ blank if not found
        else:
            row[1] = ""                         # ✅ blank if no pricelist

        corrected.append(row)

    return corrected

def parse_euro_number(value: str):
    """Parse EU-formatted numbers like '733,00' -> 733.00"""
    try:
        if value is None:
            return None
        s = str(value).strip().replace(" ", "")
        if "." in s and "," in s:
            if s.rfind(",") > s.rfind("."):
                s = s.replace(".", "").replace(",", ".")
            else:
                s = s.replace(",", "")
        else:
            s = s.replace(",", ".")
        return float(s)
    except Exception:
        return None


def parse_decimal_number(value):
    """Parse decimal strings with optional thousand separators."""
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    text = text.replace(",", "")
    try:
        return float(text)
    except Exception:
        return None


def value_after_colon(text: str) -> str:
    """Return the trimmed value after the first colon, or the full string if no colon exists."""
    if text is None:
        return ""
    raw = str(text).strip()
    if ":" not in raw:
        return raw
    return raw.split(":", 1)[1].strip()


def subtract_days_from_date(date_text: str, days: int = 2) -> str:
    """Return a date string shifted backward by `days` when the input format is recognized."""
    if not date_text:
        return date_text

    raw = str(date_text).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            parsed = datetime.strptime(raw, fmt)
            adjusted = parsed - timedelta(days=days)
            return adjusted.strftime(fmt)
        except ValueError:
            continue
    return raw


def extract_mibb_header_from_pdf(file_like) -> dict:
    """
    Extract header information from MIBB quotation PDF.
    Uses same logic as IBM header extraction.
    Returns: dict with header fields
    """

    
    try:
        doc = fitz.open(stream=file_like.read(), filetype="pdf")
     
    except Exception as e:
        return {}
    
    # Collect lines
    lines = []
    for page_num, page in enumerate(doc):
        page_text = page.get_text("text") or page.get_text()
        page_lines = []
        for l in page_text.splitlines():
            if l and l.strip():
                lines.append(l.rstrip())
                page_lines.append(l.rstrip())
        
    for idx, line in enumerate(lines[:50]):
        log_debug(f"  Line {idx:3d}: {line}")
    
    # Header fields (same as IBM)
    header_info = {
        "Customer Name": "",
        "Bid Number": "",
        
        "Select Territory": "",
        "Government Entity (GOE)": "",
        "Reseller Name": "",
        "City": "",
        "Country": "",
        "Maximum End User Price (MEP)": "",
        "Total Value Seller Revenue Opportunity": "",
        "Bid Expiration Date": ""
    }
    
    # Parse header info (same logic as IBM)
    log_debug("\nParsing header information...")
    fields_found = 0
    for i, line in enumerate(lines):
        if "Customer Name:" in line:
            value = value_after_colon(lines[i])
            header_info["Customer Name"] = value
            log_debug(f"  [Line {i}] Customer Name: '{value}'")
            fields_found += 1
        if "City:" in line:
            value = value_after_colon(lines[i])
            header_info["City"] = value
            log_debug(f"  [Line {i}] City: '{value}'")
            fields_found += 1
        if "Country:" in line:
            value = value_after_colon(lines[i])
            header_info["Country"] = value
            log_debug(f"  [Line {i}] Country: '{value}'")
            fields_found += 1
        if "Reseller Name:" in line:
            value = value_after_colon(lines[i])
            header_info["Reseller Name"] = value
            log_debug(f"  [Line {i}] Reseller Name: '{value}'")
            fields_found += 1
        if "Bid Number:" in line or "Quote Number:" in line:
            value = value_after_colon(lines[i])
            header_info["Bid Number"] = value
            fields_found += 1
        if "Business Partner of Record:" in line :
            value = value_after_colon(lines[i])
            header_info["Business Partner of Record"] = value
            fields_found += 1
       
        if "Select Territory:" in line:
            value = value_after_colon(line)
            if not value and i + 1 < len(lines):
                value = lines[i + 1].strip()
            header_info["Select Territory"] = value
            log_debug(f"  [Line {i}] Select Territory: '{value}'")
            fields_found += 1
        if "Government Entity" in line:
            value = value_after_colon(line)
            if not value and i + 1 < len(lines):
                next_line = lines[i + 1].strip()
                if ":" not in next_line and next_line.lower() not in {
                    "bid request information",
                    "subscription quotation",
                    "parts information",
                }:
                    value = next_line
            header_info["Government Entity (GOE)"] = value
            log_debug(f"  [Line {i}] Government Entity (GOE): '{value}'")
            fields_found += 1
        
        if "Bid Expiration Date:" in line or "Quote Expiration Date:" in line:
            value = value_after_colon(lines[i])
            header_info["Bid Expiration Date"] = value
            log_debug(f"  [Line {i}] Bid Expiration Date: '{value}'")
            fields_found += 1
        if (
            "Maximum End User Price" in line
            or "Total Value Seller Revenue Opportunity" in line
            or "MEP" in line
        ):
            # Look for MEP value in same line or next line
            if ":" in line:
                mep_part = line.split(":", 1)[1].strip()
                if mep_part:
                    mep_clean = re.sub(r'\s*(USD).*$', '', mep_part).strip()
                    mep_value = parse_euro_number(mep_clean)
                    if mep_value:
                        header_info["Maximum End User Price (MEP)"] = f"{mep_value:,.2f}"
                        log_debug(f"  [Line {i}] MEP found in same line: '{mep_part}' -> {mep_value:,.2f}")
                        fields_found += 1
            elif i + 1 < len(lines):
                next_line = lines[i + 1].strip()
                if "USD" in next_line or "," in next_line:
                    mep_clean = re.sub(r'\s*(USD).*$', '', next_line).strip()
                    mep_value = parse_euro_number(mep_clean)
                    if mep_value:
                        header_info["Maximum End User Price (MEP)"] = f"{mep_value:,.2f}"
                        log_debug(f"  [Line {i}] MEP found in next line: '{next_line}' -> {mep_value:,.2f}")
                        fields_found += 1
    
    log_debug(f"\nHeader extraction complete: {fields_found} fields found")
    log_debug("\nExtracted header information:")
    for key, value in header_info.items():
        log_debug(f"  {key:35s}: {value}")
    log_debug("="*80 + "\n")
    
    return header_info


def extract_mibb_table_from_pdf(file_like) -> list:
    """
    Extract table data from MIBB quotation PDF.
    Returns: list of rows [Part Number, Description, Start Date, End Date, QTY, Price USD]
    Handles tables spanning multiple pages (e.g., items continue on page 3). [1](https://midisgroup1-my.sharepoint.com/personal/z_mama_mindware_net/Documents/Microsoft%20Copilot%20Chat%20Files/WTA%20Ooredoo.pdf)
    """
    log_debug("=" * 80)
    log_debug("MIBB TABLE EXTRACTION STARTED")
    log_debug("=" * 80)

    try:
        doc = fitz.open(stream=file_like.read(), filetype="pdf")
        log_debug(f"PDF opened for table extraction: {len(doc)} pages")
    except Exception as e:
        log_debug(f"ERROR opening PDF for table extraction: {e}")
        return []

    if len(doc) == 0:
        log_debug("ERROR: PDF has 0 pages")
        return []

    # -----------------------------
    # Find candidate pages
    # -----------------------------
    marker_patterns = ["parts information", "subscription quotation", "quotation - parts information"]
    header_signals = ["part number", "coverage start", "coverage end", "quantity", "qty", "bid ext", "bid extended"]

    candidate_pages: list[tuple[int, int, int]] = []  # (page_index, marker_score, header_score)

    for page_idx in range(len(doc)):
        try:
            page_text = (doc[page_idx].get_text("text") or doc[page_idx].get_text() or "")
        except Exception as e:
            log_debug(f"[PAGE SCAN] Could not read text for page {page_idx+1}: {e}")
            continue

        text_lower = page_text.lower()
        marker_score = sum(1 for p in marker_patterns if p in text_lower)
        header_score = sum(1 for s in header_signals if s in text_lower)

        if marker_score > 0 or header_score >= 4:
            candidate_pages.append((page_idx, marker_score, header_score))
            log_debug(f"[PAGE SCAN] Candidate page {page_idx+1}: marker_score={marker_score}, header_score={header_score}")

    if candidate_pages:
        candidate_pages.sort(key=lambda t: (-t[1], -t[2], t[0]))
        pages_to_process = [doc[p[0]] for p in candidate_pages]
        log_debug(f"[PAGE SELECT] Will process pages: {[p[0] + 1 for p in candidate_pages]}")
    else:
        # fallback if scan fails
        fallback_idx = 1 if len(doc) >= 2 else 0
        pages_to_process = [doc[fallback_idx]]
        log_debug(f"[PAGE SELECT] No candidates found; falling back to page {fallback_idx+1}")

    # -----------------------------
    # Extract from each page and combine
    # -----------------------------
    all_extracted: list[list] = []
    

    for page in pages_to_process:
        page_no = page.number + 1
        log_debug(f"\n==================== PROCESSING PAGE {page_no} ====================")

        extracted_data: list[list] = []

        # -------------------------
        # STRATEGY 1: Table detection (preferred)
        # -------------------------
        try:
            log_debug(f"[STRATEGY 1] Table detection on page {page_no}...")
            tf = page.find_tables()
            tables = getattr(tf, "tables", [])
            log_debug(f"Found {len(tables)} table(s) using PyMuPDF")

            if tables:
                # ✅ Pick ONLY the "Subscription Quotation - Parts Information" table
                best_rows = None
                best_score = -1
            
                for t_idx, t in enumerate(tables):
                    r = t.extract()
                    if not r or len(r) < 2:
                        continue
            
                    header = r[0]
                    header_text = " ".join(str(x).upper() for x in header if x)
            
                    # Score based on Parts Information signals
                    score = 0
                    if "COVERAGE START" in header_text: score += 3
                    if "COVERAGE END" in header_text: score += 3
                    if "TRANSACTION TYPE" in header_text: score += 2
                    if "BID EXT SVP" in header_text or "BID EXTENDED" in header_text: score += 3
                    if "DISCOUNT%" in header_text: score += 1
                    if "ENTITLED" in header_text: score += 1
            
                    log_debug(f"[TABLE CHECK] Table #{t_idx+1}: rows={len(r)}, score={score}, header='{header_text[:120]}'")
            
                    if score > best_score:
                        best_score = score
                        best_rows = r
            
                # If we couldn't identify Parts Information, treat as failure -> fallback to Strategy 2
                if best_score < 5 or not best_rows:
                    raise Exception("Could not identify 'Parts Information' table (found tables but headers don't match)")
            
                rows = best_rows
                log_debug(f"[TABLE SELECT] Using Parts Information table score={best_score} rows={len(rows)}")
                log_debug(f"Table extracted: {len(rows)} rows found")

                if len(rows) < 2:
                    raise Exception("Insufficient rows")

                header_row = rows[0]
                part_num_col = desc_col = start_date_col = end_date_col = qty_col = bid_ext_svp_col = None

                for idx, header in enumerate(header_row):
                    h = str(header).upper() if header else ""
                    if "PART NUMBER" in h:
                        part_num_col = idx
                    elif "DESCRIPTION" in h:
                        desc_col = idx
                    elif "COVERAGE START" in h:
                        start_date_col = idx
                    elif "COVERAGE END" in h:
                        end_date_col = idx
                    elif "QUANTITY" in h or "QTY" in h:
                        qty_col = idx
                    elif "BID EXT SVP" in h or "BID EXTENDED" in h:
                        bid_ext_svp_col = idx

                for r in rows[1:]:
                    if not r:
                        continue

                    part_number = str(r[part_num_col]).strip() if part_num_col is not None and part_num_col < len(r) else ""
                    description = str(r[desc_col]).strip() if desc_col is not None and desc_col < len(r) else ""
                    start_date = str(r[start_date_col]).strip() if start_date_col is not None and start_date_col < len(r) else ""
                    end_date = str(r[end_date_col]).strip() if end_date_col is not None and end_date_col < len(r) else ""
                    qty_str = str(r[qty_col]).strip() if qty_col is not None and qty_col < len(r) else "1"
                    price_str = str(r[bid_ext_svp_col]).strip() if bid_ext_svp_col is not None and bid_ext_svp_col < len(r) else "0"

                    # Accept SKUs like E0ELXLL, E0ELHLL etc. [1](https://midisgroup1-my.sharepoint.com/personal/z_mama_mindware_net/Documents/Microsoft%20Copilot%20Chat%20Files/WTA%20Ooredoo.pdf)
                    if not part_number or not re.match(r'^[A-Z0-9]{6,12}$', part_number):
                        continue

                    # QTY integer only (handles 5,400.00)
                    try:
                        qty = int(float(qty_str.replace(",", "")))
                    except:
                        qty = 1

                    price_usd = parse_euro_number(price_str) or 0.0

                    extracted_data.append([
                        part_number,
                        description,
                        start_date.replace(" ", ""),
                        end_date.replace(" ", ""),
                        qty,
                        price_usd
                    ])

            if len(extracted_data) == 0:
                raise Exception("Strategy 1 got 0 rows")

            log_debug(f"[STRATEGY 1 SUCCESS] Extracted {len(extracted_data)} rows from page {page_no}")

        except Exception as e:
            # -------------------------
            # STRATEGY 2: Text fallback
            # -------------------------
            log_debug(f"[STRATEGY 1 FAILED] {e}")
            log_debug(f"[STRATEGY 2] Text extraction on page {page_no}...")

            page_text = page.get_text("text") or page.get_text()
            lines = [l.rstrip() for l in page_text.splitlines() if l and l.strip()]

            # ✅ Anchor Strategy 2 to "Subscription Quotation" / "Parts Information" (ignore Overage)
            start_idx = None
            for i, line in enumerate(lines):
                up = line.upper()
                if "SUBSCRIPTION QUOTATION" in up:
                    start_idx = i
                    break
            
            if start_idx is None:
                for i, line in enumerate(lines):
                    if "PARTS INFORMATION" in line.upper():
                        start_idx = i
                        break
            
            header_line_idx = None
            if start_idx is not None:
                for i in range(start_idx, min(start_idx + 60, len(lines))):
                    if "PART NUMBER" in lines[i].upper():
                        header_line_idx = i
                        break
            
            if header_line_idx is None:
                log_debug("[STRATEGY 2] Could not find 'Subscription Quotation - Parts Information' header anchor")
                extracted_data = []
                # IMPORTANT: skip processing this page in Strategy 2
                # (return extracted_data or continue depending on your structure)

            if header_line_idx is None:
                extracted_data = []
            else:
                part_number_pattern = re.compile(r'\b[A-Z][A-Z0-9]{5,11}\b')
                date_pattern = re.compile(r'\b\d{2}/\d{2}/\d{4}\b')

                extracted_data = []
                i = header_line_idx + 1

                while i < len(lines):
                    part_match = part_number_pattern.search(lines[i])
                    if not part_match:
                        i += 1
                        continue

                    part_number = part_match.group()
                    description = lines[i + 1].strip() if i + 1 < len(lines) else ""

                    # Scan wider for start/end dates (fixes missing end date case) [1](https://midisgroup1-my.sharepoint.com/personal/z_mama_mindware_net/Documents/Microsoft%20Copilot%20Chat%20Files/WTA%20Ooredoo.pdf)
                    dates_found = []
                    for j in range(i + 1, min(i + 16, len(lines))):
                        dates_found += date_pattern.findall(lines[j])
                    dates_found = list(dict.fromkeys(dates_found))
                    start_date = dates_found[0] if len(dates_found) >= 1 else ""
                    end_date = dates_found[1] if len(dates_found) >= 2 else ""

                    # QTY integer only (handles 5,400.00) [1](https://midisgroup1-my.sharepoint.com/personal/z_mama_mindware_net/Documents/Microsoft%20Copilot%20Chat%20Files/WTA%20Ooredoo.pdf)
                    qty = 1
                    for j in range(i + 6, min(i + 16, len(lines))):
                        s = lines[j].strip()
                        if s == "-":
                            break
                        m = re.match(r'^(\d{1,3}(?:,\d{3})*(?:\.\d+)?|\d+(?:\.\d+)?)$', s)
                        if m:
                            qty = int(float(m.group(1).replace(",", "")))
                            break

                    # Price in Strategy 2 is not reliable yet; prefer Strategy 1 for price.
                    price_usd = 0.0

                    extracted_data.append([part_number, description, start_date, end_date, qty, price_usd])
                    i += 1

            log_debug(f"[STRATEGY 2 COMPLETE] Extracted {len(extracted_data)} rows from page {page_no}")

        # Merge results (dedupe by part number)
        all_extracted.extend(extracted_data)

    log_debug(f"\n[FINAL] Total extracted rows from all pages: {len(all_extracted)}")
    return all_extracted


def extract_mibb_table_from_pdf(file_like) -> list:
    """
    Extract table data from MIBB quotation PDF.
    Returns: list of rows [Part Number, Description, Start Date, End Date, QTY, Price USD]
    Supports multiple separate "Parts Information" tables on the same page.
    """
    log_debug("=" * 80)
    log_debug("MIBB TABLE EXTRACTION STARTED")
    log_debug("=" * 80)

    try:
        doc = fitz.open(stream=file_like.read(), filetype="pdf")
        log_debug(f"PDF opened for table extraction: {len(doc)} pages")
    except Exception as e:
        log_debug(f"ERROR opening PDF for table extraction: {e}")
        return []

    if len(doc) == 0:
        log_debug("ERROR: PDF has 0 pages")
        return []

    marker_patterns = ["parts information", "subscription quotation", "quotation - parts information"]
    header_signals = ["part number", "coverage start", "coverage end", "quantity", "qty", "bid ext", "bid extended"]

    candidate_pages: list[tuple[int, int, int]] = []
    for page_idx in range(len(doc)):
        try:
            page_text = (doc[page_idx].get_text("text") or doc[page_idx].get_text() or "")
        except Exception as e:
            log_debug(f"[PAGE SCAN] Could not read text for page {page_idx+1}: {e}")
            continue

        text_lower = page_text.lower()
        marker_score = sum(1 for p in marker_patterns if p in text_lower)
        header_score = sum(1 for s in header_signals if s in text_lower)

        if marker_score > 0 or header_score >= 4:
            candidate_pages.append((page_idx, marker_score, header_score))
            log_debug(f"[PAGE SCAN] Candidate page {page_idx+1}: marker_score={marker_score}, header_score={header_score}")

    if candidate_pages:
        candidate_pages.sort(key=lambda t: (-t[1], -t[2], t[0]))
        pages_to_process = [doc[p[0]] for p in candidate_pages]
        log_debug(f"[PAGE SELECT] Will process pages: {[p[0] + 1 for p in candidate_pages]}")
    else:
        fallback_idx = 1 if len(doc) >= 2 else 0
        pages_to_process = [doc[fallback_idx]]
        log_debug(f"[PAGE SELECT] No candidates found; falling back to page {fallback_idx+1}")

    def extract_rows_from_table(rows: list[list], table_idx: int) -> list[list]:
        if not rows or len(rows) < 2:
            return []

        header_row = rows[0]
        header_text = " ".join(str(x).upper() for x in header_row if x)

        score = 0
        if "PART NUMBER" in header_text:
            score += 2
        if "DESCRIPTION" in header_text:
            score += 2
        if "COVERAGE START" in header_text:
            score += 3
        if "COVERAGE END" in header_text:
            score += 3
        if "TRANSACTION TYPE" in header_text or "TYPE" in header_text:
            score += 2
        if "QUANTITY" in header_text or "QTY" in header_text:
            score += 2
        if "BID EXT SVP" in header_text or "BID EXTENDED" in header_text:
            score += 3
        if "DISCOUNT%" in header_text:
            score += 1
        if "ENTITLED" in header_text:
            score += 1

        log_debug(f"[TABLE CHECK] Table #{table_idx}: rows={len(rows)}, score={score}, header='{header_text[:120]}'")
        if score < 5:
            return []

        part_num_col = desc_col = start_date_col = end_date_col = qty_col = bid_ext_svp_col = None
        for idx, header in enumerate(header_row):
            h = str(header).upper() if header else ""
            if "PART NUMBER" in h:
                part_num_col = idx
            elif "DESCRIPTION" in h:
                desc_col = idx
            elif "COVERAGE START" in h:
                start_date_col = idx
            elif "COVERAGE END" in h:
                end_date_col = idx
            elif "QUANTITY" in h or "QTY" in h:
                qty_col = idx
            elif "BID EXT SVP" in h or "BID EXTENDED" in h:
                bid_ext_svp_col = idx

        if None in (part_num_col, desc_col, start_date_col, end_date_col, qty_col, bid_ext_svp_col):
            log_debug(f"[TABLE SKIP] Table #{table_idx} missing required columns")
            return []

        extracted_rows: list[list] = []
        for r in rows[1:]:
            if not r:
                continue

            part_number = str(r[part_num_col]).strip() if part_num_col < len(r) else ""
            description = str(r[desc_col]).strip() if desc_col < len(r) else ""
            start_date = str(r[start_date_col]).strip() if start_date_col < len(r) else ""
            end_date = str(r[end_date_col]).strip() if end_date_col < len(r) else ""
            qty_str = str(r[qty_col]).strip() if qty_col < len(r) else "1"
            price_str = str(r[bid_ext_svp_col]).strip() if bid_ext_svp_col < len(r) else "0"

            if not part_number or not re.match(r'^[A-Z0-9]{6,12}$', part_number):
                continue

            try:
                qty = int(float(qty_str.replace(",", "")))
            except Exception:
                qty = 1

            price_usd = parse_euro_number(price_str) or 0.0
            extracted_rows.append([
                part_number,
                description,
                start_date.replace(" ", ""),
                end_date.replace(" ", ""),
                qty,
                price_usd
            ])

        if extracted_rows:
            log_debug(f"[TABLE SELECT] Accepted table #{table_idx} with {len(extracted_rows)} row(s)")
        return extracted_rows

    all_extracted: list[list] = []
    seen_keys = set()

    for page in pages_to_process:
        page_no = page.number + 1
        log_debug(f"\n==================== PROCESSING PAGE {page_no} ====================")

        extracted_data: list[list] = []

        try:
            log_debug(f"[STRATEGY 1] Table detection on page {page_no}...")
            tf = page.find_tables()
            tables = getattr(tf, "tables", [])
            log_debug(f"Found {len(tables)} table(s) using PyMuPDF")

            if tables:
                for t_idx, table in enumerate(tables, start=1):
                    extracted_data.extend(extract_rows_from_table(table.extract(), t_idx))

            if len(extracted_data) == 0:
                raise Exception("Strategy 1 got 0 rows")

            log_debug(f"[STRATEGY 1 SUCCESS] Extracted {len(extracted_data)} rows from page {page_no}")

        except Exception as e:
            log_debug(f"[STRATEGY 1 FAILED] {e}")
            log_debug(f"[STRATEGY 2] Text extraction on page {page_no}...")

            page_text = page.get_text("text") or page.get_text()
            lines = [l.rstrip() for l in page_text.splitlines() if l and l.strip()]

            start_idx = None
            for i, line in enumerate(lines):
                up = line.upper()
                if "SUBSCRIPTION QUOTATION" in up:
                    start_idx = i
                    break

            if start_idx is None:
                for i, line in enumerate(lines):
                    if "PARTS INFORMATION" in line.upper():
                        start_idx = i
                        break

            header_line_idx = None
            if start_idx is not None:
                for i in range(start_idx, min(start_idx + 60, len(lines))):
                    if "PART NUMBER" in lines[i].upper():
                        header_line_idx = i
                        break

            if header_line_idx is None:
                log_debug("[STRATEGY 2] Could not find 'Subscription Quotation - Parts Information' header anchor")
                extracted_data = []
            else:
                part_number_pattern = re.compile(r'\b[A-Z][A-Z0-9]{5,11}\b')
                date_pattern = re.compile(r'\b\d{2}/\d{2}/\d{4}\b')

                extracted_data = []
                i = header_line_idx + 1

                while i < len(lines):
                    part_match = part_number_pattern.search(lines[i])
                    if not part_match:
                        i += 1
                        continue

                    part_number = part_match.group()
                    description = lines[i + 1].strip() if i + 1 < len(lines) else ""

                    dates_found = []
                    for j in range(i + 1, min(i + 16, len(lines))):
                        dates_found += date_pattern.findall(lines[j])
                    dates_found = list(dict.fromkeys(dates_found))
                    start_date = dates_found[0] if len(dates_found) >= 1 else ""
                    end_date = dates_found[1] if len(dates_found) >= 2 else ""

                    qty = 1
                    for j in range(i + 6, min(i + 16, len(lines))):
                        s = lines[j].strip()
                        if s == "-":
                            break
                        m = re.match(r'^(\d{1,3}(?:,\d{3})*(?:\.\d+)?|\d+(?:\.\d+)?)$', s)
                        if m:
                            qty = int(float(m.group(1).replace(",", "")))
                            break

                    extracted_data.append([part_number, description, start_date, end_date, qty, 0.0])
                    i += 1

            log_debug(f"[STRATEGY 2 COMPLETE] Extracted {len(extracted_data)} rows from page {page_no}")

        for row in extracted_data:
            row_key = tuple(row)
            if row_key in seen_keys:
                continue
            seen_keys.add(row_key)
            all_extracted.append(row)

    log_debug(f"\n[FINAL] Total extracted rows from all pages: {len(all_extracted)}")
    return all_extracted


def get_mibb_terms_section(header_info, data, margin_pct: float = 0.0):
    """
    Generate MIBB-specific terms and conditions section.
    Returns list of (cell_address, text, style_dict) tuples.
    """
    quote_validity = subtract_days_from_date(header_info.get("Bid Expiration Date", "XXXX"), days=2)
    margin_decimal = max(0.0, min(float(margin_pct or 0), 99.0)) / 100
    original_total = sum(float(row[5]) for row in data if len(row) > 5 and row[5])
    bp_total = original_total / (1 - margin_decimal) if margin_decimal < 1 else original_total
    totalprice = bp_total

    terms = [
        ("B29", "Terms and Conditions:", {"bold": True, "size": 11, "color": "1F497D"}),

        ("C30", 
f"""• 30 Days from POE Date.
• Quote Validity: {quote_validity} as per the quote
• Mindware requires full payment of this invoice (BP Price USD {totalprice:,.2f}) if WHT is applicable on offshore payment
• Pricing valid for this transaction only.
"""
        ),

        ("C31", "Definitions", {"bold": True}),

        ("C32",
"""“Company” refers to the MIBB entity identified at the top of the first page of this Legal Quotation.
“Partner” refers to the distributor entity identified in the “Distributor Name” section on the first page of this Legal Quotation.
“End User” refers to the end-user customer entity identified in the “Customer Name” section on the first page of this Legal Quotation, which is purchasing from or through Partner for its own internal use only.
“T&M Services” refers to time-based engagements sold by half or full-day SKUs with corresponding Company SOWs.
“Packaged Services” refers to standardized offerings tied to IBM part codes and IBM service descriptions.
“Bespoke Services” refers to tailored solutions governed by SOWs through unique Company SKUs and supporting SOWs.
“SOW” refers to the applicable statement of work accompanying this Legal Quotation.

Acceptance of this Legal Quotation requires Partner to issue a valid Purchase Order (“PO”) as indicated in this Legal Quotation or, where available, to select and complete the e-sign option.
The PO must (i) reference this Legal Quotation number, (ii) include the email address of the End User contact, and (iii) include the Partner email address to which the invoice(s) will be sent (or a physical address if required by applicable law).

This Legal Quotation includes (i) the applicable contractual discount, if any, or (ii) the special pricing agreed for this transaction. Prices are exclusive of applicable taxes, which will be borne by Partner.
Invoices will be sent by email unless otherwise required by law, and shall be paid to Company within 30 days from invoice date.

Unless otherwise specified, software products will be delivered electronically and deemed accepted upon delivery of access/download availability.
Licenses under this Legal Quotation are for End User’s internal use only, unless otherwise agreed in writing.
The governing terms consist of Company’s standard distributor/partner contract terms, MIBB General Terms for IBM Cloud Offerings, and the MIBB Service Description for Ordered Cloud Services (as applicable), unless superseded by a separate signed agreement (“Governing Terms”). Software and services are sold strictly for resale.

Unless otherwise agreed in writing, products/services are purchased solely under IBM License Terms including IBM Passport Advantage and IBM Cloud Offerings (https://www.ibm.com). In the event of inconsistencies, IBM License Terms prevail.

Where applicable, and unless explicitly agreed, licenses/S&S acquired under this Legal Quotation may not be used to resolve prior non-compliance, nor authorize deployment prior to the order date.
Sub-capacity licensing details: https://www.ibm.com/software/passportadvantage/subcaplicensing.html
Container licensing details: https://www.ibm.com/software/passportadvantage/containerlicenses.html

For all professional services, SOW or service descriptions define scope, deliverables, timelines. Changes require written agreement. Scheduling depends on resource availability. Applicable expenses (travel, accommodation, subsistence) must be pre-defined in the quote or included in the SOW.

T&M Services are offered per half/full day via predefined SKUs with SOW. Packaged Services follow standard descriptions. Bespoke deliverables belong to End User unless SOW states otherwise. IBM proprietary materials remain IBM property.

Commodities included in this quotation are subject to export laws and may be delivered only to the destination shown.
Subscription licenses and software maintenance begin with delivery of keys or provisioning. SaaS, education subscriptions, and managed services begin upon provisioning. Renewal pricing may change.

Multi-year subscriptions commit Partner to the full term value, even where payment is annual. In case of non-payment beyond 30 days, all future installments become immediately due. All orders are subject to Company acceptance. Purchases are final unless explicitly provided under applicable terms.

By accepting this Legal Quotation, Partner agrees no other terms apply, including those on Partner/End User POs.
Each party shall protect confidential information using reasonable care.
Liability is capped at the aggregate fees paid; indirect or consequential damages are excluded. Limitations do not apply to IP, confidentiality, or compliance breaches where prohibited by law.

Governing law: England and Wales.
Jurisdiction: Dubai International Financial Centre (non-exclusive).
The UN Convention on Contracts for the International Sale of Goods does not apply."""
        ),
    ]

    return terms


def estimate_line_count(text, max_chars_per_line=80):
    """Estimate number of lines needed for wrapped text"""
    lines = text.split('\n')
    total_lines = 0
    for line in lines:
        if not line:
            total_lines += 1
        else:
            wrapped = len(line) // max_chars_per_line + (1 if (len(line) % max_chars_per_line) else 0)
            total_lines += max(1, wrapped)
    return total_lines


def extract_mibb_terms_from_pdf(file_like) -> str:
    """
    Extract the General Terms and Conditions section from the PDF as plain text.
    """
    try:
        raw = file_like.read()
        doc = fitz.open(stream=raw, filetype="pdf")
    except Exception:
        return ""

    terms_pages = []
    started = False
    for page in doc:
        page_text = page.get_text("text") or page.get_text() or ""
        if not started and "General Terms and Conditions" not in page_text:
            continue
        started = True
        lines = [line.rstrip() for line in page_text.splitlines() if line.strip()]
        if lines and lines[0].startswith("Page "):
            lines = lines[1:]
        terms_pages.append("\n".join(lines))

    return "\n\n".join(terms_pages).strip()


def _normalize_header_name(value) -> str:
    return re.sub(r"[^A-Z0-9]+", " ", str(value or "").upper()).strip()


def _parse_spreadsheetml_2003_rows(raw_bytes: bytes) -> list[list]:
    ns = {
        "ss": "urn:schemas-microsoft-com:office:spreadsheet",
    }
    cleaned = raw_bytes.lstrip()
    cleaned = re.sub(
        rb"&(?!amp;|lt;|gt;|apos;|quot;|#\d+;|#x[0-9A-Fa-f]+;)",
        b"&amp;",
        cleaned,
    )
    root = ET.fromstring(cleaned)
    worksheet = root.find("ss:Worksheet", ns)
    if worksheet is None:
        return []
    table = worksheet.find("ss:Table", ns)
    if table is None:
        return []

    rows = []
    for row_elem in table.findall("ss:Row", ns):
        row_values = []
        current_idx = 1
        for cell in row_elem.findall("ss:Cell", ns):
            cell_index = cell.attrib.get("{urn:schemas-microsoft-com:office:spreadsheet}Index")
            if cell_index:
                target_idx = int(cell_index)
                while current_idx < target_idx:
                    row_values.append("")
                    current_idx += 1

            data_elem = cell.find("ss:Data", ns)
            row_values.append(data_elem.text.strip() if data_elem is not None and data_elem.text else "")
            current_idx += 1

        rows.append(row_values)
    return rows


def _read_first_sheet_rows(file_like) -> list[list]:
    if hasattr(file_like, "seek"):
        file_like.seek(0)
    raw = file_like.read() if hasattr(file_like, "read") else file_like
    if hasattr(file_like, "seek"):
        file_like.seek(0)

    if not raw:
        return []

    prefix = raw[:512].lstrip()
    if prefix.startswith(b"<?xml") or b"schemas-microsoft-com:office:spreadsheet" in raw[:2048]:
        return _parse_spreadsheetml_2003_rows(raw)

    if raw[:2] == b"PK" and load_workbook is not None:
        wb = load_workbook(BytesIO(raw), data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        return [list(row) for row in ws.iter_rows(values_only=True)]

    if xlrd is not None:
        book = xlrd.open_workbook(file_contents=raw)
        sheet = book.sheet_by_index(0)
        return [sheet.row_values(i) for i in range(sheet.nrows)]

    raise ValueError("Unsupported Excel format.")


def _find_label_value(rows: list[list], target_labels: set[str]) -> str:
    for row in rows:
        normalized_row = [str(cell).strip() for cell in row]
        for idx, cell in enumerate(normalized_row):
            if cell in target_labels:
                for next_cell in normalized_row[idx + 1:]:
                    if str(next_cell).strip():
                        return str(next_cell).strip()
    return ""


def _normalize_quote_number(value: str) -> str:
    return str(value or "").strip().lstrip("0")


def check_mibb_hardware_quote_match(excel_file, pdf_bid_number):
    """
    Ensure uploaded hardware Excel belongs to the same quote as the PDF.
    """
    try:
        rows = _read_first_sheet_rows(excel_file)
        excel_quote_id = _find_label_value(rows, {"Quote Id:", "Quote number:", "Quote Id", "Quote number"})
        if _normalize_quote_number(excel_quote_id) == _normalize_quote_number(pdf_bid_number):
            return True, None
        return False, "Your uploaded PDF and Excel do not match. Please verify the quote files."
    except Exception as e:
        return False, f"Error checking quote match: {e}"


def extract_mibb_hardware_table_from_excel(file_like) -> list:
    """
    Extract hardware rows from the uploaded Excel/XML quotation.
    Returns rows in the format:
    [part_number, description, qty, list_ext_svp, bid_ext_svp]
    """
    rows = _read_first_sheet_rows(file_like)
    if not rows:
        return []

    header_idx = None
    header_map = {}
    required_headers = {
        "PART NUMBER": "part_number",
        "DESCRIPTION": "description",
        "QTY": "qty",
        "LIST EXT SVP": "list_ext_svp",
        "BID EXT SVP": "bid_ext_svp",
    }

    for idx, row in enumerate(rows):
        normalized = [_normalize_header_name(cell) for cell in row]
        if "PART NUMBER" in normalized and "DESCRIPTION" in normalized and "BID EXT SVP" in normalized:
            header_idx = idx
            for col_idx, cell in enumerate(normalized):
                if cell in required_headers:
                    header_map[required_headers[cell]] = col_idx
            break

    if header_idx is None:
        return []

    extracted = []
    empty_streak = 0
    for row in rows[header_idx + 1:]:
        part_number = str(row[header_map["part_number"]]).strip() if len(row) > header_map["part_number"] and row[header_map["part_number"]] is not None else ""
        description = str(row[header_map["description"]]).strip() if len(row) > header_map["description"] and row[header_map["description"]] is not None else ""

        if not part_number and not description:
            empty_streak += 1
            if empty_streak >= 3:
                break
            continue
        empty_streak = 0

        if _normalize_header_name(part_number).startswith("TOTAL"):
            break

        qty_raw = row[header_map["qty"]] if len(row) > header_map["qty"] else None
        list_ext_raw = row[header_map["list_ext_svp"]] if len(row) > header_map["list_ext_svp"] else None
        bid_ext_raw = row[header_map["bid_ext_svp"]] if len(row) > header_map["bid_ext_svp"] else None

        qty_value = parse_decimal_number(qty_raw) or 0
        list_ext_value = parse_decimal_number(list_ext_raw) or 0
        bid_ext_value = parse_decimal_number(bid_ext_raw) or 0

        extracted.append([
            part_number,
            description,
            int(qty_value) if float(qty_value).is_integer() else qty_value,
            list_ext_value,
            bid_ext_value,
        ])

    return extracted


def create_mibb_hardware_excel(
    data: list,
    header_info: dict,
    logo_path: str,
    output: BytesIO,
    margin_pct: float = 1.0,
    terms_text: str = "",
):
    """
    Create MIBB hardware quotation Excel file.
    Final table:
    Sl | Part Number | Description | Qty | List Ext SVP | BP Unit Price USD |
    BP Extended price USD | mindware extended cost USD | Margin
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Quotation"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("B1:C2")
    if logo_path and os.path.exists(logo_path):
        img = Image(logo_path)
        img.width = 1.87 * 96
        img.height = 0.56 * 96
        ws.add_image(img, "B1")
        ws.row_dimensions[1].height = 25
        ws.row_dimensions[2].height = 25

    ws.merge_cells("D3:G3")
    ws["D3"] = "Quotation"
    ws["D3"].font = Font(size=20, color="1F497D")
    ws["D3"].alignment = Alignment(horizontal="center", vertical="center")

    column_widths = {
        2: 8,
        3: 18,
        4: 48,
        5: 10,
        6: 16,
        7: 18,
        8: 20,
        9: 22,
        10: 12,
    }
    for col_idx, width in column_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    left_labels = ["Date:", "From:", "Email:", "Contact:", "", "Company:", "Attn:", "Email:"]
    left_values = [
        datetime.today().strftime('%d/%m/%Y'),
        "Eliana Youssef",
        "E.youssef@mindware.net",
        "+961 123 456 758",
        "",
        header_info.get("Reseller Name", "empty"),
        "empty",
        "empty",
    ]
    for row, label, value in zip([5, 6, 7, 8, 9, 10, 11, 12], left_labels, left_values):
        if label:
            ws[f"C{row}"] = label
            ws[f"C{row}"].font = Font(bold=True, color="1F497D")
        if value:
            ws[f"D{row}"] = value
            ws[f"D{row}"].font = Font(color="1F497D")

    right_labels = [
        "Customer Name:",
        "Bid Number:",
        "Business Partner of Record:",
        "Payment Terms:",
        "GOE",
        "Country",
    ]
    right_values = [
        header_info.get("Customer Name", ""),
        header_info.get("Bid Number", ""),
        header_info.get("Business Partner of Record", ""),
        "As aligned with Mindware",
        header_info.get("Government Entity (GOE)", ""),
        header_info.get("Country", ""),
    ]
    for row, label, value in zip([5, 6, 7, 8, 9, 10], right_labels, right_values):
        ws.merge_cells(f"H{row}:L{row}")
        ws[f"H{row}"] = f"{label} {value}"
        ws[f"H{row}"].font = Font(bold=True, color="1F497D")
        ws[f"H{row}"].alignment = Alignment(horizontal="left", vertical="center")

    headers = [
        "Sl",
        "Part Number",
        "Description",
        "Qty",
        "List Ext SVP",
        "BP Unit Price USD",
        "BP Extended price USD",
        "mindware extended cost USD",
        "Margin",
    ]
    header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    for col, header in enumerate(headers, start=2):
        ws.merge_cells(start_row=16, start_column=col, end_row=17, end_column=col)
        cell = ws.cell(row=16, column=col, value=header)
        cell.font = Font(bold=True, size=13, color="1F497D")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = header_fill

    row_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    start_row = 18
    margin_decimal = max(0.0, min(float(margin_pct or 0), 99.0)) / 100

    for idx, row in enumerate(data, start=1):
        excel_row = start_row + idx - 1
        part_number = row[0] if len(row) > 0 else ""
        description = row[1] if len(row) > 1 else ""
        qty = row[2] if len(row) > 2 else 0
        list_ext_svp = row[3] if len(row) > 3 else 0
        mindware_extended_cost = row[4] if len(row) > 4 else 0

        ws.cell(row=excel_row, column=2, value=idx)
        ws.cell(row=excel_row, column=3, value=part_number)
        ws.cell(row=excel_row, column=4, value=description)
        ws.cell(row=excel_row, column=5, value=qty)
        ws.cell(row=excel_row, column=6, value=list_ext_svp)
        ws.cell(row=excel_row, column=7, value=f'=IFERROR(H{excel_row}/E{excel_row},0)')
        ws.cell(row=excel_row, column=8, value=f'=IFERROR(I{excel_row}/(1-J{excel_row}),0)')
        ws.cell(row=excel_row, column=9, value=mindware_extended_cost)
        ws.cell(row=excel_row, column=10, value=margin_decimal)

        for col in range(2, 11):
            cell = ws.cell(row=excel_row, column=col)
            cell.font = Font(size=11, color="1F497D")
            cell.fill = row_fill
            if col == 4:
                cell.alignment = Alignment(wrap_text=True, horizontal="left", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        for price_col in [6, 7, 8, 9]:
            ws.cell(row=excel_row, column=price_col).number_format = '"USD"#,##0.00'
        ws.cell(row=excel_row, column=10).number_format = "0.0%"

    if data:
        data_end_row = start_row + len(data) - 1
        summary_row = data_end_row + 2
        ws.merge_cells(f"C{summary_row}:G{summary_row}")
        ws[f"C{summary_row}"] = "Total Price USD"
        ws[f"C{summary_row}"].font = Font(bold=True, color="1F497D")
        ws[f"C{summary_row}"].alignment = Alignment(horizontal="right")
        ws[f"H{summary_row}"] = f"=SUM(H{start_row}:H{data_end_row})"
        ws[f"H{summary_row}"].number_format = '"USD"#,##0.00'
        ws[f"H{summary_row}"].font = Font(bold=True, color="1F497D")
        ws[f"H{summary_row}"].fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    else:
        summary_row = start_row + 1

    terms_title_row = summary_row + 3
    ws.merge_cells(f"B{terms_title_row}:L{terms_title_row}")
    ws[f"B{terms_title_row}"] = "Shipping Freight and Clearance Charges are excluded from the BP Price"
    ws[f"B{terms_title_row}"].font = Font(bold=True, size=11, color="1F497D")
    ws[f"B{terms_title_row}"].alignment = Alignment(horizontal="left", vertical="center")

    terms_header_row = terms_title_row + 1
    ws[f"B{terms_header_row}"] = "Terms and Conditions:"
    ws[f"B{terms_header_row}"].font = Font(bold=True, size=11, color="1F497D")

    terms_body_row = terms_header_row + 1
    body_text = terms_text.strip() if terms_text else "No terms extracted from PDF."
    ws.merge_cells(f"C{terms_body_row}:L{terms_body_row + 15}")
    ws[f"C{terms_body_row}"] = body_text
    ws[f"C{terms_body_row}"].alignment = Alignment(wrap_text=True, vertical="top")
    ws[f"C{terms_body_row}"].font = Font(size=10, color="1F497D")
    ws.row_dimensions[terms_body_row].height = max(80, estimate_line_count(body_text, max_chars_per_line=95) * 15)

    last_row = terms_body_row + 15
    ws.print_area = f"A1:K{last_row}"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.15
    ws.page_margins.right = 0.15
    ws.page_margins.top = 0.25
    ws.page_margins.bottom = 0.25
    ws.page_margins.header = 0.15
    ws.page_margins.footer = 0.15
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.draft = False
    ws.page_setup.blackAndWhite = False
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    wb.calculation.fullCalcOnLoad = True
    wb.save(output)
    output.seek(0)


def create_mibb_excel(
    data: list,
    header_info: dict,
    logo_path: str,
    output: BytesIO,
    margin_pct: float = 1.0
):
    """
    Create MIBB Quotation Excel file.
    
    Args:
        data: list of rows, each row is [Part Number, Description, Start Date, End Date, QTY, Price USD]
        header_info: dict with header fields (same as IBM)
        logo_path: path to logo image
        output: BytesIO object to write Excel to
        margin_pct: margin percentage entered in the UI
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Quotation"
    ws.sheet_view.showGridLines = False

    # --- Header / Branding ---
    ws.merge_cells("B1:C2")
    if logo_path and os.path.exists(logo_path):
        img = Image(logo_path)
        img.width = 1.87 * 96
        img.height = 0.56 * 96
        ws.add_image(img, "B1")
        ws.row_dimensions[1].height = 25
        ws.row_dimensions[2].height = 25

    ws.merge_cells("D3:G3")
    ws["D3"] = "Quotation"
    ws["D3"].font = Font(size=20, color="1F497D")
    ws["D3"].alignment = Alignment(horizontal="center", vertical="center")

    # Column widths
    ws.column_dimensions[get_column_letter(2)].width = 8
    ws.column_dimensions[get_column_letter(3)].width = 15
    ws.column_dimensions[get_column_letter(4)].width = 50
    ws.column_dimensions[get_column_letter(5)].width = 10
    ws.column_dimensions[get_column_letter(6)].width = 14
    ws.column_dimensions[get_column_letter(7)].width = 14
    ws.column_dimensions[get_column_letter(8)].width = 16
    ws.column_dimensions[get_column_letter(9)].width = 16
    ws.column_dimensions[get_column_letter(10)].width = 18
    ws.column_dimensions[get_column_letter(11)].width = 12

    # Left side labels and values
    left_labels = ["Date:", "From:", "Email:", "Contact:", "", "Company:", "Attn:", "Email:"]
    left_values = [
        datetime.today().strftime('%d/%m/%Y'),
        "Eliana Youssef",
        "E.youssef@mindware.net",
        "+961 123 456 758",
        "",
        header_info.get('Reseller Name', 'empty'),
        "empty",
        "empty"
    ]
    row_positions = [5, 6, 7, 8, 9, 10, 11, 12]
    for row, label, value in zip(row_positions, left_labels, left_values):
        if label:
            ws[f"C{row}"] = label
            ws[f"C{row}"].font = Font(bold=True, color="1F497D")
        if value:
            ws[f"D{row}"] = value
            ws[f"D{row}"].font = Font(color="1F497D")

    # Right side labels and values
    right_labels = [
        "", " ",  " ","Payment Terms:"
    ]
    right_values = [
        header_info.get('Customer Name', ''),
        header_info.get('Bid Number', ''),
        
        header_info.get('Business Partner of Record', ''),
       
        "As aligned with Mindware"
    ]
    for row, label, value in zip(row_positions, right_labels, right_values):
        ws.merge_cells(f"H{row}:L{row}")
        ws[f"H{row}"] = f"{label} {value}"
        ws[f"H{row}"].font = Font(bold=True, color="1F497D")
        ws[f"H{row}"].alignment = Alignment(horizontal="left", vertical="center")

    # --- Table Headers ---
    headers = [
        "Sl",
        "Part Number",
        "Description",
        "Start Date",
        "End Date",
        "QTY",
        "unit price USD",
        "total price",
        "original total price",
        "margin"
    ]
    
    header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    for col, header in enumerate(headers, start=2):
        ws.merge_cells(start_row=16, start_column=col, end_row=17, end_column=col)
        cell = ws.cell(row=16, column=col, value=header)
        cell.font = Font(bold=True, size=13, color="1F497D")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = header_fill

    row_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    start_row = 18

    # --- Data Rows ---
    for idx, row in enumerate(data, start=1):
        excel_row = start_row + idx - 1
        
        # Extract data: [Part Number, Description, Start Date, End Date, QTY, Price USD]
        part_number = row[0] if len(row) > 0 else ""
        description = row[1] if len(row) > 1 else ""
        start_date = row[2] if len(row) > 2 else ""
        end_date = row[3] if len(row) > 3 else ""
        qty = row[4] if len(row) > 4 else 0
        price_usd = row[5] if len(row) > 5 else 0

        ws.cell(row=excel_row, column=2, value=idx).font = Font(size=11, color="1F497D")
        ws.cell(row=excel_row, column=2).alignment = Alignment(horizontal="center", vertical="center")
        
        ws.cell(row=excel_row, column=3, value=part_number).font = Font(size=11, color="1F497D")
        ws.cell(row=excel_row, column=4, value=description).font = Font(size=11, color="1F497D")
        ws.cell(row=excel_row, column=5, value=start_date).font = Font(size=11, color="1F497D")
        ws.cell(row=excel_row, column=6, value=end_date).font = Font(size=11, color="1F497D")
        # QTY (column G = 7)
        ws.cell(row=excel_row, column=7, value=qty).font = Font(size=11, color="1F497D")
        ws.cell(row=excel_row, column=7).alignment = Alignment(horizontal="center", vertical="center")
        
        # Bid extended price (column I = 9)  ✅ this is the extracted price_usd
        ws.cell(row=excel_row, column=9, value=price_usd).font = Font(size=11, color="1F497D")
        ws.cell(row=excel_row, column=9).number_format = '"USD"#,##0.00'
        ws.cell(row=excel_row, column=9).alignment = Alignment(horizontal="center", vertical="center")
        
        # Partner Price USD (column H = 8) ✅ same logic as before: 99% of Bid extended price
        partner_formula = f"=ROUNDUP(I{excel_row}*0.99, 2)"
        ws.cell(row=excel_row, column=8, value=partner_formula)
        ws.cell(row=excel_row, column=8).font = Font(size=11, color="1F497D")
        ws.cell(row=excel_row, column=8).number_format = '"USD"#,##0.00'
        ws.cell(row=excel_row, column=8).alignment = Alignment(horizontal="center", vertical="center")
        
        # Extend BP price (column J = 10) ✅ Partner Price USD * QTY
        extend_bp_formula = f"=H{excel_row}*G{excel_row}"
        ws.cell(row=excel_row, column=10, value=extend_bp_formula)
        ws.cell(row=excel_row, column=10).font = Font(size=11, color="1F497D")
        ws.cell(row=excel_row, column=10).number_format = '"USD"#,##0.00'
        ws.cell(row=excel_row, column=10).alignment = Alignment(horizontal="center", vertical="center")
        
        
        for col in range(2, 11):
            ws.cell(row=excel_row, column=col).fill = row_fill
        
        ws.cell(row=excel_row, column=4).alignment = Alignment(wrap_text=True, horizontal="left", vertical="center")
        for col in [2, 3, 5, 6, 7, 8, 9, 10]:
            ws.cell(row=excel_row, column=col).alignment = Alignment(horizontal="center", vertical="center")

    # --- Summary row (if data exists) ---
    if data:
        data_end_row = start_row + len(data) - 1
        summary_row = data_end_row + 2
        
        ws.merge_cells(f"C{summary_row}:G{summary_row}")
        ws[f"C{summary_row}"] = "Total Price USD"
        ws[f"C{summary_row}"].font = Font(bold=True, color="1F497D")
        ws[f"C{summary_row}"].alignment = Alignment(horizontal="right")
        
        # Calculate total
        total_sum = sum(float(row[5]) if len(row) > 5 and row[5] else 0 for row in data)
        total_formula = f"=SUM(I{start_row}:I{data_end_row})"
        ws[f"I{summary_row}"] = total_formula
        ws[f"I{summary_row}"].number_format = '"USD"#,##0.00'
        ws[f"I{summary_row}"].font = Font(bold=True, color="1F497D")
        ws[f"I{summary_row}"].fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    else:
        summary_row = start_row + 1

    # --- Terms and Conditions Section ---
    terms = get_mibb_terms_section(header_info, data)
    terms_start_row = summary_row + 3
    
    # Adjust terms cell addresses based on where table ends
    adjusted_terms = []
    row_offset = terms_start_row - 29  # 29 is the anchor in terms_template.py
    for cell_addr, text, *style in terms:
        try:
            if len(cell_addr) >= 2 and cell_addr[1:].isdigit():
                col_letter = cell_addr[0]
                original_row = int(cell_addr[1:])
                new_row = original_row + row_offset
                new_cell_addr = f"{col_letter}{new_row}"
                adjusted_terms.append((new_cell_addr, text, *style))
            else:
                adjusted_terms.append((cell_addr, text, *style))
        except Exception:
            adjusted_terms.append((cell_addr, text, *style))

    for cell_addr, text, *style in adjusted_terms:
        try:
            if len(cell_addr) >= 2 and cell_addr[1:].isdigit():
                row_num = int(cell_addr[1:])
                col_letter = cell_addr[0]
                merge_rows = style[0].get("merge_rows") if style else None
                end_row = row_num + (merge_rows - 1 if merge_rows else 0)
                is_bold_title = style and "bold" in style[0] and style[0].get("bold") is True
                if is_bold_title:
                    ws.merge_cells(f"{col_letter}{row_num}:L{end_row}")
                    ws.row_dimensions[row_num].height = 32
                else:
                    # Merge across full width for proper PDF wrapping
                    ws.merge_cells(f"{col_letter}{row_num}:L{end_row}")
                
                    # Better wrap estimation for PDF export
                    line_count = estimate_line_count(str(text), max_chars_per_line=55)
                    total_height = max(40, line_count * 22)
                
                    ws.row_dimensions[row_num].height = total_height
                ws[cell_addr] = text
                ws[cell_addr].alignment = Alignment(wrap_text=True, vertical="top")
                if style and "bold" in style[0]:
                    ws[cell_addr].font = Font(**style[0])
        except Exception:
            pass

    # --- Page Setup ---
    last_row = ws.max_row
    ws.print_area = f"A1:L{last_row}"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.15
    ws.page_margins.right = 0.15
    ws.page_margins.top = 0.25
    ws.page_margins.bottom = 0.25
    ws.page_margins.header = 0.15
    ws.page_margins.footer = 0.15
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.draft = False
    ws.page_setup.blackAndWhite = False
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    wb.calculation.fullCalcOnLoad = True
    wb.save(output)
    output.seek(0)


def create_mibb_excel(
    data: list,
    header_info: dict,
    logo_path: str,
    output: BytesIO,
    margin_pct: float = 1.0
):
    """
    Create MIBB quotation Excel with margin-driven formulas.
    Final table:
    Sl | Part Number | Description | Start Date | End Date | QTY |
    unit price USD | total price | original total price | margin
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Quotation"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("B1:C2")
    if logo_path and os.path.exists(logo_path):
        img = Image(logo_path)
        img.width = 1.87 * 96
        img.height = 0.56 * 96
        ws.add_image(img, "B1")
        ws.row_dimensions[1].height = 25
        ws.row_dimensions[2].height = 25

    ws.merge_cells("D3:G3")
    ws["D3"] = "Quotation"
    ws["D3"].font = Font(size=20, color="1F497D")
    ws["D3"].alignment = Alignment(horizontal="center", vertical="center")

    column_widths = {
        2: 8,
        3: 15,
        4: 50,
        5: 10,
        6: 14,
        7: 10,
        8: 16,
        9: 16,
        10: 18,
        11: 12,
        12: 18,
    }
    for col_idx, width in column_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    left_labels = ["Date:", "From:", "Email:", "Contact:", "", "Company:", "Attn:", "Email:"]
    left_values = [
        datetime.today().strftime('%d/%m/%Y'),
        "Eliana Youssef",
        "E.youssef@mindware.net",
        "+961 123 456 758",
        "",
        header_info.get("Reseller Name", "empty"),
        "empty",
        "empty",
    ]
    row_positions = [5, 6, 7, 8, 9, 10, 11, 12]
    for row, label, value in zip(row_positions, left_labels, left_values):
        if label:
            ws[f"C{row}"] = label
            ws[f"C{row}"].font = Font(bold=True, color="1F497D")
        if value:
            ws[f"D{row}"] = value
            ws[f"D{row}"].font = Font(color="1F497D")

    right_labels = [
        "Customer Name:",
        "Bid Number:",
        "Business Partner of Record:",
        "Payment Terms:",
        "GOE",
        "Country",
    ]
    right_values = [
        header_info.get("Customer Name", ""),
        header_info.get("Bid Number", ""),
        header_info.get("Business Partner of Record", ""),
        "As aligned with Mindware",
        header_info.get("Government Entity (GOE)", ""),
        header_info.get("Country", ""),
    ]
    for row, label, value in zip([5, 6, 7, 8, 9, 10], right_labels, right_values):
        ws.merge_cells(f"H{row}:L{row}")
        ws[f"H{row}"] = f"{label} {value}"
        ws[f"H{row}"].font = Font(bold=True, color="1F497D")
        ws[f"H{row}"].alignment = Alignment(horizontal="left", vertical="center")

    headers = [
        "Sl",
        "Part Number",
        "Description",
        "Start Date",
        "End Date",
        "QTY",
        "Unit BP price\nUSD",
        "Total BP price\nUSD",
        "original total price",
        "margin",
    ]
    header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    for col, header in enumerate(headers, start=2):
        ws.merge_cells(start_row=16, start_column=col, end_row=17, end_column=col)
        cell = ws.cell(row=16, column=col, value=header)
        cell.font = Font(bold=True, size=13, color="1F497D")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = header_fill

    row_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    start_row = 18
    margin_decimal = max(0.0, min(float(margin_pct or 0), 99.0)) / 100

    for idx, row in enumerate(data, start=1):
        excel_row = start_row + idx - 1
        part_number = row[0] if len(row) > 0 else ""
        description = row[1] if len(row) > 1 else ""
        start_date = row[2] if len(row) > 2 else ""
        end_date = row[3] if len(row) > 3 else ""
        qty = row[4] if len(row) > 4 else 0
        original_total_price = row[5] if len(row) > 5 else 0

        ws.cell(row=excel_row, column=2, value=idx)
        ws.cell(row=excel_row, column=3, value=part_number)
        ws.cell(row=excel_row, column=4, value=description)
        ws.cell(row=excel_row, column=5, value=start_date)
        ws.cell(row=excel_row, column=6, value=end_date)
        ws.cell(row=excel_row, column=7, value=qty)
        ws.cell(row=excel_row, column=8, value=f'=IFERROR(I{excel_row}/G{excel_row},0)')
        ws.cell(row=excel_row, column=9, value=f'=IFERROR(J{excel_row}/(1-K{excel_row}),0)')
        ws.cell(row=excel_row, column=10, value=original_total_price)
        ws.cell(row=excel_row, column=11, value=margin_decimal)

        for col in range(2, 12):
            cell = ws.cell(row=excel_row, column=col)
            cell.font = Font(size=11, color="1F497D")
            cell.fill = row_fill
            if col == 4:
                cell.alignment = Alignment(wrap_text=True, horizontal="left", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        for price_col in [8, 9, 10]:
            ws.cell(row=excel_row, column=price_col).number_format = '"USD"#,##0.00'
        ws.cell(row=excel_row, column=11).number_format = "0.0%"

    if data:
        data_end_row = start_row + len(data) - 1
        summary_row = data_end_row + 2
        ws.merge_cells(f"C{summary_row}:H{summary_row}")
        ws[f"C{summary_row}"] = "Total Price USD"
        ws[f"C{summary_row}"].font = Font(bold=True, color="1F497D")
        ws[f"C{summary_row}"].alignment = Alignment(horizontal="right")
        ws[f"I{summary_row}"] = f"=SUM(I{start_row}:I{data_end_row})"
        ws[f"I{summary_row}"].number_format = '"USD"#,##0.00'
        ws[f"I{summary_row}"].font = Font(bold=True, color="1F497D")
        ws[f"I{summary_row}"].fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    else:
        summary_row = start_row + 1

    terms = get_mibb_terms_section(header_info, data, margin_pct=margin_pct)
    total_bp_text = sum(
        (float(row[5]) / (1 - margin_decimal)) if len(row) > 5 and row[5] and margin_decimal < 1 else float(row[5])
        for row in data
        if len(row) > 5 and row[5]
    )
    if len(terms) > 1:
        terms[1] = (
            "C30",
            f"""• Payment Terms as aligned with Mindware
• Quote Validity: {subtract_days_from_date(header_info.get("Bid Expiration Date", "XXXX"), days=2)} as per the quote
• Mindware requires full payment of this invoice (Total Price USD {total_bp_text:,.2f}) if WHT is applicable on offshore payment
• Pricing valid for this transaction only.
""",
        )
    terms_start_row = summary_row + 3

    adjusted_terms = []
    row_offset = terms_start_row - 29
    for cell_addr, text, *style in terms:
        try:
            if len(cell_addr) >= 2 and cell_addr[1:].isdigit():
                col_letter = cell_addr[0]
                original_row = int(cell_addr[1:])
                new_row = original_row + row_offset
                adjusted_terms.append((f"{col_letter}{new_row}", text, *style))
            else:
                adjusted_terms.append((cell_addr, text, *style))
        except Exception:
            adjusted_terms.append((cell_addr, text, *style))

    for cell_addr, text, *style in adjusted_terms:
        try:
            if len(cell_addr) >= 2 and cell_addr[1:].isdigit():
                row_num = int(cell_addr[1:])
                col_letter = cell_addr[0]
                merge_rows = style[0].get("merge_rows") if style else None
                end_row = row_num + (merge_rows - 1 if merge_rows else 0)
                ws.merge_cells(f"{col_letter}{row_num}:L{end_row}")

                is_bold_title = style and style[0].get("bold") is True
                if is_bold_title:
                    ws.row_dimensions[row_num].height = 32
                else:
                    line_count = estimate_line_count(str(text), max_chars_per_line=55)
                    ws.row_dimensions[row_num].height = max(40, line_count * 22)

                ws[cell_addr] = text
                ws[cell_addr].alignment = Alignment(wrap_text=True, vertical="top")
                if style and "bold" in style[0]:
                    ws[cell_addr].font = Font(**style[0])
        except Exception:
            pass

    last_row = ws.max_row
    ws.print_area = f"A1:K{last_row}"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.15
    ws.page_margins.right = 0.15
    ws.page_margins.top = 0.25
    ws.page_margins.bottom = 0.25
    ws.page_margins.header = 0.15
    ws.page_margins.footer = 0.15
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.draft = False
    ws.page_setup.blackAndWhite = False
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    wb.calculation.fullCalcOnLoad = True
    wb.save(output)
    output.seek(0)


def _clean_tls_cell(value, kind="text"):
    """Normalize a raw TLS Excel cell based on the column kind (qty/price/text)."""
    if value is None:
        return 0.0 if kind == "price" else ""
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")
    if kind == "qty":
        parsed = parse_decimal_number(value)
        if parsed is None:
            return str(value).strip()
        return int(parsed) if float(parsed).is_integer() else parsed
    if kind == "price":
        return parse_decimal_number(value) or 0.0
    return str(value).strip()


def _find_tls_price_column(columns):
    """Return the index of the last column whose header contains PRICE, or None."""
    price_idx = None
    for idx, name in enumerate(columns):
        if "PRICE" in _normalize_header_name(name):
            price_idx = idx
    return price_idx


def extract_mibb_tls_from_excel(file_like):
    """
    Extract header info and the dynamic line-items table from a TLS quote Excel.
    Columns are not fixed: whatever header cells exist under "Quote Line Items"
    are captured in order, and every data cell under them is returned.
    Returns (header_info, columns, data_rows).
    """
    rows = _read_first_sheet_rows(file_like)
    if not rows:
        return {}, [], []

    marker_idx = None
    for idx, row in enumerate(rows):
        if any(_normalize_header_name(cell) == "QUOTE LINE ITEMS" for cell in row):
            marker_idx = idx
            break

    header_idx = None
    if marker_idx is not None:
        for idx in range(marker_idx + 1, len(rows)):
            non_empty = [c for c in rows[idx] if c is not None and str(c).strip()]
            if len(non_empty) >= 2:
                header_idx = idx
                break
    if header_idx is None:
        for idx, row in enumerate(rows):
            normalized = [_normalize_header_name(c) for c in row if c is not None and str(c).strip()]
            if any("QTY" in n for n in normalized) and any("PRICE" in n for n in normalized):
                header_idx = idx
                break
    if header_idx is None:
        raise ValueError(
            "Could not locate the TLS line-items header row. "
            "Make sure the quote sheet is the first sheet in the workbook."
        )

    columns = []
    col_indexes = []
    for col_idx, cell in enumerate(rows[header_idx]):
        if cell is not None and str(cell).strip():
            columns.append(str(cell).strip())
            col_indexes.append(col_idx)

    # _find_label_value treats None cells as the string "None"; blank them out first.
    label_rows = [["" if c is None else c for c in row] for row in rows[:header_idx]]
    header_info = {
        "Date": _find_label_value(label_rows, {"Date:"}),
        "Prepared For": _find_label_value(label_rows, {"Prepared For:", "Prepared For"}),
        "Quote Id": _find_label_value(label_rows, {"Quote Id:", "Quote Id", "Quote number:", "Quote number"}),
        "End User": _find_label_value(label_rows, {"End User:", "End User"}),
        "Currency": _find_label_value(label_rows, {"Currency:", "Currency"}),
    }

    price_idx = _find_tls_price_column(columns)
    kinds = []
    for idx, name in enumerate(columns):
        normalized = _normalize_header_name(name)
        if idx == price_idx:
            kinds.append("price")
        elif "QTY" in normalized or "QUANTITY" in normalized:
            kinds.append("qty")
        else:
            kinds.append("text")

    data_rows = []
    empty_streak = 0
    for row in rows[header_idx + 1:]:
        row_non_empty = [c for c in row if c is not None and str(c).strip()]
        if not row_non_empty:
            empty_streak += 1
            if empty_streak >= 3:
                break
            continue
        empty_streak = 0

        if _normalize_header_name(row_non_empty[0]).startswith("TOTAL"):
            break

        values = [row[c] if c < len(row) else None for c in col_indexes]
        if not any(v is not None and str(v).strip() for v in values):
            continue
        data_rows.append([_clean_tls_cell(v, kinds[i]) for i, v in enumerate(values)])

    return header_info, columns, data_rows


def create_mibb_tls_excel(
    data: list,
    columns: list,
    header_info: dict,
    logo_path: str,
    output: BytesIO,
    margin_pct: float = 1.0,
):
    """
    Create MIBB TLS quotation Excel file with a dynamic column layout.
    Final table: Sl | <all columns from the uploaded Excel, with the Price
    column renamed to Mindware Cost USD> | BP Price USD | Margin
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Quotation"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("B1:C2")
    if logo_path and os.path.exists(logo_path):
        img = Image(logo_path)
        img.width = 1.87 * 96
        img.height = 0.56 * 96
        ws.add_image(img, "B1")
        ws.row_dimensions[1].height = 25
        ws.row_dimensions[2].height = 25

    ws.merge_cells("D3:G3")
    ws["D3"] = "Quotation"
    ws["D3"].font = Font(size=20, color="1F497D")
    ws["D3"].alignment = Alignment(horizontal="center", vertical="center")

    price_idx = _find_tls_price_column(columns)
    out_headers = ["Sl"]
    for idx, name in enumerate(columns):
        out_headers.append("Mindware Cost USD" if idx == price_idx else name)

    cost_col = bp_col = margin_col = None
    if price_idx is not None:
        out_headers += ["BP Price USD", "Margin"]
        cost_col = 3 + price_idx
        bp_col = 3 + len(columns)
        margin_col = bp_col + 1

    last_col = 2 + len(out_headers) - 1
    last_col_letter = get_column_letter(last_col)
    merge_end_letter = last_col_letter if last_col >= 12 else "L"

    ws.column_dimensions[get_column_letter(2)].width = 6
    for col_idx, header in enumerate(out_headers[1:], start=3):
        normalized = _normalize_header_name(header)
        if "DESCRIPTION" in normalized:
            width = 38
        elif "PRICE" in normalized or "COST" in normalized:
            width = 16
        elif "DATE" in normalized:
            width = 12
        elif "QTY" in normalized or "QUANTITY" in normalized:
            width = 8
        elif "MARGIN" in normalized:
            width = 10
        else:
            width = 13
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    left_labels = ["Date:", "From:", "Email:", "Contact:", "", "Company:", "Attn:", "Email:"]
    left_values = [
        datetime.today().strftime('%d/%m/%Y'),
        "Eliana Youssef",
        "E.youssef@mindware.net",
        "+961 123 456 758",
        "",
        "",
        "empty",
        "empty",
    ]
    for row, label, value in zip([5, 6, 7, 8, 9, 10, 11, 12], left_labels, left_values):
        if label:
            ws[f"C{row}"] = label
            ws[f"C{row}"].font = Font(bold=True, color="1F497D")
        if value:
            ws[f"D{row}"] = value
            ws[f"D{row}"].font = Font(color="1F497D")

    right_labels = [
        "Customer Name:",
        "Quote Number:",
        "Currency:",
        "Payment Terms:",
        "Quote Date:",
    ]
    right_values = [
        header_info.get("End User", ""),
        header_info.get("Quote Id", ""),
        header_info.get("Currency", ""),
        "As aligned with Mindware",
        header_info.get("Date", ""),
    ]
    for row, label, value in zip([5, 6, 7, 8, 9], right_labels, right_values):
        ws.merge_cells(f"H{row}:{merge_end_letter}{row}")
        ws[f"H{row}"] = f"{label} {value}"
        ws[f"H{row}"].font = Font(bold=True, color="1F497D")
        ws[f"H{row}"].alignment = Alignment(horizontal="left", vertical="center")

    header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    for col, header in enumerate(out_headers, start=2):
        ws.merge_cells(start_row=16, start_column=col, end_row=17, end_column=col)
        cell = ws.cell(row=16, column=col, value=header)
        cell.font = Font(bold=True, size=13, color="1F497D")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = header_fill

    row_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    start_row = 18
    margin_decimal = max(0.0, min(float(margin_pct or 0), 99.0)) / 100
    cost_letter = get_column_letter(cost_col) if cost_col else None
    margin_letter = get_column_letter(margin_col) if margin_col else None

    for idx, row in enumerate(data, start=1):
        excel_row = start_row + idx - 1
        ws.cell(row=excel_row, column=2, value=idx)
        for col_offset, value in enumerate(row):
            ws.cell(row=excel_row, column=3 + col_offset, value=value)
        if price_idx is not None:
            ws.cell(
                row=excel_row,
                column=bp_col,
                value=f'=IFERROR({cost_letter}{excel_row}/(1-{margin_letter}{excel_row}),0)',
            )
            ws.cell(row=excel_row, column=margin_col, value=margin_decimal)

        for col in range(2, last_col + 1):
            cell = ws.cell(row=excel_row, column=col)
            cell.font = Font(size=11, color="1F497D")
            cell.fill = row_fill
            if "DESCRIPTION" in _normalize_header_name(out_headers[col - 2]):
                cell.alignment = Alignment(wrap_text=True, horizontal="left", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        if price_idx is not None:
            ws.cell(row=excel_row, column=cost_col).number_format = '"USD"#,##0.00'
            ws.cell(row=excel_row, column=bp_col).number_format = '"USD"#,##0.00'
            ws.cell(row=excel_row, column=margin_col).number_format = "0.0%"

    if data:
        data_end_row = start_row + len(data) - 1
        summary_row = data_end_row + 2
        if price_idx is not None:
            bp_letter = get_column_letter(bp_col)
            ws.merge_cells(f"C{summary_row}:{get_column_letter(bp_col - 1)}{summary_row}")
            ws[f"C{summary_row}"] = "Total Price USD"
            ws[f"C{summary_row}"].font = Font(bold=True, color="1F497D")
            ws[f"C{summary_row}"].alignment = Alignment(horizontal="right")
            ws[f"{bp_letter}{summary_row}"] = f"=SUM({bp_letter}{start_row}:{bp_letter}{data_end_row})"
            ws[f"{bp_letter}{summary_row}"].number_format = '"USD"#,##0.00'
            ws[f"{bp_letter}{summary_row}"].font = Font(bold=True, color="1F497D")
            ws[f"{bp_letter}{summary_row}"].fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    else:
        summary_row = start_row + 1

    terms = get_mibb_terms_section(header_info, [], margin_pct=margin_pct)
    total_cost = 0.0
    if price_idx is not None:
        for row in data:
            try:
                total_cost += float(row[price_idx] or 0)
            except (TypeError, ValueError):
                continue
    total_bp_text = total_cost / (1 - margin_decimal) if margin_decimal < 1 else total_cost
    if len(terms) > 1:
        terms[1] = (
            "C30",
            f"""• Payment Terms as aligned with Mindware
• Quote Validity: as per the quote
• Mindware requires full payment of this invoice (Total Price USD {total_bp_text:,.2f}) if WHT is applicable on offshore payment
• Pricing valid for this transaction only.
""",
        )
    terms_start_row = summary_row + 3

    adjusted_terms = []
    row_offset = terms_start_row - 29
    for cell_addr, text, *style in terms:
        try:
            if len(cell_addr) >= 2 and cell_addr[1:].isdigit():
                col_letter = cell_addr[0]
                original_row = int(cell_addr[1:])
                new_row = original_row + row_offset
                adjusted_terms.append((f"{col_letter}{new_row}", text, *style))
            else:
                adjusted_terms.append((cell_addr, text, *style))
        except Exception:
            adjusted_terms.append((cell_addr, text, *style))

    for cell_addr, text, *style in adjusted_terms:
        try:
            if len(cell_addr) >= 2 and cell_addr[1:].isdigit():
                row_num = int(cell_addr[1:])
                col_letter = cell_addr[0]
                merge_rows = style[0].get("merge_rows") if style else None
                end_row = row_num + (merge_rows - 1 if merge_rows else 0)
                ws.merge_cells(f"{col_letter}{row_num}:{merge_end_letter}{end_row}")

                is_bold_title = style and style[0].get("bold") is True
                if is_bold_title:
                    ws.row_dimensions[row_num].height = 32
                else:
                    line_count = estimate_line_count(str(text), max_chars_per_line=55)
                    ws.row_dimensions[row_num].height = max(40, line_count * 22)

                ws[cell_addr] = text
                ws[cell_addr].alignment = Alignment(wrap_text=True, vertical="top")
                if style and "bold" in style[0]:
                    ws[cell_addr].font = Font(**style[0])
        except Exception:
            pass

    last_row = ws.max_row
    ws.print_area = f"A1:{last_col_letter}{last_row}"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.15
    ws.page_margins.right = 0.15
    ws.page_margins.top = 0.25
    ws.page_margins.bottom = 0.25
    ws.page_margins.header = 0.15
    ws.page_margins.footer = 0.15
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.draft = False
    ws.page_setup.blackAndWhite = False
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    wb.calculation.fullCalcOnLoad = True
    wb.save(output)
    output.seek(0)
