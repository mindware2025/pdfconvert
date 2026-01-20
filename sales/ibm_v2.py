
from datetime import datetime
from io import BytesIO
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from terms_template import get_terms_section

def compare_mep_and_cost(header_info, data):
    """
    Compares Maximum End User Price (MEP) from PDF header_info with total cost from Excel data.
    Args:
        header_info: dict containing extracted PDF header info (should include 'Maximum End User Price (MEP)')
        data: list of lists, each row [SKU, Description, Quantity, Start Date, End Date, Cost]
    Returns:
        str: Debug message for Streamlit UI
    """
    mep_value = header_info.get("Maximum End User Price (MEP)", None)
    try:
        mep_val = float(str(mep_value).replace(',', '').replace(' ', '')) if mep_value else None
    except Exception:
        mep_val = None

    # Sum Cost column from parsed Excel rows (index 5) if numeric-like
    total_cost = 0.0
    for row in data:
        if len(row) > 5:
            val = row[5]
            try:
                if isinstance(val, str):
                    # allow "1234.56" style
                    if val.strip().replace('.', '', 1).isdigit():
                        total_cost += float(val)
                elif isinstance(val, (int, float)):
                    total_cost += float(val)
            except Exception:
                continue

    msg = f"MEP value (PDF): {mep_val if mep_val is not None else 'N/A'} | Total cost (Excel): {total_cost:.2f}"
    if mep_val is not None and abs(mep_val - total_cost) > 0.01:
        msg += "\nThere is a difference between MEP and total cost. Please check with the IT team."
    return msg


def check_bid_number_match(excel_file, pdf_bid_number):
    """
    Checks if the bid number in the Excel first sheet matches the PDF bid number.
    Args:
        excel_file: BytesIO or file path of the uploaded Excel file.
        pdf_bid_number: Bid number extracted from PDF header_info.
    Returns:
        (bool, str): (True, None) if match, (False, error_message) if not.
    """
    try:
        xls = pd.ExcelFile(excel_file)
        df = xls.parse(xls.sheet_names[0], header=None)
        # Row 13 is index 12 (0-based)
        b13 = str(df.iloc[12, 1]).strip() if df.shape[0] > 12 and df.shape[1] > 1 else ""
        c13 = str(df.iloc[12, 2]).strip() if df.shape[0] > 12 and df.shape[1] > 2 else ""
        # Normalize bid numbers by stripping leading zeros
        pdf_bid_norm = str(pdf_bid_number).lstrip('0')
        c13_norm = c13.lstrip('0')
        if b13 == "Quote number:" and c13_norm == pdf_bid_norm:
            return True, None
        else:
            return False, "Your uploaded files do not match. If you have any inquiries, reach out to IT."
    except Exception as e:
        return False, f"Error checking bid number match: {e}"


USD_TO_AED = 3.6725


def estimate_line_count(text, max_chars_per_line=80):
    lines = text.split('\n')
    total_lines = 0
    for line in lines:
        if not line:
            total_lines += 1
        else:
            wrapped = len(line) // max_chars_per_line + (1 if (len(line) % max_chars_per_line) else 0)
            total_lines += max(1, wrapped)
    return total_lines


def create_styled_excel_v2(
    data: list,
    header_info: dict,
    logo_path: str,
    output: BytesIO,
    compliance_text: str,
    ibm_terms_text: str,
    country: str = "UAE"
):
    """
    Template 1 Excel generation (for v2):
    data rows (Qatar): [SKU, Product Description, Quantity, Start Date, End Date, Unit Price in USD, Cost (USD), Partner Discount, Partner Price in USD]
    data rows (UAE):   [SKU, Product Description, Quantity, Start Date, End Date, Cost]
    Table data is provided from Excel, not PDF.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Quotation"
    ws.sheet_view.showGridLines = False

    # --- Header / Branding ---
    ws.merge_cells("B1:C2")  # Move logo to row 1-2
    if logo_path and os.path.exists(logo_path):
        img = Image(logo_path)
        img.width = 1.87 * 96
        img.height = 0.56 * 96
        ws.add_image(img, "B1")
        ws.row_dimensions[1].height = 25
        ws.row_dimensions[2].height = 25

    ws.merge_cells("D3:G3")
    ws["D3"] = "Quotation"
    ws["D3"].font = Font(size=20, color="1F497D")
    ws["D3"].alignment = Alignment(horizontal="center", vertical="center")

    ws.column_dimensions[get_column_letter(2)].width = 8
    ws.column_dimensions[get_column_letter(3)].width = 15
    ws.column_dimensions[get_column_letter(4)].width = 50
    ws.column_dimensions[get_column_letter(5)].width = 10
    ws.column_dimensions[get_column_letter(6)].width = 14
    ws.column_dimensions[get_column_letter(7)].width = 14
    ws.column_dimensions[get_column_letter(8)].width = 15
    ws.column_dimensions[get_column_letter(9)].width = 15
    ws.column_dimensions[get_column_letter(10)].width = 18
    ws.column_dimensions[get_column_letter(11)].width = 15
    ws.column_dimensions[get_column_letter(12)].width = 18

    left_labels = ["Date:", "From:", "Email:", "Contact:", "", "Company:", "Attn:", "Email:"]
    if country == "Qatar":
        left_values = [
            datetime.today().strftime('%d/%m/%Y'),
            "Eliana Youssef",
            "e.youssef@mindware.net",
            "70/519841",
            "",
            header_info.get('Reseller Name', 'empty'),
            "empty",
            "empty"
        ]
    else:
        left_values = [
            datetime.today().strftime('%d/%m/%Y'),
            "Priyansha Kapoor",
            "p.Kapoor@mindware.net",
            "+971 55 456 6650",
            "",
            header_info.get('Reseller Name', 'empty'),
            "empty",
            "empty"
        ]
    row_positions = [5, 6, 7, 8, 9, 10, 11, 12]
    for row, label, value in zip(row_positions, left_labels, left_values):
        if label:
            ws[f"C{row}"] = label
            ws[f"C{row}"].font = Font(bold=True, color="1F497D")
        if value:
            ws[f"D{row}"] = value
            ws[f"D{row}"].font = Font(color="1F497D")

    right_labels = [
        "End User:", "Bid Number:", "Agreement Number:", "PA Site Number:", "",
        "Select Territory:", "Government Entity (GOE):", "Payment Terms:"
    ]
    right_values = [
        header_info.get('Customer Name', ''),
        header_info.get('Bid Number', ''),
        header_info.get('PA Agreement Number', ''),
        header_info.get('PA Site Number', ''),
        "",
        header_info.get('Select Territory', ''),
        header_info.get('Government Entity (GOE)', ''),
        "As aligned with Mindware"
    ]
    for row, label, value in zip(row_positions, right_labels, right_values):
        ws.merge_cells(f"H{row}:L{row}")
        ws[f"H{row}"] = f"{label} {value}"
        ws[f"H{row}"].font = Font(bold=True, color="1F497D")
        ws[f"H{row}"].alignment = Alignment(horizontal="left", vertical="center")

    # --- Table Headers ---
    if country == "Qatar":
        
        headers = [
                   "Sl", "SKU", "Product Description", "Quantity", "Start Date", "End Date",
                    "MEP Unit Price in USD", "Extended MEP Price USD", "Unit Partner Price USD", "Total Partner Price in USD"
                ]
        # Expand last column for Qatar
        ws.column_dimensions[get_column_letter(11)].width = 25  # Column K (Partner Price in USD)
        ws.column_dimensions[get_column_letter(12)].width = 25  # Column L (for extra space)
    else:
        headers = [
            "Sl", "SKU", "Product Description", "Quantity", "Start Date", "End Date",
            "Unit Price in AED", "Cost (USD)", "Total Price in AED", "Partner Discount", "Partner Price in AED"
        ]

    header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    for col, header in enumerate(headers, start=2):
        ws.merge_cells(start_row=16, start_column=col, end_row=17, end_column=col)
        cell = ws.cell(row=16, column=col, value=header)
        cell.font = Font(bold=True, size=13, color="1F497D")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = header_fill

    row_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    start_row = 18

    # --- Data Rows ---
    for idx, row in enumerate(data, start=1):
        excel_row = start_row + idx - 1
        if country == "Qatar":
            # row: [SKU, Product Description, Quantity, Start Date, End Date, Unit Price USD, Cost USD, Partner Discount, Partner Price USD]
            ws.cell(row=excel_row, column=2, value=idx).font = Font(size=11, color="1F497D")
            ws.cell(row=excel_row, column=2).alignment = Alignment(horizontal="center", vertical="center")

            for col, value in enumerate(row, start=3):
                # NOTE: if 'value' is a formula string (starts with '='), openpyxl will keep it as a formula in Excel.
                ws.cell(row=excel_row, column=col, value=value)
                ws.cell(row=excel_row, column=col).font = Font(size=11, color="1F497D")
                if col == 4:
                    ws.cell(row=excel_row, column=col).alignment = Alignment(wrap_text=True, horizontal="left", vertical="center")
                else:
                    ws.cell(row=excel_row, column=col).alignment = Alignment(horizontal="center", vertical="center")

            # Currency formatting for USD columns: F..I -> (6..9) data columns correspond to 7..10 on the sheet due to "Sl" at col B
            ws.cell(row=excel_row, column=7).number_format = '"USD"#,##0.00'   # Unit Price in USD (F)
            ws.cell(row=excel_row, column=8).number_format = '"USD"#,##0.00'   # Cost (USD) (G)
            ws.cell(row=excel_row, column=9).number_format = '"USD"#,##0.00'   # Partner Discount (H)
            ws.cell(row=excel_row, column=10).number_format = '"USD"#,##0.00'  # Partner Price in USD (I)

            for col in range(2, 12):
                ws.cell(row=excel_row, column=col).fill = row_fill

        else:
            # --- UAE logic (per-row) ---
            sku = row[0] if len(row) > 0 else ""
            desc = row[1] if len(row) > 1 else ""
            qty = row[2] if len(row) > 2 else 0
            start_date = row[3] if len(row) > 3 else ""
            end_date = row[4] if len(row) > 4 else ""
            raw_cost = row[5] if len(row) > 5 else 0
            try:
                cost = float(raw_cost) if raw_cost not in (None, "", "-") else 0
            except Exception:
                cost = 0

            ws.cell(row=excel_row, column=2, value=idx).font = Font(size=11, color="1F497D")
            ws.cell(row=excel_row, column=2).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=excel_row, column=3, value=sku).font = Font(size=11, color="1F497D")
            ws.cell(row=excel_row, column=4, value=desc).font = Font(size=11, color="1F497D")
            ws.cell(row=excel_row, column=5, value=qty).font = Font(size=11, color="1F497D")
            ws.cell(row=excel_row, column=6, value=start_date).font = Font(size=11, color="1F497D")
            ws.cell(row=excel_row, column=7, value=end_date).font = Font(size=11, color="1F497D")

            for col in range(2, 13):
                ws.cell(row=excel_row, column=col).alignment = Alignment(horizontal="center", vertical="center")

            # H (col 8) = Unit Price AED (formula), I (9) = Cost USD, J (10) = Total Price AED (formula)
            unit_price_formula = f"=J{excel_row}/E{excel_row}" if qty and qty != 0 else ""
            ws.cell(row=excel_row, column=8, value=unit_price_formula)
            ws.cell(row=excel_row, column=8).font = Font(size=11, color="1F497D")

            ws.cell(row=excel_row, column=9, value=cost).number_format = '"USD"#,##0.00'
            ws.cell(row=excel_row, column=9).font = Font(size=11, color="1F497D")

            total_formula = f"=I{excel_row}*{USD_TO_AED}"
            ws.cell(row=excel_row, column=10, value=total_formula)
            ws.cell(row=excel_row, column=10).font = Font(size=11, color="1F497D")

            discount_formula = f"=ROUNDUP(H{excel_row}*0.99,2)"
            ws.cell(row=excel_row, column=11, value=discount_formula)
            ws.cell(row=excel_row, column=11).font = Font(size=11, color="1F497D")

            partner_price_formula = f"=K{excel_row}*E{excel_row}"
            ws.cell(row=excel_row, column=12, value=partner_price_formula)
            ws.cell(row=excel_row, column=12).font = Font(size=11, color="1F497D")

            for price_col in [8, 10, 11, 12]:
                ws.cell(row=excel_row, column=price_col).number_format = '"AED"#,##0.00'
            ws.cell(row=excel_row, column=9).number_format = '"USD"#,##0.00'

            for col in range(2, 2 + len(headers)):
                ws.cell(row=excel_row, column=col).fill = row_fill

            ws.cell(row=excel_row, column=4).alignment = Alignment(wrap_text=True, horizontal="left", vertical="center")

    # --- Summary sections ---
    if country == "Qatar":
        # Qatar summary (after all rows)
        if data:
            data_start_row = start_row
            data_end_row = start_row + len(data) - 1

            # --- Total MEP USD (sum of G) ---
            summary_row = data_end_row + 2
            ws.merge_cells(f"C{summary_row}:F{summary_row}")
            ws[f"C{summary_row}"] = "Total MEP USD"
            ws[f"C{summary_row}"].font = Font(bold=True, color="1F497D")
            ws[f"C{summary_row}"].alignment = Alignment(horizontal="right")

            total_mep_formula = f"=SUM(I{data_start_row}:I{data_end_row})"
            ws[f"I{summary_row}"] = total_mep_formula
            ws[f"I{summary_row}"].number_format = '"USD"#,##0.00'
            ws[f"I{summary_row}"].font = Font(bold=True, color="1F497D")
            ws[f"I{summary_row}"].fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            # --- Total Business Partner Price USD (sum of I) ---
            bp_summary_row = summary_row + 1
            ws.merge_cells(f"C{bp_summary_row}:F{bp_summary_row}")
            ws[f"C{bp_summary_row}"] = "Total Business Partner Price USD:"
            ws[f"C{bp_summary_row}"].font = Font(bold=True, color="1F497D")
            ws[f"C{bp_summary_row}"].alignment = Alignment(horizontal="right")

            bp_total_formula = f"=SUM(K{data_start_row}:K{data_end_row})"
            ws[f"K{bp_summary_row}"] = bp_total_formula
            ws[f"K{bp_summary_row}"].number_format = '"USD"#,##0.00'
            ws[f"K{bp_summary_row}"].font = Font(bold=True, color="1F497D")
            ws[f"K{bp_summary_row}"].fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            # For terms placement later, ensure we start after the second summary row
            summary_row = bp_summary_row
        else:
            summary_row = start_row + 1

    else:
        # UAE summary
        if data:
            data_start_row = start_row
            data_end_row = start_row + len(data) - 1
            summary_row = start_row + len(data) + 1  # Immediately after the table

            # TOTAL Bid Discounted Price
            ws.merge_cells(f"C{summary_row}:G{summary_row}")
            ws[f"C{summary_row}"] = "TOTAL Bid Discounted Price"
            ws[f"C{summary_row}"].font = Font(bold=True, color="1F497D")
            ws[f"C{summary_row}"].alignment = Alignment(horizontal="right")

            total_formula = f"=SUM(J{data_start_row}:J{data_end_row})"
            ws[f"J{summary_row}"] = total_formula
            ws[f"J{summary_row}"].number_format = '"AED"#,##0.00'
            ws[f"J{summary_row}"].font = Font(bold=True, color="1F497D")
            ws[f"J{summary_row}"].fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

            # TOTAL BP Special Discounted Price excluding VAT
            bp_summary_row = summary_row + 1
            ws.merge_cells(f"C{bp_summary_row}:G{bp_summary_row}")
            ws[f"C{bp_summary_row}"] = "TOTAL BP Special Discounted Price excluding VAT:"
            ws[f"C{bp_summary_row}"].font = Font(bold=True, color="1F497D")
            ws[f"C{bp_summary_row}"].alignment = Alignment(horizontal="right")

            bp_total_formula = f"=SUM(L{data_start_row}:L{data_end_row})"
            ws[f"L{bp_summary_row}"] = bp_total_formula
            ws[f"L{bp_summary_row}"].number_format = '"AED"#,##0.00'
            ws[f"L{bp_summary_row}"].font = Font(bold=True, color="1F497D")
            ws[f"L{bp_summary_row}"].fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        else:
            # If no data, set summary_row to a safe default after headers
            summary_row = start_row + 1

    # --- Everything after the table: Terms block and IBM Terms ---
    # 1) Terms block from terms_template.py
    # NOTE: We pass a numeric total to the terms section. For Qatar, Cost (USD) is at index 6 in each row,
    # which should be a numeric value (not a formula). This keeps the terms consistent.
    total_price_sum = 0.0
    for r in data:
        try:
            val = r[6] if len(r) > 6 else 0  # Use "Cost" column for a stable numeric sum
            if isinstance(val, str):
                if val.strip() and val.strip().replace('.', '', 1).isdigit():
                    total_price_sum += float(val)
            elif isinstance(val, (int, float)):
                total_price_sum += float(val)
        except Exception:
            continue

    # Ensure country is in header_info for terms logic
    if country:
        header_info['country'] = country

    terms = get_terms_section(header_info, total_price_sum)

    # Place terms below the summary rows
    terms_start_row = summary_row + 3
    adjusted_terms = []
    row_offset = terms_start_row - 29  # 29 is the anchor in terms_template.py
    for cell_addr, text, *style in terms:
        try:
            if len(cell_addr) >= 2 and cell_addr[1:].isdigit():
                col_letter = cell_addr[0]
                original_row = int(cell_addr[1:])
                new_row = original_row + row_offset
                new_cell_addr = f"{col_letter}{new_row}"
                adjusted_terms.append((new_cell_addr, text, *style))
            else:
                adjusted_terms.append((cell_addr, text, *style))
        except Exception:
            adjusted_terms.append((cell_addr, text, *style))

    for cell_addr, text, *style in adjusted_terms:
        try:
            if len(cell_addr) >= 2 and cell_addr[1:].isdigit():
                row_num = int(cell_addr[1:])
                col_letter = cell_addr[0]
                merge_rows = style[0].get("merge_rows") if style else None
                end_row = row_num + (merge_rows - 1 if merge_rows else 0)
                is_bold_title = style and "bold" in style[0] and style[0].get("bold") is True
                if is_bold_title:
                    ws.merge_cells(f"{col_letter}{row_num}:E{end_row}")
                    ws.row_dimensions[row_num].height = 24
                else:
                    ws.merge_cells(f"{col_letter}{row_num}:H{end_row}")
                    # Set row height based on wrapped line count
                    line_count = estimate_line_count(str(text), max_chars_per_line=80)
                    total_height = max(18, line_count * 16)
                    ws.row_dimensions[row_num].height = total_height
                ws[cell_addr] = text
                ws[cell_addr].alignment = Alignment(wrap_text=True, vertical="top")
                if style and "bold" in style[0]:
                    ws[cell_addr].font = Font(**style[0])
        except Exception:
            pass

    # 2) IBM Terms header and paragraphs
    last_terms_row = max([
        int(addr[1:]) + (style[0].get("merge_rows", 1) - 1)
        for addr, text, *style in adjusted_terms
        if style and len(addr) >= 2 and addr[1:].isdigit()
    ], default=terms_start_row + 10)

    current_row = last_terms_row + 3
    ibm_header_cell = ws[f"C{current_row}"]
    ibm_header_cell.value = "IBM Terms and Conditions"
    ibm_header_cell.font = Font(bold=True, size=12, color="1F497D")
    current_row += 2

    paragraphs = [p.strip() for p in ibm_terms_text.split('\n\n') if p.strip()]
    for paragraph in paragraphs:
        if paragraph:
            ws.merge_cells(f"C{current_row}:H{current_row}")
            cell = ws[f"C{current_row}"]
            cell.value = paragraph
            cell.font = Font(size=10, color="000000")
            cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
            estimated_lines = max(2, len(paragraph) // 100 + 1)
            row_height = max(25, estimated_lines * 15)
            ws.row_dimensions[current_row].height = row_height
            current_row += 1
            if "Useful/Important web resources" in paragraph:
                current_row += 2

    # 3) Printing / page setup
    last_row = ws.max_row
    ws.print_area = f"A1:L{last_row}"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.15
    ws.page_margins.right = 0.15
    ws.page_margins.top = 0.25
    ws.page_margins.bottom = 0.25
    ws.page_margins.header = 0.15
    ws.page_margins.footer = 0.15
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.draft = False
    ws.page_setup.blackAndWhite = False
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    # Ensure recalculation on open (so Excel recomputes any formulas immediately)
    wb.calculation.fullCalcOnLoad = True

    # Save to the provided BytesIO
    wb.save(output)
    output.seek(0)


def parse_uploaded_excel(file_path):
    """
    Parses the uploaded Excel log file and extracts relevant data for the table.

    Args:
        file_path (str): Path to the uploaded log file.

    Returns:
        list: Parsed data in the format [SKU, Description, Quantity, Start Date, End Date, Cost].
    """
    parsed_data = []

    xls = pd.ExcelFile(file_path)

    # Ensure the file has at least two sheets
    if len(xls.sheet_names) < 2:
        raise ValueError("The uploaded Excel file does not have a second sheet.")

    # Parse the second sheet, skipping the first 9 rows
    df = xls.parse(xls.sheet_names[1], skiprows=9, header=None)  # Skip first 9 rows, no header

    # Manually assign column names
    expected_columns = [
        'Part number', 'Part description', 'Brand', 'Part type', 'SW Value Plus product group',
        'SW Value Plus terms', 'Quantity', 'Start date', 'End date', 'Prorate months',
        'Compressed coverage months', 'Renewal quote number', 'Item points', 'Entitled unit price',
        'Total points', 'Bid unit price', 'Entitled extended price', 'End user discount',
        'Bid extended price', 'BP discount', 'BP override discount', 'BP extended price',
        'Total line discount', 'Recomm. Reseller / Entitled price', 'Is BP discount QP?'
    ]

    if len(df.columns) >= len(expected_columns):
        df.columns = expected_columns[:len(df.columns)]
    else:
        raise ValueError("The Excel file does not have enough columns to match the expected structure.")

    # Filter out rows with missing or irrelevant part numbers
    df = df[df['Part number'].notna()]

    # Select relevant columns
    columns_to_keep = [
        'Part number', 'Part description', 'Quantity', 'Start date', 'End date',
        'Bid extended price'
    ]

    missing_columns = [col for col in columns_to_keep if col not in df.columns]
    if missing_columns:
        raise ValueError(f"The following required columns are missing: {missing_columns}")

    df = df[columns_to_keep]

    # Drop summary rows (e.g., "Total for Software Parts")
    df = df[~df['Part number'].str.contains("Total", na=False)]

    # Convert DataFrame to list of lists
    for _, row in df.iterrows():
        try:
            sku = row['Part number']
            description = row['Part description']
            quantity = int(row['Quantity'])
            start_date = row['Start date']
            end_date = row['End date']
            cost = float(row['Bid extended price'])
            parsed_data.append([sku, description, quantity, start_date, end_date, cost])
        except Exception:
            # Skip rows that cannot be parsed cleanly
            pass

    return parsed_data
