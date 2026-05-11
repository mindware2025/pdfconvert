from io import BytesIO
import openpyxl

def detect_dell_template(input_bytes: bytes) -> str:
    """
    Detect Dell template type.
    Returns:
      - 'extended_services'
      - 'standard_quote'
    """
    # PDFs always go to existing logic
    if input_bytes.lstrip().startswith(b"%PDF"):
        return "standard_quote"

    try:
        wb = openpyxl.load_workbook(BytesIO(input_bytes), data_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=1, max_row=80, max_col=10):
            for cell in row:
                if isinstance(cell.value, str) and \
                   "dell extended services details" in cell.value.lower():
                    return "extended_services"

        for row in ws.iter_rows(min_row=1, max_row=20, max_col=12):
            row_text = [str(cell.value).lower() if cell.value is not None else "" for cell in row]
            if "config" in row_text and "unit selling price" in row_text and "total selling price" in row_text:
                return "standard_quote"

        for sheet_name in wb.sheetnames:
            if isinstance(sheet_name, str) and sheet_name.strip().lower().replace(" ", "") in (
                "configuration", "config", "configsheet", "configurationsheet"
            ):
                return "standard_quote"
    except Exception:
        pass

    return "standard_quote"
