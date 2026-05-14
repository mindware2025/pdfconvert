from io import BytesIO
from pathlib import Path
from typing import Optional, List, Dict
from datetime import datetime
import os
import re
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XLImage

AED_RATE = 3.68


# ---------------- Helpers ----------------

def _text(v):
    return "" if v is None else str(v).strip()


def _sanitize_filename_part(value: str) -> str:
    text = " ".join(_text(value).split()).strip()
    if not text:
        return ""
    for ch in '<>:"/\\|?*':
        text = text.replace(ch, "")
    return text.rstrip(". ")


def _strip_trailing_asterisk(value: str) -> str:
    if value is None:
        return ""
    return re.sub(r"\s*\*+$", "", _text(value)).strip()


def _usd_to_aed(val):
    try:
        return round(float(val) * AED_RATE, 2)
    except Exception:
        return 0.0


def _to_number(v):
    try:
        if v is None or v == "":
            return 0.0
        if isinstance(v, str):
            return float(v.replace(",", ""))
        return float(v)
    except Exception:
        return 0.0


def _get_local_logo_path() -> Optional[str]:
    base_dir = Path(__file__).resolve().parent
    for name in ("dell.png", "dell copy.png"):
        candidate = base_dir / name
        if candidate.exists():
            return str(candidate)
    return None


def _add_logo(ws, logo_bytes: Optional[bytes], anchor: str = "A1", width: int = 780, height: int = 52):
    if logo_bytes:
        try:
            img = XLImage(BytesIO(logo_bytes))
            img.width = width
            img.height = height
            ws.add_image(img, anchor)
            ws.merge_cells("A1:H4")
            ws.row_dimensions[1].height = 40
            ws.row_dimensions[2].height = 25
            return
        except Exception:
            pass

    local_logo = _get_local_logo_path()
    if local_logo:
        try:
            img = XLImage(local_logo)
            img.width = width
            img.height = height
            ws.add_image(img, anchor)
            ws.merge_cells("A1:H4")
            ws.row_dimensions[1].height = 40
            ws.row_dimensions[2].height = 25
        except Exception:
            pass


# ---------------- Metadata ----------------

def _extract_metadata(ws) -> Dict[str, str]:
    meta = {"quote_no": "", "date": "", "end_user": ""}

    for r in range(1, 60):
        row_values = [_text(ws.cell(r, c).value) for c in range(1, 11)]
        row_lower = [v.lower() for v in row_values]
        row_text = " ".join(row_lower)

        if "quote #:" in row_text:
            for idx, cell_text in enumerate(row_lower):
                if "quote #:" in cell_text:
                    for value in row_values[idx + 1:]:
                        if value:
                            meta["quote_no"] = value
                            break
                    if not meta["quote_no"] and r + 1 <= ws.max_row:
                        next_row_values = [_text(ws.cell(r + 1, c).value) for c in range(1, 11)]
                        for value in next_row_values:
                            if value:
                                meta["quote_no"] = value
                                break
                    break

        if any(cell_text.startswith("date") for cell_text in row_lower):
            for idx, cell_text in enumerate(row_lower):
                if cell_text.startswith("date"):
                    for value in row_values[idx + 1:]:
                        if value:
                            meta["date"] = value
                            break
                    break

        if "end user -" in row_text:
            for idx, cell_text in enumerate(row_lower):
                if "end user -" in cell_text:
                    col = idx + 1
                    next_row = r + 1
                    while next_row <= min(ws.max_row, r + 8):
                        cell_value = _text(ws.cell(next_row, col).value)
                        if not cell_value:
                            break
                        cell_lower = cell_value.lower()
                        if any(marker in cell_lower for marker in ("dell extended services details", "customer information", "terms of sale")):
                            break
                        meta["end_user"] = cell_value.strip()
                        break
                    break
            break

    return meta


def build_dell_extended_services_output_filename(input_excel_bytes: bytes) -> str:
    """Build the download filename for the extended-services workbook."""
    quote_no = ""
    end_user = ""

    try:
        src_wb = openpyxl.load_workbook(BytesIO(input_excel_bytes), data_only=True)
        src_ws = src_wb.active
        meta = _extract_metadata(src_ws)
        quote_no = meta.get("quote_no", "")
        end_user = _strip_trailing_asterisk(meta.get("end_user", ""))
    except Exception:
        pass

    parts = [
        "Mindware costing",
        _sanitize_filename_part(quote_no),
        _sanitize_filename_part(end_user),
        datetime.now().strftime("%Y-%m-%d"),
    ]
    return "- ".join(parts) + ".xlsx"


def _aed_footer_notes() -> List[str]:
    return [
        "Ø  Payment terms will be as per our finance approval.",
        "Ø  These prices are till DDP Dubai.",
        "Ø  Hardware will take 4-12 weeks delivery time from the date of Booking.",
        "Ø  These prices do not include Mindware installation of any kind.",
        "Ø  Change in Qty or partial shipment is not acceptable.",
        "Ø  PO Should be addressed to Mindware Technology Trading LLC and should be in AED.",
        "Ø  For all B2B orders complete end customer details should be mentioned on the PO.",
        "Ø  Orders once placed with Dell cannot be cancelled.",
        "Ø  Kindly also ensure to review the proposal specifications from your end and ensure that they match the requirements exactly as per the End User.",
        "Ø  Partial deliveries shall be acceptable",
        "Ø  For UAE DDP orders, the PO should be addressed to Mindware Technology Trading LLC and for Ex-Jablal Ali orders, it should be addressed to Mindware FZ.",
        "Ø  Please ensure that the PO includes the name of the end-user.",
        "Ø  Please ensure that the PO includes the Incoterms (DDP or Ex-Works Jabal Ali).",
        "Ø  Due to global market fluctuations, all prices are subject to change without prior notice, and lead times may also be affected. All quotations are non-binding and remain subject to final validation and confirmation by Dell.",
        "Ø  As the geopolitical situation in the Middle East continues to evolve, it has introduced significant instability to international shipping routes. These unforeseen and extraordinary circumstances, which remain entirely beyond our control, constitute a Force Majeure event. We are formally notifying you of the resulting impact on our current and future shipments.",
    ]


# ---------------- Table ----------------

def _extract_extended_services_rows(ws) -> List[List]:
    rows = []
    start_row = None

    for r in range(1, ws.max_row + 1):
        row_text = " ".join(_text(ws.cell(r, c).value).lower() for c in range(1, 22))
        if "dell extended services details" in row_text:
            start_row = r
            break

    if not start_row:
        return rows

    header_row = None
    for r in range(start_row + 1, min(ws.max_row, start_row + 15) + 1):
        row_text = " ".join(_text(ws.cell(r, c).value).lower() for c in range(1, 22))
        if "asset" in row_text and "price after discount" in row_text:
            header_row = r
            break

    if not header_row:
        return rows

    target_headers = [
        "Asset", "Agreement ID", "Model", "Install At/Ship To",
        "Install At/Ship To City", "Install At/Ship To State",
        "Install At/Ship To Country", "LOB or Family", "Ship Date",
        "Service Contract Expiration", "Service Contract Description",
        "Services SKU", "New Contract Start Date", "New Contract End Date",
        "Quantity", "Price After Discount", "EOSS Date", "Product Type"
    ]

    header_cols = []
    for target in target_headers:
        found_col = 0
        target_lower = target.lower()
        for c in range(1, ws.max_column + 1):
            header_text = _text(ws.cell(header_row, c).value).lower()
            if not header_text:
                continue
            if target_lower in header_text or header_text in target_lower:
                found_col = c
                break
        header_cols.append(found_col)

    data_row = header_row + 1
    while data_row <= ws.max_row:
        row = [
            _text(ws.cell(data_row, c).value) if c else ""
            for c in header_cols
        ]
        if not any(row):
            data_row += 1
            continue

        if any("total" in cell.lower() for cell in row if cell):
            break

        # Keep USD value in column 16 for later AED conversion with margin.
        row[15] = _to_number(row[15])

        rows.append(row)
        data_row += 1

    return rows


# ---------------- Main Generator ----------------

def generate_dell_extended_services_quote(
    input_excel_bytes: bytes,
    logo_bytes: Optional[bytes] = None,
    margin_percent: float = 0.0,
) -> bytes:

    src_wb = openpyxl.load_workbook(BytesIO(input_excel_bytes), data_only=True)
    src_ws = src_wb.active

    meta = _extract_metadata(src_ws)
    meta = {k: _strip_trailing_asterisk(v) for k, v in meta.items()}
    rows = _extract_extended_services_rows(src_ws)

    wb = Workbook()
    ws = wb.active
    ws.title = "Quote"
    ws.sheet_view.showGridLines = False

    # ===== HEADER: use the full banner logo across A:H =====
    ws.merge_cells("A1:H2")
    _add_logo(ws, logo_bytes, anchor="A1", width=780, height=52)

    # Contact info (AED style)
    ws.merge_cells("A5:D5")
    ws.merge_cells("A6:D6")
    ws.merge_cells("A7:D7")
    ws["A5"] = "P O Box 55609, Dubai, UAE"
    ws["A6"] = "Tel :  +9714 4500600    Fax : +9714 4500678"
    ws["A7"] = "Website :  www.mindware.net"
    for cell in ("A5", "A6", "A7"):
        ws[cell].font = Font(bold=True, size=11, color="1F497D")
        ws[cell].alignment = Alignment(horizontal="left", vertical="center")

    # ===== METADATA SECTION (AED style) =====
    section_fill = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")
    ws.merge_cells("A8:D8")
    ws["A8"] = "Quote Summary"
    ws["A8"].font = Font(bold=True, color="1F497D")
    ws["A8"].alignment = Alignment(horizontal="left", vertical="center")
    ws["A8"].fill = section_fill
    ws["A8"].border = Border(
        left=Side(style="thin", color="9FBAD0"),
        right=Side(style="thin", color="9FBAD0"),
        top=Side(style="thin", color="9FBAD0"),
        bottom=Side(style="thin", color="9FBAD0"),
    )

    summary_rows = [
        (9, "Quote Ref", meta["quote_no"]),
        (10, "Date", meta["date"]),
    ]

    for row_idx, label, value in summary_rows:
        ws[f"A{row_idx}"] = label
        ws[f"A{row_idx}"].font = Font(bold=True, color="1F497D")
        ws[f"A{row_idx}"].alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=4)
        ws[f"B{row_idx}"] = value
        ws[f"B{row_idx}"].alignment = Alignment(horizontal="left", vertical="center")

    # Customer Information section
    customer_title_row = 12
    ws.merge_cells(start_row=customer_title_row, start_column=1, end_row=customer_title_row, end_column=8)
    ws[f"A{customer_title_row}"] = "Customer Information"
    ws[f"A{customer_title_row}"].font = Font(bold=True, color="1F497D")
    ws[f"A{customer_title_row}"].alignment = Alignment(horizontal="left", vertical="center")
    ws[f"A{customer_title_row}"].fill = section_fill
    ws[f"A{customer_title_row}"].border = Border(
        left=Side(style="thin", color="9FBAD0"),
        right=Side(style="thin", color="9FBAD0"),
        top=Side(style="thin", color="9FBAD0"),
        bottom=Side(style="thin", color="9FBAD0"),
    )

    meta_rows = [
        ("End User:", meta["end_user"]),
    ]

    for idx, (label, value) in enumerate(meta_rows, start=customer_title_row + 1):
        ws[f"A{idx}"] = label
        ws[f"A{idx}"].font = Font(bold=True)
        ws[f"A{idx}"].alignment = Alignment(horizontal="left", vertical="top")
        ws.merge_cells(start_row=idx, start_column=2, end_row=idx, end_column=8)
        ws[f"B{idx}"] = value
        ws[f"B{idx}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

        text_len = len(_text(value))
        estimated_lines = max(1, min(4, (text_len // 32) + 1))
        ws.row_dimensions[idx].height = max(ws.row_dimensions[idx].height or 20, estimated_lines * 18)

    # ===== TABLE HEADER =====
    header_row = customer_title_row + 2
    ws[f"A{header_row}"] = "Sr. No."
    ws[f"B{header_row}"] = "Part Number"
    ws[f"C{header_row}"] = "Description"
    ws[f"D{header_row}"] = "Qty"
    ws[f"E{header_row}"] = "Unit Price"
    ws[f"F{header_row}"] = "Total Price (excluding vat)"
    ws[f"G{header_row}"] = "Margin"

    header_fill = PatternFill(start_color="9BC2E6", end_color="9BC2E6", fill_type="solid")
    helper_header_fill = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
    header_font = Font(bold=True, color="000000")
    border_thin = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    header_cells = [f"A{header_row}", f"B{header_row}", f"C{header_row}", f"D{header_row}", f"E{header_row}", f"F{header_row}", f"G{header_row}"]
    for addr in header_cells:
        ws[addr].fill = helper_header_fill if addr == f"G{header_row}" else header_fill
        ws[addr].font = header_font
        ws[addr].alignment = Alignment(horizontal="center", vertical="center")
        ws[addr].border = border_thin
    ws.row_dimensions[header_row].height = 20

    # ===== DATA ROWS =====
    row_ptr = header_row + 1
    sr_no = 1
    currency_fmt = '"AED" #,##0.00'
    margin_fmt = '0.00%'
    yellow = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")
    helper_body_fill = PatternFill(start_color="FCE5E5", end_color="FCE5E5", fill_type="solid")
    total_cells = []

    for row in rows:
        # Extract data from the extended services row
        services_sku = row[11]  # Services SKU column
        service_desc = row[10]  # Service Contract Description column
        qty = _to_number(row[14])  # Quantity column
        price_usd = row[15]  # Price After Discount (USD) column

        if not service_desc or qty <= 0:
            continue

        ws[f"A{row_ptr}"] = sr_no
        ws[f"B{row_ptr}"] = services_sku
        ws[f"C{row_ptr}"] = service_desc
        ws[f"D{row_ptr}"] = qty

        # Unit Price with AED conversion and margin (like Standard Quote logic)
        ws[f"E{row_ptr}"].value = f"=ROUND(({price_usd}*{AED_RATE})/(1-{margin_percent}/100),2)"
        ws[f"E{row_ptr}"].number_format = currency_fmt

        # Total Price = Qty * Unit Price
        ws[f"F{row_ptr}"].value = f"=D{row_ptr}*E{row_ptr}"
        ws[f"F{row_ptr}"].number_format = currency_fmt

        # Margin column
        ws[f"G{row_ptr}"].value = margin_percent / 100.0
        ws[f"G{row_ptr}"].number_format = margin_fmt

        # Styling
        data_cells = [f"A{row_ptr}", f"B{row_ptr}", f"C{row_ptr}", f"D{row_ptr}", f"E{row_ptr}", f"F{row_ptr}", f"G{row_ptr}"]
        for addr in data_cells:
            ws[addr].fill = helper_body_fill if addr == f"G{row_ptr}" else yellow
            ws[addr].border = border_thin
            ws[addr].alignment = Alignment(horizontal="center", vertical="top")
        ws[f"C{row_ptr}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

        total_cells.append(f"F{row_ptr}")
        sr_no += 1
        row_ptr += 1

    # ===== TOTAL ROW =====
    if total_cells:
        total_row = row_ptr
        ws.merge_cells(start_row=total_row, start_column=2, end_row=total_row, end_column=5)
        ws[f"B{total_row}"] = "Total price"
        ws[f"B{total_row}"].alignment = Alignment(horizontal="right", vertical="center")
        ws[f"B{total_row}"].font = Font(bold=True, color="1F497D")

        ws[f"F{total_row}"] = f"=SUM({','.join(total_cells)})"
        ws[f"F{total_row}"].number_format = currency_fmt
        ws[f"F{total_row}"].font = Font(bold=True, color="1F497D")
        ws[f"F{total_row}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"F{total_row}"].border = border_thin
        ws[f"G{total_row}"].fill = helper_body_fill
        ws[f"G{total_row}"].border = border_thin

    # ===== FOOTER NOTES =====
    footer_notes = _aed_footer_notes()
    notes_title_row = (total_row + 2) if total_cells else (row_ptr + 2)
    ws.merge_cells(start_row=notes_title_row, start_column=1, end_row=notes_title_row, end_column=7)
    ws.cell(notes_title_row, 1).value = "Terms and Conditions"
    ws.cell(notes_title_row, 1).font = Font(bold=True, color="1F497D")
    ws.cell(notes_title_row, 1).alignment = Alignment(horizontal="left", vertical="center")

    notes_body_row = notes_title_row + 1
    ws.merge_cells(start_row=notes_body_row, start_column=1, end_row=notes_body_row, end_column=7)
    body_cell = ws.cell(notes_body_row, 1)
    body_cell.value = "\n".join(footer_notes)
    body_cell.alignment = Alignment(wrap_text=True, vertical="top")
    body_cell.border = border_thin
    ws.row_dimensions[notes_body_row].height = max(180, min(520, len(footer_notes) * 22))

    # Column widths
    widths = {
        "A": 8, "B": 15, "C": 50, "D": 8, "E": 15, "F": 18, "G": 12
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.column_dimensions["G"].hidden = False

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()
