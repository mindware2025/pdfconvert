from io import BytesIO
from pathlib import Path
from typing import Optional, List, Dict
from datetime import datetime
import os
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XLImage

AED_RATE = 3.68
CURRENCY_FMT_AED = '"AED" #,##0.00'


# ---------------- Helpers ----------------

def _text(v):
    return "" if v is None else str(v).strip()


def _sanitize_filename_part(value: str) -> str:
    text = " ".join(_text(value).split()).strip()
    if not text:
        return ""
    for ch in '<>:"/\\|?*':
        text = text.replace(ch, "")
    return text.rstrip(". ")


def _clean_party_name(value: str) -> str:
    text = " ".join(_text(value).split()).strip()
    if not text:
        return ""
    return text.split("*", 1)[0].strip()


def _usd_to_aed(val):
    try:
        return round(float(val) * AED_RATE, 2)
    except Exception:
        return 0.0


def _to_number(v):
    try:
        if v is None or v == "":
            return 0.0
        if isinstance(v, str):
            return float(v.replace(",", ""))
        return float(v)
    except Exception:
        return 0.0


def _get_local_logo_path() -> Optional[str]:
    base_dir = Path(__file__).resolve().parent
    for name in ("dell.png", "dell copy.png"):
        candidate = base_dir / name
        if candidate.exists():
            return str(candidate)
    return None


def _add_logo(ws, logo_bytes: Optional[bytes], anchor: str = "A1", width: int = 780, height: int = 52):
    if logo_bytes:
        try:
            img = XLImage(BytesIO(logo_bytes))
            img.width = width
            img.height = height
            ws.add_image(img, anchor)
            ws.merge_cells("A1:H4")
            ws.row_dimensions[1].height = 40
            ws.row_dimensions[2].height = 25
            return
        except Exception:
            pass

    local_logo = _get_local_logo_path()
    if local_logo:
        try:
            img = XLImage(local_logo)
            img.width = width
            img.height = height
            ws.add_image(img, anchor)
            ws.merge_cells("A1:H4")
            ws.row_dimensions[1].height = 40
            ws.row_dimensions[2].height = 25
        except Exception:
            pass


# ---------------- Metadata ----------------

def _extract_metadata(ws) -> Dict[str, str]:
    meta = {"quote_no": "", "date": "", "end_user": ""}

    for r in range(1, 60):
        row_values = [_text(ws.cell(r, c).value) for c in range(1, 11)]
        row_lower = [v.lower() for v in row_values]
        row_text = " ".join(row_lower)

        if "quote #:" in row_text:
            for idx, cell_text in enumerate(row_lower):
                if "quote #:" in cell_text:
                    for value in row_values[idx + 1:]:
                        if value:
                            meta["quote_no"] = value
                            break
                    if not meta["quote_no"] and r + 1 <= ws.max_row:
                        next_row_values = [_text(ws.cell(r + 1, c).value) for c in range(1, 11)]
                        for value in next_row_values:
                            if value:
                                meta["quote_no"] = value
                                break
                    break

        if any(cell_text.startswith("date") for cell_text in row_lower):
            for idx, cell_text in enumerate(row_lower):
                if cell_text.startswith("date"):
                    for value in row_values[idx + 1:]:
                        if value:
                            meta["date"] = value
                            break
                    break

        if "end user -" in row_text:
            for idx, cell_text in enumerate(row_lower):
                if "end user -" in cell_text:
                    next_row = r + 1
                    while next_row <= min(ws.max_row, r + 8):
                        line_parts = [_text(ws.cell(next_row, c).value) for c in range(1, 11)]
                        line_parts = [part for part in line_parts if part]
                        if not line_parts:
                            break
                        joined_line = " ".join(line_parts).strip()
                        joined_line_lower = joined_line.lower()
                        if any(marker in joined_line_lower for marker in ("dell extended services details", "customer information", "terms of sale")):
                            break
                        meta["end_user"] = joined_line
                        break
                    break
            break

    return meta


def build_dell_extended_services_output_filename(input_excel_bytes: bytes) -> str:
    """Build the download filename for the extended-services workbook."""
    quote_no = ""
    end_user = ""

    try:
        src_wb = openpyxl.load_workbook(BytesIO(input_excel_bytes), data_only=True)
        src_ws = src_wb.active
        meta = _extract_metadata(src_ws)
        quote_no = meta.get("quote_no", "")
        end_user = meta.get("end_user", "")
    except Exception:
        pass

    parts = [
        "Mindware costing",
        _sanitize_filename_part(quote_no),
        _sanitize_filename_part(end_user),
        datetime.now().strftime("%Y-%m-%d"),
    ]
    return "- ".join(parts) + ".xlsx"


def _aed_footer_notes() -> List[str]:
    return [
        "Ø  Payment terms will be as per our finance approval.",
        "Ø  These prices are till DDP Dubai.",
        "Ø  Hardware will take 4-12 weeks delivery time from the date of Booking.",
        "Ø  These prices do not include Mindware installation of any kind.",
        "Ø  Change in Qty or partial shipment is not acceptable.",
        "Ø  PO Should be addressed to Mindware Technology Trading LLC and should be in AED.",
        "Ø  For all B2B orders complete end customer details should be mentioned on the PO.",
        "Ø  Orders once placed with Dell cannot be cancelled.",
        "Ø  Kindly also ensure to review the proposal specifications from your end and ensure that they match the requirements exactly as per the End User.",
        "Ø  Partial deliveries shall be acceptable",
        "Ø  For UAE DDP orders, the PO should be addressed to Mindware Technology Trading LLC and for Ex-Jablal Ali orders, it should be addressed to Mindware FZ.",
        "Ø  Please ensure that the PO includes the name of the end-user.",
        "Ø  Please ensure that the PO includes the Incoterms (DDP or Ex-Works Jabal Ali).",
        "Ø  Due to global market fluctuations, all prices are subject to change without prior notice, and lead times may also be affected. All quotations are non-binding and remain subject to final validation and confirmation by Dell.",
        "Ø  As the geopolitical situation in the Middle East continues to evolve, it has introduced significant instability to international shipping routes. These unforeseen and extraordinary circumstances, which remain entirely beyond our control, constitute a Force Majeure event. We are formally notifying you of the resulting impact on our current and future shipments.",
    ]


# ---------------- Table ----------------

def _extract_extended_services_rows(ws) -> List[List]:
    rows = []
    start_row = None

    for r in range(1, ws.max_row + 1):
        row_text = " ".join(_text(ws.cell(r, c).value).lower() for c in range(1, 22))
        if "dell extended services details" in row_text:
            start_row = r
            break

    if not start_row:
        return rows

    header_row = None
    for r in range(start_row + 1, min(ws.max_row, start_row + 15) + 1):
        row_text = " ".join(_text(ws.cell(r, c).value).lower() for c in range(1, 22))
        if "asset" in row_text and "price after discount" in row_text:
            header_row = r
            break

    if not header_row:
        return rows

    target_headers = [
        "Asset", "Agreement ID", "Model", "Install At/Ship To",
        "Install At/Ship To City", "Install At/Ship To State",
        "Install At/Ship To Country", "LOB or Family", "Ship Date",
        "Service Contract Expiration", "Service Contract Description",
        "Services SKU", "New Contract Start Date", "New Contract End Date",
        "Quantity", "Price After Discount", "EOSS Date", "Product Type"
    ]

    header_cols = []
    for target in target_headers:
        found_col = 0
        target_lower = target.lower()
        for c in range(1, ws.max_column + 1):
            header_text = _text(ws.cell(header_row, c).value).lower()
            if not header_text:
                continue
            if target_lower in header_text or header_text in target_lower:
                found_col = c
                break
        header_cols.append(found_col)

    data_row = header_row + 1
    while data_row <= ws.max_row:
        row = [
            _text(ws.cell(data_row, c).value) if c else ""
            for c in header_cols
        ]
        if not any(row):
            data_row += 1
            continue

        if any("total" in cell.lower() for cell in row if cell):
            break

        # Keep USD value in column 16 for later AED conversion with margin.
        row[15] = _to_number(row[15])

        rows.append(row)
        data_row += 1

    return rows


# ---------------- Main Generator ----------------

def generate_dell_extended_services_quote(
    input_excel_bytes: bytes,
    logo_bytes: Optional[bytes] = None,
    margin_percent: float = 0.0,
) -> bytes:

    src_wb = openpyxl.load_workbook(BytesIO(input_excel_bytes), data_only=True)
    src_ws = src_wb.active

    meta = _extract_metadata(src_ws)
    rows = _extract_extended_services_rows(src_ws)
    footer_notes = _aed_footer_notes()

    wb = Workbook()
    ws = wb.active
    ws.title = "Quote"
    ws.sheet_view.showGridLines = False
    _add_logo(ws, logo_bytes)

    widths = {
        "A": 14,
        "B": 26,
        "C": 56,
        "D": 10,
        "E": 15,
        "F": 24,
        "G": 11,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for rr in range(1, 3):
        ws.row_dimensions[rr].height = 28
    ws.row_dimensions[3].height = 12
    for rr in range(5, 18):
        ws.row_dimensions[rr].height = 20

    ws.merge_cells("A5:D5")
    ws.merge_cells("A6:D6")
    ws.merge_cells("A7:D7")
    ws["A5"] = "P O Box 55609, Dubai, UAE"
    ws["A6"] = "Tel :  +9714 4500600    Fax : +9714 4500678"
    ws["A7"] = "Website :  www.mindware.ae"
    for cell in ("A5", "A6", "A7"):
        ws[cell].font = Font(bold=True, size=11, color="1F497D")
        ws[cell].alignment = Alignment(horizontal="left", vertical="center")

    border_thin = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    table_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    # ----- Header -----
    meta_rows = [
        ("Quote No:", meta["quote_no"]),
        ("Date:", meta["date"]),
        ("Quote Validity:", "30 days"),
        ("End User:", _clean_party_name(meta["end_user"])),
    ]

    meta_start_row = 9
    for offset, (label, value) in enumerate(meta_rows):
        row_idx = meta_start_row + offset
        ws[f"A{row_idx}"] = label
        ws[f"B{row_idx}"] = value
        ws[f"A{row_idx}"].font = Font(bold=True, color="1F497D")
        ws[f"A{row_idx}"].alignment = Alignment(horizontal="left", vertical="center")
        ws[f"A{row_idx}"].fill = header_fill
        ws[f"A{row_idx}"].border = border_thin
        ws[f"B{row_idx}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws[f"B{row_idx}"].border = border_thin

        value_len = len(_text(value))
        estimated_lines = max(1, min(4, (value_len // 34) + 1))
        ws.row_dimensions[row_idx].height = max(ws.row_dimensions[row_idx].height or 20, estimated_lines * 18)

    ws.merge_cells("D14:G14")
    ws["D14"] = "Quote Details"
    ws["D14"].font = Font(bold=True, size=12, color="1F497D")
    ws["D14"].alignment = Alignment(horizontal="center", vertical="center")

    # ----- Table -----
    headers = [
        "Sr. No.",
        "Part Number",
        "Description",
        "Qty",
        "Unit Price",
        "Total Price (excluding vat)",
        "Margin",
    ]

    start = 15

    for c, h in enumerate(headers, 1):
        cell = ws.cell(start, c)
        cell.value = h
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = header_fill
        cell.border = border_thin

    r = start + 1
    total_cells = []
    for idx, row in enumerate(rows, start=1):
        services_sku = row[11]
        description = row[10]
        qty_value = _to_number(row[14])
        price_usd = _to_number(row[15])

        ws[f"A{r}"] = idx
        ws[f"B{r}"] = services_sku
        ws[f"C{r}"] = description
        ws[f"D{r}"] = qty_value
        ws[f"G{r}"] = (margin_percent or 0.0) / 100.0
        ws[f"G{r}"].number_format = '0.00%'
        ws[f"E{r}"] = f'=ROUND(({price_usd}*{AED_RATE})/(1-G{r}),2)'
        ws[f"E{r}"].number_format = CURRENCY_FMT_AED
        ws[f"F{r}"] = f"=ROUND(E{r}*D{r},2)"
        ws[f"F{r}"].number_format = CURRENCY_FMT_AED

        for col in ("A", "B", "C", "D", "E", "F", "G"):
            ws[f"{col}{r}"].border = border_thin
            ws[f"{col}{r}"].fill = table_fill
            ws[f"{col}{r}"].alignment = Alignment(horizontal="center", vertical="top")
        ws[f"C{r}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        total_cells.append(f"F{r}")
        r += 1

    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
    ws[f"C{r}"] = "Total price"
    ws[f"C{r}"].font = Font(bold=True, color="1F497D")
    ws[f"C{r}"].alignment = Alignment(horizontal="right", vertical="center")
    ws[f"F{r}"] = f"=SUM({','.join(total_cells)})" if total_cells else 0
    ws[f"F{r}"].number_format = CURRENCY_FMT_AED
    ws[f"F{r}"].font = Font(bold=True, color="1F497D")
    ws[f"F{r}"].alignment = Alignment(horizontal="center", vertical="center")
    for col in ("C", "D", "E", "F"):
        ws[f"{col}{r}"].border = border_thin

    notes_title_row = r + 2
    ws.merge_cells(start_row=notes_title_row, start_column=2, end_row=notes_title_row, end_column=6)
    ws.cell(notes_title_row, 2).value = "Terms and Conditions"
    ws.cell(notes_title_row, 2).font = Font(bold=True, color="1F497D")

    notes_body_row = notes_title_row + 1
    ws.merge_cells(start_row=notes_body_row, start_column=2, end_row=notes_body_row, end_column=6)
    body_cell = ws.cell(notes_body_row, 2)
    body_cell.value = "\n".join(footer_notes)
    body_cell.alignment = Alignment(wrap_text=True, vertical="top")
    body_cell.border = border_thin
    ws.row_dimensions[notes_body_row].height = max(180, min(520, len(footer_notes) * 22))

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()
