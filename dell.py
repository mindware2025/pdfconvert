# dell.py
from datetime import datetime
from io import BytesIO
from typing import Optional, Dict, List, Tuple
import logging
import os
import re
from logging.handlers import RotatingFileHandler

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string as colidx
from openpyxl.drawing.image import Image as XLImage


# ----------------- Logging -----------------
_LOG_FILE = os.path.join(os.path.dirname(__file__), "dell_quote.log")


def _get_logger():
    """Get a logger that writes to a rotating log file."""
    logger = logging.getLogger("dell_quote")
    if not logger.handlers:
        handler = RotatingFileHandler(_LOG_FILE, maxBytes=2_000_000, backupCount=3, encoding="utf-8")
        handler.setFormatter(logging.Formatter("%(asctime)s %(levelname)s %(message)s"))
        logger.addHandler(handler)
        logger.setLevel(logging.DEBUG)
    return logger


def _log_items(prefix: str, items: list, max_items: int = 20):
    """Log a summary of extracted items (truncates to avoid huge logs)."""
    logger = _get_logger()
    count = len(items) if items is not None else 0
    logger.debug("%s: %d items", prefix, count)
    if not items:
        return
    for i, item in enumerate(items[:max_items]):
        logger.debug("%s item %d: %s", prefix, i + 1, item)
    if count > max_items:
        logger.debug("%s: ... (skipping remaining %d items)", prefix, count - max_items)

try:
    from PIL import Image as PILImage
    from PIL import ImageChops
except Exception:
    PILImage = None
    ImageChops = None


# ================= Helpers =================

def _parse_money(val):
    """Parse strings like '$ 902.00', '902,00', '36,080.00' to float."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    s = re.sub(r"[^\d,.\-]", "", s)  # remove currency & spaces
    if "," in s and "." in s:
        # treat comma as thousands
        s = s.replace(",", "")
    else:
        # if only comma present, treat as decimal
        if "," in s and "." not in s:
            s = s.replace(",", ".")
    try:
        return float(s)
    except Exception:
        return None


def _cell_to_text(v, fallback=""):
    if v is None:
        return fallback
    if isinstance(v, datetime):
        return v.strftime("%d/%m/%Y")
    return str(v).strip()


def _normalize_text(s: str) -> str:
    """Lowercase alnum-only text used for fuzzy matching."""
    return re.sub(r"[^a-z0-9]", "", s.lower()) if s else ""


def _make_item_key(s: str, key_len: int = 70) -> str:
    return _normalize_text(s)[:key_len]


def _sanitize_filename_part(value: str) -> str:
    """Return a filesystem-safe filename segment."""
    text = re.sub(r"\s+", " ", _cell_to_text(value)).strip()
    if not text:
        return ""
    text = re.sub(r'[<>:"/\\|?*]', "", text)
    text = text.rstrip(". ")
    return text


def build_dell_output_filename(input_excel_bytes: bytes, currency_code: str = "USD") -> str:
    """Build the download filename for the generated Dell workbook."""
    currency_code = (currency_code or "USD").upper()
    if currency_code != "AED":
        return f"Dell_Quotation_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

    quote_ref_text = ""
    quote_meta: Dict[str, str] = {}

    try:
        is_pdf = input_excel_bytes.lstrip().startswith(b"%PDF")
        if is_pdf:
            _, quote_meta, _, quote_ref_text, _, _ = _extract_pdf_quote_data(input_excel_bytes)
        else:
            src_wb = openpyxl.load_workbook(BytesIO(input_excel_bytes), data_only=True)
            src_ws = src_wb.active
            quote_ref_text, _ = _extract_metadata_strict(src_ws)
            if (
                not quote_ref_text
                or "$" in quote_ref_text
                or not re.search(r"\d{6,}", quote_ref_text)
            ):
                fb_ref, _ = _extract_metadata_excel_fallback(src_ws)
                if fb_ref:
                    quote_ref_text = fb_ref
            quote_meta = _extract_quote_metadata(src_ws)
    except Exception:
        quote_ref_text = ""
        quote_meta = {}

    party_name = (
        quote_meta.get("end user", "")
        or quote_meta.get("reseller", "")
        or ""
    )

    parts = [
        "Mindware costing",
        _sanitize_filename_part(quote_ref_text),
        _sanitize_filename_part(party_name),
        datetime.now().strftime("%Y-%m-%d"),
    ]
    return "- ".join(parts) + ".xlsx"




def _row_text(ws, r, c1=1, c2=None) -> str:
    if c2 is None:
        c2 = ws.max_column
    parts = []
    for c in range(c1, c2 + 1):
        txt = _cell_to_text(ws.cell(r, c).value)
        if txt:
            parts.append(txt)
    return " ".join(parts).strip()


def _is_price_or_qty_line(text: str) -> bool:
    t = text.lower()
    if any(tok in t for tok in [
        "qty", "quantity", "unit price", "unitprice", "subtotal", "total",
        "price", "amount", "discount", "tax", "extended price", "net price",
        "grand total", "msrp", "usd", "aed", "eur", "sar"
    ]):
        return True
    if re.search(r"(\$|€|£|د\.إ|aed|usd|eur|sar)", t, flags=re.IGNORECASE):
        return True
    return False


def _trim_logo_image(pil_img):
    """Trim transparent/blank padding so logos render at a useful visible size."""
    if PILImage is None:
        return pil_img

    img = pil_img.convert("RGBA")

    # Prefer alpha-based trimming for transparent logos.
    alpha = img.getchannel("A")
    bbox = alpha.getbbox()
    if bbox:
        img = img.crop(bbox)

    # Also trim white padding when present.
    if ImageChops is not None:
        bg = PILImage.new("RGBA", img.size, (255, 255, 255, 0))
        diff = ImageChops.difference(img, bg)
        bbox = diff.getbbox()
        if bbox:
            img = img.crop(bbox)

    return img


def _pil_to_xl_image(pil_img):
    """Convert a Pillow image into an openpyxl image safely."""
    buf = BytesIO()
    pil_img.save(buf, format="PNG")
    buf.seek(0)
    return XLImage(buf)


def _get_local_logo_path() -> Optional[str]:
    """Return the first available local logo path."""
    for path in ("dell copy.png", "dell.png"):
        if os.path.exists(path):
            return path
    return None

def _extract_all_excel_quote_refs(ws, max_rows: int = 80):
    """Return all quote reference values found in the worksheet."""
    refs = []
    max_row = min(ws.max_row, max_rows)
    max_col = min(ws.max_column, 10)
    quote_ref_pattern = r"\b\d{6,}(?:\.[A-Za-z0-9]+)?[A-Za-z0-9\-]*\b"

    for r in range(1, max_row + 1):
        row_texts = [_cell_to_text(ws.cell(r, c).value).strip() for c in range(1, max_col + 1)]
        for idx, cell_text in enumerate(row_texts, start=1):
            low = cell_text.lower()
            if not low:
                continue

            if low.startswith("quote") and "quoted on" not in low:
                # Quote label may include the quote number in the same cell.
                match = re.search(quote_ref_pattern, cell_text)
                if match:
                    refs.append(match.group(0))
                    continue

                # Or the quote number may be in the next non-empty cell.
                for next_text in row_texts[idx:]:
                    if not next_text:
                        continue
                    match = re.search(quote_ref_pattern, next_text)
                    if match:
                        refs.append(match.group(0))
                        break
                continue

            if any(token in low for token in ("quote no", "quote number", "quote ref")):
                match = re.search(quote_ref_pattern, cell_text)
                if match:
                    refs.append(match.group(0))
                    continue
                for next_text in row_texts[idx:]:
                    if not next_text:
                        continue
                    match = re.search(quote_ref_pattern, next_text)
                    if match:
                        refs.append(match.group(0))
                        break

    unique_refs = []
    for ref in refs:
        normalized_ref = str(ref).strip()
        if normalized_ref and normalized_ref not in unique_refs:
            unique_refs.append(normalized_ref)
    return unique_refs


def _extract_metadata_excel_fallback(ws):
    """
    Fallback for WEB / PDF-style Dell quotes saved as Excel.
    Finds 'Quote No' and 'Quoted On' by scanning labels instead of fixed cells.
    """
    quote_refs = _extract_all_excel_quote_refs(ws)
    quote_date = ""

    for row in ws.iter_rows(min_row=1, max_row=80, max_col=10):
        row_values = [_cell_to_text(c.value) for c in row if c.value]
        row_text = " ".join(row_values).lower()

        # ---- Quote Date ----
        if "quoted on" in row_text or "quote date" in row_text:
            for cell in row:
                m = re.search(r"\d{2}/\d{2}/\d{4}", str(cell.value))
                if m:
                    quote_date = m.group(0)

    quote_ref = ", ".join(quote_refs)
    return quote_ref, quote_date


def _extract_expiry_date(ws) -> str:
    """Extract 'Expires By' by checking near the quote block first, then scanning the broader header area."""
    def _format_date_value(value) -> str:
        if isinstance(value, datetime):
            return value.strftime("%d/%m/%Y")
        if value in (None, ""):
            return ""
        text = str(value).strip()
        match = re.search(r"\d{2}/\d{2}/\d{4}", text)
        return match.group(0) if match else text

    def _scan_row(row_idx: int) -> str:
        if row_idx < 1 or row_idx > ws.max_row:
            return ""

        max_col = min(ws.max_column, 10)
        row_values = [_cell_to_text(ws.cell(row_idx, c).value) for c in range(1, max_col + 1)]
        row_text = " ".join(v for v in row_values if v).lower()
        if "expires by" not in row_text:
            return ""

        for c in range(1, max_col + 1):
            cell_value = ws.cell(row_idx, c).value
            cell_text = _cell_to_text(cell_value).lower()
            if "expires by" not in cell_text:
                continue

            same_cell = _format_date_value(cell_value)
            if re.search(r"\d{2}/\d{2}/\d{4}", same_cell):
                return same_cell

            for next_c in range(c + 1, max_col + 1):
                candidate = _format_date_value(ws.cell(row_idx, next_c).value)
                if candidate:
                    return candidate

        for value in row_values:
            match = re.search(r"\d{2}/\d{2}/\d{4}", value)
            if match:
                return match.group(0)

        return ""

    direct_expiry = _format_date_value(ws["E19"].value)
    if direct_expiry:
        return direct_expiry

    quote_date_row = 18
    for row_idx in range(max(1, quote_date_row - 4), min(ws.max_row, quote_date_row + 4) + 1):
        expiry = _scan_row(row_idx)
        if expiry:
            return expiry

    for row_idx in range(1, min(ws.max_row, 80) + 1):
        expiry = _scan_row(row_idx)
        if expiry:
            return expiry

    return ""


def _add_logo(ws, logo_bytes: Optional[bytes], anchor="A1", width: int = 180, height: int = 60):
    """Add logo from uploaded bytes or fallback to a local logo file."""
    if PILImage is not None:
        if logo_bytes:
            try:
                pil_img = _trim_logo_image(PILImage.open(BytesIO(logo_bytes)))
                img = _pil_to_xl_image(pil_img)
                img.width = width
                img.height = height
                ws.add_image(img, anchor)
                return
            except Exception:
                pass
        local_logo = _get_local_logo_path()
        if local_logo:
            try:
                pil_img = _trim_logo_image(PILImage.open(local_logo))
                img = _pil_to_xl_image(pil_img)
                img.width = width
                img.height = height
                ws.add_image(img, anchor)
                return
            except Exception:
                pass
    local_logo = _get_local_logo_path()
    if local_logo:
        try:
            img = XLImage(local_logo)
            img.width = width
            img.height = height
            ws.add_image(img, anchor)
        except Exception:
            pass


def _add_static_logo(ws, image_path: str, anchor="A1", width: int = 120, height: int = 60):
    """Add a local logo file when present."""
    if PILImage is not None:
        try:
            pil_img = _trim_logo_image(PILImage.open(image_path))
            img = _pil_to_xl_image(pil_img)
            img.width = width
            img.height = height
            ws.add_image(img, anchor)
            return
        except Exception:
            pass
    try:
        img = XLImage(image_path)
        img.width = width
        img.height = height
        ws.add_image(img, anchor)
    except Exception:
        pass


def _extract_metadata_strict(ws):
    """Extract quote ref/date from strict positions in the worksheet."""
    logger = _get_logger()
    raw_ref = ws["E15"].value
    quote_ref = "" if raw_ref is None else (
        raw_ref.strftime("%d/%m/%Y") if isinstance(raw_ref, datetime) else str(raw_ref).strip()
    )

    raw_date = ws["E18"].value
    if isinstance(raw_date, datetime):
        quote_date = raw_date.strftime("%d/%m/%Y")
    else:
        quote_date = "" if raw_date is None else str(raw_date).strip()

    quote_refs = _extract_all_excel_quote_refs(ws)
    if quote_refs:
        unique_refs = []
        if quote_ref:
            unique_refs.append(quote_ref)
        for ref in quote_refs:
            if ref not in unique_refs:
                unique_refs.append(ref)
        quote_ref = ", ".join(unique_refs)

    logger.debug("_extract_metadata_strict: quote_ref=%s, quote_date=%s", quote_ref, quote_date)
    return quote_ref, quote_date


def _find_header_row_strict_or_detect(ws):
    """
    Try to detect header in the source (within first 40 rows) by locating a row containing
    both 'Description' and 'Qty/Quantity'. If found, return the first data row (header+1)
    and the column indices for description/qty/unit price.

    Fallback (for INPUT): assume header is on row 7 and data starts on row 8 with
    columns C/D/E for Description/Qty/Unit Price.

    Returns:
      (first_data_row_index, desc_col, qty_col, unit_col)
    """
    for r in range(1, min(ws.max_row, 40) + 1):
        row_vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if not any(row_vals):
            continue
        row_text = [(_cell_to_text(v).lower()) for v in row_vals]
        if any("description" in t for t in row_text) and any(("qty" in t) or ("quantity" in t) for t in row_text):
            desc_idx = qty_idx = unit_idx = None
            for i, v in enumerate(row_vals, start=1):
                name = _cell_to_text(v).lower()
                if desc_idx is None and "description" in name:
                    desc_idx = i
                if qty_idx is None and (("qty" in name) or ("quantity" in name)):
                    qty_idx = i
                if unit_idx is None and (("unit price" in name) or ("unitprice" in name) or (name == "price")):
                    unit_idx = i
            if desc_idx is None: desc_idx = 3
            if qty_idx  is None: qty_idx  = 4
            if unit_idx is None: unit_idx = 5
            return r + 1, desc_idx, qty_idx, unit_idx
    return 8, 3, 4, 5


# ---- Pricing Summary (Sheet 1 items) ----

def _locate_pricing_summary(ws):
    B = colidx('B')
    for r in range(30, min(ws.max_row, 120) + 1):
        v = ws.cell(r, B).value
        if v and "pricing" in str(v).lower() and "summary" in str(v).lower():
            header_row = r + 1
            start_row = header_row + 2
            return header_row, start_row
    return None


def _try_extract_items_from_pricing_summary(ws):
    logger = _get_logger()
    located = _locate_pricing_summary(ws)
    if not located:
        logger.debug("_try_extract_items_from_pricing_summary: pricing summary not found")
        return None

    header_row, start_row = located
    logger.debug("_try_extract_items_from_pricing_summary: header_row=%d start_row=%d", header_row, start_row)
    A, B, K, L, N = (colidx('A'), colidx('B'), colidx('K'), colidx('L'), colidx('N'))

    items = []
    r = start_row
    while r <= ws.max_row:
        sr = ws.cell(r, A).value
        if sr is None or str(sr).strip() == "":
            break
        if not re.match(r'^\d+', str(sr).strip()):
            break

        desc_text = _cell_to_text(ws.cell(r, B).value)
        if not desc_text:
            break

        qty_raw  = ws.cell(r, K).value
        unit_raw = ws.cell(r, L).value
        sub_raw  = ws.cell(r, N).value

        try:
            qty_val = int(_parse_money(qty_raw) or 0)
        except Exception:
            qty_val = 0
        unit_val = _parse_money(unit_raw) or 0.0
        subtotal_val = _parse_money(sub_raw)
        if subtotal_val is None:
            subtotal_val = qty_val * unit_val

        if (qty_val <= 0) and (unit_val == 0.0) and (subtotal_val is None or subtotal_val == 0.0):
            break

        items.append((desc_text, qty_val, unit_val, subtotal_val))
        r += 1

    if items:
        logger.debug("_try_extract_items_from_pricing_summary: extracted %d items", len(items))
        _log_items("Pricing summary", items)
    else:
        logger.debug("_try_extract_items_from_pricing_summary: no items extracted")

    return items if items else None


# ---- Product Details -> Configuration table (Sheet 2) ----

def _find_product_details_anchor(ws) -> Optional[int]:
    
    max_c = min(ws.max_column, 40)
    for r in range(1, ws.max_row + 1):
        for c in range(1, max_c + 1):
            v = ws.cell(r, c).value
            if v and "product details" in str(v).lower():
                return r
    return None


def _find_config_table_header(ws, start_row: int, search_rows: int = 30) -> Optional[Tuple[int, Dict[str, int]]]:
    """Find the row that contains the header like 'Module | Description | SKU | Tax Type | Qty'.
    Returns (header_row, columns_map). columns_map keys: module, description, sku, tax, qty (optional if missing).
    """
    last_row = min(ws.max_row, start_row + search_rows)
    for r in range(start_row, last_row + 1):
        labels = {}
        for c in range(1, ws.max_column + 1):
            name = _cell_to_text(ws.cell(r, c).value).lower()
            if not name:
                continue
            if 'module' in name and 'module' not in labels:
                labels['module'] = c
            if 'description' in name and 'description' not in labels:
                labels['description'] = c
            normalized_name = re.sub(r"\s+", " ", name.strip())
            if (
                normalized_name in ('sku', 'part', 'part #', 'part#', 'part number', 'part no', 'part no.', 'dell part number')
                or ('sku' in normalized_name)
                or ('part' in normalized_name and 'number' in normalized_name)
            ) and 'sku' not in labels:
                labels['sku'] = c
            if 'tax' in name and 'type' in name and 'tax' not in labels:
                labels['tax'] = c
            if name.strip() in ('qty', 'quantity') and 'qty' not in labels:
                labels['qty'] = c
        # Some Dell exports omit the "Module" column and start directly with Description.
        if all(k in labels for k in ('description', 'sku')):
            labels.setdefault('module', labels['description'])
            return r, labels
    return None


def _collect_config_rows_for_product(ws, start_row: int, columns: Dict[str, int], next_product_start: Optional[int]) -> List[Tuple[str, str, str, str]]:
    """Collect configuration lines for a product starting at start_row (first line AFTER header).
    Stops before next_product_start if provided, otherwise when a strong stop condition occurs.
    Returns list of tuples (module, description, sku, tax_type). Qty/Price are intentionally ignored.
    """
    rows = []
    r = start_row
    while r <= ws.max_row and (next_product_start is None or r < next_product_start):
        # a blank row ends the table chunk
        txt_whole = _row_text(ws, r, 1, min(ws.max_column, 40))
        if not txt_whole:
            break
        # sometimes date or separator lines appear – skip those
        if _is_price_or_qty_line(txt_whole) or re.search(r"estimated delivery|ship|subtotal|total", txt_whole, re.I):
            r += 1
            continue
        m = _cell_to_text(ws.cell(r, columns.get('module', 0)).value)
        d = _cell_to_text(ws.cell(r, columns.get('description', 0)).value)
        s = _cell_to_text(ws.cell(r, columns.get('sku', 0)).value)
        t = _cell_to_text(ws.cell(r, columns.get('tax', 0)).value)
        # If the row is essentially empty, stop
        if not any([m, d, s, t]):
            break
        rows.append((m, d, s, t))
        r += 1
    return rows


def _extract_quote_metadata(ws):
    """Extract quote metadata (Company Name, Customer Name, etc.) from the input sheet.

    Dell quote layout puts labels in column B and values in column E, e.g.:
        B22: "Company Name:"   E22: "ACME"
    """
    keys = {
        "company name": "Company Name",
        "customer name": "Customer Name",
        "customer number": "Customer Number",
        "end user": "End User",
        "reseller": "Reseller",
    }

    out = {k: "" for k in keys}
    max_row = min(ws.max_row, 120)
    for r in range(1, max_row + 1):
        label = _cell_to_text(ws.cell(r, 2).value).strip().lower().rstrip(":")
        if label in keys:
            out[label] = _cell_to_text(ws.cell(r, 5).value)
    return out


def _normalize_sheet_name(name: str) -> str:
    return re.sub(r"[^a-z0-9]", "", (name or "").lower().strip())


def _is_configuration_sheet_name(name: str) -> bool:
    normalized = _normalize_sheet_name(name)
    return normalized in ("configuration", "config", "configsheet", "configurationsheet")


def _find_configuration_sheet(wb):
    for sheet_name in wb.sheetnames:
        if _is_configuration_sheet_name(sheet_name):
            return wb[sheet_name]
    return None


def _extract_config_rows_from_configuration_sheet(ws) -> List[Tuple[str, str, str, str, str, str, str]]:
    header_info = _find_config_table_header(ws, 1, search_rows=50)
    if not header_info:
        return []

    header_row, colmap = header_info
    item_col = None
    for c in range(1, ws.max_column + 1):
        header_text = _cell_to_text(ws.cell(header_row, c).value).lower()
        if header_text in (
            "item", "item#", "item #", "item no", "item number", "sr. no.", "sr no", "srno", "sr"
        ):
            item_col = c
            break

    rows = []
    current_item = "1"
    has_real_module_col = bool(colmap.get("module")) and colmap.get("module") != colmap.get("description")
    for r in range(header_row + 1, ws.max_row + 1):
        row_text = _row_text(ws, r, 1, ws.max_column)
        if not row_text:
            continue

        if item_col:
            item_value = _cell_to_text(ws.cell(r, item_col).value).strip()
            if item_value:
                current_item = item_value.rstrip(".")

        module = _cell_to_text(ws.cell(r, colmap.get("module", 0)).value) if has_real_module_col else ""
        description = _cell_to_text(ws.cell(r, colmap.get("description", 0)).value)
        sku = _cell_to_text(ws.cell(r, colmap.get("sku", 0)).value)
        tax = _cell_to_text(ws.cell(r, colmap.get("tax", 0)).value)
        qty = _cell_to_text(ws.cell(r, colmap.get("qty", 0)).value)

        if not has_real_module_col and description and not any([sku, tax, qty]):
            module = description
            description = ""

        if not any([module, description, sku, tax, qty]):
            continue

        rows.append((current_item, "", module, description, sku, tax, qty))

    return rows


def _find_grouped_config_header(ws):
    for r in range(1, min(ws.max_row, 20) + 1):
        row_text = [(_cell_to_text(ws.cell(r, c).value).lower()) for c in range(1, ws.max_column + 1)]
        if "config" in row_text and "unit selling price" in row_text and "total selling price" in row_text:
            columns = {}
            for c in range(1, ws.max_column + 1):
                name = _cell_to_text(ws.cell(r, c).value).strip().lower()
                if "description" in name and "description" not in columns:
                    columns["description"] = c
                if "sku" in name and "sku" not in columns:
                    columns["sku"] = c
                if name in ("q-ty", "qty", "quantity") and "qty" not in columns:
                    columns["qty"] = c
                if "unit selling price" in name and "unit" not in columns:
                    columns["unit"] = c
                if "total selling price" in name and "total" not in columns:
                    columns["total"] = c
            if "description" in columns and "sku" in columns:
                return r, columns
    return None


def _is_grouped_config_template(ws) -> bool:
    return _find_grouped_config_header(ws) is not None


def _extract_grouped_template_metadata(ws):
    quote_refs = _extract_all_excel_quote_refs(ws, max_rows=200)
    quote_ref = ", ".join(quote_refs)
    date_text = ""
    for r in range(1, min(ws.max_row, 200) + 1):
        first = _cell_to_text(ws.cell(r, 1).value).strip().lower()
        if first.startswith("date"):
            value = ws.cell(r, 2).value
            if isinstance(value, datetime):
                date_text = value.strftime("%d/%m/%Y")
            else:
                date_text = _cell_to_text(value)
    return quote_ref, date_text


def _is_grouped_template_summary_row(ws, row_idx, cols):
    row_text = _row_text(ws, row_idx, 1, ws.max_column).lower()
    if not row_text:
        return False

    if first := _cell_to_text(ws.cell(row_idx, 1).value).strip().lower():
        if first.startswith("quote") or first.startswith("name"):
            return True

    if "consolidation fee" in row_text:
        sku_value = _cell_to_text(ws.cell(row_idx, cols.get("sku", 0)).value)
        if not sku_value:
            return True
        return False

    if "total" in row_text and "total selling price" not in row_text:
        return True

    return False


def _extract_grouped_template_items_and_config(ws):
    header_info = _find_grouped_config_header(ws)
    if not header_info:
        return [], []

    header_row, cols = header_info
    items = []
    config_rows = []
    current_item = None
    blank_streak = 0

    for r in range(header_row + 1, ws.max_row + 1):
        row_text = _row_text(ws, r, 1, ws.max_column)
        if not row_text:
            blank_streak += 1
            if blank_streak >= 4:
                break
            continue
        blank_streak = 0

        if _is_grouped_template_summary_row(ws, r, cols):
            continue

        first_cell = _cell_to_text(ws.cell(r, 1).value).strip()
        desc = _cell_to_text(ws.cell(r, cols["description"]).value)
        sku = _cell_to_text(ws.cell(r, cols["sku"]).value)
        qty_raw = _cell_to_text(ws.cell(r, cols.get("qty", 0)).value)
        qty_val = 0
        if qty_raw:
            try:
                qty_val = int(qty_raw)
            except Exception:
                qty_val = int(_parse_money(qty_raw) or 0)

        unit_price = _parse_money(ws.cell(r, cols.get("unit", 0)).value) or 0.0
        total_price = _parse_money(ws.cell(r, cols.get("total", 0)).value) or 0.0

        if first_cell.lower().startswith("quote"):
            continue

        if first_cell:
            # Top-level item row
            if desc:
                items.append((desc, qty_val, unit_price, total_price))
                current_item = str(len(items))
            continue

        # Grouped config row belongs to the last top-level item.
        if current_item and desc:
            config_rows.append((current_item, "", "", desc, sku, "", qty_raw))

    return items, config_rows


def _find_compact_quote_header(ws):
    search_end = min(ws.max_row, 30)
    for r in range(1, search_end + 1):
        columns = {}
        for c in range(1, ws.max_column + 1):
            name = _cell_to_text(ws.cell(r, c).value).strip().lower()
            if not name:
                continue
            if name == "#" and "item" not in columns:
                columns["item"] = c
            if "sku" in name and "sku" not in columns:
                columns["sku"] = c
            if "description" in name and "description" not in columns:
                columns["description"] = c
            if name in ("q-ty", "qty", "quantity") and "qty" not in columns:
                columns["qty"] = c
            if "unit selling price" in name and "unit" not in columns:
                columns["unit"] = c
            if "total selling price" in name and "total" not in columns:
                columns["total"] = c
        if all(k in columns for k in ("item", "sku", "description", "qty", "unit", "total")):
            return r, columns
    return None


def _extract_compact_quote_items_and_config(ws):
    header_info = _find_compact_quote_header(ws)
    if not header_info:
        return [], []

    header_row, cols = header_info
    items = []
    config_rows = []
    current_item = None
    blank_streak = 0

    for r in range(header_row + 1, ws.max_row + 1):
        row_text = _row_text(ws, r, 1, ws.max_column)
        if not row_text:
            blank_streak += 1
            if blank_streak >= 2:
                break
            continue
        blank_streak = 0

        first_cell = _cell_to_text(ws.cell(r, cols["item"]).value).strip()
        sku = _cell_to_text(ws.cell(r, cols["sku"]).value)
        desc = _cell_to_text(ws.cell(r, cols["description"]).value)
        qty_raw = _cell_to_text(ws.cell(r, cols["qty"]).value)
        unit_price = _parse_money(ws.cell(r, cols["unit"]).value) or 0.0
        total_price = _parse_money(ws.cell(r, cols["total"]).value)

        if not any([first_cell, sku, desc, qty_raw, unit_price, total_price]):
            continue

        if total_price is None and any(token in row_text.lower() for token in ("total", "subtotal", "quote number", "quote name", "solution id")):
            continue

        try:
            qty_val = int(qty_raw) if qty_raw not in (None, "") else 0
        except Exception:
            qty_val = int(_parse_money(qty_raw) or 0)

        if first_cell:
            if desc and qty_val > 0:
                items.append((desc, qty_val, unit_price, total_price))
                current_item = str(len(items))
            continue

        if current_item and (sku or desc):
            config_rows.append((current_item, "", "", desc, sku, "", qty_raw))

    return items, config_rows


def _extract_excel_consolidation_fee(ws) -> float:
    """Find 'Consolidation Fee' rows in Excel and return the most relevant numeric amount."""
    logger = _get_logger()
    for row in ws.iter_rows():
        for cell in row:
            value = cell.value
            if not isinstance(value, str):
                continue
            label = value.strip().lower()
            if not re.fullmatch(r"consolidation fees?\s*:?", label):
                continue
            for next_col in range(ws.max_column, cell.column, -1):
                next_value = ws.cell(cell.row, next_col).value
                if next_value in (None, ""):
                    continue
                parsed = _parse_money(next_value)
                if parsed is not None:
                    logger.debug(
                        "Excel consolidation fee found at row=%s col=%s, value_col=%s, raw_value=%s, parsed=%s",
                        cell.row,
                        cell.column,
                        next_col,
                        next_value,
                        parsed,
                    )
                    return parsed

    logger.debug("Excel consolidation fee not found; defaulting to 0.0")
    return 0.0


def _extract_excel_shipping_fee(ws) -> float:
    """Find 'Shipping:' in Excel and read the first non-empty cell to its right."""
    logger = _get_logger()

    for row in ws.iter_rows():
        for cell in row:
            value = cell.value
            if not isinstance(value, str):
                continue
            label = value.strip().lower()
            if not re.fullmatch(r"shipping(?:\s+(?:charge|charges|cost))?\s*:?", label):
                continue

            for next_col in range(cell.column + 1, ws.max_column + 1):
                next_value = ws.cell(cell.row, next_col).value
                if next_value in (None, ""):
                    continue

                parsed = _parse_money(next_value)
                shipping_fee = parsed or 0.0
                logger.debug(
                    "Excel shipping fee found at row=%s col=%s, value_col=%s, raw_value=%s, parsed=%s",
                    cell.row,
                    cell.column,
                    next_col,
                    next_value,
                    shipping_fee,
                )
                return shipping_fee

            logger.debug(
                "Excel shipping label found at row=%s col=%s, but no value exists to the right",
                cell.row,
                cell.column,
            )
            return 0.0

    logger.debug("Excel shipping fee not found; defaulting to 0.0")
    return 0.0


def _extract_excel_service_fields(ws) -> Dict[str, Dict[str, str]]:
    """Extract Service Tag / Start Date / End Date per item from Product Details (Excel).

    Some UAE (AED) templates don't include a configuration table; instead they list:
      - Service Tag:
      - Service Start Date:
      - Service End Date:
    with values typically on the next row in the same column (or occasionally to the right).

    Returns a mapping: internal_item_no -> {"service_tag": str, "service_start_date": str, "service_end_date": str}
    """
    anchor = _find_product_details_anchor(ws)
    if not anchor:
        return {}

    logger = _get_logger()
    max_col = min(ws.max_column, 40)

    def _is_item_marker(row_idx: int) -> bool:
        item_marker = _cell_to_text(ws.cell(row_idx, 1).value)
        return bool(re.match(r"^\d+\.$", item_marker))

    def _find_next_item_row(start_row: int) -> Optional[int]:
        for row_idx in range(start_row, ws.max_row + 1):
            if _is_item_marker(row_idx):
                return row_idx
        return None

    out: Dict[str, Dict[str, str]] = {}
    item_counter = 0
    r = anchor + 1
    while r <= ws.max_row:
        if not _is_item_marker(r):
            r += 1
            continue

        item_counter += 1
        item_key = str(item_counter)
        next_item_row = _find_next_item_row(r + 1)
        search_end = (next_item_row - 1) if next_item_row is not None else ws.max_row

        fields = {"service_tag": "", "service_start_date": "", "service_end_date": ""}

        for rr in range(r + 1, min(search_end + 1, ws.max_row + 1)):
            for cc in range(1, max_col + 1):
                raw = _cell_to_text(ws.cell(rr, cc).value)
                if not raw:
                    continue
                key = _normalize_text(raw)
                if key not in ("servicetag", "servicestartdate", "serviceenddate"):
                    continue

                # Prefer value directly below; fall back to cell to the right.
                value = _cell_to_text(ws.cell(rr + 1, cc).value) if rr + 1 <= ws.max_row else ""
                if not value and cc + 1 <= max_col:
                    value = _cell_to_text(ws.cell(rr, cc + 1).value)

                if key == "servicetag" and value:
                    fields["service_tag"] = value
                elif key == "servicestartdate" and value:
                    fields["service_start_date"] = value
                elif key == "serviceenddate" and value:
                    fields["service_end_date"] = value

        if any(fields.values()):
            out[item_key] = fields
            logger.info("Excel service fields extracted for item=%s: %s", item_key, fields)

        r = next_item_row if next_item_row is not None else ws.max_row + 1

    return out


def _is_config_section_row(module: str, description: str, sku: str, tax: str) -> bool:
    """Rows like 'Components', 'Software', 'Service' act as section headers in Product Details."""
    return bool(module and not description and not sku and not tax)


def _extract_pdf_lines(pdf_bytes: bytes) -> List[str]:
    """Extract a cleaned line list from PDF using pdfplumber (best-effort).

    Falls back to pypdf extraction if pdfplumber isn't available or fails.
    """
    lines: List[str] = []
    try:
        import pdfplumber

        with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
            for page in pdf.pages:
                words = page.extract_words(use_text_flow=True)
                if not words:
                    continue
                # group words by y0 (line position)
                rows = {}
                for w in words:
                    y = round(w.get("top", 0))
                    rows.setdefault(y, []).append(w)
                for y in sorted(rows.keys()):
                    row_words = sorted(rows[y], key=lambda w: w.get("x0", 0))
                    line = " ".join(w.get("text", "") for w in row_words)
                    lines.append(line.strip())
        if lines:
            return lines
    except Exception:
        pass

    # Fallback to pypdf extraction
    try:
        from pypdf import PdfReader
    except ImportError:
        raise RuntimeError("pypdf is required to parse PDF quotes")

    reader = PdfReader(BytesIO(pdf_bytes))
    text = "\n".join((page.extract_text() or "") for page in reader.pages)
    return [l.strip() for l in text.splitlines()]


def _extract_pdf_quote_data(pdf_bytes: bytes):
    """
    FINAL VERSION — CLEAN, HYBRID (C3), PDF‑ONLY FIX
    Extracts:
      • FULL item list (13+ items across all pages)
      • FULL Product Details (C3 Hybrid)
      • Clean multi‑line merging
      • Correct item mapping
    """

    logger = _get_logger()
    lines = _extract_pdf_lines(pdf_bytes)
    logger.debug("_extract_pdf_quote_data: extracted %d lines", len(lines))

    # ---------------------- METADATA ----------------------
 # ---------------------- METADATA ----------------------
    metadata = {
        "company name": "",
        "customer name": "",
        "customer number": "",
        "end user": "",
        "reseller": "",
    }
    quote_ref_text = ""
    date_text = ""
    consolidation_fee = 0.0

    pending_keys = []
    prev_label = None
    pdf_label_aliases = {
        "authorized partner": "reseller",
    }

    MONTHS = r"(January|February|March|April|May|June|July|August|September|October|November|December)"
    DATE_RX = re.compile(rf"{MONTHS}\s+\d{{1,2}},\s+\d{{4}}")

    def extract_dates(s: str):
        return [m.group(0) for m in DATE_RX.finditer(s)]

    def _normalize_pdf_label(label: str) -> str:
        normalized = label.strip().lower().rstrip(":")
        mapped = pdf_label_aliases.get(normalized, normalized)
        if mapped != normalized:
            logger.debug("PDF metadata label alias: '%s' -> '%s'", normalized, mapped)
        return mapped

    def _extract_pdf_reseller(lines: List[str]) -> str:
        stop_markers = (
            "billing information",
            "shipping information",
            "quote summary",
            "product details",
        )

        for idx, line in enumerate(lines):
            lower_line = line.lower()
            if "authorized partner" not in lower_line:
                continue

            collected: List[str] = []

            same_line_match = re.search(r"authorized partner\s*:\s*(.+)$", line, re.IGNORECASE)
            if same_line_match and same_line_match.group(1).strip():
                collected.append(same_line_match.group(1).strip())

            next_idx = idx + 1
            while next_idx < len(lines):
                candidate = lines[next_idx].strip()
                candidate_lower = candidate.lower()
                if not candidate:
                    break
                if any(marker in candidate_lower for marker in stop_markers):
                    break
                collected.append(candidate)
                next_idx += 1

            if not collected:
                return ""

            reseller_text = " ".join(collected)
            reseller_text = re.sub(r"\s+", " ", reseller_text).strip()

            if "page name" in lower_line:
                uppercase_start = re.search(r"\b[A-Z]{2,}(?:\s+[A-Z]{2,})+", reseller_text)
                if uppercase_start:
                    reseller_text = reseller_text[uppercase_start.start():].strip()

            reseller_text = re.sub(r"\s*-\s*Authorized Partner\+?$", "", reseller_text, flags=re.IGNORECASE).strip()
            reseller_text = reseller_text.rstrip("+").strip()
            return reseller_text

        return ""

    for line in lines:
        stripped = line.strip()
        lower = stripped.lower()

        # ---------------- MULTILINE KEY HANDLING ----------------
        if pending_keys and stripped:
            line_val = stripped

            # Case: Quote number + Quote date (+ expiration) on the SAME LINE
            if {"quote number", "quote date"}.issubset(set(pending_keys)):
                m = re.match(r"^\s*(\d+)\b(.*)$", line_val)
                if m:
                    quote_ref_text = m.group(1)
                    tail = m.group(2).strip()
                else:
                    tail = line_val

                dates = extract_dates(tail)
                if len(dates) >= 1:
                    date_text = dates[0]

                pending_keys = []
                continue

            # Generic fallback: last key takes the rest
            parts = stripped.split()
            for idx, key in enumerate(pending_keys):
                if idx == len(pending_keys) - 1:
                    val = " ".join(parts[idx:])
                else:
                    val = parts[idx] if idx < len(parts) else ""

                if key == "quote number":
                    m = re.search(r"\b\d{6,}\b", val)
                    if m:
                            quote_ref_text = m.group(0)
                            logger.debug("PDF metadata multiline: quote number=%s", quote_ref_text)
                elif key == "quote date":
                    if DATE_RX.search(line_val):
                        val = DATE_RX.search(line_val).group(0)
                    date_text = val
                    logger.debug("PDF metadata multiline: quote date=%s", date_text)
                elif key in metadata:
                    metadata[key] = val
                    logger.debug("PDF metadata multiline: %s=%s", key, val)

            pending_keys = []
            continue

        # ---------------- LABEL WAITING FOR NEXT LINE ----------------
        if prev_label:
            if prev_label == "quote number":
                # extract ONLY the numeric Quote Number
                m = re.search(r"\b\d{6,}\b", stripped)
                if m:
                    quote_ref_text = m.group(0)
                    logger.debug("PDF metadata next-line: quote number=%s", quote_ref_text)
        
            elif prev_label == "quote date":
                if DATE_RX.search(stripped):
                    date_text = DATE_RX.search(stripped).group(0)
                else:
                    date_text = stripped
                logger.debug("PDF metadata next-line: quote date=%s", date_text)
        
            elif prev_label in metadata:
                metadata[prev_label] = stripped
                logger.debug("PDF metadata next-line: %s=%s", prev_label, stripped)
        
            prev_label = None
            continue

        # ---------------- KEY:VALUE ON SAME LINE ----------------
        if ":" in stripped:
            parts = [p.strip() for p in re.split(r"\s*:\s*", stripped) if p.strip()]
            normalized = [_normalize_pdf_label(p) for p in parts]

            if stripped.endswith(":"):
                for key in normalized:
                    if key in ("quote number", "quote date", "company name", "customer name", "customer number", "reseller"):
                        pending_keys.append(key)
                        logger.debug("PDF metadata pending label: %s", key)
                continue

            for i in range(0, len(parts), 2):
                key = normalized[i]
                val = parts[i+1] if i+1 < len(parts) else ""

                if key == "quote number":
                    quote_ref_text = val
                    logger.debug("PDF metadata same-line: quote number=%s", quote_ref_text)
                elif key == "quote date":
                    m = DATE_RX.search(val)
                    date_text = m.group(0) if m else val
                    logger.debug("PDF metadata same-line: quote date=%s", date_text)
                elif key in metadata:
                    metadata[key] = val
                    logger.debug("PDF metadata same-line: %s=%s", key, val)

            continue

        # ---------------- LABEL-ONLY LINE ----------------
        normalized_lower = _normalize_pdf_label(lower)
        if normalized_lower in ("quote number", "quote date", "company name", "customer name", "customer number", "reseller"):
            prev_label = normalized_lower
            logger.debug("PDF metadata label-only line detected: %s", normalized_lower)
            continue

    # FINAL FALLBACK: scan for date if still empty
    if not date_text:
        for l in lines:
            m = DATE_RX.search(l)
            if m:
                date_text = m.group(0)
                logger.debug("PDF metadata fallback date=%s", date_text)
                break

    reseller_from_layout = _extract_pdf_reseller(lines)
    if reseller_from_layout:
        metadata["reseller"] = reseller_from_layout
        logger.debug("PDF reseller layout extraction=%s", reseller_from_layout)

    # Extract consolidation fee when present; otherwise keep the zero fallback.
    for i, line in enumerate(lines):
        low = line.lower().strip()
        if "consolidation fee" not in low:
            continue

        same_line_match = re.search(r"consolidation fee[:\s]*([$€£]?\s*[\d,]+(?:\.\d+)?)", line, re.IGNORECASE)
        if same_line_match:
            consolidation_fee = _parse_money(same_line_match.group(1)) or 0.0
            logger.debug("PDF consolidation fee same-line=%s", consolidation_fee)
            break

        if i + 1 < len(lines):
            next_line_value = _parse_money(lines[i + 1])
            if next_line_value is not None:
                consolidation_fee = next_line_value
                logger.debug("PDF consolidation fee next-line=%s", consolidation_fee)
                break

    logger.debug(
        "PDF metadata summary: quote_ref=%s, date=%s, reseller=%s, company=%s, customer=%s, customer_number=%s, consolidation_fee=%s",
        quote_ref_text,
        date_text,
        metadata.get("reseller", ""),
        metadata.get("company name", ""),
        metadata.get("customer name", ""),
        metadata.get("customer number", ""),
        consolidation_fee,
    )

    # ---------------------- ITEMS (FULL 13+ extraction) ----------------------
    def _try_parse_item(line: str):
        # Example:
        #   "DELL USB-C Mobile Adapter - DA310 65.1 3 195.3"
        m = re.match(
            r"^(?P<desc>.*?)(?P<unit>\d[\d,\.]*?)\s+(?P<qty>\d+)\s+(?P<total>\d[\d,\.]*)$",
            line
        )
        if not m:
            return None
        desc = m.group("desc").strip()
        qty = int(m.group("qty"))
        unit = _parse_money(m.group("unit")) or 0.0
        total = _parse_money(m.group("total")) or 0.0
        return desc, qty, unit, total

    items = []
    in_items = False

    for line in lines:
        low = line.lower().strip()

        # Enter items region
        if "quote summary" in low:
            in_items = True
            continue
        if "unit price" in low and "qty" in low and "item total" in low:
            in_items = True
            continue

        # Stop items at Product Details
        if "product details" in low:
            break

        if not in_items:
            continue

        # Skip empty lines
        if not line.strip():
            continue

        parsed = _try_parse_item(line)
        if parsed:
            items.append(parsed)

    logger.debug("EXTRACTED PDF ITEMS = %d", len(items))

    # ---------------- CONFIGURATION (C3 HYBRID) -----------------
    config_rows = []

    try:
        pd_idx = next(i for i, l in enumerate(lines) if "product details" in l.lower())
    except StopIteration:
        pd_idx = None

    if pd_idx is not None:
        current_item = ""
        current_heading = ""
        awaiting_heading = False
        item_counter = 0

        # Helper: merge value line into previous module
        def attach_value_to_previous(text_line: str) -> bool:
            if not config_rows:
                return False
            (it, hd, mod, dsc, sku, tax) = config_rows[-1]
            if it == current_item and (dsc == "" or dsc is None):
                if ":" not in text_line and len(text_line.split()) >= 2:
                    config_rows[-1] = (it, hd, mod, text_line.strip(), sku, tax)
                    return True
            return False

        # Clean heading
        def _clean_heading_text(t: str):
            return re.sub(r"\s+\d[\d,\.]*\s+\d[\d,\.]*\s+\d[\d,\.]*$", "", t).strip()

        i = pd_idx + 1
        while i < len(lines):
            line = lines[i].strip()
            low = line.lower()

            if "unit price" in low and "qty" in low and "item total" in low:
                awaiting_heading = True
                i += 1
                continue

            if awaiting_heading:
                if line.lower() == "description":
                    i += 1
                    continue

                # Next line = heading
                parsed = _try_parse_item(line)
                if parsed:
                    desc, qty, unit, total = parsed
                    current_heading = _clean_heading_text(desc)
                else:
                    current_heading = _clean_heading_text(line)

                item_counter += 1
                current_item = str(item_counter)
                awaiting_heading = False
                i += 1
                continue

            # Stop when we captured all headings (items found)
            if item_counter >= len(items) and any(
                k in low for k in ["ship to", "subtotal", "total", "important notes"]
            ):
                break

            if not current_item:
                i += 1
                continue

            # Skip junk
            if low.startswith("page ") or low.startswith("category description"):
                i += 1
                continue

            # key:value pairs
            if ":" in line:
                parts = [p.strip() for p in re.split(r"\s*:\s*", line) if p.strip()]
                for j in range(0, len(parts), 2):
                    mod = parts[j]
                    dsc = parts[j+1] if j+1 < len(parts) else ""
                    if mod.lower() != "category description":
                        config_rows.append((current_item, current_heading, mod, dsc, "", ""))
                i += 1
                continue

            # Attach as value if previous row missing value
            if attach_value_to_previous(line):
                i += 1
                continue

            # Otherwise, treat as a module label
            config_rows.append((current_item, current_heading, line, "", "", ""))
            i += 1

    # ---------------- MERGE FRAGMENTED ROWS -----------------
    def collapse(rows):
        out = []
        i = 0
        while i < len(rows):
            item, head, mod, dsc, sku, tax = rows[i]
            mod = mod.strip()
            dsc = dsc.strip()

            # Merge rows like:
            #   "Select Power Cord Type"
            #   "UK/Irish Power Cord"
            if i + 1 < len(rows):
                ni, nh, nmod, ndsc, _, _ = rows[i+1]
                if ni == item and nh == head:
                    if dsc == "" and ":" not in nmod:
                        # second line is value
                        dsc = nmod
                        i += 1

            out.append((item, head, mod, dsc, sku, tax))
            i += 1
        return out

    config_rows = collapse(config_rows)
    logger.debug("PDF CONFIG ROWS = %d", len(config_rows))

    return items, metadata, config_rows, quote_ref_text, date_text, consolidation_fee

def _extract_all_config_rows(ws) -> List[Tuple[str, str, str, str, str, str, str]]:
    """
    FINAL CLEAN VERSION (Excel upload only)
    Extracts ALL configuration rows under Product Details from Excel files.

    Output rows have the structure:
      (item_number, item_heading, module, description, sku, tax_type, qty)

    • Preserves original Excel order
    • Skips totals / shipping / price lines
    • Merges 2-line fragments
    • Cleans up module/value pairs
    • Avoids duplicated headings
    """

    anchor = _find_product_details_anchor(ws)
    if not anchor:
        return []

    rows: List[Tuple[str, str, str, str, str, str, str]] = []
    r = anchor + 1
    max_col = min(ws.max_column, 40)

    def _clean_heading_text(text: str) -> str:
        # remove trailing price fragments like "4 $200.00 $800.00"
        return re.sub(r"\s+\d+(\.\d+)?\s+\$?[\d,\.]+\s+\$?[\d,\.]+$", "", text).strip()

    def _is_table_stop(text: str) -> bool:
        low = text.lower()
        return (
            _is_price_or_qty_line(text)
            or "estimated delivery" in low
            or "subtotal" in low
            or "total" in low
            or "ship to" in low
        )

    def _extract_item_heading(row_idx: int) -> Optional[Tuple[str, str]]:
        item_marker = _cell_to_text(ws.cell(row_idx, 1).value)
        if not re.match(r"^\d+\.$", item_marker):
            return None

        heading = _cell_to_text(ws.cell(row_idx, 2).value)
        if not heading:
            heading = _row_text(ws, row_idx, 1, max_col)

        return item_marker.rstrip("."), _clean_heading_text(heading)

    def _find_next_item_row(start_row: int) -> Optional[int]:
        for row_idx in range(start_row, ws.max_row + 1):
            if _extract_item_heading(row_idx):
                return row_idx
        return None

    item_counter = 0
    while r <= ws.max_row:
        item_info = _extract_item_heading(r)
        if not item_info:
            r += 1
            continue

        _source_item_number, current_heading = item_info
        item_counter += 1
        current_item = str(item_counter)
        next_item_row = _find_next_item_row(r + 1)
        search_end = (next_item_row - 1) if next_item_row is not None else ws.max_row

        header_info = None
        scan_row = r + 1
        while scan_row <= search_end:
            maybe_header = _find_config_table_header(ws, scan_row, search_rows=0)
            if maybe_header:
                header_info = maybe_header
                break
            scan_row += 1

        # Some items (for example accessories) have no configuration table.
        if not header_info:
            r = next_item_row if next_item_row is not None else ws.max_row + 1
            continue

        header_row, colmap = header_info
        data_row = header_row + 1
        has_real_module_col = bool(colmap.get("module")) and colmap.get("module") != colmap.get("description")

        # Skip empty rows after header
        while data_row <= search_end and not _row_text(ws, data_row, 1, max_col):
            data_row += 1

        blank_streak = 0
        while data_row <= search_end:
            row_text_all = _row_text(ws, data_row, 1, max_col)

            if not row_text_all:
                blank_streak += 1
                if blank_streak >= 2:
                    break
                data_row += 1
                continue
            blank_streak = 0

            if _is_table_stop(row_text_all):
                data_row += 1
                continue

            mod = _cell_to_text(ws.cell(data_row, colmap.get("module", 0)).value) if has_real_module_col else ""
            desc = _cell_to_text(ws.cell(data_row, colmap.get("description", 0)).value)
            sku = _cell_to_text(ws.cell(data_row, colmap.get("sku", 0)).value)
            tax = _cell_to_text(ws.cell(data_row, colmap.get("tax", 0)).value)
            qty = _cell_to_text(ws.cell(data_row, colmap.get("qty", 0)).value)

            if not has_real_module_col and desc and not any([sku, tax, qty]):
                mod = desc
                desc = ""

            if not any([mod, desc, sku, tax, qty]):
                break

            rows.append((current_item, current_heading, mod, desc, sku, tax, qty))
            data_row += 1

        r = next_item_row if next_item_row is not None else ws.max_row + 1

    # -----------------------
    # MERGE FRAGMENTED ROWS
    # -----------------------
    cleaned = []
    i = 0
    while i < len(rows):
        item, head, mod, desc, sku, tax, qty = rows[i]
        mod, desc = mod.strip(), desc.strip()

        # CASE: 2-line module label (common in Dell exports)
        if i + 1 < len(rows):
            ni, nh, nmod, ndesc, nsku, ntax, nqty = rows[i + 1]
            if ni == item and nh == head:
                if desc == "" and ndesc == "" and ":" not in mod and ":" not in nmod:
                    # join "Smart" + "Dock SD25TB5"
                    mod = f"{mod} {nmod}".strip()
                    i += 1

        cleaned.append((item, head, mod, desc, sku, tax, qty))
        i += 1

    return cleaned


def _extract_product_detail_headings(ws) -> Dict[str, str]:
    """Extract item heading lines from Product Details for Excel uploads."""
    anchor = _find_product_details_anchor(ws)
    if not anchor:
        return {}

    headings: Dict[str, str] = {}
    r = anchor + 1
    max_col = min(ws.max_column, 40)

    def _clean_heading_text(text: str) -> str:
        return re.sub(r"\s+\d+(\.\d+)?\s+\$?[\d,\.]+\s+\$?[\d,\.]+$", "", text).strip()

    item_counter = 0
    while r <= ws.max_row:
        item_marker = _cell_to_text(ws.cell(r, 1).value)
        if re.match(r"^\d+\.$", item_marker):
            item_counter += 1
            item_no = str(item_counter)
            heading = _cell_to_text(ws.cell(r, 2).value)
            if not heading:
                heading = _row_text(ws, r, 1, max_col)
            headings[item_no] = _clean_heading_text(heading)
        r += 1

    return headings


def _extract_product_detail_display_numbers(ws) -> Dict[str, str]:
    """Map internal sequential item ids to the source-visible item numbers from Product Details."""
    anchor = _find_product_details_anchor(ws)
    if not anchor:
        return {}

    display_numbers: Dict[str, str] = {}
    r = anchor + 1
    item_counter = 0

    while r <= ws.max_row:
        item_marker = _cell_to_text(ws.cell(r, 1).value)
        if re.match(r"^\d+\.$", item_marker):
            item_counter += 1
            display_numbers[str(item_counter)] = item_marker.rstrip(".")
        r += 1

    return display_numbers
# ================= Main =================

CURRENCY_CONVERSION_RATES = {
    # Input files are in USD; convert to selected output currency using fixed rates.
    "USD": 1.0,
    "QAR": 3.64,
    "AED": 3.68,
}

CURRENCY_NUMBER_FORMATS = {
    "USD": '"$"#,##0.00',
    "QAR": '"QAR" #,##0.00',
    "AED": '"AED" #,##0.00',
}

def generate_dell_quote(
    input_excel_bytes: bytes,
    logo_bytes: Optional[bytes] = None,
    margin_percent: float = 0.0,
    currency_code: str = "USD",
) -> bytes:
    """
    Generate a 2-sheet workbook from either:
      - Dell quote Excel file (as bytes) OR
      - Dell quote PDF (as bytes)

      Output:
      - 'Quote' formatted with strict template:
          - A1:B4 merged for logo
          - Address block in D1:F3 (merged per row)
          - 'Quote Ref' shown at C5 with value at D5 (read from INPUT E15 or PDF "Quote number")
          - 'Date' shown at C6 with value at D6 (read from INPUT E18 or PDF "Quote date")
          - Table header at row 8; data from row 9
      - 'Configuration' sheet that replicates the 'Product Details' configuration table(s)
        PER PRODUCT, preserving columns Module, Description, SKU, Tax Type (dropping Qty/Unit Price/Subtotal/date lines).
    """
    logger = _get_logger()
    logger.info("Generating Dell quote (bytes=%d)", len(input_excel_bytes) if input_excel_bytes is not None else 0)
    
    currency_code = (currency_code or "USD").upper()
    conversion_rate = CURRENCY_CONVERSION_RATES.get(currency_code, 1.0)

    # Missing consolidation fee should be treated as zero for both Excel and PDF uploads.
    consolidation_fee = 0.0
    item_display_numbers_by_item: Dict[str, str] = {}
    service_fields_by_item: Dict[str, Dict[str, str]] = {}
    expiry_text = ""

    # ---- Load source ----
    is_pdf = input_excel_bytes.lstrip().startswith(b"%PDF")

    if is_pdf:
        items, quote_meta, config_rows, quote_ref_text, date_text, consolidation_fee = _extract_pdf_quote_data(input_excel_bytes)
        logger.info("Parsed PDF quote: %d items, quote_ref=%s, date=%s", len(items), quote_ref_text, date_text)
        _log_items("PDF items", items)
        item_descs_order = [it[0] for it in items]
    else:
        src_wb = openpyxl.load_workbook(BytesIO(input_excel_bytes), data_only=True)
        src_ws = src_wb.active
        src_config_ws = _find_configuration_sheet(src_wb)
        if src_config_ws is not None:
            logger.info("Detected separate configuration sheet: %s", src_config_ws.title)
        logger.info("Parsed Excel quote (sheets=%d, active=%s)", len(src_wb.sheetnames), src_ws.title)

        # ---- Extract metadata (STRICT E15/E18) ----
        quote_ref_text, date_text = _extract_metadata_strict(src_ws)
        expiry_text = _extract_expiry_date(src_ws)
        logger.info(
            "Extracted metadata (strict): quote_ref=%s, date=%s, expires_by=%s",
            quote_ref_text,
            date_text,
            expiry_text,
        )
        
        # ---- FALLBACK for web/PDF-style Excel quotes ----
        if (
            not quote_ref_text
            or "$" in quote_ref_text
            or not re.search(r"\d{6,}", quote_ref_text)
        ):
            fb_ref, fb_date = _extract_metadata_excel_fallback(src_ws)
            if fb_ref:
                quote_ref_text = fb_ref
            if fb_date:
                date_text = fb_date
        
            logger.info(
                "Fallback Excel metadata applied: quote_ref=%s, date=%s",
                quote_ref_text,
                date_text,
            )
        
        quote_meta = _extract_quote_metadata(src_ws)

        if _is_grouped_config_template(src_ws):
            logger.info("Detected grouped config Excel template")
            quote_ref_text, date_text = _extract_grouped_template_metadata(src_ws)
            items, grouped_config_rows = _extract_grouped_template_items_and_config(src_ws)
            if src_config_ws is not None:
                config_rows = _extract_config_rows_from_configuration_sheet(src_config_ws)
                logger.info("Used separate configuration sheet for grouped template: %d config rows", len(config_rows))
            else:
                config_rows = grouped_config_rows
            item_descs_order = [it[0] for it in items]
            item_headings_by_item = {str(i + 1): items[i][0] for i in range(len(items))}
            item_display_numbers_by_item = {str(i + 1): f"Option {i + 1}" for i in range(len(items))}
            consolidation_fee = _extract_excel_consolidation_fee(src_ws)
            logger.info("Grouped template items extracted: %d items, %d config rows", len(items), len(config_rows))
        else:
            # ---- Extract items (Pricing Summary layout first; else generic) ----
            extracted_config_rows = None
            items_ps = _try_extract_items_from_pricing_summary(src_ws)
            if items_ps:
                items = items_ps
                logger.info("Found %d items via Pricing Summary extraction", len(items))
                _log_items("Pricing summary items", items)
            else:
                compact_items, compact_config_rows = _extract_compact_quote_items_and_config(src_ws)
                if compact_items:
                    items = compact_items
                    extracted_config_rows = compact_config_rows
                    logger.info("Found %d items via compact quote extraction", len(items))
                    _log_items("Compact quote items", items)
                else:
                    first_data_row, desc_col, qty_col, unit_col = _find_header_row_strict_or_detect(src_ws)
                    logger.info("Using generic item extraction starting at row %d (desc_col=%d, qty_col=%d, unit_col=%d)", first_data_row, desc_col, qty_col, unit_col)
                    items = []
                    r = first_data_row
                    while r <= src_ws.max_row:
                        desc = src_ws.cell(r, desc_col).value
                        qty  = src_ws.cell(r, qty_col).value
                        unit = src_ws.cell(r, unit_col).value

                        desc_text = _cell_to_text(desc)
                        if not desc_text or desc_text.lower().startswith("total"):
                            break
                        try:
                            qty_val = int(qty) if qty not in (None, "") else 0
                        except Exception:
                            qty_val = int(_parse_money(qty) or 0)
                        unit_val = _parse_money(unit) or 0.0
                        if qty_val > 0:
                            items.append((desc_text, qty_val, unit_val, None))
                        r += 1
                    logger.info("Extracted %d items via generic table parsing", len(items))
                    _log_items("Parsed table items", items)

            item_descs_order = [it[0] for it in items]
            if src_config_ws is not None:
                config_rows = _extract_config_rows_from_configuration_sheet(src_config_ws)
                item_headings_by_item = _extract_product_detail_headings(src_config_ws) or _extract_product_detail_headings(src_ws)
                item_display_numbers_by_item = _extract_product_detail_display_numbers(src_config_ws) or _extract_product_detail_display_numbers(src_ws)
                logger.info("Using separate configuration sheet for product detail extraction: %d config rows", len(config_rows))
            else:
                config_rows = extracted_config_rows if extracted_config_rows is not None else _extract_all_config_rows(src_ws)
                item_headings_by_item = _extract_product_detail_headings(src_ws)
                item_display_numbers_by_item = _extract_product_detail_display_numbers(src_ws)
            consolidation_fee = _extract_excel_consolidation_fee(src_ws)
            if currency_code == "AED":
                shipping_fee = _extract_excel_shipping_fee(src_ws)
                if shipping_fee:
                    logger.info(
                        "AED: adding Excel shipping fee to consolidation fee (shipping=%s, consolidation_before=%s)",
                        shipping_fee,
                        consolidation_fee,
                    )
                consolidation_fee += shipping_fee
                # If this template doesn't have a config table, we may still have service fields in Product Details.
                if not config_rows:
                    service_fields_by_item = _extract_excel_service_fields(src_ws)

    # Use today's date on the generated quote instead of the source file date.
    date_text = datetime.now().strftime("%d/%m/%Y")

    if conversion_rate != 1.0:
        items = [
            (
                desc_text,
                qty_val,
                (unit_val or 0.0) * conversion_rate,
                (subtotal_val * conversion_rate) if subtotal_val is not None else None,
            )
            for (desc_text, qty_val, unit_val, subtotal_val) in items
        ]
        consolidation_fee *= conversion_rate
        
            
    


    def _extract_part_number_from_description(text: str) -> str:
        text = _cell_to_text(text)
        if not text:
            return ""

        matches = re.findall(r"\(([^()]+)\)", text)
        for candidate in reversed(matches):
            normalized = candidate.strip()
            normalized = normalized.replace("–", "-").replace("—", "-").replace("−", "-")
            normalized = re.sub(r"\s*-\s*", "-", normalized)
            normalized = re.sub(r"\s+", " ", normalized).strip()

            # Dell files use multiple part number formats, including 210-XXXX and AD123456.
            if re.fullmatch(r"(?:\d{3}-[A-Z0-9]{4,5}|[A-Z]{2}\d{6,})", normalized, re.I):
                return normalized
        return ""

    include_part_number = not is_pdf
    part_numbers_by_item: Dict[str, str] = {}
    item_headings_by_item = item_headings_by_item if include_part_number else {}
    item_display_numbers_by_item = item_display_numbers_by_item if include_part_number else {}
    if include_part_number:
        for row_data in config_rows:
            if len(row_data) >= 7:
                item_no, _heading, _module, _desc, sku, _tax, _qty = row_data
            else:
                item_no, _heading, _module, _desc, sku, _tax = row_data
            if sku and item_no not in part_numbers_by_item:
                part_numbers_by_item[item_no] = sku

    # ---- Build output workbook ----
    wb = Workbook()
    ws = wb.active
    ws.title = "Quote"
    ws.sheet_view.showGridLines = False
    
    
    helper_unit_col = "J" if include_part_number else "I"
    helper_margin_col = "K" if include_part_number else "J"
    desc_col = "C" if include_part_number else "B"
    qty_col = "D" if include_part_number else "C"
    unit_price_col = "E" if include_part_number else "D"
    total_price_col = "F" if include_part_number else "E"

    # ---- Step 2: Write Consolidation Fee and Factor ----
    # The helper columns store the consolidation fee and per-line margin inputs.
    ws[f"{helper_unit_col}2"] = consolidation_fee
    ws[f"{helper_unit_col}2"].font = Font(bold=True, color="1F497D")
    ws[f"{helper_unit_col}2"].alignment = Alignment(horizontal="center", vertical="center")
    
    ws[f"{helper_unit_col}3"].value = ""
    ws[f"{helper_unit_col}3"].font = Font(bold=True, color="1F497D")
    ws[f"{helper_unit_col}3"].alignment = Alignment(horizontal="center", vertical="center")

    ws[f"{helper_margin_col}2"].font = Font(bold=True, color="1F497D")
    ws[f"{helper_margin_col}2"].alignment = Alignment(horizontal="center", vertical="center")
    
    
    # Give long descriptions more horizontal space so they don't create very tall rows.
    longest_description = max((len(_cell_to_text(desc_text)) for desc_text, _, _, _ in items), default=0)
    description_width = min(max(44, int(longest_description * 0.55)), 68)

    # Column widths for visible quote columns, metadata, and helper pricing columns.
    widths = {"A": 10, "B": 18, "C": 14, "D": 14, "E": 14, "F": 14, "G": 16, "H": 26, "I": 14, "J": 16, "K": 14}
    if include_part_number:
        widths["B"] = 15
        widths["C"] = description_width
        widths["D"] = 8
        widths["E"] = 15
        widths["F"] = 17
    else:
        widths["B"] = description_width
        widths["C"] = 8
        widths["D"] = 15
        widths["E"] = 17
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.column_dimensions[helper_unit_col].hidden = True
    ws.column_dimensions[helper_margin_col].hidden = True

    # Header rows height
    for rr in range(1, 3):
        ws.row_dimensions[rr].height = 28
    ws.row_dimensions[3].height = 12
    for rr in range(5, 11):
        ws.row_dimensions[rr].height = 20

    # ===== HEADER: use the full banner logo across A:H =====
    ws.merge_cells("A1:H2")
    _add_logo(ws, logo_bytes, anchor="A1", width=780, height=52)

    ws.merge_cells("A5:D5")
    ws.merge_cells("A6:D6")
    ws.merge_cells("A7:D7")
    if currency_code == "QAR":
        ws["A5"] = "Mindware SA, PO Box 22421, D-Ring Road"
        ws["A6"] = "Next to Doha bank, Doha, Qatar"
        ws["A7"] = "Tel : +974 44405000    Website : www.midisglobal.com"
    else:
        ws["A5"] = "P O Box 55609, Dubai, UAE"
        ws["A6"] = "Tel :  +9714 4500600    Fax : +9714 4500678"
        ws["A7"] = "Website :  www.mindware.ae"
    for cell in ("A5", "A6", "A7"):
        ws[cell].font = Font(bold=True, size=11, color="1F497D")
        ws[cell].alignment = Alignment(horizontal="left", vertical="center")

    # ===== META =====
    ws["C8"] = "Quote Ref"
    ws["C8"].font = Font(bold=True, color="1F497D")
    ws["D8"] = quote_ref_text

    ws["C9"] = "Date"
    ws["C9"].font = Font(bold=True, color="1F497D")
    ws["D9"] = date_text

    has_aed_expiry = currency_code == "AED" and bool(expiry_text)
    if has_aed_expiry:
        ws["C10"] = "Expires By"
        ws["C10"].font = Font(bold=True, color="1F497D")
        ws["D10"] = expiry_text

    # ---- Quote metadata (varies by country/template)
    if currency_code == "AED":
        meta_rows = [
            ("End User:", quote_meta.get("end user", "")),
            ("Reseller:", quote_meta.get("reseller", "")),
        ]
    else:
        meta_rows = [
            ("Company Name:", quote_meta.get("company name", "")),
            ("Customer Name:", quote_meta.get("customer name", "")),
            ("Customer Number:", quote_meta.get("customer number", "")),
            ("End User:", quote_meta.get("end user", "")),
            ("Reseller:", quote_meta.get("reseller", "")),
        ]

    meta_label_col = "F" if currency_code == "AED" else "G"
    meta_value_col = "G" if currency_code == "AED" else "H"
    meta_value_end_col = "H" if currency_code == "AED" else None

    if currency_code == "AED":
        ws.column_dimensions["F"].width = 14
        ws.column_dimensions["G"].width = 24
        ws.column_dimensions["H"].width = 24

    for idx, (label, value) in enumerate(meta_rows, start=5):
        label_addr = f"{meta_label_col}{idx}"
        value_addr = f"{meta_value_col}{idx}"

        ws[label_addr] = label
        ws[label_addr].font = Font(bold=True)
        ws[label_addr].alignment = Alignment(horizontal="left", vertical="top")

        if meta_value_end_col:
            merge_range = f"{meta_value_col}{idx}:{meta_value_end_col}{idx}"
            if merge_range not in ws.merged_cells:
                ws.merge_cells(merge_range)

        ws[value_addr] = value
        ws[value_addr].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

        if currency_code == "AED":
            text_len = len(_cell_to_text(value))
            estimated_lines = max(1, min(4, (text_len // 28) + 1))
            ws.row_dimensions[idx].height = max(ws.row_dimensions[idx].height or 20, estimated_lines * 18)

    # ===== TABLE HEADER at row 8; data from row 9 =====
    header_row = 11 if has_aed_expiry else 10
    ws[f"A{header_row}"] = "Sr. No."
    if include_part_number:
        ws[f"B{header_row}"] = "Part Number"
    ws[f"{desc_col}{header_row}"] = "Description"
    ws[f"{qty_col}{header_row}"] = "Qty"
    ws[f"{unit_price_col}{header_row}"] = "Unit Price"
    ws[f"{total_price_col}{header_row}"] = (
        "Total Price (excluding vat)" if currency_code == "AED" else "Total Price"
    )
    ws[f"{helper_unit_col}{header_row}"] = "Original Unit Price"
    ws[f"{helper_margin_col}{header_row}"] = "Margin"
    header_fill = PatternFill(start_color="9BBB59", end_color="9BBB59", fill_type="solid")
    header_font = Font(bold=True, color="000000")
    border_thin = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    header_cells = [f"A{header_row}", f"{desc_col}{header_row}", f"{qty_col}{header_row}", f"{unit_price_col}{header_row}", f"{total_price_col}{header_row}", f"{helper_unit_col}{header_row}", f"{helper_margin_col}{header_row}"]
    if include_part_number:
        header_cells.insert(1, f"B{header_row}")
    for addr in header_cells:
        ws[addr].fill = header_fill
        ws[addr].font = header_font
        ws[addr].alignment = Alignment(horizontal="center", vertical="center")
        ws[addr].border = border_thin
    ws.row_dimensions[header_row].height = 20

    # ===== DATA ROWS (start at 9) =====
    row_ptr = header_row + 1
    sr_no = 1
    currency_fmt = CURRENCY_NUMBER_FORMATS.get(currency_code, f'"{currency_code}" #,##0.00')
    margin_fmt = '0.00\\%'
    yellow = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    total_cells = []

    for (desc_text, qty_val, unit_val, subtotal_val) in items:
        ws[f"A{row_ptr}"] = sr_no
        if include_part_number:
            part_number_from_config = part_numbers_by_item.get(str(sr_no), "")
            part_number_from_heading = _extract_part_number_from_description(item_headings_by_item.get(str(sr_no), ""))
            part_number = part_number_from_heading or part_number_from_config
            ws[f"B{row_ptr}"] = part_number
            logger.debug(
                "Part number for item %s resolved to '%s' (heading='%s', config='%s', pricing_description='%s')",
                sr_no,
                part_number,
                item_headings_by_item.get(str(sr_no), ""),
                part_number_from_config,
                desc_text,
            )
        ws[f"{desc_col}{row_ptr}"] = desc_text
        ws[f"{qty_col}{row_ptr}"] = qty_val
    
        # Helper columns keep the original unit price and per-unit adjustment.
        ws[f"{helper_unit_col}{row_ptr}"].value = unit_val
        ws[f"{helper_unit_col}{row_ptr}"].number_format = currency_fmt

        # The margin helper column stays editable for each line item.
        ws[f"{helper_margin_col}{row_ptr}"].value = margin_percent
        ws[f"{helper_margin_col}{row_ptr}"].number_format = margin_fmt
    
        # ---- Unit Price shows the adjusted unit price
        ws[f"{unit_price_col}{row_ptr}"].value = f"=ROUND(((({helper_unit_col}{row_ptr}*${helper_margin_col}$2)+{helper_unit_col}{row_ptr})/(1-{helper_margin_col}{row_ptr}/100)),2)"
        ws[f"{unit_price_col}{row_ptr}"].number_format = currency_fmt
    
        # ---- Total Price = Qty * adjusted unit price
        ws[f"{total_price_col}{row_ptr}"].value = f"={qty_col}{row_ptr}*{unit_price_col}{row_ptr}"
        ws[f"{total_price_col}{row_ptr}"].number_format = currency_fmt
    
        # Styling
        data_cells = [f"A{row_ptr}", f"{desc_col}{row_ptr}", f"{qty_col}{row_ptr}", f"{unit_price_col}{row_ptr}", f"{total_price_col}{row_ptr}", f"{helper_unit_col}{row_ptr}", f"{helper_margin_col}{row_ptr}"]
        if include_part_number:
            data_cells.insert(1, f"B{row_ptr}")
        for addr in data_cells:
            ws[addr].fill = yellow
            ws[addr].border = border_thin
            ws[addr].alignment = Alignment(horizontal="center", vertical="top")
        ws[f"{desc_col}{row_ptr}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
        total_cells.append(f"{total_price_col}{row_ptr}")
        sr_no += 1
        row_ptr += 1

    first_data_row = header_row + 1
    last_data_row = row_ptr - 1
    if last_data_row >= first_data_row:
        ws[f"{helper_margin_col}2"] = f"=IFERROR(${helper_unit_col}$2/SUMPRODUCT(${qty_col}${first_data_row}:${qty_col}${last_data_row},${helper_unit_col}${first_data_row}:${helper_unit_col}${last_data_row}),0)"
    else:
        ws[f"{helper_margin_col}2"] = 0

    # ===== TOTAL ROW =====
    total_label_col = "C" if include_part_number else "B"
    total_value_col = total_price_col
    if include_part_number:
        ws.merge_cells(start_row=row_ptr, start_column=3, end_row=row_ptr, end_column=5)
    else:
        ws.merge_cells(start_row=row_ptr, start_column=2, end_row=row_ptr, end_column=4)
    ws[f"{total_label_col}{row_ptr}"] = "Total price"
    ws[f"{total_label_col}{row_ptr}"].alignment = Alignment(horizontal="right", vertical="center")
    ws[f"{total_label_col}{row_ptr}"].font = Font(bold=True, color="1F497D")

    ws[f"{total_value_col}{row_ptr}"] = f"=SUM({','.join(total_cells)})" if total_cells else 0
    ws[f"{total_value_col}{row_ptr}"].number_format = currency_fmt
    ws[f"{total_value_col}{row_ptr}"].font = Font(bold=True, color="1F497D")
    ws[f"{total_value_col}{row_ptr}"].alignment = Alignment(horizontal="center", vertical="center")
    ws[f"{total_value_col}{row_ptr}"].border = border_thin

    # Footer notes
    if currency_code == "AED":
        notes = [
            "Ø  Payment terms will be as per our finance approval.",
            "Ø  These prices are till DDP Dubai.",
            "Ø  Hardware will take 4-12 weeks delivery time from the date of Booking.",
            "Ø  These prices do not include Mindware installation of any kind.",
            "Ø  Change in Qty or partial shipment is not acceptable.",
            "Ø  PO Should be addressed to Mindware Technology Trading LLC and should be in AED.",
            "Ø  For all B2B orders complete end customer details should be mentioned on the PO.",
            "Ø  Orders once placed with Dell cannot be cancelled.",
            "Ø  Kindly also ensure to review the proposal specifications from your end and ensure that they match the requirements exactly as per the End User.",
            "Ø  Partial deliveries shall be acceptable",
            "Ø  For UAE DDP orders, the PO should be addressed to Mindware Technology Trading LLC and for Ex-Jablal Ali orders, it should be addressed to Mindware FZ.",
            "Ø  Please ensure that the PO includes the name of the end-user.",
            "Ø  Please ensure that the PO includes the Incoterms (DDP or Ex-Works Jabal Ali).",
            "Ø  Due to global market fluctuations, all prices are subject to change without prior notice, and lead times may also be affected. All quotations are non-binding and remain subject to final validation and confirmation by Dell.",
            "Ø  As the geopolitical situation in the Middle East continues to evolve, it has introduced significant instability to international shipping routes. These unforeseen and extraordinary circumstances, which remain entirely beyond our control, constitute a Force Majeure event. We are formally notifying you of the resulting impact on our current and future shipments.",
        ]
    else:
        notes = [
            "Incoterms:",
            "",
            "Payment Terms:",
            "",
            "Quote validity:",
            "",
            "Estimated Delivery Time from the date of booking:",
            "",
            "These prices do not include installation of any kind",
            "Change in Qty or partial shipment is not acceptable",
            "For all B2B orders complete end customer details should be mentioned on the PO",
            f"PO Should be addressed to Mindware FZ LLC and should be in {currency_code}",
            "Orders once placed with Dell cannot be cancelled",
            "",
            "And as an important note – All items are not proposed with any Professional Services to cater for installation.",
            "",
            "Please note that these prices are granted for this QTY and deal & Dell cannot guarantee same prices if QTY is reduced , it will also have to be one shot order.",
            "",
            "Kindly also ensure to review the proposal specifications from your end and ensure that they match the requirements exactly as per the End User.",
        ]
    footer_row = max(row_ptr + 2, 22)
    for line in notes:
        ws.merge_cells(start_row=footer_row, start_column=2, end_row=footer_row, end_column=6)
        ws.cell(footer_row, 2).value = line
        ws.cell(footer_row, 2).alignment = Alignment(wrap_text=True, vertical="top")
        footer_row += 1

    ws.freeze_panes = None

    # ===== Sheet 2: Configuration =====
    ws2 = wb.create_sheet("Configuration")
    ws2.sheet_view.showGridLines = False

    use_service_layout = bool(service_fields_by_item) and (currency_code == "AED") and (not is_pdf) and (not config_rows)
    if use_service_layout:
        ws2.column_dimensions["A"].width = 14  # Item #
        ws2.column_dimensions["B"].width = 18  # Module
        ws2.column_dimensions["C"].width = 80  # Description
        ws2.column_dimensions["D"].width = 18  # Service Tag
        ws2.column_dimensions["E"].width = 18  # Service Start Date
        ws2.column_dimensions["F"].width = 18  # Service End Date
        ws2.column_dimensions["G"].width = 10  # Qty
    else:
        ws2.column_dimensions["A"].width = 22  # Item #
        ws2.column_dimensions["B"].width = 70  # Module
        ws2.column_dimensions["C"].width = 100  # Description
        ws2.column_dimensions["D"].width = 20  # SKU
        ws2.column_dimensions["E"].width = 14  # Tax Type
        ws2.column_dimensions["F"].width = 10  # Qty

    r2 = 1
    title_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    section_fill = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")
    thin_gray = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )

    def write_table_header(row_index: int):
        if use_service_layout:
            headers = [
                ("A", "Item #"),
                ("B", "Module"),
                ("C", "Description"),
                ("D", "Service Tag:"),
                ("E", "Service Start Date:"),
                ("F", "Service End Date:"),
                ("G", "Qty"),
            ]
        else:
            headers = [
                ("A", "Item #"),
                ("B", "Module"),
                ("C", "Description"),
                ("D", "SKU"),
                ("E", "Tax Type"),
                ("F", "Qty"),
            ]

        for col, label in headers:
            addr = f"{col}{row_index}"
            ws2[addr] = label
            ws2[addr].font = Font(bold=True)
            ws2[addr].fill = title_fill
            ws2[addr].alignment = Alignment(horizontal="center", vertical="center")
            ws2[addr].border = Border(
                left=Side(style="thin", color="000000"),
                right=Side(style="thin", color="000000"),
                top=Side(style="thin", color="000000"),
                bottom=Side(style="thin", color="000000"),
            )
        ws2.row_dimensions[row_index].height = 20

    # Header row
    write_table_header(r2)
    r2 += 1

    if use_service_layout:
        qty_by_item: Dict[str, str] = {}
        for idx, it in enumerate(items, start=1):
            try:
                qty_by_item[str(idx)] = str(int(it[1]))
            except Exception:
                qty_by_item[str(idx)] = _cell_to_text(it[1])

        total_items = max(
            len(item_descs_order),
            len(service_fields_by_item),
            len(item_headings_by_item) if include_part_number else 0,
        )
        if total_items == 0:
            ws2.merge_cells(start_row=r2, start_column=1, end_row=r2, end_column=7)
            ws2[f"A{r2}"] = "(No service details found)"
            ws2[f"A{r2}"].alignment = Alignment(horizontal="left", vertical="center")
            r2 += 1
        else:
            for idx in range(1, total_items + 1):
                item_key = str(idx)
                display_item_no = item_display_numbers_by_item.get(item_key, str(idx))
                heading = item_headings_by_item.get(item_key, "")
                if not heading and idx - 1 < len(item_descs_order):
                    heading = item_descs_order[idx - 1]
                details = service_fields_by_item.get(item_key, {})
                ws2[f"A{r2}"] = display_item_no
                ws2[f"B{r2}"] = ""
                ws2[f"C{r2}"] = heading
                ws2[f"D{r2}"] = details.get("service_tag", "")
                ws2[f"E{r2}"] = details.get("service_start_date", "")
                ws2[f"F{r2}"] = details.get("service_end_date", "")
                ws2[f"G{r2}"] = qty_by_item.get(item_key, "")

                for col in ("A", "B", "C", "D", "E", "F", "G"):
                    ws2[f"{col}{r2}"].border = thin_gray
                    ws2[f"{col}{r2}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                r2 += 1
    elif not config_rows:
        ws2.merge_cells(start_row=r2, start_column=1, end_row=r2, end_column=6)
        ws2[f"A{r2}"] = "(No configuration details found)"
        ws2[f"A{r2}"].alignment = Alignment(horizontal="left", vertical="center")
        r2 += 1
    else:
        config_rows_by_item: Dict[str, List[tuple]] = {}
        for row in config_rows:
            config_rows_by_item.setdefault(row[0], []).append(row)

        total_items = max(
            len(item_descs_order),
            len(config_rows_by_item),
            len(item_headings_by_item) if include_part_number else 0,
        )
        for idx in range(1, total_items + 1):
            item_key = str(idx)
            rows_for_item = config_rows_by_item.get(item_key, [])
            display_item_no = item_display_numbers_by_item.get(item_key, str(idx))
            fallback_heading = item_headings_by_item.get(item_key, "")
            if not fallback_heading and idx - 1 < len(item_descs_order):
                fallback_heading = item_descs_order[idx - 1]
            if not fallback_heading:
                fallback_heading = f"Item {idx}"
            extracted_heading = rows_for_item[0][1] if rows_for_item and rows_for_item[0][1] else ""
            item_heading = fallback_heading if extracted_heading in ("", f"Item {idx}") else extracted_heading

            ws2.merge_cells(start_row=r2, start_column=1, end_row=r2, end_column=6)
            ws2[f"A{r2}"] = f"Item {display_item_no}"
            ws2[f"A{r2}"].font = Font(bold=True, color="1F497D")
            ws2[f"A{r2}"].alignment = Alignment(horizontal="left", vertical="center")
            r2 += 1

            ws2.merge_cells(start_row=r2, start_column=1, end_row=r2, end_column=6)
            ws2[f"A{r2}"] = item_heading
            ws2[f"A{r2}"].font = Font(italic=True, color="1F497D")
            ws2[f"A{r2}"].alignment = Alignment(horizontal="left", vertical="center")
            r2 += 1

            if not rows_for_item:
                ws2.merge_cells(start_row=r2, start_column=2, end_row=r2, end_column=6)
                ws2[f"B{r2}"] = "(No configuration details found for this item)"
                ws2[f"B{r2}"].font = Font(italic=True, color="7F7F7F")
                ws2[f"B{r2}"].alignment = Alignment(horizontal="left", vertical="center")
                for col in ("A", "B", "C", "D", "E", "F"):
                    ws2[f"{col}{r2}"].border = thin_gray
                r2 += 1
                continue

            for row_data in rows_for_item:
                if len(row_data) >= 7:
                    _, _, module, dsc, sku, tax, qty = row_data
                else:
                    _, _, module, dsc, sku, tax = row_data
                    qty = ""
                if _is_config_section_row(module, dsc, sku, tax):
                    ws2[f"A{r2}"] = ""
                    ws2.merge_cells(start_row=r2, start_column=2, end_row=r2, end_column=6)
                    ws2[f"B{r2}"] = module
                    ws2[f"B{r2}"].font = Font(bold=True, color="1F1F1F")
                    ws2[f"B{r2}"].fill = section_fill
                    ws2[f"B{r2}"].alignment = Alignment(horizontal="left", vertical="center")
                    for col in ("A", "B", "C", "D", "E", "F"):
                        ws2[f"{col}{r2}"].border = thin_gray
                    r2 += 1
                    continue

                ws2[f"A{r2}"] = ""
                ws2[f"B{r2}"] = module
                ws2[f"C{r2}"] = dsc
                ws2[f"D{r2}"] = sku
                ws2[f"E{r2}"] = tax
                ws2[f"F{r2}"] = qty
                for col in ("A", "B", "C", "D", "E", "F"):
                    ws2[f"{col}{r2}"].alignment = Alignment(vertical="top", wrap_text=True)
                    ws2[f"{col}{r2}"].border = thin_gray
                r2 += 1

            r2 += 1
            

    ws2.freeze_panes = None
    

    # Save to bytes
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()
