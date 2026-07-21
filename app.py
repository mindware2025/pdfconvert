
import base64
import csv
import logging
import os
import random
from pathlib import Path
import zipfile
import streamlit as st
import pandas as pd
import io
import traceback
import hashlib
from datetime import datetime
from openpyxl.styles import PatternFill
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import gspread
from google.oauth2.service_account import Credentials
from extractors.ibm_cn_ksa import build_ibm_ksa_output_filename, prepare_ibm_ksa_excel_bytes, process_ibm_ksa_pdfs
from sales.dell_southcomp import render_southcomp_tool
from extractors.barcodeper50 import barcode_tooll
from extractors.aws import AWS_OUTPUT_COLUMNS, build_dnts_cnts_rows, process_multiple_aws_pdfs
from extractors.google_dnts import extract_invoice_info, extract_table_from_text, make_dnts_header_row, DNTS_HEADER_COLS, DNTS_ITEM_COLS
from extractors.insurance import process_insurance_excel
from extractors.insurance2  import process_grouped_customer_files
#from sales.mibb_quotation import create_mibb_excel, extract_mibb_header_from_pdf, extract_mibb_table_from_pdf
from extractors.oracle import prepare_excel_bytes, process_oracle_pdfs_cached
from extractors.lenovo_cn import build_ksa_output_filename, build_output_filename, prepare_ksa_excel_bytes, process_lenovo_credit_pdfs, process_lenovo_ksa_pdfs
from extractors.freight_forwarder_processor import (
    JVConfig,
    create_excel_file as create_freight_forwarder_excel_file,
    process_freight_forwarder_pdfs,
)
from amal.processor import build_output_workbook as build_comm_generator_workbook
from amal.processor import process_uploaded_pairs as process_comm_generator_pairs

from sales.dell_extended_services import generate_dell_extended_services_quote
from sales.dell import build_dell_output_filename, detect_dell_standard_variant, generate_dell_quote
from sales.dell_extended_services import build_dell_extended_services_output_filename
from sales.dell_orion import build_dell_orion_output_filename, generate_orion_quote
from sales.lenovo import build_lenovo_output_filename, generate_lenovo_quote, parse_lenovo_quote_pdf
from sales.mibb import check_mibb_hardware_quote_match, correct_mibb_descriptions, create_mibb_excel, create_mibb_hardware_excel, create_mibb_tls_excel, extract_mibb_hardware_table_from_excel, extract_mibb_header_from_pdf, extract_mibb_table_from_pdf, extract_mibb_terms_from_pdf, extract_mibb_tls_from_excel
from sales.quotetemplate import detect_dell_template
from utils.helpers import format_amount, format_invoice_date, format_month_year
from dotenv import load_dotenv
from ibm import extract_ibm_data_from_pdf, create_styled_excel, create_styled_excel_template2, correct_descriptions, extract_last_page_text
from ibm_template2 import extract_ibm_template2_from_pdf, get_extraction_debug
from sales.ibm_v2 import compare_mep_and_cost
from template_detector import detect_ibm_template
import logging
logging.basicConfig(level=logging.INFO)
load_dotenv()
from extractors.google_invoice import extract_table_from_text as extract_invoice_table, extract_invoice_info as extract_invoice_info_invoice, GOOGLE_INVOICE_COLS
from extractors.dell_invoice import (
    extract_invoice_info as extract_dell_invoice_info,
    extract_table_from_text as extract_dell_table,
    DELL_INVOICE_COLS,
    PRE_ALERT_HEADERS,
    build_pre_alert_rows,
    read_master_mapping,
)

from oauth2client.service_account import ServiceAccountCredentials
from claims_automation import (
    build_output_rows_from_source1,
    write_output_excel,
    read_source1_rows,
    read_master1_map,
    read_source2_rows,
    build_debit_rows_from_source2,
    read_master2_entries,
    derive_defaults_from_source1,
)
import plotly.express as px
from dashboard import CATALOG_HEADERS, TOOL_CATALOG_SHEET_NAME, render_dashboard

SHEET_JSON = "tool-mindware-0d87ca5562ad.json"  # Path to your downloaded JSON
SHEET_NAME = "mindware tool"

scope = ["https://www.googleapis.com/auth/spreadsheets", "https://spreadsheets.google.com/feeds",
         "https://www.googleapis.com/auth/drive"]

# Read credentials from Streamlit secrets
service_account_info = st.secrets["gcp_service_account"]
creds = Credentials.from_service_account_info(service_account_info, scopes=scope)

# ✅ Authorize and open sheet
gc = gspread.authorize(creds)

workbook = gc.open(SHEET_NAME)
tool_sheet = workbook.worksheet("Sheet1")     # Main usage sheet
env = st.secrets.get("env", "live")  # Default to live if not set

USAGE_HEADERS = ["Tool", "Month", "Usage Count", "Team", "PDF Count", "Excel Count"]
EXCEL_EXTENSIONS = (".xlsx", ".xls", ".xlsm")
RUN_LOG_SHEET_NAME = "Run Log"
RUN_LOG_HEADERS = ["Timestamp", "Tool", "Month", "Team", "PDF Count", "Excel Count"]


def get_or_create_worksheet(title, rows=1000, cols=10):
    try:
        return workbook.worksheet(title)
    except gspread.WorksheetNotFound:
        return workbook.add_worksheet(title=title, rows=rows, cols=cols)


run_log_sheet = get_or_create_worksheet(RUN_LOG_SHEET_NAME)

# "Tool Catalog": hand-maintained in Sheets (replaces the manual Excel handoff
# to management) — one row per tool with its time-saved assumptions. The app
# never writes rows here, only makes sure the tab and header row exist.
catalog_sheet = get_or_create_worksheet(TOOL_CATALOG_SHEET_NAME, rows=200, cols=len(CATALOG_HEADERS))


def ensure_tool_catalog_columns():
    headers = catalog_sheet.row_values(1)
    if not headers:
        catalog_sheet.append_row(CATALOG_HEADERS)
        return
    for idx, header in enumerate(CATALOG_HEADERS, start=1):
        if len(headers) < idx or headers[idx - 1] != header:
            catalog_sheet.update_cell(1, idx, header)


ensure_tool_catalog_columns()


def ensure_usage_sheet_columns():
    headers = tool_sheet.row_values(1)
    if not headers:
        tool_sheet.append_row(USAGE_HEADERS)
        return

    for idx, header in enumerate(USAGE_HEADERS, start=1):
        if len(headers) < idx or headers[idx - 1] != header:
            tool_sheet.update_cell(1, idx, header)


def count_uploaded_files(files, extensions):
    if not files:
        return 0

    file_list = files if isinstance(files, list) else [files]
    return sum(
        1 for uploaded_file in file_list
        if getattr(uploaded_file, "name", "").lower().endswith(extensions)
    )


def ensure_run_log_sheet_columns():
    headers = run_log_sheet.row_values(1)
    if not headers:
        run_log_sheet.append_row(RUN_LOG_HEADERS)
        return

    for idx, header in enumerate(RUN_LOG_HEADERS, start=1):
        if len(headers) < idx or headers[idx - 1] != header:
            run_log_sheet.update_cell(1, idx, header)


def log_run_details(tool_name, team, pdf_count, excel_count):
    ensure_run_log_sheet_columns()
    now = datetime.now()
    run_log_sheet.append_row([
        now.strftime("%Y-%m-%d %H:%M:%S"),
        tool_name,
        now.strftime("%b-%Y"),
        team,
        pdf_count,
        excel_count,
    ])


def update_usage(tool_name, team, pdf_count=0, excel_count=0):
    ensure_usage_sheet_columns()
    log_run_details(tool_name, team, pdf_count, excel_count)
    month = datetime.today().strftime("%b-%Y")
    all_rows = tool_sheet.get_all_records()
    found = False

    for i, row in enumerate(all_rows, start=2):
        if row.get("Tool") == tool_name and row.get("Month") == month and row.get("Team") == team:
            current_count = row.get("Usage Count", 0) or 0
            current_pdf_count = row.get("PDF Count", 0) or 0
            current_excel_count = row.get("Excel Count", 0) or 0
            tool_sheet.update_cell(i, 3, current_count + 1)
            tool_sheet.update_cell(i, 5, current_pdf_count + pdf_count)
            tool_sheet.update_cell(i, 6, current_excel_count + excel_count)
            found = True
            break

    if not found:
        tool_sheet.append_row([tool_name, month, 1, team, pdf_count, excel_count])




st.set_page_config(
    page_title="Mindware Tool",
    layout="wide",
    initial_sidebar_state="collapsed"
)
# ----------- Constants -----------
DEFAULTS = {
    "supp_code": "SDIG005",
    "curr_code": "USD",
    "form_code": 0,
    "doc_src_locn": "UJ000",
    "location_code": "UJ200"
}
#CORRECT_USERNAME = "admin"
#CORRECT_PASSWORD = "admin"
CORRECT_USERNAME = os.getenv("NAME")
CORRECT_PASSWORD = os.getenv("PASSWORD")

if "login_state" not in st.session_state:
    st.session_state.login_state = "login" 
def show_login():
    streams = "".join(
        f'<div class="mw-stream" style="'
        f'left:{random.uniform(1, 99):.2f}%;'
        f'--h:{random.uniform(90, 240):.0f}px;'
        f'--delay:{random.uniform(0, 9):.2f}s;'
        f'--duration:{random.uniform(6, 14):.1f}s;'
        f'--o:{random.uniform(0.2, 0.55):.2f};'
        f'"></div>'
        for _ in range(18)
    )
    letters = "".join(
        f'<span class="mw-letter" style="--i:{i}">{ch}</span>'
        for i, ch in enumerate("MINDBOT")
    )

    st.markdown(f"""
    <style>
    :root {{
        --mw-navy: #0c1930;
        --mw-steel: #6ea8ff;
        --mw-glow: #7cc7ff;
        --mw-white: #f2f7ff;
        --mw-dim: rgba(157,193,255,0.65);
    }}

    /* Make Streamlit chrome transparent / out of the way (login only) */
    [data-testid="stAppViewContainer"],
    [data-testid="stApp"],
    [data-testid="stHeader"] {{
        background: transparent !important;
    }}
    [data-testid="stToolbar"], #MainMenu {{
        display: none !important;
    }}
    .block-container {{
        padding-top: 6vh !important;
    }}

    @keyframes mw-fall {{ to {{ transform: translateY(calc(100vh + 300px)); }} }}
    @keyframes mw-scan {{
        0%   {{ top: -2%; }}
        60%  {{ top: 102%; }}
        100% {{ top: 102%; }}
    }}
    @keyframes mw-bar {{
        0%, 100% {{ transform: scaleY(1); }}
        50%      {{ transform: scaleY(0.45); }}
    }}
    @keyframes mw-letter-in {{
        to {{ opacity: 1; transform: none; }}
    }}
    @keyframes mw-card-in {{
        from {{ opacity: 0; transform: translateY(24px); }}
        to   {{ opacity: 1; transform: none; }}
    }}
    @keyframes mw-card-scan {{
        from {{ top: -70px; }}
        to   {{ top: 110%; }}
    }}
    @keyframes mw-pulse {{
        0%, 100% {{ opacity: 1; transform: scale(1); }}
        50%      {{ opacity: .3; transform: scale(.65); }}
    }}
    @keyframes mw-type {{ from {{ width: 0; }} to {{ width: 28ch; }} }}
    @keyframes mw-caret {{ 50% {{ border-color: transparent; }} }}

    /* ---- Fullscreen scene: deep Mindware navy ---- */
    .mw-scene {{
        position: fixed;
        inset: 0;
        z-index: -5;
        overflow: hidden;
        background: radial-gradient(1200px 700px at 50% -10%, #14294a 0%, #0c1930 48%, #060e1c 100%);
    }}
    .mw-dots {{
        position: absolute;
        inset: 0;
        background-image: radial-gradient(rgba(126,167,255,0.10) 1px, transparent 1.4px);
        background-size: 28px 28px;
        -webkit-mask-image: radial-gradient(ellipse at 50% 35%, black 0%, transparent 75%);
        mask-image: radial-gradient(ellipse at 50% 35%, black 0%, transparent 75%);
    }}
    .mw-stream {{
        position: absolute;
        top: -260px;
        width: 1px;
        height: var(--h);
        opacity: var(--o);
        background: linear-gradient(180deg, transparent, rgba(110,168,255,0.8) 50%, transparent);
        animation: mw-fall var(--duration) linear var(--delay) infinite;
    }}
    .mw-scanline {{
        position: absolute;
        left: 0; right: 0;
        height: 2px;
        background: linear-gradient(90deg, transparent 10%, rgba(124,199,255,0.4), transparent 90%);
        filter: blur(1px);
        animation: mw-scan 10s ease-in-out infinite;
    }}
    .mw-vignette {{
        position: absolute; inset: 0;
        background: radial-gradient(ellipse at center, transparent 45%, rgba(4,9,20,0.6) 100%);
    }}

    /* ---- Brand header: live signal-bars mark + wordmark ---- */
    .mw-header {{
        max-width: 460px;
        margin: 3vh auto 24px auto;
        text-align: center;
        animation: mw-card-in .6s ease both;
    }}
    .mw-mark {{
        display: flex;
        align-items: flex-end;
        justify-content: center;
        gap: 6px;
        height: 54px;
        margin: 0 auto 18px auto;
    }}
    .mw-mark .mw-bar {{
        width: 6px;
        border-radius: 3px;
        background: linear-gradient(180deg, #ffffff, #9cc3ff);
        box-shadow: 0 0 14px rgba(124,199,255,0.55);
        transform-origin: bottom;
        animation: mw-bar 1.7s ease-in-out infinite;
    }}
    .mw-mark .mw-bar:nth-child(1) {{ height: 100%; animation-delay: 0s;   }}
    .mw-mark .mw-bar:nth-child(2) {{ height: 42%;  animation-delay: .17s; }}
    .mw-mark .mw-bar:nth-child(3) {{ height: 72%;  animation-delay: .34s; }}
    .mw-mark .mw-bar:nth-child(4) {{ height: 42%;  animation-delay: .51s; }}
    .mw-mark .mw-bar:nth-child(5) {{ height: 100%; animation-delay: .68s; }}
    .mw-wordmark {{
        display: flex;
        justify-content: center;
        margin: 0 0 8px 0;
        padding-left: 12px; /* offset trailing letter-spacing so it stays centered */
    }}
    .mw-letter {{
        font-size: 34px;
        font-weight: 600;
        letter-spacing: 12px;
        color: var(--mw-white);
        text-shadow: 0 0 26px rgba(124,199,255,0.35);
        opacity: 0;
        transform: translateY(12px);
        animation: mw-letter-in .5s cubic-bezier(.2,.9,.3,1) forwards;
        animation-delay: calc(var(--i) * 70ms + .25s);
    }}
    .mw-subtitle {{
        color: var(--mw-dim);
        margin: 0 0 16px 0;
        font-weight: 500;
        font-size: 12px;
        letter-spacing: 5px;
        text-transform: uppercase;
    }}
    .mw-status {{
        display: flex; align-items: center; justify-content: center; gap: 8px;
    }}
    .mw-status-dot {{
        width: 7px; height: 7px; border-radius: 50%;
        background: var(--mw-glow);
        box-shadow: 0 0 8px 2px rgba(124,199,255,0.7);
        animation: mw-pulse 1.8s ease-in-out infinite;
    }}
    .mw-typewriter {{
        font-family: 'Consolas', 'Courier New', monospace;
        font-size: 12.5px;
        color: var(--mw-dim);
        white-space: nowrap;
        overflow: hidden;
        border-right: 2px solid var(--mw-glow);
        width: 0;
        animation: mw-type 2.1s steps(28, end) .5s forwards, mw-caret .75s step-end infinite;
    }}

    /* ---- Login card: secure-terminal window ---- */
    div[data-testid="stForm"] {{
        max-width: 420px;
        margin: 0 auto;
        position: relative;
        overflow: hidden;
        background: rgba(10,21,40,0.78);
        backdrop-filter: blur(20px) saturate(150%);
        -webkit-backdrop-filter: blur(20px) saturate(150%);
        border-radius: 14px;
        padding: 22px 26px 16px 26px;
        border: 1px solid rgba(110,168,255,0.22);
        box-shadow: 0 26px 60px rgba(2,6,16,0.65), inset 0 1px 0 rgba(255,255,255,0.05);
        animation: mw-card-in .7s cubic-bezier(.16,1,.3,1) .15s both;
        transition: border-color .3s ease, box-shadow .3s ease;
    }}
    div[data-testid="stForm"]:focus-within {{
        border-color: rgba(124,199,255,0.5);
        box-shadow: 0 26px 60px rgba(2,6,16,0.65), 0 0 24px rgba(79,143,247,0.25), inset 0 1px 0 rgba(255,255,255,0.05);
    }}
    div[data-testid="stForm"]::after {{
        content: "";
        position: absolute;
        left: 0; right: 0;
        top: -70px;
        height: 60px;
        background: linear-gradient(180deg, transparent, rgba(124,199,255,0.07), transparent);
        animation: mw-card-scan 7s linear 1.5s infinite;
        pointer-events: none;
    }}

    /* Terminal title bar inside the card */
    .mw-term {{
        display: flex;
        align-items: center;
        gap: 10px;
        padding-bottom: 12px;
        margin-bottom: 4px;
        border-bottom: 1px solid rgba(110,168,255,0.14);
    }}
    .mw-term-dots {{ display: flex; gap: 5px; }}
    .mw-term-dots i {{
        width: 8px; height: 8px; border-radius: 50%;
        background: rgba(110,168,255,0.25);
    }}
    .mw-term-dots i:nth-child(2) {{ background: rgba(110,168,255,0.45); }}
    .mw-term-dots i:nth-child(3) {{ background: rgba(124,199,255,0.8); }}
    .mw-term-title {{
        font-family: 'Consolas', 'Courier New', monospace;
        font-size: 11px;
        letter-spacing: 1px;
        color: var(--mw-dim);
    }}

    /* Labels */
    label p, label {{
        color: rgba(210,228,255,0.85) !important;
        font-weight: 600 !important;
        font-size: 13px !important;
    }}

    /* Inputs */
    div[data-testid="stTextInput"] input {{
        background: rgba(255,255,255,0.05) !important;
        border-radius: 10px !important;
        border: 1px solid rgba(110,168,255,0.2) !important;
        color: #ffffff !important;
        caret-color: var(--mw-glow) !important;
        padding: 13px 16px !important;
        transition: box-shadow .25s ease, border-color .25s ease, background .25s ease;
    }}
    div[data-testid="stTextInput"] input::placeholder {{
        color: rgba(210,228,255,0.35) !important;
    }}
    div[data-testid="stTextInput"] input:focus {{
        border-color: #4f8ff7 !important;
        background: rgba(255,255,255,0.08) !important;
        box-shadow: 0 0 0 4px rgba(79,143,247,0.16), 0 0 16px rgba(79,143,247,0.3) !important;
    }}

    /* Fix browser autofill */
    input:-webkit-autofill,
    input:-webkit-autofill:hover,
    input:-webkit-autofill:focus {{
        -webkit-text-fill-color: #ffffff !important;
        transition: background-color 9999s ease-in-out 0s;
        box-shadow: 0 0 0px 1000px rgba(255,255,255,0.06) inset !important;
        border: 1px solid rgba(110,168,255,0.2) !important;
    }}

    /* Button */
    div[data-testid="stFormSubmitButton"] button {{
        width: 100%;
        border: none !important;
        border-radius: 10px !important;
        padding: 13px 18px !important;
        font-weight: 700 !important;
        text-transform: uppercase;
        letter-spacing: 2.5px;
        font-size: 13px !important;
        color: white !important;
        background: linear-gradient(90deg, #1d4ed8, #2f6ff2, #4f8ff7, #1d4ed8) !important;
        background-size: 300% auto !important;
        box-shadow: 0 12px 26px rgba(37,99,235,0.35);
        transition: background-position .6s ease, transform .15s ease, box-shadow .3s ease;
    }}
    div[data-testid="stFormSubmitButton"] button:hover {{
        background-position: 100% center !important;
        transform: translateY(-2px);
        box-shadow: 0 16px 32px rgba(79,143,247,0.45);
    }}
    div[data-testid="stFormSubmitButton"] button:active {{
        transform: translateY(0) scale(.98);
    }}

    .mw-heart {{
        display: inline-block;
        font-size: 13px;
        animation: mw-heartbeat 1.6s ease-in-out infinite;
    }}
    @keyframes mw-heartbeat {{
        0%, 48%, 100% {{ transform: scale(1); }}
        12% {{ transform: scale(1.3); }}
        24% {{ transform: scale(1); }}
        36% {{ transform: scale(1.25); }}
    }}

    /* ---- Cute mini-bot mascot ---- */
    .mw-bot {{
        width: 44px;
        margin: 18px auto 8px auto;
        animation: mw-bob 3s ease-in-out infinite;
    }}
    .mw-bot-antenna {{
        width: 2px; height: 9px;
        margin: 0 auto;
        background: rgba(110,168,255,0.6);
        position: relative;
    }}
    .mw-bot-antenna::before {{
        content: "";
        position: absolute;
        top: -6px; left: 50%;
        transform: translateX(-50%);
        width: 6px; height: 6px;
        border-radius: 50%;
        background: var(--mw-glow);
        box-shadow: 0 0 8px 2px rgba(124,199,255,0.7);
        animation: mw-pulse 1.8s ease-in-out infinite;
    }}
    .mw-bot-head {{
        width: 40px; height: 30px;
        margin: 0 auto;
        background: rgba(16,30,55,0.9);
        border: 1px solid rgba(110,168,255,0.4);
        border-radius: 10px;
        display: flex;
        align-items: center;
        justify-content: center;
        gap: 7px;
        position: relative;
        box-shadow: 0 4px 14px rgba(2,6,16,0.5), inset 0 1px 0 rgba(255,255,255,0.06);
    }}
    .mw-bot-eye {{
        width: 6px; height: 6px;
        border-radius: 50%;
        background: var(--mw-glow);
        box-shadow: 0 0 6px rgba(124,199,255,0.8);
        margin-top: -6px;
        animation: mw-blink 4.2s ease-in-out infinite;
    }}
    .mw-bot-mouth {{
        position: absolute;
        bottom: 5px; left: 50%;
        transform: translateX(-50%);
        width: 12px; height: 6px;
        border: 2px solid rgba(124,199,255,0.75);
        border-top: none;
        border-radius: 0 0 12px 12px;
    }}
    @keyframes mw-bob {{
        0%, 100% {{ transform: translateY(0); }}
        50%      {{ transform: translateY(-4px); }}
    }}
    @keyframes mw-blink {{
        0%, 90%, 96%, 100% {{ transform: scaleY(1); }}
        93% {{ transform: scaleY(0.1); }}
    }}

    /* Rotating cute footer messages */
    .mw-rotator {{
        position: relative;
        height: 20px;
        max-width: 420px;
        margin: 0 auto;
        text-align: center;
        font-size: 12px;
        color: rgba(157,193,255,0.6);
        letter-spacing: .5px;
    }}
    .mw-rotator span {{
        position: absolute;
        left: 0; right: 0;
        opacity: 0;
        animation: mw-msg 9s ease-in-out infinite;
    }}
    .mw-rotator span:nth-child(2) {{ animation-delay: 3s; }}
    .mw-rotator span:nth-child(3) {{ animation-delay: 6s; }}
    @keyframes mw-msg {{
        0%  {{ opacity: 0; transform: translateY(8px); }}
        4%  {{ opacity: 1; transform: none; }}
        30% {{ opacity: 1; }}
        36% {{ opacity: 0; transform: translateY(-8px); }}
        100% {{ opacity: 0; }}
    }}

    @media (prefers-reduced-motion: reduce) {{
        * {{ animation: none !important; transition: none !important; }}
        .mw-letter {{ opacity: 1; transform: none; }}
        .mw-typewriter {{ width: 28ch; }}
        .mw-rotator span {{ opacity: 0; }}
        .mw-rotator span:first-child {{ opacity: 1; }}
    }}
    </style>

    <div class="mw-scene">
        <div class="mw-dots"></div>
        {streams}
        <div class="mw-scanline"></div>
        <div class="mw-vignette"></div>
    </div>

    <div class="mw-header">
        <div class="mw-mark">
            <div class="mw-bar"></div>
            <div class="mw-bar"></div>
            <div class="mw-bar"></div>
            <div class="mw-bar"></div>
            <div class="mw-bar"></div>
        </div>
        <div class="mw-wordmark">{letters}</div>
        <div class="mw-subtitle">Productivity Suite</div>
        
    </div>
    """, unsafe_allow_html=True)

    with st.form("login_form", clear_on_submit=False):
        st.markdown(
            '<div class="mw-term">'
            '<span class="mw-term-dots"><i></i><i></i><i></i></span>'
            '<span class="mw-term-title">mindbot</span>'
            '</div>',
            unsafe_allow_html=True,
        )
        username = st.text_input("Username", key="login_user", placeholder="Enter your username")
        password = st.text_input("Password", type="password", key="login_pass", placeholder="Enter your password")
        submitted = st.form_submit_button("Sign In", type="primary")

    st.markdown(
        '<div class="mw-bot">'
        '<div class="mw-bot-antenna"></div>'
        '<div class="mw-bot-head">'
        '<span class="mw-bot-eye"></span><span class="mw-bot-eye"></span>'
        '<div class="mw-bot-mouth"></div>'
        '</div>'
        '</div>'
        '<div class="mw-rotator">'
        '<span>Made with <b class="mw-heart">💙</b> by Mindware</span>'
        '<span>Fueled by coffee ☕ and a few thousand PDFs 📄</span>'
        '<span>Beep boop — happy to see you! 🤖</span>'
        '</div>',
        unsafe_allow_html=True,
    )

    if submitted:
        if username == CORRECT_USERNAME and password == CORRECT_PASSWORD:
            st.session_state.login_state = "success"
        else:
            st.session_state.login_state = "fail"
        st.rerun()



def show_fail():
    st.error("Oops! Wrong credentials... Nice try, but no entry! 😜")
    # Display image from local file
    img_path = Path("img.png")  # or just Path("image.png") if in the root
    if img_path.exists():
        st.image(str(img_path), caption="Access Denied", use_column_width=True)
    else:
        st.warning("image.png not found. Please check the path.")

    if st.button("Back to Login", key="back_login"):
        st.session_state.login_state = "login"

# --- TV / wall dashboard: read-only view for office screens (?view=dashboard).
# Placed before the login gate on purpose: it shows numbers only, no tools.
if st.query_params.get("view") == "dashboard":
    render_dashboard(run_log_sheet, catalog_sheet, tv_mode=True)
    st.stop()

if st.session_state.login_state == "login":
    show_login()
    st.stop()
elif st.session_state.login_state == "fail":
    show_fail()
    st.stop()

# Main app CSS (only after login — avoids flash/flicker when typing on login page)
st.markdown("""
    <style>
    /* Hide Share, GitHub, Settings icons on top-right */
    [data-testid="stToolbar"] {
        display: none !important;
    }
    /* Ensure sidebar is always visible */
    [data-testid="stSidebar"] {
        visibility: visible !important;
        min-width: 280px !important;
    }
    /* Optional: adjust sidebar content font */
    [data-testid="stSidebar"] * {
        font-family: 'Google Sans', sans-serif !important;
    }
    /* General styling */
    html, body, [class*="css"] {
        font-family: 'Google Sans', sans-serif !important;
        background: #f6f8fa;
        color: #202124;
    }
    h1, h2, h3 {
        color: #1a73e8;
        font-weight: 700;
        letter-spacing: -1px;
        margin-bottom: 0.5em;
    }
    .stButton > button, .stDownloadButton > button {
        background: linear-gradient(90deg, #1a73e8, #188038);
        color: white;
        border-radius: 8px;
        font-weight: 600;
        font-size: 16px;
        border: none;
        padding: 12px 28px;
        margin-top: 10px;
        margin-bottom: 10px;
        box-shadow: 0 2px 8px rgba(26, 115, 232, 0.08);
        transition: background 0.2s;
    }
    .stButton > button:hover, .stDownloadButton > button:hover {
        background: linear-gradient(90deg, #188038, #1a73e8);
        color: #fff;
    }
    .stDataFrame {
        background: #fff;
        border-radius: 10px;
        box-shadow: 0 2px 8px rgba(26, 115, 232, 0.06);
        margin-bottom: 2em;
    }
    </style>
""", unsafe_allow_html=True)

# --- In-app management dashboard view (behind login) ---
if st.session_state.get("app_view") == "dashboard":
    if st.button("⬅️ Back to tools", key="dash_back"):
        st.session_state.app_view = "tools"
        st.rerun()
    render_dashboard(run_log_sheet, catalog_sheet, tv_mode=False)
    st.stop()

# 🎯 Team Selection Section
TEAMS = [
    {
        "key": "Finance",
        "icon": "💰",
        "accent": "#1a73e8",
        "tint": "#eaf2fe",
        "desc": "Invoices, credit notes, claims & freight.",
    },
    {
        "key": "Operations",
        "icon": "⚙️",
        "accent": "#0f9d58",
        "tint": "#e9f7ef",
        "desc": "Dell invoices, barcodes & packing lists.",
    },
    {
        "key": "Credit",
        "icon": "📊",
        "accent": "#f29900",
        "tint": "#fef3e1",
        "desc": "AR ageing, EDD files & Coface uploads.",
    },
    {
        "key": "Sales",
        "icon": "📈",
        "accent": "#8430ce",
        "tint": "#f5ecfb",
        "desc": "Quotations for IBM, MIBB, Dell & Lenovo.",
    },
]

st.session_state.setdefault("selected_team", TEAMS[0]["key"])

st.markdown("""
    <style>
    .team-card {
        position: relative;
        border-radius: 18px;
        padding: 1.3rem 1.2rem 1.1rem;
        background: #ffffff;
        border: 1px solid #e3e8ef;
        box-shadow: 0 1px 3px rgba(16,24,40,0.06);
        transition: transform 0.18s ease, box-shadow 0.18s ease, border-color 0.18s ease;
        margin-bottom: 0.6rem;
    }
    .team-card:hover {
        transform: translateY(-3px);
        box-shadow: 0 10px 24px rgba(16,24,40,0.10);
        border-color: var(--accent);
    }
    .team-card.selected {
        border-color: var(--accent);
        box-shadow: 0 6px 20px rgba(16,24,40,0.10);
    }
    .team-card .accent-bar {
        position: absolute; top: -1px; left: -1px; right: -1px; height: 4px;
        border-radius: 18px 18px 0 0;
        background: var(--accent);
        opacity: 0; transition: opacity 0.18s ease;
    }
    .team-card.selected .accent-bar, .team-card:hover .accent-bar { opacity: 1; }
    .team-card .badge-selected {
        position: absolute; top: 0.9rem; right: 1rem;
        font-size: 0.68rem; font-weight: 700; color: var(--accent);
        background: var(--tint);
        padding: 0.15rem 0.55rem; border-radius: 999px;
        letter-spacing: 0.02em;
    }
    .team-card .icon-badge {
        width: 44px; height: 44px; border-radius: 12px;
        display: flex; align-items: center; justify-content: center;
        font-size: 1.4rem; margin-bottom: 0.65rem;
        background: var(--tint);
    }
    .team-card h3 {
        margin: 0 0 0.2rem; font-size: 1.05rem; font-weight: 700;
        color: #1f2430; font-family: 'Google Sans', sans-serif;
    }
    .team-card p { margin: 0; font-size: 0.82rem; color: #6b7280; line-height: 1.4; }
    div[data-testid="column"] .stButton > button {
        box-shadow: none !important;
    }
    </style>
""", unsafe_allow_html=True)

header_l, header_r = st.columns([5, 1])
with header_l:
    st.markdown("""
        <div style='margin: 0.4rem 0 1.2rem;'>
            <h2 style='color:#1a73e8; font-family:"Google Sans", sans-serif; font-weight:700; letter-spacing:-1px; margin-bottom:0.15rem;'>Choose your team</h2>
            <p style='font-size:0.95rem; color:#6b7280; margin:0;'>Pick your department to see the tools built for you</p>
        </div>
    """, unsafe_allow_html=True)
with header_r:
    st.markdown("<div style='height:0.6rem'></div>", unsafe_allow_html=True)
    if st.button("📊 Dashboard", key="open_dashboard", use_container_width=True):
        st.session_state.app_view = "dashboard"
        st.rerun()

cards = st.columns(4, gap="medium")
for col, t in zip(cards, TEAMS):
    is_selected = st.session_state.selected_team == t["key"]
    with col:
        # Built as one unbroken line on purpose: a blank line in the middle of
        # raw HTML passed to st.markdown silently ends the HTML block (CommonMark
        # rule), dumping everything after it as literal text — which happened
        # here when the "✓ Active" badge conditional collapsed to an empty line.
        badge_html = '<div class="badge-selected">✓ Active</div>' if is_selected else ""
        card_html = (
            f'<div class="team-card{" selected" if is_selected else ""}" style="--accent:{t["accent"]}; --tint:{t["tint"]}">'
            f'<div class="accent-bar"></div>'
            f'{badge_html}'
            f'<div class="icon-badge">{t["icon"]}</div>'
            f'<h3>{t["key"]}</h3>'
            f'<p>{t["desc"]}</p>'
            f'</div>'
        )
        st.markdown(card_html, unsafe_allow_html=True)
        if st.button(
            "✓ Selected" if is_selected else "Select",
            key=f"team_btn_{t['key']}",
            use_container_width=True,
            disabled=is_selected,
        ):
            st.session_state.selected_team = t["key"]
            st.rerun()

team = st.session_state.selected_team

def load_master_map(master_file):
    df = pd.read_excel(master_file) if master_file.name.endswith(".xlsx") else pd.read_csv(master_file)
    df = df.iloc[:, :2]
    df.columns = ["part", "desc"]
    df["part"] = df["part"].astype(str).str.upper().str.replace(" ", "").str.replace("-", "")
    df["desc"] = df["desc"].fillna("").astype(str)
    return dict(zip(df["part"], df["desc"]))
def extractor_workflow(
    extractor_name,
    extractor_info,
    file_uploader_label,
    extract_invoice_info_func,
    extract_table_func,
    table_columns,
    file_name_template,
    show_header_df_func=None,
    header_columns=None,
    item_row_builder=None,
    item_columns=None
):
    st.title(f"PDF TO EXCEL ({extractor_name})")
    st.write(extractor_info)
    uploaded_file = st.file_uploader(file_uploader_label, type=["pdf"], accept_multiple_files=False, key=f"uploader_{extractor_name}")
    if uploaded_file:
        invoice_num, invoice_date = extract_invoice_info_func(uploaded_file)
        if invoice_num and invoice_date:
            file_date = format_month_year(invoice_date)
            file_name = file_name_template.format(invoice_num=invoice_num, file_date=file_date)
        else:
            file_name = file_name_template.format(invoice_num='unknown', file_date='unknown')
        rows = extract_table_func(uploaded_file)
        if rows:
            df = pd.DataFrame(rows, columns=table_columns)
            if show_header_df_func and header_columns and item_row_builder and item_columns:
                today_str = datetime.today().strftime("%d/%m/%Y")
                remarks = f"GOOGLE INV-{invoice_num}" if invoice_num else "GOOGLE INV-UNKNOWN"
                header_df = pd.DataFrame([
                    show_header_df_func(invoice_num, invoice_date, today_str, remarks)
                ], columns=header_columns)
                dnts_item_data = [item_row_builder(idx, *row, invoice_num) for idx, row in enumerate(rows, 1)]
                dnts_item_df = pd.DataFrame(dnts_item_data, columns=item_columns)
               
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    header_df.to_excel(writer, sheet_name='DNTS_HEADER', index=False)
                    dnts_item_df.to_excel(writer, sheet_name='DNTS_ITEM', index=False)
                output.seek(0)
                
                st.download_button(
                    label=f"⬇️ Download DNTS Excel",
                    data=output.getvalue(),
                    file_name=file_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    on_click=lambda: (
                            update_usage("Google Automation", team, pdf_count=1)
                        ),
                    key=f"download_{extractor_name}"
                )
            else:
                
                towrite = io.BytesIO()
                df.to_excel(towrite, index=False, engine='openpyxl')
                towrite.seek(0)
                st.download_button(
                    label="Download as Excel",
                    data=towrite,
                    file_name=file_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    on_click=lambda: (
                            update_usage("google Automation", team, pdf_count=1)
                        ),
                )
        else:
            st.warning("No table data found in the uploaded PDF.")
    else:
        st.info(f"Please upload a {extractor_name} PDF file to get started.")
        

# ----------- Tool Selector UI -----------
st.markdown("""
    <div style='text-align:center; margin-top:2rem; margin-bottom:1.5rem;'>
        <h2 style='color:#1a73e8; font-family:Google Sans, sans-serif; font-weight:700; letter-spacing:-1px;'>🛠️ Tool Selection</h2>
        <p style='font-size:1.2rem; color:#444;'>Choose the tool you want to use for your PDF extraction.</p>
    </div>
""", unsafe_allow_html=True)
if team == "Finance":
    TOOL_OPTIONS = [
        "-- Select a tool --",
        "🟦 Google DNTS Extractor",
        "🟩 Google Invoice Extractor",
        "📄 Claims Automation",
        "🟨 AWS Invoice Tool",
        "🟧 Oracle Invoice Tool",
        "🟥 Lenovo CNTS Tool - KSA",
        "🟪 Lenovo Credit Note Tool - UAE",
        "🚚 Freight Forwarder JV Tool",
        "🟪 IBM Credit Note Automation (KSA)"
        
    ]
elif team == "Operations":
    TOOL_OPTIONS = [
        "-- Select a tool --",
        "💻 Dell Invoice Extractor",
        "📦 Barcode PDF Generator grouped",
        "CI and Packing list - IBM"
    ]
elif team == "Credit":
    TOOL_OPTIONS = [
        "AR to EDD file",
        "Coface CSV Uploader"
    ]
elif team == "Sales":
    TOOL_OPTIONS = [
        "IBM Quotation",
        "MIBB Quotations",
        "💻 Dell Quotation",
        "💻 Dell Quotation (Orion)",
        "💻 Dell Quotation Southcomp Polaris",
        "💻 Lenovo Quotation"
    ]
else:
    TOOL_OPTIONS = ["-- Select a tool --"]
tool = st.selectbox(
    "Select a tool:",
    TOOL_OPTIONS,
    key="tool_selector"
)
if tool == "-- Select a tool --":
    st.info("Please select a tool above to get started.")
elif tool == "🟦 Google DNTS Extractor":
    def dnts_item_row(idx, domain, customer_id, amount, invoice_num):
        formatted_amount = format_amount(amount)
        item_name = (
            f"GOOGLE INV-{invoice_num} / DOMAIN NAME : {domain} / CUSTOMER ID : {customer_id} / AMOUNT - USD - {formatted_amount}"
        ).upper()
        return [
            idx, 1, "NS", item_name, "NA", "NA", "NOS", 1, 0, formatted_amount, 14401, "SDIG005", "PUHO", "GEN", "ZZ-COMM"
        ]
    extractor_workflow(
        extractor_name="Google DNTS Extractor",
        extractor_info="Upload one PDF containing a **'Summary of costs by domain'** table. The app will extract the table and let you download it as Excel.",
        file_uploader_label="Choose your Google DNTS Invoice PDF",
        extract_invoice_info_func=extract_invoice_info,
        extract_table_func=extract_table_from_text,
        table_columns=DNTS_ITEM_COLS[:3],
        file_name_template="{invoice_num}-{file_date}.xlsx",
        show_header_df_func=make_dnts_header_row,
        header_columns=DNTS_HEADER_COLS,
        item_row_builder=dnts_item_row,
        item_columns=DNTS_ITEM_COLS
    )
elif tool == "🟩 Google Invoice Extractor":
    extractor_workflow(
        extractor_name="Google Invoice Extractor",
        extractor_info="Upload a Google Invoice PDF. The app will extract the relevant data and let you download it as Excel.",
        file_uploader_label="Choose your Google Invoice PDF",
        extract_invoice_info_func=extract_invoice_info_invoice,
        extract_table_func=extract_invoice_table,
        table_columns=GOOGLE_INVOICE_COLS,
        file_name_template="{invoice_num}-{file_date}.xlsx"
    )
elif tool == "📄 Claims Automation":
    st.title("Claims Automation")
    
    st.header("📁 Upload Files")
    source1_file = st.file_uploader("JV Orion from SAP (.xlsx)", type=["xlsx"], accept_multiple_files=False, key="claims_source1")
    master1_file = st.file_uploader("User information (.xlsx)", type=["xlsx"], accept_multiple_files=False, key="claims_master1")
    source2_file = st.file_uploader("Employee benefits (.xlsx)", type=["xlsx"], accept_multiple_files=False, key="claims_source2")
    master2_file = st.file_uploader("Main acc file (.xlsx)", type=["xlsx"], accept_multiple_files=False, key="claims_master2")
    
    st.markdown("---")
    run_clicked = st.button("🚀 Generate Output", key="claims_run", use_container_width=True)
    if run_clicked:
        if not source1_file:
            st.error("Please upload Source File 1.")
            st.stop()
        try:
            src_rows = read_source1_rows(source1_file)
            master_map = read_master1_map(master1_file) if master1_file else None
            src2_rows = read_source2_rows(source2_file) if source2_file else None
            master2_entries = read_master2_entries(master2_file) if master2_file else None
            credit_rows = build_output_rows_from_source1(
                src_rows,
                master1_map=master_map,
                source2_rows=src2_rows,
                user_id_col="Sub Acct",
            )
            src1_defaults = derive_defaults_from_source1(src_rows)
            doc_ref = ""
            for r in src_rows:
                val = r.get("Doc Ref.", "")
                if val and str(val).strip():
                    doc_ref = str(val).strip()
                    break
            debit_rows = build_debit_rows_from_source2(
                src2_rows,
                master2_entries=master2_entries,
                master1_map=master_map,
                default_div=src1_defaults.get("Div", ""),
                default_dept=src1_defaults.get("Dept", ""),
                default_anly1=src1_defaults.get("Anly1", ""),
                default_anly2=src1_defaults.get("Anly2", ""),
                default_currency=src_rows[0].get("Currency", "") if src_rows else "",
                doc_ref=doc_ref,
            )
            out_rows = credit_rows + debit_rows
            output_buffer = io.BytesIO()
            write_output_excel(output_buffer, out_rows)
            output_buffer.seek(0)
            st.download_button(
                label="Download claims_output.xlsx",
                data=output_buffer,
                file_name="claims_output.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                on_click=lambda: (
                            update_usage(
                                "Claims Automation",
                                team,
                                excel_count=count_uploaded_files(
                                    [source1_file, master1_file, source2_file, master2_file],
                                    EXCEL_EXTENSIONS,
                                ),
                            )
                        ),
                key="claims_download"
            )
        except Exception as e:
            st.error(f"Error: {e}")

elif tool == "🚚 Freight Forwarder JV Tool":
    st.title("Freight Forwarder JV Tool")
    st.write("Upload freight forwarder PDF invoices and download the JV upload file.")

    uploaded_files = st.file_uploader(
        "Choose Freight Forwarder PDF(s)",
        type=["pdf"],
        accept_multiple_files=True,
        key="freight_forwarder_upload",
    )

    if uploaded_files:
            output_df, _, errors = process_freight_forwarder_pdfs(uploaded_files, JVConfig())

            if errors:
                st.error("\n".join(errors))

            if not output_df.empty:
                output_excel = create_freight_forwarder_excel_file(output_df)
                st.download_button(
                    label="⬇️ Download Freight Forwarder JV file",
                    data=output_excel.getvalue(),
                    file_name=f"Expeditors-{datetime.now().strftime('%Y-%m-%d')}.xlsx",
                    on_click=lambda: (
                            update_usage(
                                "freight_forwarder_Expeditor",
                                team,
                                pdf_count=count_uploaded_files(uploaded_files, (".pdf",)),
                            )
                        ),
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
elif tool == "📦 Barcode PDF Generator grouped":
    st.write("Upload a CSV file with PalletID and IMEIs to generate barcode PDF.")

    pdf_bytes, success = barcode_tooll()

    if success and pdf_bytes:
        st.success("✅ Barcode PDF is ready!")

        # Create ZIP buffer and write PDF into it
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
            zip_file.writestr("pallet_barcodes_fullpage.pdf", pdf_bytes)
        zip_buffer.seek(0)

        st.download_button(
            label="📥 Download Full-Page Barcode PDF (Zipped)",
            data=pdf_bytes,
            file_name="pallet_barcodes_fullpage.zip",
            mime="application/zip",
            on_click=lambda: (
                            update_usage("barcode Automation", team)
                        ),

        )

elif tool == "💻 Dell Invoice Extractor":
    st.title("Dell Invoice Extractor (Pre-Alert Upload)")
    st.write("Upload one or more Dell invoice PDFs. We'll generate a single Excel with sheet 'PRE ALERT UPLOAD'.")
    uploaded_files = st.file_uploader("Choose Dell invoice PDF(s)", type=["pdf"], accept_multiple_files=True, key="dell_upload")
    master_file = st.file_uploader("Master Excel (starts header at row 9)", type=["xlsx"], key="dell_master")
    if uploaded_files:
            from datetime import datetime, timedelta
            tomorrow_date = (datetime.today() + timedelta(days=1)).strftime("%d/%m/%Y")
            all_rows = []
            master_lookup = None
            supplier_counts = None
            orion_counts = None
            supplier_index = None
            orion_index = None
            po_price_index = None
            if master_file is not None:
                try:
                    master_lookup, supplier_counts, orion_counts, supplier_index, orion_index, po_price_index = read_master_mapping(master_file)
                except Exception as e:
                    st.warning(f"Could not read master file: {e}")
            diag: list[dict] = []
            import os
            log_path = os.path.abspath('pdf_extract_debug.log')
            import tempfile
            for f in uploaded_files:
                st.info(f"DEBUG: Processing file: {getattr(f, 'name', str(f))} (type: {type(f)})")
                # Save UploadedFile to a temp file for processing
                with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp:
                    tmp.write(f.read())
                    tmp_path = tmp.name
                try:
                    rows = build_pre_alert_rows(
                        tmp_path,
                        tomorrow_date,
                        master_lookup=master_lookup,
                        supplier_counts=supplier_counts,
                        orion_counts=orion_counts,
                        supplier_index=supplier_index,
                        orion_index=orion_index,
                        po_price_index=po_price_index,
                        diagnostics=diag,
                    )
                    all_rows.extend(rows)
                except Exception as e:
                    st.warning(f"Failed to parse {getattr(f, 'name', 'file')}: {e}")
                finally:
                    try:
                        os.remove(tmp_path)
                    except Exception:
                        pass
            st.info(f"PDF extraction debug log saved at: {log_path}")
            if all_rows:
                df = pd.DataFrame(all_rows, columns=PRE_ALERT_HEADERS)
                
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name='PRE ALERT UPLOAD', index=False)
                    ws = writer.sheets['PRE ALERT UPLOAD']
                    from datetime import datetime as _dt
                    for r in range(2, len(df) + 2):
                        for c in (4, 11, 12):  # D, K, L
                            cell = ws.cell(row=r, column=c)
                            val = cell.value
                            if isinstance(val, str):
                                try:
                                    d = _dt.strptime(val, "%d/%m/%Y")
                                    cell.value = d
                                    cell.number_format = 'dd/mm/yyyy'
                                except Exception:
                                    pass
                    # ====== NEW: Highlight Unit Rate (Q) vs Orion Unit Price (U) in PRE ALERT UPLOAD ======
                    from openpyxl.styles import PatternFill
                    red_fill = PatternFill(fill_type='solid', start_color='FFFFC7CE', end_color='FFFFC7CE')  # light red
                    green_fill = PatternFill(fill_type='solid', start_color='FFC6EFCE', end_color='FFC6EFCE')  # light green
                    
                    def safe_float(val):
                        try:
                            return float(str(val).replace(',', '').strip())
                        except Exception:
                            return None
                    # Columns: Q = 17, U = 21
                    for row_idx in range(2, len(df) + 2):
                        unit_rate_cell = ws.cell(row=row_idx, column=17)
                        orion_price_cell = ws.cell(row=row_idx, column=21)
                        unit_rate_val = safe_float(unit_rate_cell.value)
                        orion_price_val = safe_float(orion_price_cell.value)
                        if unit_rate_val is not None and orion_price_val is not None:
                            if abs(unit_rate_val - orion_price_val) < 1e-6:
                                # Match: highlight green
                                unit_rate_cell.fill = green_fill
                                orion_price_cell.fill = green_fill
                            else:
                                # Mismatch: highlight red
                                unit_rate_cell.fill = red_fill
                                orion_price_cell.fill = red_fill
                    # ====== EXISTING color highlights on PRE ALERT UPLOAD based on diagnostics ======
                    if diag:
                        red = PatternFill(fill_type='solid', start_color='FFFF0000', end_color='FFFF0000')
                        yellow = PatternFill(fill_type='solid', start_color='FFFFFF00', end_color='FFFFFF00')
                        green = PatternFill(fill_type='solid', start_color='FF00FF00', end_color='FF00FF00')
                        for i, d in enumerate(diag, start=2):
                            h = str(d.get('highlight', 'none')).lower()
                            fill = None
                            if h == 'red':
                                fill = red
                            elif h == 'yellow':
                                fill = yellow
                            elif h == 'green':
                                fill = green
                            if fill is not None:
                                ws.cell(row=i, column=13).fill = fill
                                ws.cell(row=i, column=14).fill = fill
                    # Add an extremely user-friendly REVIEW sheet (emoji + clear actions)
                    if diag:
                        def _status_texts(h: str, status: str) -> tuple[str, str, str, int]:
                            h = (h or 'none').lower()
                            status = status or ''
                            if h == 'red':
                                if status in ('C_orion_price_single', 'C_po_price_single'):
                                    return (
                                        '❌ No match (price suggested)',
                                        'Use the suggested Orion code. Confirm Qty & Unit Price.',
                                        'We did not find a supplier match, but price uniquely matched one item.',
                                        0,
                                    )
                                return (
                                    '❌ No match',
                                    'Ask to add mapping in MASTER or correct Supplier Item code in PDF.',
                                    'We could not find this PO + Supplier Item in MASTER.',
                                    0,
                                )
                            if h == 'yellow':
                                if status == 'B_price_single':
                                    return (
                                        '✅ Price matched (needs quick check)',
                                        'Quick check the suggested Orion Code, Qty, and Unit Price.',
                                        'Multiple entries in MASTER; price selected exactly one.',
                                        1,
                                    )
                                return (
                                    '⚠️ Many matches',
                                    'Pick the correct Orion code or ask to refine MASTER.',
                                    'Multiple MASTER entries matched; price did not decide one.',
                                    1,
                                )
                            return (
                                '✅ All good',
                                'No action. Just review and proceed.',
                                'Single clear match found in MASTER.',
                                2,
                            )
                        simple_rows = []
                        for idx0, d in enumerate(diag):
                            h = str(d.get('highlight', 'none')).lower()
                            status = str(d.get('status', '')).strip()
                            label, action, why, sort_key = _status_texts(h, status)
                            try:
                                row_df = df.iloc[idx0]
                            except Exception:
                                row_df = None
                            qty_val = row_df['Qty'] if row_df is not None and 'Qty' in row_df else ''
                            pdf_price = row_df['Unit Rate'] if row_df is not None and 'Unit Rate' in row_df else d.get('pdf_unit_price', '')
                            use_price = d.get('out_orion_unit_price', '') or pdf_price
                            orion_price_u = d.get('out_orion_unit_price', '')
                            use_code = d.get('out_orion_item_code', '') or d.get('mapped_item_code', '')
                            pre_alert_row_num = idx0 + 2  # header row is 1 in PRE ALERT
                            simple_rows.append({
                                'Row in PRE ALERT': pre_alert_row_num,
                                'Status': label,
                                'What to do now': action,
                                'Why this happened': why,
                                'PO': d.get('po', ''),
                                'Supplier Item': d.get('supplier_item_code', ''),
                                'Copy: Orion Code': use_code,
                                'Copy: Qty': qty_val,
                                'Copy: Unit Price': use_price,
                                'Orion Unit Price (col U)': orion_price_u,
                                '_sort': sort_key,
                                '_highlight': h,
                            })
                        simple_rows.sort(key=lambda x: x.get('_sort', 1))
                        for r in simple_rows:
                            r.pop('_sort', None)
                        review_columns = [
                            'Row in PRE ALERT',
                            'Status',
                            'What to do now',
                            'Why this happened',
                            'Copy: Orion Code',
                            'Copy: Qty',
                            'Copy: Unit Price',
                            'Orion Unit Price (col U)',
                            'PO',
                            'Supplier Item',
                        ]
                        review_df = pd.DataFrame(simple_rows, columns=review_columns + ['_highlight'])
                        review_df.to_excel(writer, sheet_name='REVIEW', index=False)
                        ws_review = writer.sheets['REVIEW']
                        ws_review.freeze_panes = 'A2'
                        widths = {
                            'A': 16, 'B': 30, 'C': 38, 'D': 42,
                            'E': 20, 'F': 12, 'G': 18, 'H': 12, 'I': 20,
                        }
                        for col, w in widths.items():
                            try:
                                ws_review.column_dimensions[col].width = w
                            except Exception:
                                pass
                        try:
                            from openpyxl.styles import Alignment
                            wrap_align = Alignment(wrap_text=True, vertical='top')
                            for r in ws_review.iter_rows(min_row=1, max_row=ws_review.max_row, min_col=1, max_col=ws_review.max_column):
                                for cell in r:
                                    if cell.column in (2, 3, 4):  # B:Status, C:What to do, D:Why
                                        cell.alignment = wrap_align
                        except Exception:
                            pass
                        red_fill_row = PatternFill(fill_type='solid', start_color='FFFFE5E5', end_color='FFFFE5E5')
                        yellow_fill_row = PatternFill(fill_type='solid', start_color='FFFFFBE6', end_color='FFFFFBE6')
                        green_fill_row = PatternFill(fill_type='solid', start_color='FFE9FBE9', end_color='FFE9FBE9')
                        for idx1, row_data in enumerate(simple_rows, start=2):
                            hval = str(row_data.get('_highlight', '')).lower()
                            fill = None
                            if hval == 'red':
                                fill = red_fill_row
                            elif hval == 'yellow':
                                fill = yellow_fill_row
                            elif hval == 'green':
                                fill = green_fill_row
                            if fill is not None:
                                for c in range(1, ws_review.max_column + 1):
                                    ws_review.cell(row=idx1, column=c).fill = fill
                        header_map = {str(ws_review.cell(row=1, column=c).value): c for c in range(1, ws_review.max_column + 1)}
                        orion_u_col = header_map.get('Orion Unit Price (col U)')
                        red_cell = PatternFill(fill_type='solid', start_color='FFFFC7CE', end_color='FFFFC7CE')
                        green_cell = PatternFill(fill_type='solid', start_color='FFC6EFCE', end_color='FFC6EFCE')
                        def _as_float(x):
                            try:
                                return float(str(x).replace(',', '').strip())
                            except Exception:
                                return None
                        if orion_u_col:
                            for idx0 in range(len(simple_rows)):
                                r_excel = idx0 + 2
                                try:
                                    row_df = df.iloc[idx0]
                                except Exception:
                                    row_df = None
                                pdf_price_val = _as_float(row_df['Unit Rate']) if row_df is not None and 'Unit Rate' in row_df else _as_float(simple_rows[idx0].get('Copy: Unit Price', ''))
                                orion_price_val = _as_float(simple_rows[idx0].get('Orion Unit Price (col U)', ''))
                                if orion_price_val is not None and pdf_price_val is not None and abs(orion_price_val - pdf_price_val) < 1e-6:
                                    ws_review.cell(row=r_excel, column=orion_u_col).fill = green_cell
                                else:
                                    ws_review.cell(row=r_excel, column=orion_u_col).fill = red_cell
                    # Create COMPONENT UPLOAD sheet with only the header row
                    component_headers = [
                        'PO Txn Code',
                        'PO Number',
                        'Parent Item Code',
                        'Component Item Code',
                        'UOM',
                        'Qty',
                        'Rate',
                    ]
                    pd.DataFrame(columns=component_headers).to_excel(
                        writer, sheet_name='COMPONENT UPLOAD', index=False
                    )
                    # Write master file content as sheet 2 if provided
                    if master_file is not None:
                        try:
                            master_file.seek(0)
                        except Exception:
                            pass
                        try:
                            df_master = pd.read_excel(master_file, header=8)
                            df_master.to_excel(writer, sheet_name='MASTER', index=False)
                        except Exception:
                            pass
                output.seek(0)
                st.download_button(
                    label="⬇️ Download PRE ALERT UPLOAD",
                    data=output.getvalue(),
                    file_name="pre_alert_upload.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="download_dell_pre_alert",
                    on_click=lambda: (
                            update_usage(
                                "Dell Automation",
                                team,
                                pdf_count=count_uploaded_files(uploaded_files, (".pdf",)),
                                excel_count=count_uploaded_files(master_file, EXCEL_EXTENSIONS),
                            )
                        ),
                )
            else:
                st.warning("No items found in the uploaded PDF(s).")

elif tool == "MIBB Quotations":
    st.header("MIBB Quotations")
    quote_type = st.radio("Quotation type", ["Software", "Hardware", "TLS"], horizontal=True)
    if quote_type == "TLS":
        st.info("Upload the TLS quote Excel. Header information and line items are extracted automatically.")
    else:
        st.info("Upload a MIBB quotation PDF. The tool will extract header information and table data automatically.")

    logo_path = "image.png"
    margin_pct = st.number_input(
        "Margin (%)",
        min_value=0.0,
        max_value=99.0,
        value=1.0,
        step=0.1,
        help="Used in the generated Excel formulas.",
    )

    if quote_type != "TLS":
        st.subheader("Upload MIBB Quotation PDF")
        uploaded_pdf = st.file_uploader(
            "Upload MIBB Quotation PDF (.pdf)",
            type=["pdf"],
            help="Upload a MIBB quotation PDF. The tool will extract header information and table data automatically.",
        )

    if quote_type == "Software":
        st.subheader("Upload Pricelist / Master File (Descriptions)")
        master_file = st.file_uploader(
            "Upload (.csv or .xlsx) - only first 2 columns used",
            type=["csv", "xlsx"],
        )

        if uploaded_pdf:
            pdf_bytes = io.BytesIO(uploaded_pdf.getbuffer())
            header_info = extract_mibb_header_from_pdf(pdf_bytes)

            pdf_bytes.seek(0)
            table_data = extract_mibb_table_from_pdf(pdf_bytes)

            if master_file:
                master_map = load_master_map(master_file)
            else:
                master_map = None
                st.warning("please upload pricelist")

            table_data = correct_mibb_descriptions(table_data, master_map)

            missing = []
            if master_map:
                for row in table_data:
                    part = str(row[0]).strip().upper()
                    if part not in master_map:
                        missing.append(part)

            missing = list(dict.fromkeys(missing))
            if missing:
                st.warning(
                    "Some part numbers were not found in the master file. "
                    "Descriptions were kept blank in Excel. Please double-check:\n\n"
                    + ", ".join(missing)
                )

            if table_data:
                output = io.BytesIO()
                create_mibb_excel(
                    data=table_data,
                    header_info=header_info,
                    logo_path=logo_path,
                    output=output,
                    margin_pct=margin_pct,
                )
                st.success("Excel file generated successfully!")
                st.download_button(
                    label="Download MIBB Quotation Excel",
                    data=output.getvalue(),
                    file_name="MIBB_Quotation.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    on_click=lambda: update_usage(f"MIBB Quotations-{quote_type}", team, pdf_count=1),
                )
        else:
            st.info("Please upload a MIBB quotation PDF to get started.")
    elif quote_type == "Hardware":
        st.subheader("Upload Hardware Quotation Excel")
        uploaded_hardware_excel = st.file_uploader(
            "Upload Hardware Quote Excel (.xlsx, .xlsm, .xls)",
            type=["xlsx", "xlsm", "xls"],
            help="Upload the hardware quotation Excel or XML-based .xls export.",
        )

        if uploaded_pdf and uploaded_hardware_excel:
            pdf_bytes = io.BytesIO(uploaded_pdf.getbuffer())
            header_info = extract_mibb_header_from_pdf(pdf_bytes)

            pdf_bytes.seek(0)
            terms_text = extract_mibb_terms_from_pdf(pdf_bytes)

            excel_bytes = io.BytesIO(uploaded_hardware_excel.getbuffer())
            is_match, match_error = check_mibb_hardware_quote_match(
                excel_bytes,
                header_info.get("Bid Number", ""),
            )

            if not is_match:
                st.error(match_error)
            else:
                excel_bytes.seek(0)
                table_data = extract_mibb_hardware_table_from_excel(excel_bytes)
                if not table_data:
                    st.error("No hardware rows were found in the uploaded Excel.")
                else:
                    output = io.BytesIO()
                    create_mibb_hardware_excel(
                        data=table_data,
                        header_info=header_info,
                        logo_path=logo_path,
                        output=output,
                        margin_pct=margin_pct,
                        terms_text=terms_text,
                    )
                    st.success("Hardware quotation Excel generated successfully!")
                    st.download_button(
                        label="Download MIBB Hardware Quotation Excel",
                        data=output.getvalue(),
                        file_name="MIBB_Hardware_Quotation.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        on_click=lambda: update_usage(
                            f"MIBB Quotations-{quote_type}", team, pdf_count=1, excel_count=1
                        ),
                    )
        else:
            st.info("Please upload both the MIBB PDF and hardware Excel to get started.")
    else:
        st.subheader("Upload TLS Quotation Excel")
        uploaded_tls_excel = st.file_uploader(
            "Upload TLS Quote Excel (.xlsx, .xls)",
            type=["xlsx", "xls"],
            help="Upload the TLS quote Excel. All columns are carried into the output.",
        )

        if uploaded_tls_excel:
            excel_bytes = io.BytesIO(uploaded_tls_excel.getbuffer())
            try:
                header_info, tls_columns, table_data = extract_mibb_tls_from_excel(excel_bytes)
            except ValueError as e:
                st.error(str(e))
            else:
                if not table_data:
                    st.error("No line items were found in the uploaded Excel.")
                else:
                    if not any("price" in str(c).lower() for c in tls_columns):
                        st.warning("No Price column detected — margin columns were omitted.")
                    output = io.BytesIO()
                    create_mibb_tls_excel(
                        data=table_data,
                        columns=tls_columns,
                        header_info=header_info,
                        logo_path=logo_path,
                        output=output,
                        margin_pct=margin_pct,
                    )
                    st.success("TLS quotation Excel generated successfully!")
                    st.download_button(
                        label="Download MIBB TLS Quotation Excel",
                        data=output.getvalue(),
                        file_name="MIBB_TLS_Quotation.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        on_click=lambda: update_usage(f"MIBB Quotations-{quote_type}", team, excel_count=1),
                    )
        else:
            st.info("Please upload the TLS quote Excel to get started.")

elif tool == "🟨 AWS Invoice Tool":
        st.title("AWS Invoice Tool")
        st.write("Upload AWS invoice PDF(s) and download the extracted data as Excel.")
    
        uploaded_files = st.file_uploader(
            "Choose AWS invoice PDF(s)", type=["pdf"], key="aws_upload", accept_multiple_files=True
        )
    
        if uploaded_files:
            rows, template_map, text_map = process_multiple_aws_pdfs(uploaded_files)
            if rows:
                df = pd.DataFrame(rows, columns=AWS_OUTPUT_COLUMNS)
                
    
                output_original = io.BytesIO()
                with pd.ExcelWriter(output_original, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name='AWS_INVOICE', index=False)
                output_original.seek(0)
                st.download_button(
                    label="⬇️ Download Extracted AWS Invoice Data",
                    data=output_original.getvalue(),
                    file_name="aws_invoice_data.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    
                )
    
                output_files = build_dnts_cnts_rows(rows, template_map, text_map)
                zip_buffer = io.BytesIO()
                with zipfile.ZipFile(zip_buffer, "w") as zip_file:
                    for file_key, data in output_files.items():
                        bill_to, file_type = file_key.split("__")
                        safe_bill_to = bill_to.replace(" ", "_").replace(".", "").replace(",", "")
                        file_name = f"{file_type}_{safe_bill_to}.xlsx"
                
                        output = io.BytesIO()
                        with pd.ExcelWriter(output, engine='openpyxl') as writer:
                            pd.DataFrame(data["header"], columns=[
                                "S.No", "Date - (dd/MM/yyyy)", "Supp_Code", "Curr_Code", "Form_Code",
                                "Doc_Src_Locn", "Location_Code", "Remarks", "Supplier_Ref", "Supplier_Ref_Date - (dd/MM/yyyy)"
                            ]).to_excel(writer, sheet_name=f"{file_type}_HEADER", index=False)
                
                            pd.DataFrame(data["item"], columns=[
                                "S.No", "Ref. Key", "Item_Code", "Item_Name", "Grade1", "Grade2", "UOM",
                                "Qty", "Qty_Ls", "Rate", "Main_Account", "Sub_Account", "Division", "Department", "Analysis2"
                            ]).to_excel(writer, sheet_name=f"{file_type}_ITEM", index=False)
                
                        output.seek(0)
                        zip_file.writestr(file_name, output.read())
                
                zip_buffer.seek(0)
                
                st.download_button(
                    label="⬇️ Download All DNTS/CNTS Files as ZIP",
                    data=zip_buffer.getvalue(),
                    file_name="aws_dnts_cnts_files.zip",
                    mime="application/zip",
             
                    on_click=lambda: (
                            update_usage(
                                "AWS Automation",
                                team,
                                pdf_count=count_uploaded_files(uploaded_files, (".pdf",)),
                            )
                        ),
                )
            else:
                st.warning("No data extracted from the uploaded AWS PDFs.")
        else:
            st.info("Please upload one or more AWS invoice PDFs to begin.")
            
elif tool == "Coface CSV Uploader":
    
    st.write("Upload an Excel file with customer invoice data to generate grouped outputs by customer code.")
    
    uploaded_file = st.file_uploader("📤 Upload Excel File", type=["xlsx"])
    
    if uploaded_file:
        st.success("✅ File uploaded successfully.")
        zip_output = process_grouped_customer_files(uploaded_file)
    
        st.download_button(
            label="⬇️ Download All Customer Files (ZIP)",
            data=zip_output.getvalue(),
            file_name="customer_outputs.zip",
            mime="application/zip",
            on_click=lambda: (
                            update_usage(
                                "credit format by customer",
                                team,
                                excel_count=count_uploaded_files(uploaded_file, EXCEL_EXTENSIONS),
                            )
                        ),
        )
elif tool == "AR to EDD file":
    st.title("AR to EDD file")
    st.write("Upload the insurance Excel file (starting from row 16) to filter and extract relevant data.")

    ageing_min_threshold = st.number_input(
        label="📅 Minimum Ageing Threshold (days)",
        min_value=0,
        value=200,
        step=1,
        help="Only include records with ageing greater than this number"
    )
    
    ageing_max_threshold = st.number_input(
        label="⏱️ Maximum Ageing Threshold (days)",
        min_value=0,
        value=270,  # default as requested
        step=1,
        help="Only include records with ageing less than or equal to this number"
    )


    uploaded_file = st.file_uploader(
        "Choose Insurance Excel File", type=["xlsx"], key="insurance_upload"
    )

    if uploaded_file:
        
        if ageing_min_threshold > ageing_max_threshold:
            st.error("Minimum ageing threshold cannot be greater than the maximum threshold.")
        else:

           output_excel = process_insurance_excel(
            uploaded_file,
            ageing_filter=True,
            
            ageing_min_threshold=ageing_min_threshold,
            ageing_max_threshold=ageing_max_threshold

        )

        st.download_button(
            label="⬇️ Download AR to EDD file",
            data=output_excel.getvalue(),
            file_name="EDD.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            on_click=lambda: (
                            update_usage(
                                "credit Automation ",
                                team,
                                excel_count=count_uploaded_files(uploaded_file, EXCEL_EXTENSIONS),
                            )
                            
                        ),
        )


elif tool == "CI and Packing list - IBM":
    st.title("CI and Packing list - IBM")

    if "comm_generator_pair_count" not in st.session_state:
        st.session_state.comm_generator_pair_count = 1

    pair_inputs = []
    incomplete_pairs = []

    for index in range(st.session_state.comm_generator_pair_count):
        pair_number = index + 1
        st.subheader(f"Shipment Pair {pair_number}")
        left_col, right_col = st.columns(2)
        with left_col:
            sob_file = st.file_uploader(
                "Upload SOB PDF",
                type=["pdf"],
                key=f"comm_generator_sob_pdf_{index}",
            )
        with right_col:
            ibm_file = st.file_uploader(
                "Upload IBM PO / Commercial Invoice PDF",
                type=["pdf"],
                key=f"comm_generator_ibm_pdf_{index}",
            )

        if sob_file or ibm_file:
            if sob_file and ibm_file:
                pair_inputs.append((sob_file, ibm_file))
            else:
                incomplete_pairs.append(pair_number)

    if st.button("Add another pair", key="comm_generator_add_pair"):
        st.session_state.comm_generator_pair_count += 1
        st.rerun()

    if incomplete_pairs:
        st.error(
            "Each pair must include both files before workbook generation. "
            f"Incomplete pair(s): {', '.join(str(value) for value in incomplete_pairs)}"
        )

    if pair_inputs and not incomplete_pairs:
        with st.spinner("Preparing workbook..."):
            result = process_comm_generator_pairs(pair_inputs)
            workbook_bytes = build_comm_generator_workbook(result)

        st.success("Workbook prepared successfully.")

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        st.download_button(
            label="Download Excel Workbook",
            data=workbook_bytes.getvalue(),
            file_name=f"output_{timestamp}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="comm_generator_download",
            on_click=lambda: update_usage(
                "IBM Packing List Automation",
                team,
                pdf_count=count_uploaded_files([file_obj for pair in pair_inputs for file_obj in pair], (".pdf",)),
                excel_count=0,
            ),
        )
    else:
        st.caption("Upload at least one complete SOB + IBM pair to generate the workbook.")
elif tool == "🟧 Oracle Invoice Tool":
    st.title("Oracle Invoice Tool")
    st.write("Upload Oracle invoice PDF(s) and download the extracted data as Excel.")

    uploaded_files = st.file_uploader(
        "Choose Oracle invoice PDF(s)",
        type=["pdf"],
        accept_multiple_files=True,
    )

    if uploaded_files:
        file_blobs = [(f.name, f.read()) for f in uploaded_files]
        df, text_map = process_oracle_pdfs_cached(file_blobs)

        if not df.empty:
            excel_bytes = prepare_excel_bytes(df)
            st.download_button(
                label="⬇️ Download Extracted Oracle Invoice Data",
                data=excel_bytes,
                file_name="oracle_invoice_data.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            on_click=lambda: (
                            update_usage(
                                "Oracle Automation ",
                                team,
                                pdf_count=count_uploaded_files(uploaded_files, (".pdf",)),
                            )
                            
                        ),
            )

            with st.expander("Preview extracted data"):
                st.dataframe(df)
        else:
            st.warning("No data extracted.")
    else:
        st.info("Please upload Oracle invoices.")

elif tool == "IBM Quotation":

    st.header("🆕 IBM Excel to Excel + PDF to Excel (Combo)")
    st.info("Upload an IBM quotation PDF and (optionally) an Excel file. The tool will auto-detect the template and use the best logic for each.")

    # Country selection
    country = st.selectbox("Choose a country:", ["UAE", "Qatar", "Levant", "KSA"])

    logo_path = "image.png"
    compliance_text = ""  # Add compliance text if needed

    st.subheader("📤 Upload IBM Quotation Files")

    uploaded_pdf = st.file_uploader(
        "Upload IBM Quotation PDF (.pdf)",
        type=["pdf"],
        help="Supports .pdf files. The tool will extract header information from the PDF."
    )

    uploaded_excel = st.file_uploader(
        "Upload IBM Quotation Excel (.xlsx, .xlsm, .xls)",
        type=["xlsx", "xlsm", "xls"],
        help="Supports .xlsx, .xlsm, and .xls files. The tool will extract line items from the second sheet."
    )

    if uploaded_pdf:
        from sales.ibm_v2_combo import process_ibm_combo
        import io
        pdf_bytes = io.BytesIO(uploaded_pdf.getbuffer())
        excel_bytes = io.BytesIO(uploaded_excel.getbuffer()) if uploaded_excel else None
        result = process_ibm_combo(pdf_bytes, excel_bytes, country=country)

        if result['error']:
            st.error(f"❌ {result['error']}")
        else:
            st.success(f"✅ Detected Template: {result['template']}")
            if result['mep_cost_msg']:
                st.info(result['mep_cost_msg'])
            if result['bid_number_error']:
                st.error(result['bid_number_error'])
            if result.get('date_validation_msg'):
                st.info(f"📅 Date Validation:\n{result['date_validation_msg']}")
            if result['data']:
                if result.get('columns'):
                    st.dataframe(pd.DataFrame(result['data'], columns=result['columns']))
                else:
                    st.dataframe(pd.DataFrame(result['data']))
            if result.get('excel_bytes'):
                # Normalize template to "1" or "2" for usage label
                t = result['template']
                temp_num = "1" if t in ('1', 'template1') else "2"
                st.download_button(
                    label="📥 Download Styled Excel File",
                    data=result['excel_bytes'],
                    file_name="Styled_Quotation.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    on_click=lambda co=country, tn=temp_num: update_usage(
                        f"IBM Automation ({co}) temp {tn}",
                        team,
                        pdf_count=count_uploaded_files(uploaded_pdf, (".pdf",)),
                        excel_count=count_uploaded_files(uploaded_excel, EXCEL_EXTENSIONS),
                    ),
                )


elif tool == "🟪 Lenovo Credit Note Tool - UAE":

    st.title("Lenovo Credit Note Tool - UAE")
    st.write("Upload Lenovo UAE credit note PDF(s) and download the matching Excel output.")

    uploaded_files = st.file_uploader(
        "Choose Lenovo UAE credit note PDF(s)",
        type=["pdf"],
        accept_multiple_files=True,
        key="lenovo_cn_uae_upload",
    )

    if uploaded_files:
        file_blobs = [(f.name, f.read()) for f in uploaded_files]
        df = process_lenovo_credit_pdfs(file_blobs)

        if not df.empty:
            excel_bytes = prepare_excel_bytes(df)

            st.download_button(
                label="Download Lenovo UAE Excel",
                data=excel_bytes,
                file_name=build_output_filename(),
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="lenovo_uae_download_btn",
                            on_click=lambda: (
                            update_usage(
                                "Lenovo Credit Note Tool - UAE",
                                team,
                                pdf_count=count_uploaded_files(uploaded_files, (".pdf",)),
                            )
                            
                        ),
            )

            with st.expander("Preview extracted rows"):
                st.dataframe(df, use_container_width=True)
        else:
            st.warning("No rows produced. Check the UAE PDF format.")

    else:
        st.info("Upload Lenovo UAE credit note PDFs to begin.")

elif tool == "🟥 Lenovo CNTS Tool - KSA":

    st.title("Lenovo CNTS Tool - KSA")
    st.write("Upload Lenovo KSA credit note PDF(s) and download the CNTS workbook output.")

    uploaded_files = st.file_uploader(
        "Choose Lenovo KSA credit note PDF(s)",
        type=["pdf"],
        accept_multiple_files=True,
        key="lenovo_cnts_ksa_upload",
    )

    if uploaded_files:
        file_blobs = [(f.name, f.read()) for f in uploaded_files]
        header_df, item_df = process_lenovo_ksa_pdfs(file_blobs)

        if not header_df.empty and not item_df.empty:
            excel_bytes = prepare_ksa_excel_bytes(header_df, item_df)

            st.download_button(
                label="Download Lenovo KSA CNTS Excel",
                data=excel_bytes,
                file_name=build_ksa_output_filename(),
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                on_click=lambda: update_usage(
                    "Lenovo Credit Note Tool - KSA",
                    team,
                    pdf_count=count_uploaded_files(uploaded_files, (".pdf",)),
                ),
                key="lenovo_ksa_download_btn",
            )

        else:
            st.warning("No rows produced. Check the KSA PDF format.")

    else:
        st.info("Upload Lenovo KSA credit note PDFs to begin.")
        
elif tool == "💻 Dell Quotation":

    st.title("💼 Dell Quotation Tool")

    uploaded = st.file_uploader(
        "Upload Dell BOQ Excel or PDF",
        type=["xlsx", "xlsm", "xls", "pdf"],
        key="dell_uploader"
    )

    margin_percent = st.number_input(
        "Default Margin %",
        min_value=0.0,
        max_value=100.0,
        value=5.0,
        step=0.5
    )

    currency_code = st.radio(
        "Currency",
        ["USD", "QAR", "AED","SAR"],
        horizontal=True
    )

    DELL_MW_SALES_PERSONS = [
        "S.Abdulsalam@mindware.net",
        "C.BouKhaled@mindware.net",
        "R.Kaddoura@mindware.net",
        "N.Xavier@mindware.net",
        "S.Jammoul@mindware.net",
        "m.ali@mindware.net",
        "R.AlKhateeb@mindware.net",
    ]
    mw_sales_person = ""
    if currency_code == "AED":
        sales_person_choice = st.selectbox(
            "MW Sales Person",
            ["— Select —"] + DELL_MW_SALES_PERSONS,
            key="dell_mw_sales_person",
        )
        if sales_person_choice != "— Select —":
            mw_sales_person = sales_person_choice

    if "dell_output_bytes" not in st.session_state:
        st.session_state["dell_output_bytes"] = None
    if "dell_output_name" not in st.session_state:
        st.session_state["dell_output_name"] = None
    if "dell_generation_done" not in st.session_state:
        st.session_state["dell_generation_done"] = False
    if "dell_generation_success" not in st.session_state:
        st.session_state["dell_generation_success"] = False
    if "dell_last_error" not in st.session_state:
        st.session_state["dell_last_error"] = None
    if "dell_uploaded_hash" not in st.session_state:
        st.session_state["dell_uploaded_hash"] = None
    if "dell_uploaded_bytes" not in st.session_state:
        st.session_state["dell_uploaded_bytes"] = None
    if "dell_last_uploaded_name" not in st.session_state:
        st.session_state["dell_last_uploaded_name"] = None
    if "dell_last_currency_code" not in st.session_state:
        st.session_state["dell_last_currency_code"] = None
    if "dell_last_margin_percent" not in st.session_state:
        st.session_state["dell_last_margin_percent"] = None
    if "dell_last_template_label" not in st.session_state:
        st.session_state["dell_last_template_label"] = None
    if "dell_last_sales_person" not in st.session_state:
        st.session_state["dell_last_sales_person"] = None

    if uploaded is not None:
        uploaded_bytes = uploaded.getvalue()
        uploaded_hash = hashlib.sha256(uploaded_bytes).hexdigest()
        if (
            st.session_state["dell_last_uploaded_name"] != uploaded.name
            or st.session_state["dell_uploaded_hash"] != uploaded_hash
        ):
            st.session_state["dell_output_bytes"] = None
            st.session_state["dell_output_name"] = None
            st.session_state["dell_generation_done"] = False
            st.session_state["dell_generation_success"] = False
            st.session_state["dell_last_error"] = None
            st.session_state["dell_last_template_label"] = None
            st.session_state["dell_uploaded_hash"] = uploaded_hash
            st.session_state["dell_uploaded_bytes"] = uploaded_bytes
            st.session_state["dell_last_uploaded_name"] = uploaded.name
    else:
        uploaded_bytes = st.session_state.get("dell_uploaded_bytes")

    if (
        st.session_state.get("dell_last_currency_code") not in (None, currency_code)
        or st.session_state.get("dell_last_margin_percent") not in (None, margin_percent)
        or st.session_state.get("dell_last_sales_person") not in (None, mw_sales_person)
    ):
        st.session_state["dell_output_bytes"] = None
        st.session_state["dell_output_name"] = None
        st.session_state["dell_generation_done"] = False
        st.session_state["dell_generation_success"] = False
        st.session_state["dell_last_error"] = None
        st.session_state["dell_last_template_label"] = None


    col1, col2 = st.columns([1, 1])
    
    with col1:
        generate_clicked = st.button("🚀 Generate Quotation", key="generate_dell_quote_btn", use_container_width=True)
    
    with col2:
        if st.session_state.get("dell_output_bytes"):
            st.download_button(
                label="⬇️ Download quotation",
                data=st.session_state.get("dell_output_bytes"),
                file_name=st.session_state.get("dell_output_name", "quotation.xlsx"),
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="download_dell_quote_top",
                on_click=lambda: (
                    update_usage(
                        f"dell quotation-{st.session_state.get('dell_last_template_label', 'unknown')}-{st.session_state.get('dell_last_currency_code', currency_code)}",
                        team,
                        pdf_count=1 if str(st.session_state.get("dell_last_uploaded_name", "")).lower().endswith(".pdf") else 0,
                        excel_count=1 if str(st.session_state.get("dell_last_uploaded_name", "")).lower().endswith(EXCEL_EXTENSIONS) else 0,
                    )
                ),
                use_container_width=True
            )

    if st.session_state.get("dell_generation_success", False):
        st.success("✅ Quotation generated successfully! Download button is ready above.")
        st.session_state["dell_generation_success"] = False

    if generate_clicked:
        if uploaded is None and st.session_state.get("dell_uploaded_bytes") is None:
            st.warning("Please upload a file first.")
        else:
            try:
                st.session_state["dell_last_error"] = None
                input_bytes = st.session_state.get("dell_uploaded_bytes") or uploaded.getvalue()
                if input_bytes is None:
                    raise ValueError("Uploaded file bytes are missing.")

                with st.spinner("⚙️ Generating quotation..."):
                    template_type = detect_dell_template(input_bytes)
                    if template_type == "extended_services":
                        template_label = "extended_services"
                        out_bytes = generate_dell_extended_services_quote(
                            input_excel_bytes=input_bytes,
                            margin_percent=margin_percent,
                            currency_code=currency_code,
                            mw_sales_person=mw_sales_person,
                        )
                        output_name = build_dell_extended_services_output_filename(
                            input_bytes,
                            currency_code=currency_code,
                        )
                    else:
                        template_label = detect_dell_standard_variant(input_bytes)
                        out_bytes = generate_dell_quote(
                            input_excel_bytes=input_bytes,
                            margin_percent=margin_percent,
                            currency_code=currency_code,
                            mw_sales_person=mw_sales_person,
                        )
                        output_name = build_dell_output_filename(
                            input_bytes,
                            currency_code=currency_code,
                        )

                    if isinstance(out_bytes, io.BytesIO):
                        out_bytes = out_bytes.getvalue()

                    if not out_bytes:
                        raise ValueError("Quotation generation completed but produced no file data.")

                    st.session_state["dell_output_bytes"] = out_bytes
                    st.session_state["dell_output_name"] = output_name
                    st.session_state["dell_generation_done"] = True
                    st.session_state["dell_generation_success"] = True
                    st.session_state["dell_last_currency_code"] = currency_code
                    st.session_state["dell_last_margin_percent"] = margin_percent
                    st.session_state["dell_last_template_label"] = template_label
                    st.session_state["dell_last_sales_person"] = mw_sales_person

                st.success("✅ Quotation generated successfully.")
            except Exception as e:
                st.session_state["dell_last_error"] = str(e)
                st.error(str(e))
                st.exception(e)

    if st.session_state.get("dell_output_bytes"):
        st.markdown("### Download your file")
        st.download_button(
            label="⬇️ Download quotation",
            data=st.session_state.get("dell_output_bytes"),
            file_name=st.session_state.get("dell_output_name", "quotation.xlsx"),
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="download_dell_quote_bottom",
            on_click=lambda: (
                update_usage(
                    f"dell quotation-{st.session_state.get('dell_last_template_label', 'unknown')}-{st.session_state.get('dell_last_currency_code', currency_code)}",
                    team,
                    pdf_count=1 if str(st.session_state.get("dell_last_uploaded_name", "")).lower().endswith(".pdf") else 0,
                    excel_count=1 if str(st.session_state.get("dell_last_uploaded_name", "")).lower().endswith(EXCEL_EXTENSIONS) else 0,
                )
            ),
            use_container_width=True
        )
        st.caption(f"Prepared file: {st.session_state.get('dell_output_name', 'quotation.xlsx')}")

    if uploaded is None and not st.session_state.get("dell_output_bytes"):
        st.info("Upload Dell BOQ Excel or PDF, then click Generate Quotation.")
elif tool == "💻 Dell Quotation (Orion)":
    st.title("💼 Dell Quotation (Orion)")
    st.markdown("Upload a Dell quotation Excel or PDF and generate the Orion export.")
    uploaded = st.file_uploader(
        "Upload Dell BOQ Excel or PDF",
        type=["xlsx", "xlsm", "xls", "pdf"],
        accept_multiple_files=False,
        key="dell_orion_upload",
    )
    currency_code = st.radio(
        "Currency",
        ["USD", "QAR", "AED", "SAR", "EUR"],
        horizontal=True,
        key="dell_orion_currency",
    )
    if st.button("Generate Dell Orion Quotation", key="generate_dell_orion_btn"):
        if not uploaded:
            st.warning("Please upload a file first.")
        else:
            input_bytes = uploaded.getvalue()
            with st.spinner("Generating Orion quotation..."):
                try:
                    out_bytes = generate_orion_quote(
                        input_excel_bytes=input_bytes,
                        currency_code=currency_code,
                    )
                    output_name = build_dell_orion_output_filename(input_excel_bytes=input_bytes)
                    st.download_button(
                        "⬇️ Download Orion quotation",
                        data=out_bytes,
                        file_name=output_name,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="download_dell_orion_quote",
                        on_click=lambda: update_usage(
                            f"dell orion-{currency_code}",
                            team,
                            pdf_count=1 if str(uploaded.name).lower().endswith(".pdf") else 0,
                            excel_count=1 if str(uploaded.name).lower().endswith(EXCEL_EXTENSIONS) else 0,
                        ),
                    )
                    st.success("Done ✅")
                except Exception as e:
                    st.error(f"Generation failed: {e}")
                    st.exception(e)
                    

elif tool == "💻 Dell Quotation Southcomp Polaris":
    render_southcomp_tool(team, update_usage)
    
    
elif tool == "🟪 IBM Credit Note Automation (KSA)":

    st.title("IBM Credit Note Automation (KSA)")
    st.write("Upload IBM credit note PDF(s) and download the CNTS-UPLOAD Excel.")

    uploaded_files = st.file_uploader(
        "Choose IBM credit note PDF(s)",
        type=["pdf"],
        accept_multiple_files=True,
        key="ibm_cn_ksa_upload",
    )

    if uploaded_files:
        file_blobs = [(f.name, f.read()) for f in uploaded_files]

        ibm_header_df, ibm_item_df = process_ibm_ksa_pdfs(file_blobs)

        if not ibm_header_df.empty and not ibm_item_df.empty:
            excel_bytes = prepare_ibm_ksa_excel_bytes(ibm_header_df, ibm_item_df)

            st.download_button(
                label="Download CNTS-UPLOAD Excel",
                data=excel_bytes,
                file_name=build_ibm_ksa_output_filename(),
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="ibm_cn_ksa_download_btn",
                on_click=lambda: update_usage(
                    "IBM Credit Note Automation (KSA)",
                    team,
                    pdf_count=count_uploaded_files(uploaded_files, (".pdf",)),
                ),
            )

            with st.expander("Preview CNTS_HEADER"):
                st.dataframe(ibm_header_df, use_container_width=True)
            with st.expander("Preview CNTS_ITEM"):
                st.dataframe(ibm_item_df, use_container_width=True)
        else:
            st.warning("No rows produced. Check the IBM credit note PDF format.")

    else:
        st.info("Upload IBM credit note PDFs to begin.")

elif tool == "💻 Lenovo Quotation":
    st.title("💼 Lenovo Quotation Tool")
    st.write("Upload a Lenovo quotation PDF, set the margin, and download the Mindware quotation Excel file.")

    lenovo_uploaded = st.file_uploader(
        "Upload Lenovo quote PDF",
        type=["pdf"],
        key="lenovo_quote_uploader",
    )

    lenovo_margin_percent = st.number_input(
        "Default Margin %",
        min_value=0.0,
        max_value=99.0,
        value=5.0,
        step=0.5,
        key="lenovo_margin_percent",
    )

    lenovo_partner = st.text_input(
        "Partner name",
        key="lenovo_partner_name",
        placeholder="e.g. CITG",
    )

    if "lenovo_output_bytes" not in st.session_state:
        st.session_state["lenovo_output_bytes"] = None
    if "lenovo_output_name" not in st.session_state:
        st.session_state["lenovo_output_name"] = None

    if lenovo_uploaded is not None:
        lenovo_input_hash = hashlib.sha256(lenovo_uploaded.getvalue()).hexdigest()
        if st.session_state.get("lenovo_input_hash") != lenovo_input_hash:
            st.session_state["lenovo_input_hash"] = lenovo_input_hash
            st.session_state["lenovo_output_bytes"] = None
            st.session_state["lenovo_output_name"] = None
            st.session_state["lenovo_meta"] = None

    if st.button("🚀 Generate Quotation", key="generate_lenovo_quote_btn", use_container_width=True):
        if lenovo_uploaded is None:
            st.warning("Please upload a Lenovo quote PDF first.")
        else:
            try:
                with st.spinner("⚙️ Generating quotation..."):
                    input_bytes = lenovo_uploaded.getvalue()
                    lenovo_meta = st.session_state.get("lenovo_meta")
                    if lenovo_meta is None:
                        lenovo_meta = parse_lenovo_quote_pdf(input_bytes)
                        st.session_state["lenovo_meta"] = lenovo_meta
                    out_bytes = generate_lenovo_quote(
                        input_bytes,
                        margin_percent=lenovo_margin_percent,
                        partner=lenovo_partner.strip(),
                        meta=lenovo_meta,
                    )
                    st.session_state["lenovo_output_bytes"] = out_bytes
                    st.session_state["lenovo_output_name"] = build_lenovo_output_filename(input_bytes, meta=lenovo_meta)
                items_total = sum(qty * unit for _, _, _, qty, unit in lenovo_meta["items"])
                pdf_grand_total = lenovo_meta.get("grand_total")
                st.success(f"✅ Quotation generated successfully — {len(lenovo_meta['items'])} items extracted.")
                if pdf_grand_total is not None and abs(items_total - pdf_grand_total) > 0.01:
                    st.warning(
                        f"⚠️ Extraction check (before margin): the items read from the PDF add up to "
                        f"{items_total:,.2f} {lenovo_meta['currency']}, but the PDF Grand Total is "
                        f"{pdf_grand_total:,.2f} {lenovo_meta['currency']}. "
                        "Some items may be missing — please review the output against the PDF."
                    )
                elif pdf_grand_total is not None:
                    st.caption(
                        f"Extraction check (before margin): all items captured — cost total equals the "
                        f"PDF Grand Total ({pdf_grand_total:,.2f} {lenovo_meta['currency']}) ✓. "
                        "The margin is applied on top of this in the Excel file."
                    )
            except Exception as e:
                st.session_state["lenovo_output_bytes"] = None
                st.session_state["lenovo_output_name"] = None
                st.error(str(e))

    if st.session_state.get("lenovo_output_bytes"):
        st.download_button(
            label="⬇️ Download quotation",
            data=st.session_state["lenovo_output_bytes"],
            file_name=st.session_state.get("lenovo_output_name") or "Lenovo_Quotation.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="download_lenovo_quote",
            use_container_width=True,
            on_click=lambda: update_usage(
                "Lenovo quotation",
                team,
                pdf_count=count_uploaded_files(lenovo_uploaded, (".pdf",)),
            ),
        )

    if lenovo_uploaded is None and not st.session_state.get("lenovo_output_bytes"):
        st.info("Upload a Lenovo quote PDF, then click Generate Quotation.")


st.markdown("""
<footer style='text-align:center; margin-top:3rem; color:#1a73e8; font-size:20px; font-weight:bold; font-family: Google Sans, sans-serif;'>
    Made with ❤️ by Mindware | © 2025
</footer>
""", unsafe_allow_html=True)
