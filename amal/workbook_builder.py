import io

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


PURPLE_FILL = PatternFill(fill_type="solid", fgColor="3F3D9E")
WHITE_BOLD_FONT = Font(color="FFFFFF", bold=True)
BOLD_FONT = Font(bold=True)
THIN_SIDE = Side(style="thin", color="000000")
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=False)
RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=False)
TOP_LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)
TOP_CENTER = Alignment(horizontal="center", vertical="top", wrap_text=True)

SUPPLIER_TEXT = (
    "Mindware Fz LLC\n"
    "P.O.Box 55609\n"
    "Jabel Ali, United Arab Emirates\n"
    "Tel : + 971 4500600\n"
    "Email: outboundjebelali@mindware.ae\n"
    "VAT TRN No : 100019912300003"
)


def style_range(worksheet, cell_range: str, fill=None, font=None, border=None, alignment=None) -> None:
    for row in worksheet[cell_range]:
        for cell in row:
            if fill:
                cell.fill = fill
            if font:
                cell.font = font
            if border:
                cell.border = border
            if alignment:
                cell.alignment = alignment


def apply_border_to_range(worksheet, start_row: int, end_row: int, start_col: int, end_col: int) -> None:
    """Apply THIN_BORDER to every cell in the given range."""
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            worksheet.cell(row=row, column=col).border = THIN_BORDER


# Keep old name as alias for backward compatibility
set_outer_border = apply_border_to_range


def get_comm_inv_footer_rows(item_count: int) -> tuple[int, int, int, int]:
    visible_item_count = max(item_count, 1)
    freight_row = 18 + visible_item_count
    total_row = freight_row + 1
    net_total_row = freight_row + 2
    total_in_words_row = freight_row + 4
    return freight_row, total_row, net_total_row, total_in_words_row


def to_number_if_possible(value):
    if isinstance(value, (int, float)):
        return value
    if not isinstance(value, str):
        return value

    candidate = value.replace(",", "").strip()
    if not candidate:
        return ""

    try:
        return float(candidate)
    except ValueError:
        return value


def set_merged_block_row_heights(worksheet, start_row: int, end_row: int, total_lines: int) -> None:
    line_count = max(total_lines, end_row - start_row + 1)
    total_height = max(26 * (end_row - start_row + 1), line_count * 24)
    base_height = total_height / (end_row - start_row + 1)
    for row in range(start_row, end_row + 1):
        worksheet.row_dimensions[row].height = base_height


def count_display_lines(value: str) -> int:
    if not value:
        return 1
    return len([line for line in str(value).splitlines() if line.strip()]) or 1


def build_comm_inv_sheet(worksheet) -> None:
    worksheet.title = "comm-inv"

    widths = {
        "A": 21,
        "B": 36,
        "C": 16,
        "D": 11,
        "E": 15,
        "F": 14,
        "G": 15,
        "H": 16,
    }
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width

    for row, height in {
        2: 28,
        4: 18,
        8: 18,
    }.items():
        worksheet.row_dimensions[row].height = height

    # Row 2: Title
    worksheet.merge_cells("A2:H2")
    worksheet["A2"] = "Commercial Invoice"
    worksheet["A2"].font = Font(bold=True, size=16)
    worksheet["A2"].alignment = CENTER

    # Row 4: Purple divider bar
    worksheet.merge_cells("A4:H4")
    style_range(worksheet, "A4:H4", fill=PURPLE_FILL, border=THIN_BORDER)

    # Rows 5-7: Payment/IncoTerms/CustomerPO  |  Invoice No / Date / Currency
    for row in (5, 6, 7):
        worksheet.row_dimensions[row].height = 18

    worksheet.merge_cells("A5:F5")
    worksheet.merge_cells("A6:F6")
    worksheet.merge_cells("A7:F7")
    apply_border_to_range(worksheet, 5, 7, 1, 6)

    worksheet.merge_cells("G5:H5")
    worksheet.merge_cells("G6:H6")
    worksheet.merge_cells("G7:H7")
    apply_border_to_range(worksheet, 5, 7, 7, 8)

    for cell_addr in ("A5", "A6", "A7"):
        worksheet[cell_addr].font = BOLD_FONT
        worksheet[cell_addr].alignment = LEFT

    for cell_addr in ("G5", "G6", "G7"):
        worksheet[cell_addr].font = BOLD_FONT
        worksheet[cell_addr].alignment = LEFT

    # Row 8: Bill To / Ship To / Supplier header
    worksheet.merge_cells("A8:D8")
    worksheet.merge_cells("E8:F8")
    worksheet.merge_cells("G8:H8")
    worksheet["A8"] = "Bill To"
    worksheet["E8"] = "Ship To"
    worksheet["G8"] = "Supplier"
    style_range(worksheet, "A8:H8", fill=PURPLE_FILL, font=WHITE_BOLD_FONT, border=THIN_BORDER)
    worksheet["A8"].alignment = LEFT
    worksheet["E8"].alignment = LEFT
    worksheet["G8"].alignment = LEFT

    # Rows 9-16: Address blocks
    worksheet.merge_cells("A9:D16")
    worksheet.merge_cells("E9:F16")
    worksheet.merge_cells("G9:H16")
    apply_border_to_range(worksheet, 9, 16, 1, 8)

    worksheet["G9"] = SUPPLIER_TEXT
    worksheet["G9"].font = BOLD_FONT
    worksheet["G9"].alignment = TOP_LEFT

    # Row 17: Column headers
    worksheet.row_dimensions[17].height = 22
    headers = ["Item Code", "Desc", "Case#", "Origin", "HS Code", "Qty", "Unit Price", "Amount"]
    for index, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=17, column=index)
        cell.value = header
        cell.fill = PURPLE_FILL
        cell.font = WHITE_BOLD_FONT
        cell.border = THIN_BORDER
        cell.alignment = LEFT if header in {"Item Code", "Desc"} else CENTER

    # NOTE: item rows, footer rows are built dynamically in fill_comm_inv_items / fill_comm_inv_sheet


def fill_comm_inv_sheet(worksheet, fields: dict, item_count: int) -> None:
    freight_row, total_row, net_total_row, total_in_words_row = get_comm_inv_footer_rows(item_count)

    # --- Rows 5-7: header info ---
    worksheet["A5"] = f"Payment Term: {fields.get('payment_term', '')}"
    worksheet["A6"] = f"Inco Terms: {fields.get('inco_terms', '')}"
    worksheet["A7"] = f"Customer PO: {fields.get('customer_po', '')}"
    for cell_addr in ("A5", "A6", "A7"):
        worksheet[cell_addr].font = BOLD_FONT
        worksheet[cell_addr].alignment = LEFT

    worksheet["G5"] = f"Commercial Invoice No : {fields.get('commercial_invoice_no', '')}"
    worksheet["G6"] = f"Date: {fields.get('date', '')}"
    worksheet["G7"] = f"Currency: {fields.get('currency', '')}"
    for cell_addr in ("G5", "G6", "G7"):
        worksheet[cell_addr].font = BOLD_FONT
        worksheet[cell_addr].alignment = LEFT

    # --- Address blocks ---
    worksheet["A9"] = fields.get("bill_to", "")
    worksheet["A9"].alignment = TOP_LEFT

    worksheet["E9"] = fields.get("ship_to", "")
    worksheet["E9"].alignment = TOP_LEFT

    address_block_lines = max(
        count_display_lines(fields.get("bill_to", "")),
        count_display_lines(fields.get("ship_to", "")),
        count_display_lines(SUPPLIER_TEXT),
    )
    set_merged_block_row_heights(worksheet, 9, 16, address_block_lines)

    # --- Freight Charges row ---
    # Col G = "Freight Charges" label (right-aligned bold), Col H = value
    worksheet.cell(row=freight_row, column=7).value = "Freight Charges"
    worksheet.cell(row=freight_row, column=7).font = BOLD_FONT
    worksheet.cell(row=freight_row, column=7).alignment = RIGHT
    worksheet.cell(row=freight_row, column=7).border = THIN_BORDER
    # merge cols 1-6 for freight row (empty/label span)
    worksheet.merge_cells(start_row=freight_row, start_column=1, end_row=freight_row, end_column=6)
    apply_border_to_range(worksheet, freight_row, freight_row, 1, 6)
    worksheet.cell(row=freight_row, column=8).border = THIN_BORDER
    worksheet.cell(row=freight_row, column=8).alignment = RIGHT

    if fields.get("freight_charges", "") != "":
        worksheet.cell(row=freight_row, column=8).value = to_number_if_possible(fields.get("freight_charges", ""))

    # --- Total Amount row ---
    worksheet.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=6)
    apply_border_to_range(worksheet, total_row, total_row, 1, 6)
    worksheet.cell(row=total_row, column=7).value = "Total Amount"
    worksheet.cell(row=total_row, column=7).font = BOLD_FONT
    worksheet.cell(row=total_row, column=7).alignment = RIGHT
    worksheet.cell(row=total_row, column=7).border = THIN_BORDER
    worksheet.cell(row=total_row, column=8).border = THIN_BORDER
    worksheet.cell(row=total_row, column=8).alignment = RIGHT
    worksheet.cell(row=total_row, column=8).font = BOLD_FONT

    item_end_row = 17 + max(item_count, 1)
    worksheet.cell(row=total_row, column=8).value = f"=SUM(H18:H{item_end_row})+H{freight_row}"

    # --- Net Total row (yellow fill) ---
    worksheet.merge_cells(start_row=net_total_row, start_column=1, end_row=net_total_row, end_column=6)
    apply_border_to_range(worksheet, net_total_row, net_total_row, 1, 6)
    worksheet.cell(row=net_total_row, column=7).value = "Net Total"
    worksheet.cell(row=net_total_row, column=7).font = BOLD_FONT
    worksheet.cell(row=net_total_row, column=7).alignment = RIGHT
    worksheet.cell(row=net_total_row, column=7).border = THIN_BORDER
    worksheet.cell(row=net_total_row, column=8).font = BOLD_FONT
    worksheet.cell(row=net_total_row, column=8).border = THIN_BORDER
    worksheet.cell(row=net_total_row, column=8).alignment = RIGHT
    worksheet.cell(row=net_total_row, column=8).value = f"=H{total_row}"

    # --- Blank separator row (total_in_words_row - 1) ---
    worksheet.row_dimensions[total_in_words_row - 1].height = 8

    # --- Total in Words row ---
    worksheet.merge_cells(
        start_row=total_in_words_row, start_column=1, end_row=total_in_words_row + 1, end_column=6
    )
    apply_border_to_range(worksheet, total_in_words_row, total_in_words_row + 1, 1, 6)

    if fields.get("total_in_words", "") != "":
        words_text = f"Total in Words : {fields.get('total_in_words', '')}"
        worksheet.cell(row=total_in_words_row, column=1).value = words_text
    else:
        worksheet.cell(row=total_in_words_row, column=1).value = "Total in Words :"

    worksheet.cell(row=total_in_words_row, column=1).font = BOLD_FONT
    worksheet.cell(row=total_in_words_row, column=1).alignment = TOP_LEFT

    # --- Mindware FZ LLC signature block ---
    worksheet.merge_cells(
        start_row=total_in_words_row, start_column=7, end_row=total_in_words_row + 1, end_column=8
    )
    apply_border_to_range(worksheet, total_in_words_row, total_in_words_row + 1, 7, 8)
    worksheet.cell(row=total_in_words_row, column=7).value = "Mindware FZ LLC"
    worksheet.cell(row=total_in_words_row, column=7).font = BOLD_FONT
    worksheet.cell(row=total_in_words_row, column=7).alignment = CENTER


def ensure_comm_inv_item_rows(worksheet, item_count: int) -> None:
    item_start_row = 18
    default_item_rows = 1
    footer_start_row = 24

    if item_count > default_item_rows:
        worksheet.insert_rows(footer_start_row, item_count - default_item_rows)

    item_end_row = item_start_row + max(item_count, default_item_rows) - 1
    apply_border_to_range(worksheet, item_start_row, item_end_row, 1, 8)


def fill_comm_inv_items(worksheet, items: list[dict]) -> None:
    ensure_comm_inv_item_rows(worksheet, len(items))

    for offset, item in enumerate(items):
        row = 18 + offset

        # Item Code (col A) — left aligned
        c = worksheet.cell(row=row, column=1)
        c.value = item.get("item_code", "")
        c.alignment = LEFT
        c.border = THIN_BORDER

        # Desc (col B) — left aligned, wrap if long
        c = worksheet.cell(row=row, column=2)
        c.value = item.get("desc", "")
        desc_val = str(item.get("desc", ""))
        c.alignment = TOP_LEFT if ("\n" in desc_val or len(desc_val) > 45) else LEFT
        c.border = THIN_BORDER
        worksheet.row_dimensions[row].height = 34 if "\n" in desc_val or len(desc_val) > 45 else 22

        # Case# (col C) — center
        c = worksheet.cell(row=row, column=3)
        c.value = item.get("case_no", "")
        c.alignment = CENTER
        c.border = THIN_BORDER

        # Origin (col D) — center
        c = worksheet.cell(row=row, column=4)
        c.value = item.get("origin", "")
        c.alignment = CENTER
        c.border = THIN_BORDER

        # HS Code (col E) — center
        c = worksheet.cell(row=row, column=5)
        c.value = item.get("hs_code", "")
        c.alignment = CENTER
        c.border = THIN_BORDER

        # Qty (col F) — center
        c = worksheet.cell(row=row, column=6)
        c.value = item.get("qty", "")
        c.alignment = CENTER
        c.border = THIN_BORDER

        # Unit Price (col G) — right
        c = worksheet.cell(row=row, column=7)
        c.value = item.get("unit_price", "")
        c.alignment = RIGHT
        c.border = THIN_BORDER

        # Amount (col H) — right
        c = worksheet.cell(row=row, column=8)
        c.value = item.get("amount", "")
        c.alignment = RIGHT
        c.border = THIN_BORDER


def fill_comm_inv_unmatched_items(worksheet, items: list[dict], item_count: int) -> None:
    if not items:
        return

    _, main_total_row, net_total_row, total_in_words_row = get_comm_inv_footer_rows(item_count)
    start_row = total_in_words_row + 4

    # Header row
    hdr_row = start_row - 1
    worksheet.cell(row=hdr_row, column=6).value = "SOB No"
    worksheet.cell(row=hdr_row, column=6).font = BOLD_FONT
    worksheet.cell(row=hdr_row, column=6).border = THIN_BORDER
    worksheet.cell(row=hdr_row, column=6).alignment = CENTER
    worksheet.cell(row=hdr_row, column=7).value = "Other SOB Items"
    worksheet.cell(row=hdr_row, column=7).font = BOLD_FONT
    worksheet.cell(row=hdr_row, column=7).border = THIN_BORDER
    worksheet.cell(row=hdr_row, column=7).alignment = CENTER
    worksheet.cell(row=hdr_row, column=8).value = "Amount"
    worksheet.cell(row=hdr_row, column=8).font = BOLD_FONT
    worksheet.cell(row=hdr_row, column=8).border = THIN_BORDER
    worksheet.cell(row=hdr_row, column=8).alignment = CENTER

    for offset, item in enumerate(items):
        row = start_row + offset
        worksheet.cell(row=row, column=6).value = item.get("sob_reference", "")
        worksheet.cell(row=row, column=6).border = THIN_BORDER
        worksheet.cell(row=row, column=6).alignment = CENTER
        worksheet.cell(row=row, column=7).value = item.get("item_code", "")
        worksheet.cell(row=row, column=7).border = THIN_BORDER
        worksheet.cell(row=row, column=7).alignment = LEFT
        worksheet.cell(row=row, column=8).value = item.get("amount", "")
        worksheet.cell(row=row, column=8).border = THIN_BORDER
        worksheet.cell(row=row, column=8).alignment = RIGHT

    other_total_row = start_row + len(items)
    worksheet.cell(row=other_total_row, column=7).value = "Total"
    worksheet.cell(row=other_total_row, column=7).font = BOLD_FONT
    worksheet.cell(row=other_total_row, column=7).border = THIN_BORDER
    worksheet.cell(row=other_total_row, column=7).alignment = RIGHT
    worksheet.cell(row=other_total_row, column=8).value = f"=SUM(H{start_row}:H{other_total_row - 1})"
    worksheet.cell(row=other_total_row, column=8).font = BOLD_FONT
    worksheet.cell(row=other_total_row, column=8).border = THIN_BORDER
    worksheet.cell(row=other_total_row, column=8).alignment = RIGHT

    # Net Total in H{net_total_row} now sums main total + other total
    worksheet.cell(row=net_total_row, column=8).value = f"=H{main_total_row}+H{other_total_row}"

    apply_border_to_range(worksheet, hdr_row, other_total_row, 6, 8)


def build_pack_list_sheet(worksheet) -> None:
    worksheet.title = "pack_list"

    widths = {
        "A": 21,
        "B": 34,
        "C": 15,
        "D": 15,
        "E": 15,
        "F": 14,
        "G": 13,
        "H": 10,
    }
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width

    for row, height in {
        2: 28,
        4: 18,
        7: 18,
    }.items():
        worksheet.row_dimensions[row].height = height

    worksheet.merge_cells("A2:H2")
    worksheet["A2"] = "Packing List"
    worksheet["A2"].font = Font(bold=True, size=16)
    worksheet["A2"].alignment = CENTER

    worksheet.merge_cells("A4:H4")
    style_range(worksheet, "A4:H4", fill=PURPLE_FILL, border=THIN_BORDER)

    worksheet.merge_cells("A5:E6")
    worksheet.merge_cells("F5:H5")
    worksheet.merge_cells("F6:H6")
    apply_border_to_range(worksheet, 5, 6, 1, 8)

    worksheet["F5"] = "No. :"
    worksheet["F6"] = "Date :"
    worksheet["F5"].font = BOLD_FONT
    worksheet["F6"].font = BOLD_FONT
    worksheet["F5"].alignment = LEFT
    worksheet["F6"].alignment = LEFT

    worksheet["A7"] = "Bill To"
    worksheet["C7"] = "Ship To"
    worksheet["F7"] = "Supplier"
    style_range(worksheet, "A7:H7", fill=PURPLE_FILL, font=WHITE_BOLD_FONT, border=THIN_BORDER)

    worksheet.merge_cells("A7:B7")
    worksheet.merge_cells("C7:E7")
    worksheet.merge_cells("F7:H7")
    worksheet["A7"].alignment = LEFT
    worksheet["C7"].alignment = LEFT
    worksheet["F7"].alignment = LEFT

    worksheet.merge_cells("A8:B15")
    worksheet.merge_cells("C8:E15")
    worksheet.merge_cells("F8:H15")
    apply_border_to_range(worksheet, 8, 15, 1, 8)

    worksheet["F8"] = SUPPLIER_TEXT
    worksheet["F8"].font = BOLD_FONT
    worksheet["F8"].alignment = TOP_LEFT

    # Row 16: column headers
    worksheet.row_dimensions[16].height = 22
    headers = ["Item Code", "Desc", "Case#", "Origin", "HS Code", "Qty", "Weight", "Package"]
    for index, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=16, column=index)
        cell.value = header
        cell.fill = PURPLE_FILL
        cell.font = WHITE_BOLD_FONT
        cell.border = THIN_BORDER
        cell.alignment = LEFT if header in {"Item Code", "Desc"} else CENTER

    # Placeholder item rows + total row built dynamically
    apply_border_to_range(worksheet, 17, 18, 1, 8)

    # Total row placeholder
    worksheet.cell(row=19, column=5).value = "Total"
    worksheet.cell(row=19, column=5).alignment = CENTER
    worksheet.cell(row=19, column=5).font = BOLD_FONT
    apply_border_to_range(worksheet, 19, 19, 6, 8)

    # Summary section placeholder
    apply_border_to_range(worksheet, 21, 22, 2, 4)
    worksheet["B21"] = "Total No of Cases"
    worksheet["B22"] = "Total Gross Weight"
    worksheet["B21"].font = BOLD_FONT
    worksheet["B22"].font = BOLD_FONT

    apply_border_to_range(worksheet, 24, 25, 2, 4)
    worksheet["B24"] = "CASE #"
    worksheet["D24"] = "Dimension In Cms"
    worksheet["B24"].font = BOLD_FONT
    worksheet["D24"].font = BOLD_FONT

    worksheet["H20"] = "Mindware FZ LLC"
    worksheet["H20"].font = BOLD_FONT
    worksheet["H20"].alignment = CENTER


def ensure_pack_list_rows(worksheet, item_count: int) -> tuple[int, int]:
    default_item_rows = 2
    if item_count > default_item_rows:
        insert_count = item_count - default_item_rows
        worksheet.insert_rows(19, insert_count)

    total_row = 17 + max(item_count, default_item_rows)
    summary_start = total_row + 2
    case_header_row = summary_start + 3
    case_data_start = case_header_row + 1

    apply_border_to_range(worksheet, 17, total_row - 1, 1, 8)
    apply_border_to_range(worksheet, total_row, total_row, 6, 8)
    apply_border_to_range(worksheet, summary_start, summary_start + 1, 2, 4)
    apply_border_to_range(worksheet, case_header_row, case_header_row, 2, 4)
    return total_row, case_data_start


def fill_pack_list_sheet(worksheet, fields: dict) -> None:
    worksheet["F5"] = '="No. : " & \'comm-inv\'!G5'
    worksheet["F6"] = '="Date : " & \'comm-inv\'!G6'
    worksheet["A8"] = "='comm-inv'!A9"
    worksheet["C8"] = "='comm-inv'!E9"
    worksheet["A8"].alignment = TOP_LEFT
    worksheet["C8"].alignment = TOP_LEFT
    set_merged_block_row_heights(
        worksheet,
        8,
        15,
        max(
            count_display_lines(fields.get("bill_to", "")),
            count_display_lines(fields.get("ship_to", "")),
            count_display_lines(SUPPLIER_TEXT),
        ),
    )
    worksheet["D21"] = fields.get("total_packages", "")
    worksheet["D22"] = fields.get("total_gross_weight", "")


def fill_pack_list_items(worksheet, items: list[dict]) -> None:
    total_row, case_data_start = ensure_pack_list_rows(worksheet, len(items))

    for offset, item in enumerate(items):
        row = 17 + offset
        desc_val = str(item.get("desc", ""))

        c = worksheet.cell(row=row, column=1)
        c.value = item.get("item_code", "")
        c.alignment = LEFT
        c.border = THIN_BORDER

        c = worksheet.cell(row=row, column=2)
        c.value = item.get("desc", "")
        c.alignment = TOP_LEFT if ("\n" in desc_val or len(desc_val) > 45) else LEFT
        c.border = THIN_BORDER
        worksheet.row_dimensions[row].height = 34 if "\n" in desc_val or len(desc_val) > 45 else 22

        c = worksheet.cell(row=row, column=3)
        c.value = item.get("case_no", "")
        c.alignment = CENTER
        c.border = THIN_BORDER

        c = worksheet.cell(row=row, column=4)
        c.value = item.get("origin", "")
        c.alignment = CENTER
        c.border = THIN_BORDER

        c = worksheet.cell(row=row, column=5)
        c.value = item.get("hs_code", "")
        c.alignment = CENTER
        c.border = THIN_BORDER

        c = worksheet.cell(row=row, column=6)
        c.value = item.get("qty", "")
        c.alignment = CENTER
        c.border = THIN_BORDER

        c = worksheet.cell(row=row, column=7)
        c.value = item.get("gross_weight", "")
        c.alignment = RIGHT
        c.border = THIN_BORDER

        c = worksheet.cell(row=row, column=8)
        c.value = item.get("package", "")
        c.alignment = CENTER
        c.border = THIN_BORDER

    qty_total = round(
        sum(item["qty"] for item in items if isinstance(item.get("qty"), (int, float))),
        2,
    )
    weight_total = round(
        sum(item["gross_weight"] for item in items if isinstance(item.get("gross_weight"), (int, float))),
        2,
    )
    package_total = round(
        sum(item["package"] for item in items if isinstance(item.get("package"), (int, float))),
        2,
    )

    # Total row styling
    total_label_cell = worksheet.cell(row=total_row, column=5)
    total_label_cell.value = "Total"
    total_label_cell.alignment = CENTER
    total_label_cell.font = BOLD_FONT
    total_label_cell.border = THIN_BORDER

    c = worksheet.cell(row=total_row, column=6)
    c.value = qty_total
    c.alignment = CENTER
    c.font = BOLD_FONT
    c.border = THIN_BORDER

    c = worksheet.cell(row=total_row, column=7)
    c.value = weight_total
    c.alignment = RIGHT
    c.font = BOLD_FONT
    c.border = THIN_BORDER

    c = worksheet.cell(row=total_row, column=8)
    c.value = package_total
    c.alignment = CENTER
    c.font = BOLD_FONT
    c.border = THIN_BORDER

    for offset, item in enumerate(items):
        row = case_data_start + offset
        c = worksheet.cell(row=row, column=2)
        c.value = item.get("case_no", "")
        c.border = THIN_BORDER
        c.alignment = CENTER

        c = worksheet.cell(row=row, column=4)
        c.value = item.get("dimensions_cm", "")
        c.border = THIN_BORDER
        c.alignment = CENTER

    if items:
        apply_border_to_range(worksheet, case_data_start, case_data_start + len(items) - 1, 2, 4)


def write_dataframe_to_sheet(worksheet, df: pd.DataFrame) -> None:
    worksheet.append(list(df.columns))
    for row in df.itertuples(index=False, name=None):
        worksheet.append(list(row))


def create_workbook_bytes(
    comm_inv_fields: dict,
    comm_inv_items: list[dict],
    comm_inv_unmatched_items: list[dict],
    pack_list_fields: dict,
    pack_list_items: list[dict],
    comm_inv_df: pd.DataFrame,
    pack_list_df: pd.DataFrame,
) -> io.BytesIO:
    workbook = Workbook()

    comm_sheet = workbook.active
    build_comm_inv_sheet(comm_sheet)
    fill_comm_inv_items(comm_sheet, comm_inv_items)
    fill_comm_inv_sheet(comm_sheet, comm_inv_fields, len(comm_inv_items))
    fill_comm_inv_unmatched_items(comm_sheet, comm_inv_unmatched_items, len(comm_inv_items))

    pack_sheet = workbook.create_sheet("pack_list")
    build_pack_list_sheet(pack_sheet)
    fill_pack_list_sheet(pack_sheet, pack_list_fields)
    fill_pack_list_items(pack_sheet, pack_list_items)

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output