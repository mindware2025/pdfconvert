"""Lenovo Quotation tool.

Parses a Lenovo Special Bid quotation PDF (PRODUCT AND SERVICE DETAILS table)
and generates a Mindware-branded quotation Excel file with the margin applied
to each unit price (price = cost / (1 - margin/100)), matching the agreed
"Quote" template layout.
"""

import os
import re
from datetime import datetime, timedelta, timezone
from io import BytesIO
from typing import Dict, List, Optional, Tuple

import fitz  # PyMuPDF
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

MINDWARE_LOGO_CANDIDATES = ["mindware_quote_logo.png"]
LENOVO_LOGO_CANDIDATES = ["lenovo_quote_logo.png", "lenovologo.png"]

NAVY = "0E2841"        # titles / total-price column text
LABEL_TEAL = "0A3041"  # Partner / Customer / Date labels and values
TOTAL_FONT = "171717"  # grand total figure

UNIT_PRICE_FMT = '_(* #,##0.00_);_(* \\(#,##0.00\\);_(* "-"??_);_(@_)'
GRAND_TOTAL_FMT = '_("$"* #,##0.00_);_("$"* \\(#,##0.00\\);_("$"* "-"??_);_(@_)'
DATE_FMT = "[$-409]mmmm\\ d\\,\\ yyyy;@"

TERMS_AND_CONDITIONS = [
    "Payment terms: as per Mindware credit policies.",
    "Incoterms: Ex-Jabel Ali.",
    "Validity: quote is valid until {validity}, after that there will be a new prices due to cost increases.",
    "Delivery Time: It will take 10-14 weeks delivery time from the date of Booking.",
    "These prices do not include installation of any kind.",
    "Change in Qty is not acceptable.",
    "Full end user details to be mentioned in your PO",
    "PO Should be addressed to Mindware FZ LLC",
    "Orders once placed with Lenovo cannot be cancelled.",
]

# A pricing row, e.g.:
#   1 13BACTO1WW ThinkCentre neo 50t 2032 988.00 25-Apr-2026 2,007,616.00
_MAIN_ROW_RE = re.compile(
    r"^(?P<line>\d{1,4})\s+"
    r"(?P<part>[A-Z0-9]{5,})\s*"
    r"(?P<desc>.*?)\s*"
    r"(?P<qty>\d[\d,]*)\s+"
    r"(?P<unit>\d[\d,]*\.\d{2})\s+"
    r"(?P<end_date>\d{1,2}-[A-Za-z]{3}-\d{4})\s+"
    r"(?P<total>\d[\d,]*\.\d{2})$"
)

# Vertical distance (pt) within which a wrapped description line is considered
# part of the nearest pricing row.
_WRAP_DISTANCE_PT = 25.0


def _parse_number(text: str) -> float:
    return float(str(text).replace(",", "").strip() or 0)


class _Page:
    """pdfplumber-shaped view over a PyMuPDF page.

    The extraction logic below consumes pdfplumber-style word dicts
    ("text"/"x0"/"x1"/"top"/"bottom") and rect dicts; PyMuPDF is only the
    faster reading engine underneath, so it is adapted to the same shape.
    """

    def __init__(self, page):
        self._page = page
        self.height = page.rect.height
        self._words: Optional[List[dict]] = None

    def extract_words(self) -> List[dict]:
        if self._words is None:
            self._words = [
                {"text": w[4], "x0": w[0], "top": w[1], "x1": w[2], "bottom": w[3]}
                for w in self._page.get_text("words")
            ]
        return self._words

    def extract_text(self) -> str:
        lines = _cluster_words_into_lines(self.extract_words(), tolerance=3.0)
        return "\n".join(text for _, text, _ in lines)

    @property
    def rects(self) -> List[dict]:
        rects = []
        for drawing in self._page.get_drawings():
            for item in drawing["items"]:
                if item[0] == "re":
                    r = item[1]
                    rects.append({"x0": r.x0, "top": r.y0, "x1": r.x1, "bottom": r.y1})
        return rects


class _Document:
    def __init__(self, pdf_bytes: bytes):
        self._doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        self.pages = [_Page(p) for p in self._doc]

    def __enter__(self):
        return self

    def __exit__(self, *exc):
        self._doc.close()


def _cluster_words_into_lines(words: List[dict], tolerance: float = 2.0) -> List[Tuple[float, str, List[dict]]]:
    """Group pdfplumber words into text lines ordered top-to-bottom."""
    buckets: Dict[float, List[dict]] = {}
    for word in words:
        placed = False
        for top in buckets:
            if abs(top - word["top"]) <= tolerance:
                buckets[top].append(word)
                placed = True
                break
        if not placed:
            buckets[word["top"]] = [word]
    lines = []
    for top in sorted(buckets):
        line_words = sorted(buckets[top], key=lambda w: w["x0"])
        lines.append((top, " ".join(w["text"] for w in line_words), line_words))
    return lines


def _extract_config_grid(page, min_top: float) -> Tuple[List[Tuple[float, float]], List[float]]:
    """Return the config table grid on a page: row (top, bottom) bands and
    column left-edge x positions, both taken from the drawn cell rects."""
    stripes: List[Tuple[float, float]] = []
    edge_clusters: List[List[float]] = []  # [x0, count]
    for rect in page.rects:
        height = rect["bottom"] - rect["top"]
        if rect["top"] < min_top or not 3 < height < 60 or rect["x1"] - rect["x0"] < 20:
            continue
        for band in stripes:
            if abs(band[0] - rect["top"]) <= 1.5:
                break
        else:
            stripes.append((rect["top"], rect["bottom"]))
        for cluster in edge_clusters:
            if abs(cluster[0] - rect["x0"]) <= 2:
                cluster[1] += 1
                break
        else:
            edge_clusters.append([rect["x0"], 1])
    stripes.sort()
    # The config rows dominate the page, so their 4 cell edges are the most
    # frequent x positions — this keeps stray tables (e.g. the Distributor
    # block) from shifting the column boundaries.
    edge_clusters.sort(key=lambda c: -c[1])
    col_edges = sorted(c[0] for c in edge_clusters[:4])

    # Rows without their own cell rects (unshaded zebra rows) live in the gaps
    # between consecutive drawn rows.
    bands = []
    for i, (top, bottom) in enumerate(stripes):
        bands.append((top, bottom))
        if i + 1 < len(stripes):
            gap_top, gap_bottom = bottom, stripes[i + 1][0]
            if 4 <= gap_bottom - gap_top <= 60:
                bands.append((gap_top, gap_bottom))
    return bands, col_edges


def _parse_config_section(pdf, start_page_idx: int, start_top: float) -> List[dict]:
    """Parse CONFIGURATION DETAILS into per-item component lists."""
    config_items: List[dict] = []
    current: Optional[dict] = None

    for page_idx in range(start_page_idx, len(pdf.pages)):
        page = pdf.pages[page_idx]
        words = page.extract_words()
        min_top = start_top if page_idx == start_page_idx else 0.0

        # A page whose text lacks the repeated table header is past the
        # configuration section.
        header_texts = {w["text"] for w in words}
        if not {"Components", "Description", "Qty"} <= header_texts:
            if config_items:
                break
            continue

        bands, col_edges = _extract_config_grid(page, min_top)
        if len(col_edges) < 4:
            continue
        comp_x0, desc_x0, qty_x0 = col_edges[1], col_edges[2], col_edges[3]

        # The partner blocks ("Authorized Partners" / "Distributor", each with
        # "Partner Number | Partner Address" headers) follow the configuration
        # section — ignore everything from the first such header downwards.
        cut_top: Optional[float] = None
        partner_tops = sorted(w["top"] for w in words if w["text"] == "Partner")
        for a, b in zip(partner_tops, partner_tops[1:]):
            if abs(a - b) <= 3:
                cut_top = a
                break

        for band_top, band_bottom in bands:
            if cut_top is not None and band_bottom > cut_top - 2:
                break
            cols: Dict[str, List[dict]] = {"item": [], "comp": [], "desc": [], "qty": []}
            for w in words:
                center = (w["top"] + w["bottom"]) / 2
                if not band_top <= center <= band_bottom:
                    continue
                if w["x0"] < comp_x0 - 2:
                    cols["item"].append(w)
                elif w["x0"] < desc_x0 - 2:
                    cols["comp"].append(w)
                elif w["x0"] < qty_x0 - 2:
                    cols["desc"].append(w)
                else:
                    cols["qty"].append(w)
            joined = {
                key: " ".join(w["text"] for w in sorted(ws_, key=lambda w: (round(w["top"]), w["x0"])))
                for key, ws_ in cols.items()
            }
            row_text = " ".join(joined.values())
            if "Distributor" in row_text and "Partner" in row_text:
                # The Distributor / Partner block follows the configuration
                # section — everything after it is not configuration data.
                return config_items
            if joined["comp"] == "Components" and joined["desc"] == "Description":
                continue
            if joined["item"].replace("Line", "").strip().isdigit() and joined["comp"]:
                current = {
                    "line_no": int(joined["item"].replace("Line", "").strip()),
                    "part_number": joined["comp"],
                    "title": joined["desc"],
                    "qty": int(_parse_number(joined["qty"])) if joined["qty"] else None,
                    "components": [],
                }
                config_items.append(current)
            elif current is not None and (joined["comp"] or joined["desc"]):
                current["components"].append((joined["comp"], joined["desc"], joined["qty"]))

    return config_items


def parse_lenovo_quote_pdf(pdf_bytes: bytes) -> dict:
    """Extract metadata and pricing items from a Lenovo quotation PDF.

    Returns a dict with: customer, bid_number, currency, price_end_date
    (datetime or None), items — a list of (line_no, part_number, description,
    qty, unit_price) — and config, the CONFIGURATION DETAILS per product.
    """
    items: List[Tuple[int, str, str, int, float]] = []
    meta = {
        "customer": "",
        "bid_number": "",
        "currency": "USD",
        "price_end_date": None,
        "grand_total": None,
        "config": [],
    }
    config_start: Optional[Tuple[int, float]] = None

    with _Document(pdf_bytes) as pdf:
        first_text = pdf.pages[0].extract_text() or ""
        m = re.search(r"Customer Name:\s*(.+)", first_text)
        if m:
            meta["customer"] = m.group(1).strip()
        m = re.search(r"Bid Request No\.?\s*:?\s*([A-Z0-9]+)", first_text)
        if m:
            meta["bid_number"] = m.group(1).strip()
        m = re.search(r"Price End Date\s*:?\s*(\d{1,2}/\d{1,2}/\d{2,4})", first_text)
        if m:
            raw = m.group(1)
            for fmt in ("%d/%m/%y", "%d/%m/%Y"):
                try:
                    meta["price_end_date"] = datetime.strptime(raw, fmt)
                    break
                except ValueError:
                    continue

        in_table = False
        done = False
        desc_x0: Optional[float] = None
        desc_x1: Optional[float] = None
        page_offset = 0.0

        for page_idx, page in enumerate(pdf.pages):
            lines = _cluster_words_into_lines(page.extract_words())
            main_rows = []   # (global_top, match)
            wrap_lines = []  # (global_top, text)

            for top, text, line_words in lines:
                global_top = page_offset + top
                if not in_table:
                    if "PRODUCT AND SERVICE DETAILS" in text:
                        in_table = True
                    continue
                if "CONFIGURATION DETAILS" in text:
                    config_start = (page_idx, top)
                    done = True
                    break
                # The Grand Total line repeats at the bottom of every pricing
                # page, so it does not end the table — CONFIGURATION DETAILS does.
                grand = re.match(r"^Grand Total\s+([A-Z]{3})\b\s*([\d,]+\.\d{2})?", text)
                if grand:
                    meta["currency"] = grand.group(1)
                    if grand.group(2):
                        meta["grand_total"] = _parse_number(grand.group(2))
                    continue
                if text.startswith("Line Item"):
                    # Column header: capture the Description column bounds.
                    for idx, w in enumerate(line_words):
                        if w["text"] == "Description":
                            desc_x0 = w["x0"]
                        elif w["text"] == "Qty" and desc_x0 is not None:
                            desc_x1 = w["x0"]
                    continue
                m = _MAIN_ROW_RE.match(text)
                if m:
                    main_rows.append((global_top, m))
                    continue
                # Wrapped description fragments live inside the Description column.
                if desc_x0 is not None and desc_x1 is not None:
                    if all(w["x0"] >= desc_x0 - 40 and w["x1"] <= desc_x1 + 5 for w in line_words):
                        wrap_lines.append((global_top, text))
                else:
                    wrap_lines.append((global_top, text))

            fragments: Dict[int, List[Tuple[float, str]]] = {i: [] for i in range(len(main_rows))}
            for top, text in wrap_lines:
                if not main_rows:
                    break
                nearest = min(range(len(main_rows)), key=lambda i: abs(main_rows[i][0] - top))
                if abs(main_rows[nearest][0] - top) <= _WRAP_DISTANCE_PT:
                    fragments[nearest].append((top, text))

            for i, (top, m) in enumerate(main_rows):
                before = [t for frag_top, t in fragments[i] if frag_top < top]
                after = [t for frag_top, t in fragments[i] if frag_top > top]
                inline = m.group("desc").strip()
                desc = "\n".join(before + ([inline] if inline else []) + after)
                items.append(
                    (
                        int(m.group("line")),
                        m.group("part"),
                        desc,
                        int(_parse_number(m.group("qty"))),
                        _parse_number(m.group("unit")),
                    )
                )

            page_offset += page.height
            if done:
                break

        if config_start is not None:
            meta["config"] = _parse_config_section(pdf, config_start[0], config_start[1])

    meta["items"] = items
    return meta


def build_lenovo_output_filename(pdf_bytes: bytes, meta: Optional[dict] = None) -> str:
    try:
        if meta is None:
            meta = parse_lenovo_quote_pdf(pdf_bytes)
        bid = meta.get("bid_number") or ""
    except Exception:
        bid = ""
    stamp = datetime.now().strftime("%Y%m%d_%H%M")
    if bid:
        return f"Lenovo_Quotation_{bid}_{stamp}.xlsx"
    return f"Lenovo_Quotation_{stamp}.xlsx"


def _find_logo_path(candidates: List[str]) -> Optional[str]:
    base_dir = os.path.dirname(os.path.abspath(__file__))
    for name in candidates:
        for candidate in (os.path.join(base_dir, name), name):
            if os.path.exists(candidate):
                return candidate
    return None


def _add_logo(ws, path: str, anchor: str, max_width: int, max_height: int) -> None:
    img = XLImage(path)
    scale = min(max_width / img.width, max_height / img.height)
    img.width = int(img.width * scale)
    img.height = int(img.height * scale)
    ws.add_image(img, anchor)


def generate_lenovo_quote(
    pdf_bytes: bytes,
    margin_percent: float = 5.0,
    partner: str = "",
    meta: Optional[dict] = None,
) -> bytes:
    """Generate the Mindware Lenovo quotation workbook as xlsx bytes."""
    if meta is None:
        meta = parse_lenovo_quote_pdf(pdf_bytes)
    items = meta["items"]
    if not items:
        raise ValueError(
            "No pricing rows found in the PDF. Please make sure this is a Lenovo "
            "quotation with a 'PRODUCT AND SERVICE DETAILS' table."
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Quote"
    # White page look: no gridlines, only the table carries borders.
    ws.sheet_view.showGridLines = False

    for col, width in zip("ABCDEF", (7.9, 19.9, 92.1, 15.7, 14.4, 21.7)):
        ws.column_dimensions[col].width = width

    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Logos sit in rows 1-4, above the "Quotation" title: Mindware on the
    # left (column B), Lenovo on the right (column F).
    mindware_path = _find_logo_path(MINDWARE_LOGO_CANDIDATES)
    if mindware_path:
        _add_logo(ws, mindware_path, "B1", max_width=200, max_height=68)
    lenovo_path = _find_logo_path(LENOVO_LOGO_CANDIDATES)
    if lenovo_path:
        _add_logo(ws, lenovo_path, "F1", max_width=155, max_height=73)

    ws["C5"] = "Quotation"
    ws["C5"].font = Font(name="Aptos Display", size=18, color=NAVY)
    ws["C5"].alignment = center_wrap
    ws.row_dimensions[5].height = 23.4

    label_font = Font(name="Aptos Narrow", size=11, bold=True, color=LABEL_TEAL)
    value_font = Font(name="Calibri", size=11, color=LABEL_TEAL)
    left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws["B8"] = "Partner"
    ws["B8"].font = label_font
    ws["B8"].alignment = Alignment(vertical="center")
    ws["C8"] = partner or ""
    ws["C8"].font = value_font
    ws["C8"].alignment = left_wrap

    ws["B10"] = "Customer:"
    ws["B10"].font = label_font
    ws["C10"] = meta.get("customer") or ""
    ws["C10"].font = value_font
    ws["C10"].alignment = left_wrap
    ws["E10"] = "Date"
    ws["E10"].font = value_font
    ws["E10"].alignment = left_wrap
    # Quote date in Gulf Standard Time: the deployed server runs on UTC, which
    # is a day behind the local date in the evening hours.
    gst_now = datetime.now(timezone.utc) + timedelta(hours=4)
    ws["F10"] = gst_now.replace(tzinfo=None, hour=0, minute=0, second=0, microsecond=0)
    ws["F10"].font = Font(name="Aptos Narrow", size=11, color=LABEL_TEAL)
    ws["F10"].number_format = DATE_FMT
    ws["F10"].alignment = Alignment(horizontal="center")

    # ===== Table header (rows 12:13, merged per column) =====
    header_row = 12
    headers = ["Item", "Product", "Product Description", "Unit Price", "QTY", "Total Price"]
    header_font = Font(name="Aptos Narrow", size=14, bold=True, color="000000")
    for col_idx, title in enumerate(headers, start=1):
        col = get_column_letter(col_idx)
        ws.merge_cells(f"{col}{header_row}:{col}{header_row + 1}")
        for r in (header_row, header_row + 1):
            cell = ws[f"{col}{r}"]
            cell.font = header_font
            cell.alignment = center_wrap
            cell.border = border_thin
        ws[f"{col}{header_row}"] = title
    ws.row_dimensions[header_row].height = 9.6
    ws.row_dimensions[header_row + 1].height = 30.6

    # ===== Item rows =====
    if not 0 <= margin_percent < 100:
        raise ValueError("Margin % must be between 0 and 99.99.")
    margin_factor = 1 - (margin_percent / 100.0)

    data_font = Font(name="Calibri", size=11, bold=True)
    total_col_font = Font(name="Aptos Narrow", size=11, bold=True, color=NAVY)

    row = header_row + 2
    first_data_row = row
    for line_no, part_number, description, qty, unit_cost in items:
        ws[f"A{row}"] = line_no
        ws[f"B{row}"] = part_number
        ws[f"C{row}"] = description
        ws[f"D{row}"] = unit_cost / margin_factor
        ws[f"D{row}"].number_format = UNIT_PRICE_FMT
        ws[f"E{row}"] = qty
        ws[f"F{row}"] = f"=D{row}*E{row}"
        ws[f"F{row}"].number_format = UNIT_PRICE_FMT
        for col in "ABCDEF":
            cell = ws[f"{col}{row}"]
            cell.border = border_thin
            cell.alignment = center_wrap if col in ("B", "C") else center
            cell.font = total_col_font if col == "F" else data_font
        line_count = description.count("\n") + 1
        ws.row_dimensions[row].height = max(49.2, line_count * 15 + 4)
        row += 1
    last_data_row = row - 1

    # ===== Grand total row =====
    total_row = row
    ws.merge_cells(f"A{total_row}:E{total_row}")
    for col in "ABCDEF":
        cell = ws[f"{col}{total_row}"]
        cell.border = border_thin
        cell.alignment = center
    ws[f"A{total_row}"].font = Font(name="Aptos Narrow", size=11, bold=True, color=NAVY)
    ws[f"F{total_row}"] = f"=SUM(F{first_data_row}:F{last_data_row})"
    ws[f"F{total_row}"].number_format = GRAND_TOTAL_FMT
    ws[f"F{total_row}"].font = Font(name="Aptos Narrow", size=12, bold=True, color=TOTAL_FONT)
    ws.row_dimensions[total_row].height = 15.6

    # ===== Terms & Conditions =====
    validity_text = ""
    if meta.get("price_end_date"):
        d = meta["price_end_date"]
        validity_text = f"{d.day}-{d.strftime('%b-%Y')}"
    terms_start = total_row + 2
    terms_rows = ["Terms & Conditions"] + [
        t.format(validity=validity_text or "the price end date shown on the Lenovo quote")
        for t in TERMS_AND_CONDITIONS
    ]
    for offset, text in enumerate(terms_rows):
        r = terms_start + offset
        ws.merge_cells(f"A{r}:F{r}")
        cell = ws[f"A{r}"]
        cell.value = text
        if offset == 0:
            cell.font = Font(name="Times New Roman", size=12, bold=True)
        elif text.startswith("Validity:"):
            cell.font = Font(name="Times New Roman", size=12, color="FFFF0000")
        else:
            cell.font = Font(name="Times New Roman", size=12)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[r].height = 15.6

    if meta.get("config"):
        _add_configuration_sheet(wb, meta["config"], border_thin)

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def _add_configuration_sheet(wb, config_items: List[dict], border_thin: Border) -> None:
    """Add a 'Configuration' sheet: one block per configured product."""
    ws = wb.create_sheet("Configuration")
    ws.sheet_view.showGridLines = False
    for col, width in zip("ABC", (35, 80, 10)):
        ws.column_dimensions[col].width = width

    left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)
    center = Alignment(horizontal="center", vertical="center")
    item_font = Font(name="Aptos Narrow", size=12, bold=True, color=NAVY)
    col_header_font = Font(name="Calibri", size=11, bold=True)
    body_font = Font(name="Calibri", size=11)

    ws["A1"] = "Configuration Details"
    ws["A1"].font = Font(name="Aptos Display", size=16, color=NAVY)

    row = 3
    for cfg in config_items:
        qty_text = f"   |   Qty: {cfg['qty']}" if cfg.get("qty") else ""
        ws.merge_cells(f"A{row}:C{row}")
        title_cell = ws[f"A{row}"]
        title_cell.value = f"Item {cfg['line_no']}   |   {cfg['part_number']}   |   {cfg['title']}{qty_text}"
        title_cell.font = item_font
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        for col in "ABC":
            ws[f"{col}{row}"].border = border_thin
        ws.row_dimensions[row].height = 20
        row += 1

        for col, header in zip("ABC", ("Component", "Description", "Qty")):
            cell = ws[f"{col}{row}"]
            cell.value = header
            cell.font = col_header_font
            cell.alignment = center
            cell.border = border_thin
        row += 1

        for comp, desc, qty in cfg["components"]:
            ws[f"A{row}"] = comp
            ws[f"A{row}"].font = body_font
            ws[f"A{row}"].alignment = left_wrap
            ws[f"B{row}"] = desc
            ws[f"B{row}"].font = body_font
            ws[f"B{row}"].alignment = left_wrap
            ws[f"C{row}"] = qty
            ws[f"C{row}"].font = body_font
            ws[f"C{row}"].alignment = center
            for col in "ABC":
                ws[f"{col}{row}"].border = border_thin
            row += 1

        row += 1  # blank row between products
