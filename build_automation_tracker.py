"""
Builds the Automation Tracker Excel file — single sheet, internal development (cost = 0), ROI-style metrics.
Run: python build_automation_tracker.py
Output: Automation_Tracker.xlsx
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

TOOLS = [
    {
        "name": "Google DNTS Extractor",
        "team": "Finance",
        "oneline": "Turns Google Cloud invoice PDFs into the exact Excel format required for our finance system (DNTS/Orion) in one step.",
        "desc": "Google sends invoices as PDFs with a 'Summary of costs by domain' table. Previously, staff retyped domain names, customer IDs, and amounts into Excel and then adjusted layout and codes for the finance system. This tool reads the PDF, pulls invoice number and date, extracts every cost line, and produces a ready-to-upload Excel in the correct format. This cuts manual data entry, speeds up posting, and reduces errors that could affect reconciliation and reporting.",
        "manual_min": 42, "auto_min": 4, "runs_month": 24, "freq": "Weekly", "dev_hrs": 36, "months_use": 12,
        "risk": "High error rate; delayed posting",
        "quality": "Fewer posting errors; faster reconciliation",
        "status": "Live", "priority": "High", "notes": "DNTS/Orion format",
    },
    {
        "name": "Google Invoice Extractor",
        "team": "Finance",
        "oneline": "Quickly pulls domain-level costs from Google invoice PDFs into a simple Excel for analysis or checks.",
        "desc": "When the full DNTS format is not needed—for example for a quick check or a report—staff still need the cost breakdown (domain, customer ID, amount) from the PDF. This tool does that in one go: upload the Google invoice PDF and get an Excel with Domain name, Customer ID, and Amount. It saves time on ad-hoc extractions and avoids copy-paste mistakes.",
        "manual_min": 25, "auto_min": 2, "runs_month": 12, "freq": "Monthly", "dev_hrs": 20, "months_use": 12,
        "risk": "Copy-paste errors; inconsistent reports",
        "quality": "Consistent extraction format",
        "status": "Live", "priority": "Medium", "notes": "Lightweight extraction",
    },
    {
        "name": "Claims Automation",
        "team": "Finance",
        "oneline": "Combines four different claims-related files (SAP export, user list, benefits, account mapping) into one ready-to-use claims file.",
        "desc": "Claims processing used to depend on four separate files: SAP journal data, employee master, benefit details, and account mapping. Staff had to merge them, match employees to accounts, and build narrations by hand. This tool takes all four as input and produces a single, consistent claims file with the right account codes and narrations. It reduces reconciliation errors, speeds up month-end, and keeps audit trails clear.",
        "manual_min": 120, "auto_min": 15, "runs_month": 1, "freq": "Monthly", "dev_hrs": 80, "months_use": 12,
        "risk": "Reconciliation errors; delayed month-end",
        "quality": "Clear audit trail; correct account mapping",
        "status": "Live", "priority": "High", "notes": "4 inputs → 1 output",
    },
    {
        "name": "AWS Invoice Tool",
        "team": "Finance",
        "oneline": "Reads AWS invoice and credit-note PDFs and produces both a summary Excel and the correct files for our finance system (DNTS/CNTS) by entity.",
        "desc": "AWS sends different PDF layouts (tax invoice, credit note, direct vs marketplace). Staff used to re-enter amounts, VAT, and billing details and then build DNTS/CNTS files per entity (e.g. Mindware UAE vs FZ). The tool detects the PDF type, extracts amounts and dates, calculates VAT where needed, and generates both a summary Excel and the right DNTS or CNTS files in one ZIP. This reduces re-keying, ensures correct entity and VAT treatment, and speeds up AWS invoice posting.",
        "manual_min": 55, "auto_min": 6, "runs_month": 18, "freq": "Weekly", "dev_hrs": 45, "months_use": 12,
        "risk": "Wrong entity/VAT; delayed posting",
        "quality": "Correct VAT and entity split",
        "status": "Live", "priority": "High", "notes": "Multi-entity DNTS/CNTS",
    },
    {
        "name": "Dell Invoice Extractor (Pre-Alert Upload)",
        "team": "Operations",
        "oneline": "Turns Dell invoice PDFs into the pre-alert upload sheet and highlights lines that need review or master-data updates.",
        "desc": "Pre-alert requires data from Dell PDFs (PO, invoice number, dates, line items, shipping, consolidation) to be entered into our template and matched to internal item codes and prices. Doing this by hand is slow and error-prone. The tool reads one or more Dell PDFs (and an optional master file), fills the pre-alert sheet, matches items where possible, and adds a review sheet that flags 'no match,' 'price match,' or 'ok' so staff can focus on exceptions. This shortens pre-alert turnaround and improves accuracy of item and price matching.",
        "manual_min": 90, "auto_min": 12, "runs_month": 15, "freq": "Weekly", "dev_hrs": 65, "months_use": 10,
        "risk": "Wrong item/price match; delayed pre-alert",
        "quality": "Faster pre-alert; fewer matching errors",
        "status": "Live", "priority": "High", "notes": "Optional master file; REVIEW sheet",
    },
    {
        "name": "Cloud Invoice Tool",
        "team": "Operations",
        "oneline": "Converts the cloud billing export (after CB Convert) into standard cloud invoice Excel, negative-invoice file, and SRCL file with correct codes and versioning.",
        "desc": "Cloud billing data comes from a CB export that must be converted and then mapped to our locations, tax codes, item codes (e.g. M365, Azure), LPO, and end user. Staff used to do this in Excel with a lot of manual lookup and versioning. The tool takes the converted file, applies location and tax rules, derives item codes from descriptions, extracts LPO and end user, and produces the main cloud invoice workbook (with versioning), the negative-invoices file, and the SRCL file. It keeps coding consistent and reduces manual parsing and versioning errors.",
        "manual_min": 180, "auto_min": 20, "runs_month": 1, "freq": "Monthly", "dev_hrs": 100, "months_use": 12,
        "risk": "Wrong codes/versioning; audit issues",
        "quality": "Consistent location/tax/item codes",
        "status": "Live", "priority": "High", "notes": "CB Convert required first",
    },
    {
        "name": "Barcode PDF Generator",
        "team": "Operations",
        "oneline": "Generates pallet barcode labels from a simple list of Pallet ID and IMEI so warehouse can print and apply labels without manual barcode creation.",
        "desc": "Shipping and warehouse need barcode labels per pallet, with each label encoding a set of IMEIs. Creating these manually is tedious and error-prone. Staff upload a CSV with Pallet ID and IMEI; the tool checks that IMEIs are valid and grouped correctly, then produces a PDF of barcode pages (one page per group) with the pallet ID shown. This speeds up labeling and reduces wrong or missing barcodes.",
        "manual_min": 60, "auto_min": 5, "runs_month": 30, "freq": "Daily", "dev_hrs": 28, "months_use": 12,
        "risk": "Wrong/missing barcodes; shipping delays",
        "quality": "Accurate pallet labels",
        "status": "Live", "priority": "Medium", "notes": "CSV: PalletID, IMEI",
    },
    {
        "name": "AR to EDD File",
        "team": "Credit",
        "oneline": "Filters the insurance/AR export to the right ageing window and adds the columns needed for EDD and reconciliation in one go.",
        "desc": "Credit needs to work from the insurance file with only relevant rows (e.g. positive limit, balance, and ageing between 200–270 days) and standard columns for EDD (status, reason, paid amount, payment date, over-due days). Doing this manually with filters and new columns is repetitive and easy to get wrong. The tool reads the export (from row 16), applies the business rules and ageing range, and outputs a clean Excel ready for EDD and reconciliation. This keeps the process consistent and saves time each cycle.",
        "manual_min": 75, "auto_min": 8, "runs_month": 1, "freq": "Monthly", "dev_hrs": 32, "months_use": 12,
        "risk": "Wrong ageing filter; inconsistent EDD",
        "quality": "Consistent EDD format",
        "status": "Live", "priority": "High", "notes": "Ageing window configurable",
    },
    {
        "name": "Coface CSV Uploader",
        "team": "Credit",
        "oneline": "Splits one customer-invoice export into separate, correctly formatted CSV files per customer for Coface (or similar) submission.",
        "desc": "Coface (or similar) often requires one file per customer in a specific format: document number, dates, balance, status, paid amount, payment date, reason (semicolon-separated). Manually splitting and formatting from one big export is slow and risks wrong formatting or wrong customer split. The tool checks status rules (e.g. UNPAID must have zero paid amount and blank payment date), then groups by customer code and outputs one CSV per customer in the required format, zipped. This ensures consistent format and reduces manual splitting and copy-paste errors.",
        "manual_min": 90, "auto_min": 10, "runs_month": 2, "freq": "Monthly", "dev_hrs": 40, "months_use": 8,
        "risk": "Wrong format; wrong customer split",
        "quality": "Consistent Coface format",
        "status": "Live", "priority": "Medium", "notes": "One CSV per customer; ZIP",
    },
    {
        "name": "IBM Quotation",
        "team": "Sales",
        "oneline": "Takes an IBM quotation PDF (and optional Excel), detects the quotation type, and produces a single styled quotation Excel in the right currency for UAE, Qatar, or KSA.",
        "desc": "IBM sends quotation PDFs in different formats (parts with coverage dates vs subscription/SaaS). Sales used to copy data between PDF and Excel, set currency (AED/SAR/USD) by country, and keep terms and layout consistent. The tool lets the user choose country (UAE, Qatar, KSA), uploads the PDF and optionally the Excel, and automatically detects whether it is a 'parts' or 'subscription' quotation. It then builds one styled quotation Excel (with logo and terms), checks total price vs PDF where relevant, and aligns dates between PDF and Excel. This reduces manual rework, keeps pricing and compliance text consistent across countries, and speeds up quotation turnaround.",
        "manual_min": 45, "auto_min": 5, "runs_month": 35, "freq": "Daily", "dev_hrs": 72, "months_use": 12,
        "risk": "Wrong currency; inconsistent terms",
        "quality": "Consistent pricing and compliance",
        "status": "Live", "priority": "High", "notes": "UAE, Qatar, KSA; Template 1 & 2",
    },
    {
        "name": "MIBB Quotations",
        "team": "Sales",
        "oneline": "(Planned) Same idea as IBM Quotation but for MIBB quotation PDFs—one-click extraction and styled Excel with correct terms.",
        "desc": "This tool is planned for MIBB quotation PDFs: extract header and line table (part number, description, dates, qty, price), optionally apply a description master, and produce a styled quotation Excel with MIBB terms. It is not yet connected in the app; list it in the tracker as Planned and use it for roadmap and capacity discussions.",
        "manual_min": 50, "auto_min": 6, "runs_month": 8, "freq": "Monthly", "dev_hrs": 55, "months_use": 0,
        "risk": "Manual rework until live",
        "quality": "TBD when live",
        "status": "Planned", "priority": "Medium", "notes": "Not yet in app",
    },
]

HEADER_FILL = PatternFill(start_color="1a73e8", end_color="1a73e8", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SUMMARY_FILL = PatternFill(start_color="e8f0fe", end_color="e8f0fe", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "Automation Tracker"

    n = len(TOOLS)
    # Data table: header at row 18, data at rows 19..19+n-1
    data_first_row = 19
    last_data_row = data_first_row + n - 1

    # ---- Top: title and assumptions ----
    ws["A1"] = "Automation Tracker — Management View"
    ws["A1"].font = Font(bold=True, size=16, color="1a73e8")
    ws.merge_cells("A1:D1")
    ws["A2"] = "Cost per hour (USD)"
    ws["B2"] = 50
    ws["A2"].fill = SUMMARY_FILL
    ws["B2"].fill = SUMMARY_FILL
    ws["A3"] = "Year"
    ws["B3"] = 2025
    ws["A3"].fill = SUMMARY_FILL
    ws["B3"].fill = SUMMARY_FILL
    ws["A4"] = "Development: Internal (no third-party cost)"
    ws.merge_cells("A4:D4")
    ws["A4"].fill = SUMMARY_FILL
    ws["A4"].font = Font(italic=True)

    # ---- Summary block ----
    r = 6
    ws.cell(row=r, column=1, value="Total hours saved this year").fill = SUMMARY_FILL
    ws.cell(row=r, column=2, value=f"=SUM(L{data_first_row}:L{last_data_row})").fill = SUMMARY_FILL
    r += 1
    ws.cell(row=r, column=1, value="Total value saved this year (USD)").fill = SUMMARY_FILL
    ws.cell(row=r, column=2, value=f"=SUM(P{data_first_row}:P{last_data_row})").fill = SUMMARY_FILL
    r += 1
    ws.cell(row=r, column=1, value="Number of tools (Live)").fill = SUMMARY_FILL
    ws.cell(row=r, column=2, value=f'=COUNTIF(V{data_first_row}:V{last_data_row},"Live")').fill = SUMMARY_FILL
    r += 1
    ws.cell(row=r, column=1, value="Average payback (months)").fill = SUMMARY_FILL
    ws.cell(row=r, column=2, value=f"=AVERAGE(N{data_first_row}:N{last_data_row})").fill = SUMMARY_FILL
    r += 1
    ws.cell(row=r, column=1, value="Return per dev hour (USD) — overall").fill = SUMMARY_FILL
    ws.cell(row=r, column=2, value=f"=IF(SUM(M{data_first_row}:M{last_data_row})=0,\"\",SUM(P{data_first_row}:P{last_data_row})/SUM(M{data_first_row}:M{last_data_row}))").fill = SUMMARY_FILL
    r += 2

    # ---- By team ----
    ws.cell(row=r, column=1, value="By team").font = Font(bold=True, size=12)
    r += 1
    teams = ["Finance", "Operations", "Credit", "Sales"]
    for c, h in enumerate(["Team", "Hours saved/year", "Value saved/year (USD)", "Return per dev hour (USD)"], 1):
        ws.cell(row=r, column=c, value=h).fill = HEADER_FILL
        ws.cell(row=r, column=c).font = HEADER_FONT
    r += 1
    for team in teams:
        ws.cell(row=r, column=1, value=team)
        ws.cell(row=r, column=2, value=f'=SUMIF(B{data_first_row}:B{last_data_row},"{team}",L{data_first_row}:L{last_data_row})')
        ws.cell(row=r, column=3, value=f'=SUMIF(B{data_first_row}:B{last_data_row},"{team}",P{data_first_row}:P{last_data_row})')
        ws.cell(row=r, column=4, value=f'=IF(SUMIF(B{data_first_row}:B{last_data_row},"{team}",M{data_first_row}:M{last_data_row})=0,"",SUMIF(B{data_first_row}:B{last_data_row},"{team}",P{data_first_row}:P{last_data_row})/SUMIF(B{data_first_row}:B{last_data_row},"{team}",M{data_first_row}:M{last_data_row}))')
        r += 1
    r += 1

    # ---- Main table headers ----
    headers = [
        "Tool", "Owner team", "One-line summary", "Description",
        "Manual (min)", "Automated (min)", "Time saved per run (min)", "Frequency used", "Runs per month", "Months in use",
        "Hours saved per month", "Time saved per year (hours)", "Development (hours)", "Payback (months)",
        "Cost per hour (USD)", "Value saved per year (USD)", "Development cost (USD)", "Return per dev hour (USD)", "Hours saved per dev hour",
        "Risk without automation", "Quality impact", "Status", "Priority", "Notes",
    ]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    r += 1
    data_start_row = r

    # ---- Data rows ----
    for i, t in enumerate(TOOLS):
        row = data_start_row + i
        ws.cell(row=row, column=1, value=t["name"]).border = THIN_BORDER
        ws.cell(row=row, column=2, value=t["team"]).border = THIN_BORDER
        ws.cell(row=row, column=3, value=t["oneline"]).border = THIN_BORDER
        ws.cell(row=row, column=4, value=t["desc"]).border = THIN_BORDER
        ws.cell(row=row, column=5, value=t["manual_min"]).border = THIN_BORDER
        ws.cell(row=row, column=6, value=t["auto_min"]).border = THIN_BORDER
        ws.cell(row=row, column=7, value=f"=E{row}-F{row}").border = THIN_BORDER
        ws.cell(row=row, column=8, value=t["freq"]).border = THIN_BORDER
        ws.cell(row=row, column=9, value=t["runs_month"]).border = THIN_BORDER
        ws.cell(row=row, column=10, value=t["months_use"]).border = THIN_BORDER
        ws.cell(row=row, column=11, value=f"=(G{row}*I{row})/60").border = THIN_BORDER
        ws.cell(row=row, column=12, value=f"=K{row}*J{row}").border = THIN_BORDER
        ws.cell(row=row, column=13, value=t["dev_hrs"]).border = THIN_BORDER
        ws.cell(row=row, column=14, value=f"=IF(K{row}=0,\"\",ROUND(M{row}/K{row},1))").border = THIN_BORDER
        ws.cell(row=row, column=15, value=f"=$B$2").border = THIN_BORDER
        ws.cell(row=row, column=16, value=f"=L{row}*O{row}").border = THIN_BORDER
        ws.cell(row=row, column=17, value=0).border = THIN_BORDER  # Development cost = 0 (internal)
        ws.cell(row=row, column=18, value=f"=IF(M{row}=0,\"\",ROUND(P{row}/M{row},0))").border = THIN_BORDER  # Return per dev hour (USD)
        ws.cell(row=row, column=19, value=f"=IF(M{row}=0,\"\",ROUND(L{row}/M{row},1))").border = THIN_BORDER   # Hours saved per dev hour
        ws.cell(row=row, column=20, value=t["risk"]).border = THIN_BORDER
        ws.cell(row=row, column=21, value=t["quality"]).border = THIN_BORDER
        ws.cell(row=row, column=22, value=t["status"]).border = THIN_BORDER
        ws.cell(row=row, column=23, value=t["priority"]).border = THIN_BORDER
        ws.cell(row=row, column=24, value=t["notes"]).border = THIN_BORDER

    # ---- Column widths and alignment ----
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 48
    ws.column_dimensions["D"].width = 62
    for col in range(5, 25):
        ws.column_dimensions[get_column_letter(col)].width = 14
    for row in range(data_start_row, data_start_row + n):
        for col in range(1, 25):
            ws.cell(row=row, column=col).alignment = Alignment(vertical="top", wrap_text=True)

    # Freeze panes: keep title/summary and column A visible when scrolling
    ws.freeze_panes = ws.cell(row=data_start_row, column=2)

    out_path = "Automation_Tracker.xlsx"
    wb.save(out_path)
    print(f"Saved: {out_path}")


if __name__ == "__main__":
    main()
