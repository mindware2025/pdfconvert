# southcomp_engine.py
# Standalone EUR-only quote engine for Dell Quotation Southcomp Polaris.
# No imports from dell.py or any other dell_* module.

from datetime import datetime, timedelta
from io import BytesIO
from typing import Dict, List, Optional, Tuple
import os
import re

import openpyxl
from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter, column_index_from_string as colidx

try:
    from PIL import Image as PILImage, ImageChops
except ImportError:
    PILImage = ImageChops = None


# ==================== CONSTANTS ====================

CURRENCY_FORMATS = {
    "EUR": '"€"#,##0.00',
    "USD": '"$"#,##0.00',
}

# EUR is the base; USD conversion keeps original USD prices from the BOQ
CONVERSION_RATES: Dict[str, float] = {
    "EUR": 0.92,
    "USD": 1.0,
}


# ==================== HELPERS ====================

def _parse_money(val) -> Optional[float]:
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = re.sub(r"[^\d,.\-]", "", str(val).strip())
    if "," in s and "." in s:
        s = s.replace(",", "")
    elif "," in s and "." not in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except Exception:
        return None


def _cell_to_text(v, fallback: str = "") -> str:
    if v is None:
        return fallback
    if isinstance(v, datetime):
        return v.strftime("%d/%m/%Y")
    return _sanitize_excel_text(str(v).strip())


def _sanitize_excel_text(value: str) -> str:
    if value is None:
        return ""
    return ILLEGAL_CHARACTERS_RE.sub("", str(value))[:32767]


def _normalize_text(s: str) -> str:
    return re.sub(r"[^a-z0-9]", "", s.lower()) if s else ""


def _strip_trailing_asterisk(value: str) -> str:
    if value is None:
        return ""
    text = _cell_to_text(value).split("|")[0]
    return re.sub(r"\s*\*+", "", text).strip()


def _row_text(ws, r: int, c1: int = 1, c2: Optional[int] = None) -> str:
    if c2 is None:
        c2 = ws.max_column
    return " ".join(
        _cell_to_text(ws.cell(r, c).value)
        for c in range(c1, c2 + 1)
        if ws.cell(r, c).value
    ).strip()


def _is_price_or_qty_line(text: str) -> bool:
    t = text.lower()
    return any(tok in t for tok in [
        "qty", "quantity", "unit price", "subtotal", "total", "price",
        "amount", "discount", "tax", "grand total", "msrp", "usd", "aed", "eur", "sar",
    ]) or bool(re.search(r"(\$|€|£|aed|usd|eur|sar)", t, re.IGNORECASE))


def _sanitize_filename_part(value: str) -> str:
    text = re.sub(r"\s+", " ", _cell_to_text(value)).strip()
    if not text:
        return ""
    text = re.sub(r'[<>:"/\\|?*]', "", text)
    return text.rstrip(". ")


_MONTH_NAMES = {
    "january": 1, "february": 2, "march": 3, "april": 4,
    "may": 5, "june": 6, "july": 7, "august": 8,
    "september": 9, "october": 10, "november": 11, "december": 12,
}
_MONTH_DATE_PAT = re.compile(
    r"([A-Za-z]+)\s+(\d{1,2}),?\s+(\d{4})"
)


def _parse_month_date(text: str) -> str:
    """Parse 'July 15, 2026' → '15/07/2026'. Returns '' on failure."""
    m = _MONTH_DATE_PAT.search(text)
    if not m:
        return ""
    month_str, day_str, year_str = m.groups()
    month_num = _MONTH_NAMES.get(month_str.lower())
    if not month_num:
        return ""
    return f"{int(day_str):02d}/{month_num:02d}/{year_str}"


# ==================== LOGO ====================

def _trim_logo(pil_img):
    if PILImage is None:
        return pil_img
    img = pil_img.convert("RGBA")
    alpha = img.getchannel("A")
    bbox = alpha.getbbox()
    if bbox:
        img = img.crop(bbox)
    if ImageChops is not None:
        bg = PILImage.new("RGBA", img.size, (255, 255, 255, 0))
        diff = ImageChops.difference(img, bg)
        bbox = diff.getbbox()
        if bbox:
            img = img.crop(bbox)
    return img


def _pil_to_xl(pil_img):
    buf = BytesIO()
    pil_img.save(buf, format="PNG")
    buf.seek(0)
    return XLImage(buf)


def _logo_path() -> Optional[str]:
    base = os.path.dirname(os.path.abspath(__file__))
    for name in ["dell spc.png", "dell copy.png", "dell.png"]:
        for directory in [base, os.path.dirname(base)]:
            p = os.path.join(directory, name)
            if os.path.exists(p):
                return p
    return None


def _add_logo(ws, anchor: str = "A1", width: int = 780, height: int = 52) -> None:
    path = _logo_path()
    if not path:
        return
    if PILImage is not None:
        try:
            img = _pil_to_xl(_trim_logo(PILImage.open(path)))
            img.width = width
            img.height = height
            ws.add_image(img, anchor)
            return
        except Exception:
            pass
    try:
        img = XLImage(path)
        img.width = width
        img.height = height
        ws.add_image(img, anchor)
    except Exception:
        pass


# ==================== TEMPLATE DETECTION ====================

def detect_template_type(input_bytes: bytes) -> str:
    """Return 'extended_services' or 'standard_quote'."""
    if input_bytes.lstrip().startswith(b"%PDF"):
        return "standard_quote"
    try:
        wb = openpyxl.load_workbook(BytesIO(input_bytes), data_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=1, max_row=80, max_col=10):
            for cell in row:
                if isinstance(cell.value, str) and "dell extended services details" in cell.value.lower():
                    return "extended_services"
    except Exception:
        pass
    return "standard_quote"


# ==================== HEADER / COLUMN DETECTION ====================

def _find_compact_header(ws) -> Optional[Tuple[int, Dict[str, int]]]:
    for r in range(1, min(ws.max_row, 40) + 1):
        cols: Dict[str, int] = {}
        for c in range(1, ws.max_column + 1):
            name = _cell_to_text(ws.cell(r, c).value).strip().lower()
            if not name:
                continue
            if name == "#" and "item" not in cols:
                cols["item"] = c
            if "sku" in name and "sku" not in cols:
                cols["sku"] = c
            if "description" in name and "description" not in cols:
                cols["description"] = c
            if name in ("q-ty", "qty", "quantity") and "qty" not in cols:
                cols["qty"] = c
            if ("unit selling price" in name or "unit price" in name) and "unit" not in cols:
                cols["unit"] = c
            if ("total selling price" in name or "total price" in name) and "total" not in cols:
                cols["total"] = c
        if all(k in cols for k in ("description", "qty", "total")) and ("sku" in cols or "item" in cols):
            return r, cols
    return None


def _find_grouped_header(ws) -> Optional[Tuple[int, Dict[str, int]]]:
    for r in range(1, min(ws.max_row, 40) + 1):
        row_values = [_cell_to_text(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
        normalized = [re.sub(r"\s+", " ", v.strip().lower()) for v in row_values]
        has_desc = any("description" in n for n in normalized)
        has_sku = any("sku" in n or "part number" in n or "part no" in n for n in normalized)
        has_qty = any(n in ("qty", "quantity", "q-ty") for n in normalized)
        has_unit = any("unit selling price" in n or "unit price" in n for n in normalized)
        has_total = any("total selling price" in n or "total price" in n for n in normalized)
        if has_desc and has_sku and has_qty and has_unit and has_total:
            cols: Dict[str, int] = {}
            for c, n in enumerate(normalized, start=1):
                if "description" in n and "description" not in cols:
                    cols["description"] = c
                if ("sku" in n or "part number" in n or "part no" in n) and "sku" not in cols:
                    cols["sku"] = c
                if n in ("qty", "quantity", "q-ty") and "qty" not in cols:
                    cols["qty"] = c
                if ("unit selling price" in n or "unit price" in n) and "unit" not in cols:
                    cols["unit"] = c
                if ("total selling price" in n or "total price" in n) and "total" not in cols:
                    cols["total"] = c
            if "description" in cols and "sku" in cols:
                return r, cols
    return None


def _find_generic_header(ws) -> Tuple[int, int, int, int]:
    """Return (first_data_row, desc_col, qty_col, unit_col)."""
    for r in range(1, min(ws.max_row, 40) + 1):
        row_vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if not any(row_vals):
            continue
        texts = [_cell_to_text(v).lower() for v in row_vals]
        if any("description" in t for t in texts) and any("qty" in t or "quantity" in t for t in texts):
            desc_idx = qty_idx = unit_idx = None
            for i, v in enumerate(row_vals, start=1):
                name = _cell_to_text(v).lower()
                if desc_idx is None and "description" in name:
                    desc_idx = i
                if qty_idx is None and ("qty" in name or "quantity" in name):
                    qty_idx = i
                if unit_idx is None and ("unit price" in name or "unitprice" in name or name == "price"):
                    unit_idx = i
            return r + 1, desc_idx or 3, qty_idx or 4, unit_idx or 5
    return 8, 3, 4, 5


# ==================== METADATA EXTRACTION ====================

def _scan_all_quote_refs(ws, max_rows: int = 80) -> List[str]:
    pat = r"\b\d{6,}(?:\.[A-Za-z0-9]+)?[A-Za-z0-9\-]*\b"
    refs = []
    for r in range(1, min(ws.max_row, max_rows) + 1):
        for c in range(1, min(ws.max_column, 10) + 1):
            text = _cell_to_text(ws.cell(r, c).value)
            low = text.lower()
            if low.startswith("quote") and "quoted on" not in low:
                m = re.search(pat, text)
                if m:
                    refs.append(m.group(0))
            elif any(tok in low for tok in ("quote no", "quote number", "quote ref")):
                m = re.search(pat, text)
                if m:
                    refs.append(m.group(0))
                else:
                    row_texts = [_cell_to_text(ws.cell(r, cc).value) for cc in range(1, ws.max_column + 1)]
                    for t in row_texts:
                        m = re.search(pat, t)
                        if m:
                            refs.append(m.group(0))
                            break
    seen = []
    for ref in refs:
        if ref not in seen:
            seen.append(ref)
    return seen


def _find_label_value(ws, labels: Tuple[str, ...], max_rows: int = 60, max_cols: int = 10) -> str:
    for r in range(1, min(ws.max_row, max_rows) + 1):
        for c in range(1, min(ws.max_column, max_cols) + 1):
            text = _cell_to_text(ws.cell(r, c).value).strip().lower()
            if not text:
                continue
            if any(label in text for label in labels):
                for nc in range(c + 1, min(ws.max_column, max_cols) + 1):
                    candidate = _cell_to_text(ws.cell(r, nc).value).strip()
                    if candidate:
                        return candidate
                for nr in range(r + 1, min(ws.max_row, max_rows) + 1):
                    candidate = _cell_to_text(ws.cell(nr, c).value).strip()
                    if candidate:
                        return candidate
    return ""


def _extract_metadata(ws) -> Tuple[str, str]:
    """Return (quote_ref, date)."""
    ref = _find_label_value(ws, ("quote no", "quote number", "quote ref", "quotation no"))
    if not ref:
        raw = ws["E15"].value
        ref = "" if raw is None else (raw.strftime("%d/%m/%Y") if isinstance(raw, datetime) else str(raw).strip())

    date = _find_label_value(ws, ("quote date", "quoted on", "date"))
    if not date:
        raw_d = ws["E18"].value
        if isinstance(raw_d, datetime):
            date = raw_d.strftime("%d/%m/%Y")
        else:
            date = "" if raw_d is None else str(raw_d).strip()

    all_refs = _scan_all_quote_refs(ws)
    if all_refs:
        combined = []
        if ref:
            combined.append(ref)
        for r in all_refs:
            if r not in combined:
                combined.append(r)
        ref = ", ".join(combined)

    # Fallback: scan for quoted-on date
    if not date:
        for row in ws.iter_rows(min_row=1, max_row=80, max_col=10):
            row_text = " ".join(_cell_to_text(c.value) for c in row).lower()
            if "quoted on" in row_text or "quote date" in row_text:
                for cell in row:
                    m = re.search(r"\d{2}/\d{2}/\d{4}", str(cell.value))
                    if m:
                        date = m.group(0)
                        break

    return ref, date


def _extract_expiry(ws) -> str:
    def _fmt(value) -> str:
        if isinstance(value, datetime):
            return value.strftime("%d/%m/%Y")
        text = str(value or "").strip()
        m = re.search(r"\d{2}/\d{2}/\d{4}", text)
        return m.group(0) if m else text

    def _adjust(value: str) -> str:
        if not value:
            return ""
        try:
            return (datetime.strptime(value, "%d/%m/%Y") - timedelta(days=2)).strftime("%d/%m/%Y")
        except Exception:
            return value

    # Try strict position first
    direct = _fmt(ws["E19"].value)
    if direct:
        return _adjust(direct)

    for r in range(1, min(ws.max_row, 80) + 1):
        row_text = " ".join(_cell_to_text(ws.cell(r, c).value) for c in range(1, min(ws.max_column, 10) + 1)).lower()
        if "expires by" not in row_text:
            continue
        for c in range(1, min(ws.max_column, 10) + 1):
            if "expires by" in _cell_to_text(ws.cell(r, c).value).lower():
                for nc in range(c + 1, min(ws.max_column, 10) + 1):
                    candidate = _fmt(ws.cell(r, nc).value)
                    if candidate:
                        return _adjust(candidate)
        m = re.search(r"\d{2}/\d{2}/\d{4}", row_text)
        if m:
            return _adjust(m.group(0))
    return ""


def _extract_quote_metadata(ws) -> Dict[str, str]:
    keys = {"company name", "customer name", "customer number", "end user", "reseller"}
    out = {k: "" for k in keys}
    max_row = min(ws.max_row, 120)
    for r in range(1, max_row + 1):
        label = _cell_to_text(ws.cell(r, 2).value).strip().lower().rstrip(":")
        if label in keys:
            out[label] = _cell_to_text(ws.cell(r, 5).value)
    # Shipping information block (PDF-style Excel)
    for r in range(1, max_row + 1):
        row_values = [_cell_to_text(ws.cell(r, c).value) for c in range(1, 11)]
        if any("shipping information" in v.lower() for v in row_values):
            for idx, v in enumerate(row_values):
                if "shipping information" in v.lower():
                    col = idx + 1
                    lines = []
                    for nr in range(r + 1, min(ws.max_row, r + 12) + 1):
                        cv = _cell_to_text(ws.cell(nr, col).value)
                        if not cv:
                            break
                        if any(m in cv.lower() for m in ("quote summary", "payment details", "product details")):
                            break
                        lines.append(cv.strip())
                    if lines:
                        out["end user"] = "\n".join(lines)
                    break
            break
    return out


def _extract_grouped_metadata(ws) -> Tuple[str, str]:
    refs = _scan_all_quote_refs(ws, max_rows=200)
    ref = ", ".join(refs)
    date = ""
    for r in range(1, min(ws.max_row, 200) + 1):
        first = _cell_to_text(ws.cell(r, 1).value).strip().lower()
        if first.startswith("date"):
            v = ws.cell(r, 2).value
            date = v.strftime("%d/%m/%Y") if isinstance(v, datetime) else _cell_to_text(v)
    return ref, date


# ==================== ITEMS EXTRACTION ====================

def _extract_items_compact(ws) -> Tuple[List, List]:
    header_info = _find_compact_header(ws)
    if not header_info:
        return [], []
    header_row, cols = header_info
    items: List[Tuple] = []
    config_rows: List[Tuple] = []
    current_item: Optional[str] = None
    blank_streak = 0
    for r in range(header_row + 1, ws.max_row + 1):
        row_text = _row_text(ws, r, 1, ws.max_column)
        if not row_text:
            blank_streak += 1
            if blank_streak >= 2:
                break
            continue
        blank_streak = 0
        first_cell = _cell_to_text(ws.cell(r, cols.get("item", 1)).value).strip()
        sku = _cell_to_text(ws.cell(r, cols["sku"]).value) if "sku" in cols else ""
        desc = _cell_to_text(ws.cell(r, cols["description"]).value)
        qty_raw = _cell_to_text(ws.cell(r, cols["qty"]).value)
        unit_col = cols.get("unit")
        unit_price = (_parse_money(ws.cell(r, unit_col).value) or 0.0) if unit_col else None
        total_price = _parse_money(ws.cell(r, cols["total"]).value)
        if not any([first_cell, sku, desc, qty_raw, unit_price, total_price]):
            continue
        if total_price is None and any(t in row_text.lower() for t in ("total", "subtotal", "quote number", "solution id")):
            continue
        try:
            qty_val = int(qty_raw) if qty_raw not in (None, "") else 0
        except Exception:
            qty_val = int(_parse_money(qty_raw) or 0)
        if unit_col is None and total_price is not None:
            unit_price = total_price
        elif unit_price == 0.0 and qty_val > 0 and total_price is not None:
            unit_price = total_price / qty_val
        if first_cell:
            if desc and qty_val > 0:
                items.append((desc, qty_val, unit_price, total_price))
                current_item = str(len(items))
            continue
        if current_item and (sku or desc):
            config_rows.append((current_item, "", "", desc, "", qty_raw))
    return items, config_rows


def _is_grouped_summary_row(ws, r: int, cols: Dict) -> bool:
    row_text = _row_text(ws, r, 1, ws.max_column).lower()
    if not row_text:
        return False
    first = _cell_to_text(ws.cell(r, 1).value).strip().lower()
    if first.startswith(("quote", "name")):
        return True
    if "consolidation fee" in row_text:
        return not _cell_to_text(ws.cell(r, cols.get("sku", 0)).value)
    return "total" in row_text and "total selling price" not in row_text


def _extract_items_grouped(ws) -> Tuple[List, List]:
    header_info = _find_grouped_header(ws)
    if not header_info:
        return [], []
    header_row, cols = header_info
    items: List[Tuple] = []
    config_rows: List[Tuple] = []
    current_item: Optional[str] = None
    blank_streak = 0
    for r in range(header_row + 1, ws.max_row + 1):
        row_text = _row_text(ws, r, 1, ws.max_column)
        if not row_text:
            blank_streak += 1
            if blank_streak >= 4:
                break
            continue
        blank_streak = 0
        if _is_grouped_summary_row(ws, r, cols):
            continue
        first_cell = _cell_to_text(ws.cell(r, 1).value).strip()
        desc = _cell_to_text(ws.cell(r, cols["description"]).value)
        sku = _cell_to_text(ws.cell(r, cols["sku"]).value)
        qty_raw = _cell_to_text(ws.cell(r, cols.get("qty", 0)).value)
        try:
            qty_val = int(qty_raw) if qty_raw else 0
        except Exception:
            qty_val = int(_parse_money(qty_raw) or 0)
        unit_price = _parse_money(ws.cell(r, cols.get("unit", 0)).value) or 0.0
        total_price = _parse_money(ws.cell(r, cols.get("total", 0)).value) or 0.0
        if first_cell.lower().startswith("quote"):
            continue
        if first_cell:
            if desc:
                items.append((desc, qty_val, unit_price, total_price))
                current_item = str(len(items))
            continue
        if current_item and desc:
            config_rows.append((current_item, "", "", desc, sku, qty_raw))
    return items, config_rows


def _locate_pricing_summary(ws) -> Optional[Tuple[int, int]]:
    B = colidx("B")
    for r in range(30, min(ws.max_row, 120) + 1):
        v = ws.cell(r, B).value
        if v and "pricing" in str(v).lower() and "summary" in str(v).lower():
            return r + 1, r + 3
    return None


def _extract_items_pricing_summary(ws) -> Optional[List[Tuple]]:
    located = _locate_pricing_summary(ws)
    if not located:
        return None
    _, start_row = located
    A, B, K, L, N = colidx("A"), colidx("B"), colidx("K"), colidx("L"), colidx("N")
    items = []
    r = start_row
    while r <= ws.max_row:
        sr = ws.cell(r, A).value
        if sr is None or not re.match(r"^\d+", str(sr).strip()):
            break
        desc = _cell_to_text(ws.cell(r, B).value)
        if not desc:
            break
        qty_val = int(_parse_money(ws.cell(r, K).value) or 0)
        unit_val = _parse_money(ws.cell(r, L).value) or 0.0
        sub_val = _parse_money(ws.cell(r, N).value)
        if sub_val is None:
            sub_val = qty_val * unit_val
        if qty_val <= 0 and unit_val == 0.0 and (sub_val is None or sub_val == 0.0):
            break
        items.append((desc, qty_val, unit_val, sub_val))
        r += 1
    return items if items else None


def _extract_pdf_metadata_by_position(pdf_bytes: bytes) -> Dict[str, str]:
    """
    Extract customer metadata from page 1 of a Dell portal PDF using word X-positions.
    The PDF uses a 2-column layout; the right column (x >= ~200) holds the actual values.
    Returns keys: quote_creator, end_user (shipping address), quote_name.
    """
    out = {"quote_creator": "", "end_user": "", "quote_name": ""}
    try:
        import pdfplumber
        with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
            page = pdf.pages[0]
            words = page.extract_words(use_text_flow=True)
    except Exception:
        return out

    # Group words by y position
    rows: Dict[int, List] = {}
    for w in words:
        y = round(w.get("top", 0))
        rows.setdefault(y, []).append(w)

    # Detect the right-column x boundary from the "Quote Creator:" label
    col2_x = 200.0
    for y in sorted(rows):
        row_words = sorted(rows[y], key=lambda w: w.get("x0", 0))
        line = " ".join(w["text"] for w in row_words).lower()
        if "quote creator" in line and "quote name" in line:
            for w in row_words:
                if "quote" in w["text"].lower() and w.get("x0", 0) > 100:
                    col2_x = w.get("x0", 200.0)
                    break
            break

    # State machine over sorted rows
    next_row_is_quote_name_creator = False
    next_row_is_reseller = False
    in_shipping = False

    for y in sorted(rows):
        row_words = sorted(rows[y], key=lambda w: w.get("x0", 0))
        line = " ".join(w["text"] for w in row_words).strip()
        low = line.lower()

        # Stop at "Quote Summary" or "Custom Fields"
        if any(stop in low for stop in ("quote summary", "custom fields")):
            break

        left_words = [w["text"] for w in row_words if w.get("x0", 0) < col2_x]
        right_words = [w["text"] for w in row_words if w.get("x0", 0) >= col2_x]
        left_text = " ".join(left_words).strip().lower().rstrip(":")
        right_text = " ".join(right_words).strip()

        # "Quote Name:" (left) / "Quote Creator:" (right) — label row
        if "quote name" in left_text and "quote creator" in right_text.lower():
            next_row_is_quote_name_creator = True
            in_shipping = False
            continue

        if next_row_is_quote_name_creator:
            next_row_is_quote_name_creator = False
            out["quote_name"] = " ".join(left_words).strip()
            out["quote_creator"] = right_text
            continue

        if next_row_is_reseller:
            next_row_is_reseller = False
            reseller_raw = right_text.rstrip("+").strip()
            reseller_raw = re.sub(r"\s*[-–]\s*Authorized Partner\+?$", "", reseller_raw, flags=re.I).strip()
            out["reseller"] = reseller_raw
            continue

        # "Page Name:" (left) / "Authorized Partner:" (right) — label row → next row has reseller
        if "authorized partner" in right_text.lower() and not out.get("reseller"):
            next_row_is_reseller = True
            in_shipping = False
            continue

        # "Billing Information:" (left) / "Shipping Information:" (right) — label row
        if "billing information" in left_text and "shipping information" in right_text.lower():
            in_shipping = True
            continue

        if in_shipping:
            # Right column = shipping address; left column is billing (usually "-", skip)
            if right_text and right_text != "-":
                if out["end_user"]:
                    out["end_user"] += "\n" + right_text
                else:
                    out["end_user"] = right_text

    return out


def _extract_items_generic(ws) -> List[Tuple]:
    first_data_row, desc_col, qty_col, unit_col = _find_generic_header(ws)
    items = []
    r = first_data_row
    while r <= ws.max_row:
        desc = _cell_to_text(ws.cell(r, desc_col).value)
        qty = ws.cell(r, qty_col).value
        unit = ws.cell(r, unit_col).value
        if not desc or desc.lower().startswith("total"):
            break
        try:
            qty_val = int(qty) if qty not in (None, "") else 0
        except Exception:
            qty_val = int(_parse_money(qty) or 0)
        unit_val = _parse_money(unit) or 0.0
        if qty_val > 0:
            items.append((desc, qty_val, unit_val, None))
        r += 1
    return items


def _extract_items_pdf(pdf_bytes: bytes) -> Tuple[List, Dict, List, str, str, str, float]:
    """Extract items, metadata, config_rows, quote_ref, date, expiry, consolidation_fee from PDF."""
    lines = _extract_pdf_lines(pdf_bytes)
    metadata = {"company name": "", "customer name": "", "customer number": "",
                 "end user": "", "reseller": "", "quote creator": "", "shipping info": ""}
    quote_ref = date_text = expiry_text = ""
    consolidation_fee = 0.0

    pending_keys: List[str] = []
    prev_label: Optional[str] = None
    in_items = False
    items: List[Tuple] = []
    config_rows: List[Tuple] = []
    # Flag: next non-empty line holds the values for quote number / date / expiry
    _next_line_is_quote_values = False

    for line in lines:
        low = line.lower().strip()
        if not low:
            continue

        # ---- Parse the value line that follows "Quote number: Quote date: Quote expiration:" ----
        if _next_line_is_quote_values:
            _next_line_is_quote_values = False
            # Pattern: <ref_number> <Month DD, YYYY> <Month DD, YYYY>
            # Find all date tokens in "Month DD, YYYY" form
            all_dates = _MONTH_DATE_PAT.findall(line)
            all_nums = re.findall(r"\b(\d{6,})\b", line)
            if not quote_ref and all_nums:
                quote_ref = all_nums[0]
            dates_fmt = []
            for month_s, day_s, year_s in all_dates:
                mn = _MONTH_NAMES.get(month_s.lower())
                if mn:
                    dates_fmt.append(f"{int(day_s):02d}/{mn:02d}/{year_s}")
            if not date_text and len(dates_fmt) >= 1:
                date_text = dates_fmt[0]
            if not expiry_text and len(dates_fmt) >= 2:
                expiry_text = dates_fmt[1]
            continue

        # Detect the combined label line (Dell PDF portal format)
        if "quote number" in low and "quote date" in low:
            _next_line_is_quote_values = True
            continue

        # Metadata extraction
        if not in_items:
            for key in ("company name", "customer name", "customer number", "end user", "reseller", "quote creator", "shipping info"):
                if low.rstrip(":") == key or low.startswith(key + ":"):
                    pending_keys = [key]
                    prev_label = key
                    rest = line[len(key):].lstrip(":").strip()
                    if rest:
                        metadata[key] = rest
                    break
            else:
                if pending_keys and prev_label:
                    if not any(low.rstrip(":") == k or low.startswith(k + ":") for k in metadata):
                        metadata[prev_label] = (metadata[prev_label] + " " + line.strip()).strip()
                    else:
                        pending_keys = []
                        prev_label = None

            # Fallback: inline date search (handles "dd/mm/yyyy" format)
            if not quote_ref:
                m = re.search(r"\b\d{6,}(?:\.[A-Za-z0-9]+)?[A-Za-z0-9\-]*\b", line)
                if m:
                    quote_ref = m.group(0)
            if not date_text:
                m = re.search(r"\d{2}/\d{2}/\d{4}", line)
                if m and any(t in low for t in ("quote date", "quoted on", "date")):
                    date_text = m.group(0)
            if not expiry_text:
                if any(t in low for t in ("quote expiration", "expiry", "expires", "expiration date")):
                    m = re.search(r"\d{2}/\d{2}/\d{4}", line)
                    if m:
                        expiry_text = m.group(0)
                    else:
                        d = _parse_month_date(line)
                        if d:
                            expiry_text = d

        if "quote summary" in low:
            in_items = True
            continue

        if in_items:
            # Stop extracting when we hit another major section
            if any(stop in low for stop in (
                "payment details", "product details", "ship to:", "subtotal:",
            )):
                in_items = False
                continue

            # Skip page footer lines (e.g. "Page 1", "Page 2")
            if re.match(r"^page\s+\d+$", low.strip()):
                continue

            # Match: description $unit qty $total  OR  description unit qty total
            m = re.search(
                r"^(.+?)\s+[$]?([\d,]+[.]?\d*)\s+(\d+)\s+[$]?([\d,]+[.]?\d*)\s*$",
                line,
            )
            if m:
                desc_s, unit_s, qty_s, total_s = m.groups()
                qty_val = int(qty_s)
                unit_val = _parse_money(unit_s) or 0.0
                total_val = _parse_money(total_s) or (qty_val * unit_val)
                items.append((desc_s.strip(), qty_val, unit_val, total_val))
            elif items and not _is_price_or_qty_line(line) and not re.match(r"item\s+unit", low):
                # Continuation line — append to last item description
                old_desc, qty, unit, total = items[-1]
                if old_desc.endswith(","):
                    joined = old_desc.rstrip(",").strip() + ", " + line.strip()
                else:
                    joined = old_desc + " " + line.strip()
                items[-1] = (joined, qty, unit, total)
            elif re.search(r"consolidation fee", low, re.I):
                m2 = re.search(r"[\d,]+\.?\d*", line)
                if m2:
                    consolidation_fee = _parse_money(m2.group(0)) or 0.0

    return items, metadata, config_rows, quote_ref, date_text, expiry_text, consolidation_fee


def _extract_config_from_pdf(pdf_bytes: bytes) -> List[Tuple]:
    """
    Parse the 'Product Details' section of a Dell portal PDF and return config rows.
    Each tuple: (item_number_str, "", category, description, "", "")
    Uses the x-position split (Category col < desc_x, Description col >= desc_x).
    """
    config_rows: List[Tuple] = []
    in_product_details = False
    in_config_section = False
    item_number = 0
    desc_x = 134.0  # will be detected from "Category Description" header

    try:
        import pdfplumber
        with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
            for page in pdf.pages:
                words = page.extract_words(use_text_flow=True)
                if not words:
                    continue
                rows: Dict[int, List] = {}
                for w in words:
                    y = round(w.get("top", 0))
                    rows.setdefault(y, []).append(w)

                for y in sorted(rows):
                    row_words = sorted(rows[y], key=lambda w: w.get("x0", 0))
                    line = " ".join(w["text"] for w in row_words).strip()
                    low = line.lower().strip()

                    if not in_product_details:
                        if "product details" in low:
                            in_product_details = True
                        continue

                    # Item block header — each product's detail section starts with this
                    if "unit price" in low and "qty" in low and "item total" in low:
                        item_number += 1
                        in_config_section = False
                        continue

                    # Stop entirely at end-of-document sections
                    if any(stop in low for stop in (
                        "ship to:", "important notes", "governing terms",
                        "sincerely,", "thanks for shopping", "all orders are subject",
                    )):
                        in_config_section = False
                        in_product_details = False
                        break

                    # Skip page footers, catalog numbers, standalone "Description" header
                    if re.match(r"^page\s+\d+$", low):
                        continue
                    if low.startswith("catalog number"):
                        in_config_section = False
                        continue
                    if low == "description":
                        continue

                    # Skip item price lines inside product details
                    if re.search(r"[$][\d,]+[.]\d+\s+\d+\s+[$][\d,]+", line):
                        continue

                    # "Category Description" header — captures x boundary for this page
                    if "category" in low and "description" in low and len(line.split()) <= 4:
                        for w in row_words:
                            if "description" in w["text"].lower():
                                desc_x = w.get("x0", 134.0)
                                break
                        in_config_section = True
                        continue

                    if not in_config_section:
                        continue

                    # Split by x boundary into category vs description
                    cat_words = [w["text"] for w in row_words if w.get("x0", 0) < desc_x]
                    dsc_words = [w["text"] for w in row_words if w.get("x0", 0) >= desc_x]
                    cat_part = " ".join(cat_words).strip()
                    dsc_part = " ".join(dsc_words).strip()

                    if cat_part and dsc_part:
                        config_rows.append((str(item_number), "", cat_part, dsc_part, "", ""))
                    elif cat_part and not dsc_part and config_rows and config_rows[-1][0] == str(item_number):
                        # Category name wraps to next line — append to previous row's category
                        last = config_rows[-1]
                        config_rows[-1] = (last[0], last[1], last[2] + " " + cat_part, last[3], last[4], last[5])
    except Exception:
        pass

    return config_rows


def _extract_pdf_lines(pdf_bytes: bytes) -> List[str]:
    try:
        import pdfplumber
        with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
            lines = []
            for page in pdf.pages:
                words = page.extract_words(use_text_flow=True)
                if not words:
                    continue
                rows: Dict[int, List] = {}
                for w in words:
                    y = round(w.get("top", 0))
                    rows.setdefault(y, []).append(w)
                for y in sorted(rows):
                    row_words = sorted(rows[y], key=lambda w: w.get("x0", 0))
                    lines.append(" ".join(w.get("text", "") for w in row_words).strip())
        if lines:
            return lines
    except Exception:
        pass
    try:
        from pypdf import PdfReader
    except ImportError:
        raise RuntimeError("pypdf is required to parse PDF quotes")
    reader = PdfReader(BytesIO(pdf_bytes))
    text = "\n".join(page.extract_text() or "" for page in reader.pages)
    return [l.strip() for l in text.splitlines()]


# ==================== CONFIGURATION SHEET ====================

def _find_config_sheet(wb) -> Optional[object]:
    normalized_names = {
        "configuration", "config", "configsheet", "configurationsheet",
        "configurationdetails", "configdetails", "productdetails",
    }
    for name in wb.sheetnames:
        if re.sub(r"[^a-z0-9]", "", name.lower().strip()) in normalized_names:
            return wb[name]
    return None


def _find_config_table_header(ws, start_row: int = 1) -> Optional[Tuple[int, Dict[str, int]]]:
    for r in range(start_row, min(ws.max_row, start_row + 50) + 1):
        labels: Dict[str, int] = {}
        for c in range(1, ws.max_column + 1):
            name = _cell_to_text(ws.cell(r, c).value).lower()
            if not name:
                continue
            name_n = re.sub(r"\s+", " ", name.strip())
            if "module" in name and "module" not in labels:
                labels["module"] = c
            if "description" in name and "description" not in labels:
                labels["description"] = c
            if ("sku" in name_n or "part" in name_n and "number" in name_n) and "sku" not in labels:
                labels["sku"] = c
            if name_n in ("qty", "quantity") and "qty" not in labels:
                labels["qty"] = c
        if "description" in labels and "sku" in labels:
            labels.setdefault("module", labels["description"])
            return r, labels
    return None


def _extract_config_rows(ws) -> List[Tuple]:
    header_info = _find_config_table_header(ws)
    if not header_info:
        return []
    header_row, colmap = header_info
    has_module = bool(colmap.get("module")) and colmap.get("module") != colmap.get("description")
    rows = []
    current_item = "1"
    item_col: Optional[int] = None
    for c in range(1, ws.max_column + 1):
        if _cell_to_text(ws.cell(header_row, c).value).lower() in ("item", "item#", "sr no", "sr"):
            item_col = c
            break
    for r in range(header_row + 1, ws.max_row + 1):
        row_text = _row_text(ws, r, 1, ws.max_column)
        if not row_text:
            continue
        if item_col:
            iv = _cell_to_text(ws.cell(r, item_col).value).strip()
            if iv:
                current_item = iv.rstrip(".")
        module = _cell_to_text(ws.cell(r, colmap.get("module", 0)).value) if has_module else ""
        description = _cell_to_text(ws.cell(r, colmap.get("description", 0)).value)
        sku = _cell_to_text(ws.cell(r, colmap.get("sku", 0)).value)
        qty = _cell_to_text(ws.cell(r, colmap.get("qty", 0)).value)
        if not has_module and description and not any([sku, qty]):
            module, description = description, ""
        if not any([module, description, sku, qty]):
            continue
        rows.append((current_item, "", module, description, sku, qty))
    return rows


def _extract_all_config_rows(ws) -> List[Tuple]:
    """Extract all config rows from the main sheet (no separate config sheet)."""
    header_info = _find_config_table_header(ws)
    if header_info:
        return _extract_config_rows(ws)
    # Collect from product details block
    rows = []
    anchor = None
    for r in range(1, ws.max_row + 1):
        for c in range(1, min(ws.max_column, 40) + 1):
            v = ws.cell(r, c).value
            if v and "product details" in str(v).lower():
                anchor = r
                break
        if anchor:
            break
    if not anchor:
        return []
    header_info2 = _find_config_table_header(ws, anchor)
    if not header_info2:
        return []
    return _extract_config_rows(ws)


def _extract_consolidation_fee(ws) -> float:
    for row in ws.iter_rows():
        for cell in row:
            if not isinstance(cell.value, str):
                continue
            if not re.fullmatch(r"consolidation fees?\s*:?", cell.value.strip().lower()):
                continue
            for next_col in range(ws.max_column, cell.column, -1):
                nv = ws.cell(cell.row, next_col).value
                if nv in (None, ""):
                    continue
                parsed = _parse_money(nv)
                if parsed is not None:
                    return 0.0 if abs(parsed) < 1e-9 else parsed
    return 0.0


def _extract_shipping_fee(ws) -> float:
    for row in ws.iter_rows():
        for cell in row:
            if not isinstance(cell.value, str):
                continue
            if not re.fullmatch(r"shipping(?:\s+(?:charge|charges|cost))?\s*:?", cell.value.strip().lower()):
                continue
            for next_col in range(cell.column + 1, ws.max_column + 1):
                nv = ws.cell(cell.row, next_col).value
                if nv in (None, ""):
                    continue
                fee = _parse_money(nv) or 0.0
                return 0.0 if abs(fee) < 1e-9 else fee
    return 0.0


def _extract_part_numbers(ws) -> Dict[str, str]:
    """Return {item_no: part_number}."""
    out: Dict[str, str] = {}
    sku_pat = re.compile(r"\b[A-Z]{2,4}[0-9]{2,}[A-Z0-9\-]*\b|"
                         r"\b[0-9]{3,}[A-Z]{2,}[0-9A-Z\-]*\b", re.IGNORECASE)
    current_item = "1"
    for r in range(1, ws.max_row + 1):
        for c in range(1, min(ws.max_column, 5) + 1):
            v = _cell_to_text(ws.cell(r, c).value).strip()
            if re.match(r"^\d+\.$", v):
                current_item = v.rstrip(".")
                # Look for a SKU in the rest of the row
                for rc in range(c + 1, min(ws.max_column, 20) + 1):
                    rv = _cell_to_text(ws.cell(r, rc).value).strip()
                    m = sku_pat.match(rv)
                    if m:
                        out.setdefault(current_item, m.group(0))
                        break
    return out


# ==================== QUOTE GENERATION ====================

def _build_quote_workbook(
    items: List[Tuple],
    config_rows: List[Tuple],
    quote_ref: str,
    date_text: str,
    expiry_text: str,
    quote_meta: Dict[str, str],
    currency_code: str,
    exchange_rate: float,
    consolidation_fee: float,
    margin_percent: float,
    is_pdf: bool,
    part_numbers: Optional[Dict[str, str]] = None,
) -> bytes:
    """Build the EUR-style 2-sheet workbook (Quote + Configuration)."""
    currency_code = currency_code.upper()
    conversion_rate = exchange_rate if currency_code == "EUR" else 1.0
    margin_decimal = margin_percent / 100.0

    # Convert items to selected currency; keep originals for the USD helper columns
    original_usd_items = list(items)
    if conversion_rate != 1.0:
        items = [
            (desc, qty, (unit or 0.0) * conversion_rate,
             ((sub or (qty * (unit or 0.0))) * conversion_rate))
            for desc, qty, unit, sub in items
        ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Quote"
    ws.sheet_view.showGridLines = False

    include_part_number = bool(part_numbers)

    # Column layout (EUR-style: always has helper + USD columns)
    if include_part_number:
        desc_col, qty_col, unit_col, total_col = "C", "D", "E", "F"
        helper_unit_col, helper_fee_col = "G", "H"
        usd_unit_col, usd_total_col = "I", "J"
        helper_margin_col = "K"
    else:
        desc_col, qty_col, unit_col, total_col = "B", "C", "D", "E"
        helper_unit_col, helper_fee_col = "F", "G"
        usd_unit_col, usd_total_col = "H", "I"
        helper_margin_col = "J"

    # Styling constants
    section_fill = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")
    helper_header_fill = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
    helper_body_fill = PatternFill(start_color="FCE5E5", end_color="FCE5E5", fill_type="solid")
    header_fill = PatternFill(start_color="9BC2E6", end_color="9BC2E6", fill_type="solid")
    helper_font = Font(bold=True, color="9C0006")
    header_font = Font(bold=True, color="000000")
    border_thin = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    def _style_section_title(addr: str) -> None:
        ws[addr].font = Font(bold=True, color="1F497D")
        ws[addr].alignment = Alignment(horizontal="left", vertical="center")
        ws[addr].fill = section_fill
        ws[addr].border = Border(
            left=Side(style="thin", color="9FBAD0"),
            right=Side(style="thin", color="9FBAD0"),
            top=Side(style="thin", color="9FBAD0"),
            bottom=Side(style="thin", color="9FBAD0"),
        )

    # --- Column widths ---
    description_width = min(max(44, int(max((len(_cell_to_text(it[0])) for it in items), default=0) * 0.55)), 68)
    widths: Dict[str, float] = {"A": 11}
    if include_part_number:
        widths.update({"B": 16, "C": min(max(42, description_width), 56), "D": 8, "E": 15, "F": 17, "G": 17, "H": 12})
    else:
        widths.update({"B": min(max(42, description_width), 56), "C": 8, "D": 15, "E": 18, "F": 17, "G": 12})
    widths[helper_fee_col] = 12
    widths[usd_unit_col] = 18
    widths[usd_total_col] = 18
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Row heights
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 26
    for rr in range(3, 11):
        ws.row_dimensions[rr].height = 20

    # --- Logo ---
    ws.merge_cells("A1:H2")
    _add_logo(ws, anchor="A1", width=780, height=52)

    # --- Address block (French / Southcomp) ---
    def _write_address(start_row: int, end_row: int, lines: List[str], merge: bool = True) -> None:
        if merge:
            rng = f"A{start_row}:D{end_row}"
            ws.merge_cells(rng)
            ws.unmerge_cells(rng)
        for offset, text in enumerate(lines):
            cell = ws.cell(row=start_row + offset, column=1, value=text)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        if merge:
            ws.merge_cells(f"A{start_row}:D{end_row}")

    _write_address(5, 8, [
        "14, rue du Bas Marin",
        "94537 Orly cedex - France",
        "DL:     +33 1 49 79 42 24",
        "Fax:   +33 1 49 79 45 33",
    ], merge=False)
    address_end_row = 8
    for addr in ("A5", "A6", "A7", "A8"):
        ws[addr].font = Font(bold=True, size=11, color="1F497D")
        ws[addr].alignment = Alignment(horizontal="left", vertical="center")

    # --- Quote Summary ---
    has_expiry = bool(expiry_text)
    summary_title_row = 9
    ws.merge_cells(f"A{summary_title_row}:D{summary_title_row}")
    ws[f"A{summary_title_row}"] = "Quote Summary"
    _style_section_title(f"A{summary_title_row}")

    summary_rows = [(summary_title_row + 1, "Quote Ref", quote_ref),
                    (summary_title_row + 2, "Date", date_text)]
    if has_expiry:
        summary_rows.append((summary_title_row + 3, "Expires By", expiry_text))

    for row_idx, label, value in summary_rows:
        ws[f"A{row_idx}"] = label
        ws[f"A{row_idx}"].font = Font(bold=True, color="1F497D")
        ws[f"A{row_idx}"].alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=4)
        ws[f"B{row_idx}"] = value
        ws[f"B{row_idx}"].alignment = Alignment(horizontal="left", vertical="center")
        if label == "Expires By":
            ws[f"B{row_idx}"].font = Font(bold=True)

    customer_title_row = (summary_title_row + 4) if has_expiry else (summary_title_row + 3)

    # --- Quote metadata ---
    if is_pdf:
        meta_rows = [
            ("End Customer:", quote_meta.get("end user", "")),
            ("Reseller:", quote_meta.get("reseller", "")),
            ("Quote Creator:", quote_meta.get("quote creator", "")),
        ]
        if quote_meta.get("shipping info"):
            meta_rows.append(("Shipping Information:", quote_meta.get("shipping info", "")))
    else:
        meta_rows = [
            ("Company Name:", quote_meta.get("company name", "")),
            ("Customer Name:", quote_meta.get("customer name", "")),
            ("End User:", quote_meta.get("end user", "")),
            ("Reseller:", quote_meta.get("reseller", "")),
        ]

    ws.merge_cells(f"A{customer_title_row}:D{customer_title_row}")
    ws[f"A{customer_title_row}"] = "Customer Information"
    _style_section_title(f"A{customer_title_row}")

    for i, (label, value) in enumerate(meta_rows, start=1):
        row_idx = customer_title_row + i
        ws[f"A{row_idx}"] = label
        ws[f"A{row_idx}"].font = Font(bold=True, color="1F497D")
        ws[f"A{row_idx}"].alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=4)
        ws[f"B{row_idx}"] = value
        ws[f"B{row_idx}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        explicit_newlines = value.count("\n")
        text_len = len(str(value))
        estimated_lines = max(1, explicit_newlines + 1 + max(0, text_len // 32))
        ws.row_dimensions[row_idx].height = max(ws.row_dimensions[row_idx].height or 20, min(estimated_lines, 12) * 18)

    # --- Recalculate helper row positions ---
    last_meta_row = customer_title_row + len(meta_rows)
    helper_value_row = last_meta_row + 1
    helper_aux_row = helper_value_row + 1

    # F/G col of helper row: consolidation fee total (editable by user in Excel)
    ws[f"{helper_unit_col}{helper_value_row}"] = consolidation_fee
    ws[f"{helper_unit_col}{helper_value_row}"].font = helper_font
    ws[f"{helper_unit_col}{helper_value_row}"].alignment = Alignment(horizontal="center", vertical="center")
    ws[f"{helper_unit_col}{helper_value_row}"].fill = helper_body_fill
    ws[f"{helper_unit_col}{helper_value_row}"].border = border_thin

    # Marge col of helper row: margin % as a static decimal — NO circular formula
    ws[f"{helper_margin_col}{helper_value_row}"] = margin_decimal
    ws[f"{helper_margin_col}{helper_value_row}"].number_format = "0.00%"
    ws[f"{helper_margin_col}{helper_value_row}"].font = helper_font
    ws[f"{helper_margin_col}{helper_value_row}"].alignment = Alignment(horizontal="center", vertical="center")
    ws[f"{helper_margin_col}{helper_value_row}"].fill = helper_body_fill
    ws[f"{helper_margin_col}{helper_value_row}"].border = border_thin

    # --- Table header ---
    header_row = helper_aux_row + 1
    ws[f"A{header_row}"] = "N°"
    if include_part_number:
        ws[f"B{header_row}"] = "N° de pièce"
    ws[f"{desc_col}{header_row}"] = "Description"
    ws[f"{qty_col}{header_row}"] = "Qté"
    ws[f"{unit_col}{header_row}"] = "Prix unitaire"
    ws[f"{total_col}{header_row}"] = "Prix total"
    ws[f"{helper_unit_col}{header_row}"] = "Prix unitaire d'origine"
    ws[f"{helper_fee_col}{header_row}"] = "Fees"
    ws[f"{usd_unit_col}{header_row}"] = "Unit Price USD original"
    ws[f"{usd_total_col}{header_row}"] = "Total Price USD original"
    ws[f"{helper_margin_col}{header_row}"] = "Marge"

    header_cells = [f"A{header_row}", f"{desc_col}{header_row}", f"{qty_col}{header_row}",
                    f"{unit_col}{header_row}", f"{total_col}{header_row}",
                    f"{helper_unit_col}{header_row}", f"{helper_fee_col}{header_row}",
                    f"{usd_unit_col}{header_row}", f"{usd_total_col}{header_row}",
                    f"{helper_margin_col}{header_row}"]
    if include_part_number:
        header_cells.insert(1, f"B{header_row}")
    helper_header_cells = (f"{helper_unit_col}{header_row}", f"{helper_fee_col}{header_row}",
                           f"{usd_unit_col}{header_row}", f"{usd_total_col}{header_row}",
                           f"{helper_margin_col}{header_row}")
    for addr in header_cells:
        ws[addr].fill = helper_header_fill if addr in helper_header_cells else header_fill
        ws[addr].font = header_font
        ws[addr].alignment = Alignment(horizontal="center", vertical="center")
        ws[addr].border = border_thin
    ws.row_dimensions[header_row].height = 20

    # --- Data rows ---
    row_ptr = header_row + 1
    sr_no = 1
    currency_fmt = CURRENCY_FORMATS.get(currency_code, f'"{currency_code}" #,##0.00')
    usd_fmt = CURRENCY_FORMATS["USD"]
    yellow = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")
    total_cells = []

    for idx, (desc_text, qty_val, unit_val, subtotal_val) in enumerate(items):
        original_usd_unit = original_usd_items[idx][2] if idx < len(original_usd_items) else None

        ws[f"A{row_ptr}"] = sr_no
        if include_part_number and part_numbers:
            ws[f"B{row_ptr}"] = _sanitize_excel_text(part_numbers.get(str(sr_no), ""))
        ws[f"{desc_col}{row_ptr}"] = _sanitize_excel_text(desc_text)
        ws[f"{qty_col}{row_ptr}"] = qty_val or 0
        unit_val = unit_val or 0.0

        # "Prix unitaire d'origine" — static original cost price (in output currency, pre-margin)
        orig_helper = f"{helper_unit_col}{row_ptr}"
        ws[orig_helper] = unit_val
        ws[orig_helper].font = helper_font
        ws[orig_helper].fill = helper_body_fill
        ws[orig_helper].number_format = currency_fmt
        ws[orig_helper].border = border_thin
        ws[orig_helper].alignment = Alignment(horizontal="center", vertical="center")

        # "Fees" per unit — static 0, user can edit in Excel
        fee_helper = f"{helper_fee_col}{row_ptr}"
        ws[fee_helper] = 0
        ws[fee_helper].font = helper_font
        ws[fee_helper].fill = helper_body_fill
        ws[fee_helper].number_format = currency_fmt
        ws[fee_helper].border = border_thin
        ws[fee_helper].alignment = Alignment(horizontal="center", vertical="center")

        # Per-item margin cell: defaults to J17 but user can override individually
        ws[f"{helper_margin_col}{row_ptr}"] = f"={helper_margin_col}${helper_value_row}"
        ws[f"{helper_margin_col}{row_ptr}"].number_format = "0.00%"
        ws[f"{helper_margin_col}{row_ptr}"].font = helper_font
        ws[f"{helper_margin_col}{row_ptr}"].fill = helper_body_fill
        ws[f"{helper_margin_col}{row_ptr}"].border = border_thin
        ws[f"{helper_margin_col}{row_ptr}"].alignment = Alignment(horizontal="center", vertical="center")

        # "Prix unitaire" — selling price = (original + fees) * (1 + this row's margin)
        ws[f"{unit_col}{row_ptr}"] = f"=({helper_unit_col}{row_ptr}+{helper_fee_col}{row_ptr})*(1+{helper_margin_col}{row_ptr})"
        ws[f"{unit_col}{row_ptr}"].number_format = currency_fmt
        ws[f"{unit_col}{row_ptr}"].border = border_thin
        ws[f"{unit_col}{row_ptr}"].alignment = Alignment(horizontal="center", vertical="center")

        # "Prix total"
        total_addr = f"{total_col}{row_ptr}"
        ws[total_addr] = f"={unit_col}{row_ptr}*{qty_col}{row_ptr}"
        ws[total_addr].number_format = currency_fmt
        ws[total_addr].border = border_thin
        ws[total_addr].alignment = Alignment(horizontal="center", vertical="center")
        total_cells.append(total_addr)

        # USD original columns (always USD values from original BOQ)
        usd_unit = original_usd_unit or (unit_val / conversion_rate if conversion_rate and conversion_rate != 1.0 else unit_val)
        ws[f"{usd_unit_col}{row_ptr}"] = usd_unit
        ws[f"{usd_unit_col}{row_ptr}"].number_format = usd_fmt
        ws[f"{usd_unit_col}{row_ptr}"].fill = helper_body_fill
        ws[f"{usd_unit_col}{row_ptr}"].border = border_thin
        ws[f"{usd_unit_col}{row_ptr}"].alignment = Alignment(horizontal="center", vertical="center")

        ws[f"{usd_total_col}{row_ptr}"] = usd_unit * (qty_val or 0)
        ws[f"{usd_total_col}{row_ptr}"].number_format = usd_fmt
        ws[f"{usd_total_col}{row_ptr}"].fill = helper_body_fill
        ws[f"{usd_total_col}{row_ptr}"].border = border_thin
        ws[f"{usd_total_col}{row_ptr}"].alignment = Alignment(horizontal="center", vertical="center")

        for addr in [f"A{row_ptr}", f"{desc_col}{row_ptr}", f"{qty_col}{row_ptr}"]:
            ws[addr].border = border_thin
            ws[addr].alignment = Alignment(horizontal="center" if addr == f"A{row_ptr}" else "left", vertical="center", wrap_text=True)
        if include_part_number:
            ws[f"B{row_ptr}"].border = border_thin
            ws[f"B{row_ptr}"].alignment = Alignment(horizontal="center", vertical="center")

        ws[f"A{row_ptr}"].fill = yellow
        if include_part_number:
            ws[f"B{row_ptr}"].fill = yellow
        ws[f"{desc_col}{row_ptr}"].fill = yellow
        ws[f"{qty_col}{row_ptr}"].fill = yellow
        ws[f"{qty_col}{row_ptr}"].alignment = Alignment(horizontal="center", vertical="center")

        row_ptr += 1
        sr_no += 1

    # --- Total row ---
    total_label_col = "C" if include_part_number else "B"
    if include_part_number:
        ws.merge_cells(start_row=row_ptr, start_column=3, end_row=row_ptr, end_column=5)
    else:
        ws.merge_cells(start_row=row_ptr, start_column=2, end_row=row_ptr, end_column=4)
    ws[f"{total_label_col}{row_ptr}"] = "Prix total"
    ws[f"{total_label_col}{row_ptr}"].alignment = Alignment(horizontal="right", vertical="center")
    ws[f"{total_label_col}{row_ptr}"].font = Font(bold=True, color="1F497D")
    ws[f"{total_col}{row_ptr}"] = f"=SUM({','.join(total_cells)})" if total_cells else 0
    ws[f"{total_col}{row_ptr}"].number_format = currency_fmt
    ws[f"{total_col}{row_ptr}"].font = Font(bold=True, color="1F497D")
    ws[f"{total_col}{row_ptr}"].alignment = Alignment(horizontal="center", vertical="center")
    ws[f"{total_col}{row_ptr}"].border = border_thin
    ws[f"{helper_unit_col}{row_ptr}"].fill = helper_body_fill
    ws[f"{helper_margin_col}{row_ptr}"].fill = helper_body_fill
    ws[f"{helper_unit_col}{row_ptr}"].border = border_thin
    ws[f"{helper_margin_col}{row_ptr}"].border = border_thin

    # --- Configuration sheet (same layout as dell.py AED output) ---
    ws2 = wb.create_sheet("Configuration")
    ws2.sheet_view.showGridLines = False

    ws2.column_dimensions["A"].width = 22   # Item #
    ws2.column_dimensions["B"].width = 70   # Module
    ws2.column_dimensions["C"].width = 100  # Description
    ws2.column_dimensions["D"].width = 10   # Qty

    title_fill2 = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    thin_gray = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )

    # Table header
    r2 = 1
    for col, label in (("A", "Item #"), ("B", "Module"), ("C", "Description"), ("D", "Qty")):
        ws2[f"{col}{r2}"] = label
        ws2[f"{col}{r2}"].font = Font(bold=True)
        ws2[f"{col}{r2}"].fill = title_fill2
        ws2[f"{col}{r2}"].alignment = Alignment(horizontal="center", vertical="center")
        ws2[f"{col}{r2}"].border = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000"),
        )
    r2 += 1

    # Group config rows by item number
    config_by_item: Dict[str, List] = {}
    for row in config_rows:
        config_by_item.setdefault(row[0], []).append(row)

    # Item descriptions for headings (from the items list)
    original_descs = [_sanitize_excel_text(it[0]) for it in original_usd_items]

    total_items = max(len(original_descs), len(config_by_item)) if (original_descs or config_by_item) else 0
    for idx in range(1, total_items + 1):
        item_key = str(idx)
        rows_for_item = config_by_item.get(item_key, [])
        heading = original_descs[idx - 1] if idx - 1 < len(original_descs) else f"Item {idx}"

        # "Item N" row
        ws2.merge_cells(start_row=r2, start_column=1, end_row=r2, end_column=4)
        ws2[f"A{r2}"] = f"Item {idx}"
        ws2[f"A{r2}"].font = Font(bold=True, color="1F497D")
        ws2[f"A{r2}"].alignment = Alignment(horizontal="left", vertical="center")
        r2 += 1

        # Item description heading
        ws2.merge_cells(start_row=r2, start_column=1, end_row=r2, end_column=4)
        ws2[f"A{r2}"] = heading
        ws2[f"A{r2}"].font = Font(italic=True, color="1F497D")
        ws2[f"A{r2}"].alignment = Alignment(horizontal="left", vertical="center")
        r2 += 1

        if not rows_for_item:
            ws2.merge_cells(start_row=r2, start_column=2, end_row=r2, end_column=4)
            ws2[f"B{r2}"] = "(No configuration details found for this item)"
            ws2[f"B{r2}"].font = Font(italic=True, color="7F7F7F")
            ws2[f"B{r2}"].alignment = Alignment(horizontal="left", vertical="center")
            for col in ("A", "B", "C", "D"):
                ws2[f"{col}{r2}"].border = thin_gray
            r2 += 1
        else:
            for row_data in rows_for_item:
                _, _, module, dsc, sku, qty = (row_data + ("", "", "", ""))[:6]
                ws2[f"A{r2}"] = ""
                ws2[f"B{r2}"] = _sanitize_excel_text(module)
                ws2[f"C{r2}"] = _sanitize_excel_text(dsc)
                ws2[f"D{r2}"] = _sanitize_excel_text(qty)
                for col in ("A", "B", "C", "D"):
                    ws2[f"{col}{r2}"].alignment = Alignment(vertical="top", wrap_text=True)
                    ws2[f"{col}{r2}"].border = thin_gray
                r2 += 1

        r2 += 1  # blank gap between items

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


# ==================== PUBLIC API ====================

def generate_southcomp_quote(
    input_bytes: bytes,
    margin_percent: float,
    currency_code: str,
    exchange_rate: float,
) -> bytes:
    """
    Generate a Southcomp Polaris EUR-style quote workbook.
    currency_code: 'EUR' or 'USD'
    exchange_rate: EUR/USD rate (used when currency_code='EUR')
    Returns raw xlsx bytes.
    """
    currency_code = (currency_code or "EUR").upper()
    effective_rate = exchange_rate if currency_code == "EUR" else 1.0

    is_pdf = input_bytes.lstrip().startswith(b"%PDF")
    items: List[Tuple] = []
    config_rows: List[Tuple] = []
    quote_ref = date_text = expiry_text = ""
    quote_meta: Dict[str, str] = {}
    consolidation_fee = 0.0
    part_numbers: Dict[str, str] = {}

    if is_pdf:
        items, raw_meta, config_rows, quote_ref, date_text, expiry_text, consolidation_fee = _extract_items_pdf(input_bytes)
        config_rows = _extract_config_from_pdf(input_bytes)
        pos_meta = _extract_pdf_metadata_by_position(input_bytes)
        quote_meta = raw_meta
        if pos_meta.get("quote_creator"):
            quote_meta["quote creator"] = pos_meta["quote_creator"]
        if pos_meta.get("end_user"):
            quote_meta["end user"] = pos_meta["end_user"]
        if pos_meta.get("reseller"):
            quote_meta["reseller"] = pos_meta["reseller"]
        if pos_meta.get("quote_name") and not quote_meta.get("company name"):
            quote_meta["company name"] = pos_meta["quote_name"]
    else:
        src_wb = openpyxl.load_workbook(BytesIO(input_bytes), data_only=True)
        src_ws = src_wb.active
        config_ws = _find_config_sheet(src_wb)

        # Try grouped template
        if _find_grouped_header(src_ws) is not None:
            quote_ref, date_text = _extract_grouped_metadata(src_ws)
            items, grp_config_rows = _extract_items_grouped(src_ws)
            config_rows = _extract_config_rows(config_ws) if config_ws else grp_config_rows
        else:
            # Metadata
            quote_ref, date_text = _extract_metadata(src_ws)
            expiry_text = _extract_expiry(src_ws)
            quote_meta = _extract_quote_metadata(src_ws)

            # Items: pricing summary → compact → generic
            items_ps = _extract_items_pricing_summary(src_ws)
            if items_ps:
                items = items_ps
                config_rows = _extract_config_rows(config_ws) if config_ws else _extract_all_config_rows(src_ws)
            else:
                compact_items, compact_config = _extract_items_compact(src_ws)
                if compact_items:
                    items = compact_items
                    config_rows = _extract_config_rows(config_ws) if config_ws else compact_config
                else:
                    items = _extract_items_generic(src_ws)
                    config_rows = _extract_config_rows(config_ws) if config_ws else _extract_all_config_rows(src_ws)

            consolidation_fee = _extract_consolidation_fee(src_ws) + _extract_shipping_fee(src_ws)
            part_numbers = _extract_part_numbers(config_ws or src_ws)

        if not quote_meta:
            quote_meta = _extract_quote_metadata(src_ws)

    quote_meta = {k: _strip_trailing_asterisk(v) for k, v in quote_meta.items()}

    # Apply margin factor to consolidation fee
    margin_factor = margin_percent / 100.0
    adjusted_consolidation_fee = consolidation_fee * (1 + margin_factor)

    return _build_quote_workbook(
        items=items,
        config_rows=config_rows,
        quote_ref=quote_ref,
        date_text=date_text,
        expiry_text=expiry_text,
        quote_meta=quote_meta,
        currency_code=currency_code,
        exchange_rate=effective_rate,
        consolidation_fee=adjusted_consolidation_fee,
        margin_percent=margin_percent,
        is_pdf=is_pdf,
        part_numbers=part_numbers or None,
    )


def build_output_filename(currency_code: str = "EUR") -> str:
    return f"Southcomp_Polaris_{currency_code.upper()}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
