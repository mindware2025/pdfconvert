from __future__ import annotations

import re
from datetime import datetime, timedelta
from io import BytesIO

import pandas as pd
import pdfplumber
from openpyxl import Workbook

OUTPUT_HEADERS = [
    "INV NUMBER",
    "INV DATE",
    "Job No",
    "Base Value for Vat",
    "W/o Vat",
    "NARRATION",
    "Due date",
    "",
    "Vat",
    "Inv ",
    "",
    "Text",
    "Text",
    "Text",
    "Description of Charges",
    "Entity",
]

CURRENCY_FORMAT = '_(* #,##0.00_);_(* \\(#,##0.00\\);_(* "-"??_);_(@_)'
DATE_FORMAT = "mm-dd-yy"
AMOUNT_FORMAT = "#,##0.00"

INV_LABEL = "JOINTLINE INV #"
JOB_LABEL = "JOB# "


def _find_cell_containing(table: list[list], needle: str) -> str:
    for row in table:
        for cell in row:
            if cell and needle in cell:
                return cell
    return ""


def _search_line(pattern: str, text: str) -> str:
    match = re.search(pattern, text, re.MULTILINE)
    return match.group(1).strip() if match else ""


def _find_charges_table(file_bytes: bytes) -> list[list]:
    with pdfplumber.open(BytesIO(file_bytes)) as pdf:
        page = pdf.pages[0]
        tables = page.extract_tables()
    for table in tables:
        for row in table:
            if row and (row[0] or "").strip() == "Sr" and (row[1] or "").strip() == "Description of Charges":
                return table
    return tables[0] if tables else []


def extract_invoice_data(file_name: str, file_bytes: bytes) -> tuple[dict | None, list[str]]:
    errors: list[str] = []
    table = _find_charges_table(file_bytes)
    if not table:
        return None, [f"{file_name}: could not read a table from the PDF."]

    info_left = _find_cell_containing(table, "Customer")
    info_right = _find_cell_containing(table, "Job No.")

    entity = _search_line(r"^Customer\s+(.+?)\s*$", info_left)
    invoice_no_raw = _search_line(r"^I\s*n\s*v\s*o\s*i\s*c\s*e\s*No\.\s*(\d+)", info_right)
    invoice_date_raw = _search_line(r"^Date\s+(\d{2}/\d{2}/\d{4})", info_right)
    job_no_raw = _search_line(r"^Job No\.\s*(\d+)", info_right)

    items: list[tuple[str, float, float]] = []
    in_items = False
    for row in table:
        if not row:
            continue
        cell0 = (row[0] or "").strip()
        if cell0 == "Sr":
            in_items = True
            continue
        if not in_items:
            continue
        if not cell0.isdigit():
            break
        description = (row[1] or "").strip()
        tax_pct_raw = row[8] if len(row) > 8 else ""
        pre_tax_raw = row[10] if len(row) > 10 else ""
        try:
            tax_pct = float((tax_pct_raw or "0").replace(",", ""))
            pre_tax_amt = float((pre_tax_raw or "0").replace(",", ""))
        except ValueError:
            errors.append(f"{file_name}: could not parse amount for line '{description}'.")
            continue
        items.append((description, tax_pct, pre_tax_amt))

    if not entity:
        errors.append(f"{file_name}: could not find customer/entity name.")
    if not invoice_no_raw:
        errors.append(f"{file_name}: could not find invoice number.")
    if not invoice_date_raw:
        errors.append(f"{file_name}: could not find invoice date.")
    if not job_no_raw:
        errors.append(f"{file_name}: could not find job number.")
    if not items:
        errors.append(f"{file_name}: no line items found in the charges table.")

    if errors:
        return None, errors

    invoice_date = datetime.strptime(invoice_date_raw, "%d/%m/%Y")
    base_value_vat = round(sum(amt for _, tax, amt in items if tax > 0), 2)
    wo_vat = round(sum(amt for _, tax, amt in items if tax == 0), 2)
    narration = "; ".join(description for description, _, _ in items)

    return {
        "file_name": file_name,
        "invoice_no": int(invoice_no_raw),
        "invoice_date": invoice_date,
        "job_no": int(job_no_raw),
        "base_value_vat": base_value_vat,
        "wo_vat": wo_vat,
        "narration": narration,
        "entity": entity,
    }, []


def process_joint_line_pdfs(uploaded_files: list) -> tuple[list[dict], list[str]]:
    invoices: list[dict] = []
    errors: list[str] = []

    for uploaded_file in uploaded_files:
        file_bytes = uploaded_file.getvalue()
        invoice_data, parse_errors = extract_invoice_data(uploaded_file.name, file_bytes)
        if parse_errors:
            errors.extend(parse_errors)
            continue
        invoices.append(invoice_data)

    invoices.sort(key=lambda inv: inv["invoice_no"])
    return invoices, errors


def build_preview_dataframe(invoices: list[dict]) -> pd.DataFrame:
    rows = []
    for inv in invoices:
        due_date = inv["invoice_date"] + timedelta(days=30)
        vat = round(inv["base_value_vat"] * 0.05, 2)
        total = round(vat + inv["wo_vat"] + inv["base_value_vat"], 2)
        description = (
            f'{INV_LABEL}{inv["invoice_no"]} {inv["narration"]}  {JOB_LABEL}{inv["job_no"]}'
        )
        rows.append(
            {
                "INV NUMBER": inv["invoice_no"],
                "INV DATE": inv["invoice_date"].date(),
                "Job No": inv["job_no"],
                "Base Value for Vat": inv["base_value_vat"],
                "W/o Vat": inv["wo_vat"],
                "NARRATION": inv["narration"],
                "Due date": due_date.date(),
                "Vat": vat,
                "Inv ": total,
                "Description of Charges": description,
                "Entity": inv["entity"],
            }
        )
    return pd.DataFrame(rows)


def create_joint_line_excel_file(invoices: list[dict]) -> BytesIO:
    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(OUTPUT_HEADERS)

    for idx, inv in enumerate(invoices, start=2):
        ws.cell(row=idx, column=1, value=inv["invoice_no"])
        date_cell = ws.cell(row=idx, column=2, value=inv["invoice_date"])
        date_cell.number_format = DATE_FORMAT
        ws.cell(row=idx, column=3, value=inv["job_no"])
        ws.cell(row=idx, column=4, value=inv["base_value_vat"])
        wo_vat_cell = ws.cell(row=idx, column=5, value=inv["wo_vat"])
        wo_vat_cell.number_format = CURRENCY_FORMAT
        ws.cell(row=idx, column=6, value=inv["narration"])
        due_date_cell = ws.cell(row=idx, column=7, value=f"=B{idx}+30")
        due_date_cell.number_format = DATE_FORMAT
        vat_cell = ws.cell(row=idx, column=9, value=f"=D{idx}*5%")
        vat_cell.number_format = AMOUNT_FORMAT
        inv_cell = ws.cell(row=idx, column=10, value=f"=I{idx}+E{idx}+D{idx}")
        inv_cell.number_format = AMOUNT_FORMAT
        ws.cell(row=idx, column=12, value=INV_LABEL)
        ws.cell(row=idx, column=13, value=JOB_LABEL)
        ws.cell(
            row=idx,
            column=15,
            value=f'=L{idx}&A{idx}&" "&F{idx}&" "&N{idx}&" "&M{idx}&C{idx}',
        )
        ws.cell(row=idx, column=16, value=inv["entity"])

    for column_cells in ws.columns:
        max_length = max((len(str(cell.value or "")) for cell in column_cells), default=0)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max_length + 2, 60)

    wb.save(output)
    output.seek(0)
    return output
