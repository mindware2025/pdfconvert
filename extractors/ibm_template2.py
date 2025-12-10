# extractors/ibm_template2.py
import fitz
import re
from datetime import datetime
import logging
from pathlib import Path
from io import BytesIO
from terms_template import get_terms_section

# Configure detailed logging for template 2
log_file_path = 'template2_extraction_debug.log'

# Clear existing handlers to avoid duplicates
for handler in logging.root.handlers[:]:
    logging.root.removeHandler(handler)

# Create file handler with UTF-8 encoding
file_handler = logging.FileHandler(log_file_path, mode='w', encoding='utf-8')
file_handler.setLevel(logging.DEBUG)

# Create console handler
console_handler = logging.StreamHandler()
console_handler.setLevel(logging.DEBUG)

# Create formatter
formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
file_handler.setFormatter(formatter)
console_handler.setFormatter(formatter)

# Configure logger
logger = logging.getLogger('ibm_template2')
logger.setLevel(logging.DEBUG)
logger.addHandler(file_handler)
logger.addHandler(console_handler)

# Prevent propagation to avoid duplicate messages
logger.propagate = False

# Use the same debug system as ibm.py
debug_info = []

# Constants
USD_TO_AED = 3.6725  # Fixed conversion rate

def add_debug(message):
    """Add debug info to both in-memory list and log file"""
    debug_info.append(message)
    logger.debug(message)
    
    # Also write directly to file to ensure it's saved
    try:
        with open('template2_extraction_debug.log', 'a', encoding='utf-8') as f:
            f.write(f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')} - DEBUG - {message}\n")
            f.flush()
    except Exception as e:
        print(f"Warning: Could not write to log file: {e}")
    
    if len(debug_info) > 500:
        debug_info.pop(0)

def get_extraction_debug():
    """Get collected debug info"""
    return debug_info.copy()

def clear_debug():
    """Clear debug info and prepare log file"""
    debug_info.clear()
    # Clear the log file for fresh start
    try:
        with open('template2_extraction_debug.log', 'w', encoding='utf-8') as f:
            f.write(f"=== TEMPLATE 2 EXTRACTION LOG - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} ===\n\n")
    except Exception as e:
        print(f"Warning: Could not clear log file: {e}")

def save_debug_to_file():
    """Save all debug messages to file"""
    try:
        with open('template2_extraction_debug.log', 'w', encoding='utf-8') as f:
            f.write(f"=== TEMPLATE 2 EXTRACTION LOG - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} ===\n\n")
            for message in debug_info:
                f.write(f"{message}\n")
            f.flush()
        logger.info(f"Debug log saved to template2_extraction_debug.log with {len(debug_info)} messages")
    except Exception as e:
        logger.error(f"Failed to save debug log: {e}")

# Constants
USD_TO_AED = 3.6725

def parse_number(value: str):
    """Parse numbers with various formats (including European: 107.856,00)"""
    try:
        if value is None:
            return None
        s = str(value).strip().replace(" ", "")
        
        # Handle European format: 107.856,00 or 1.000,50
        # Period is thousands separator, comma is decimal
        if ',' in s and '.' in s:
            # Has both - period is thousands separator, comma is decimal
            s = s.replace(".", "").replace(",", ".")
        elif ',' in s:
            # Only comma - it's the decimal separator
            s = s.replace(",", ".")
        # else: only period or neither - keep as is
        
        result = float(s)
        return result
    except Exception:
        return None

def extract_ibm_template2_from_pdf(file_like) -> tuple[list, dict]:
    """
    Extract data from IBM Template 2 (Software as a Service / Subscription format)
    Returns: (extracted_data, header_info)
    """
    clear_debug()
    
    try:
        add_debug("="*80)
        add_debug("STARTING TEMPLATE 2 EXTRACTION")
        add_debug("="*80)
        logger.info("Template 2 extraction started")
    except Exception as e:
        print(f"Error in initial logging: {e}")
        return [], {}
    
    try:
        doc = fitz.open(stream=file_like.read(), filetype="pdf")
        add_debug(f"PDF opened successfully: {len(doc)} pages")
    except Exception as e:
        add_debug(f"ERROR opening PDF: {e}")
        logger.error(f"PDF opening failed: {e}")
        return [], {}
    
    # Collect all text
    lines = []
    for page_num, page in enumerate(doc):
        page_text = page.get_text("text") or page.get_text()
        page_lines = []
        for line in page_text.splitlines():
            if line and line.strip():
                lines.append(line.strip())
                page_lines.append(line.strip())
        add_debug(f"Page {page_num + 1}: Extracted {len(page_lines)} lines")
    
    add_debug(f"\n[TOTAL LINES] Extracted {len(lines)} non-empty lines from PDF")
    add_debug("\n" + "="*80)
    add_debug("RAW TEXT DUMP (First 100 lines)")
    add_debug("="*80)
    for idx, line in enumerate(lines[:100]):
        add_debug(f"Line {idx:3d}: {line}")
    
    # Header info extraction
    header_info = {
        "Customer Name": "",
        "Bid Number": "",
        "PA Agreement Number": "",
        "PA Site Number": "",
        "Select Territory": "",
        "Government Entity (GOE)": "",
        "IBM Opportunity Number": "",
        "Reseller Name": "",
        "City": "",
        "Country": "",
        "Bid Expiration Date": "",
        "Maximum End User Price (MEP)": ""
    }
    
    # Parse header fields
    for i, line in enumerate(lines):
        if "Customer Name:" in line:
            header_info["Customer Name"] = lines[i + 1] if i + 1 < len(lines) else ""
        elif "City:" in line:
            header_info["City"] = lines[i + 1] if i + 1 < len(lines) else ""
        elif "Country:" in line:
            header_info["Country"] = lines[i + 1] if i + 1 < len(lines) else ""
        elif "Bid Number:" in line or "Quote Number:" in line:
            header_info["Bid Number"] = lines[i + 1] if i + 1 < len(lines) else ""
        elif "Bid Expiration Date:" in line or "Quote Expiration Date:" in line:
            header_info["Bid Expiration Date"] = lines[i + 1] if i + 1 < len(lines) else ""
        elif "IBM Agreement Number:" in line or "PA Agreement Number:" in line:
            header_info["PA Agreement Number"] = lines[i + 1] if i + 1 < len(lines) else ""
        elif "IBM Site Number:" in line or "PA Site Number:" in line:
            header_info["PA Site Number"] = lines[i + 1] if i + 1 < len(lines) else ""
        elif "Select Territory:" in line:
            header_info["Select Territory"] = lines[i + 1] if i + 1 < len(lines) else ""
        elif "Government Entity" in line:
            header_info["Government Entity (GOE)"] = lines[i + 1] if i + 1 < len(lines) else ""
        elif "Reseller Name:" in line:
            header_info["Reseller Name"] = lines[i + 1] if i + 1 < len(lines) else ""
        elif "Maximum End User Price" in line or "MEP" in line:
            # Skip if this is just a yes/no question (like "MEP (Maximum End User Price): Yes")
            if ":" in line:
                after_colon = line.split(":", 1)[1].strip().lower()
                if after_colon in ["yes", "no"]:
                    # This is just a yes/no field, skip it - the actual value will be on another line
                    add_debug(f"[MEP] Skipping MEP question line (yes/no): {line}")
                    continue
            
            # Look for MEP value in same line or next line
            mep_value = None
            if ":" in line:
                mep_part = line.split(":", 1)[1].strip()
                if mep_part and mep_part.lower() not in ["yes", "no"]:
                    mep_clean = re.sub(r'\s*(USD|AED|EUR).*$', '', mep_part).strip()
                    mep_value = parse_number(mep_clean)
                    if mep_value:
                        add_debug(f"[MEP] Found MEP value on same line: {mep_value}")
            
            # If not found in same line, check next line
            if not mep_value and i + 1 < len(lines):
                next_line = lines[i + 1].strip()
                if "USD" in next_line or "," in next_line:
                    mep_clean = re.sub(r'\s*(USD|AED|EUR).*$', '', next_line).strip()
                    mep_value = parse_number(mep_clean)
                    if mep_value:
                        add_debug(f"[MEP] Found MEP value on next line: {mep_value}")
            
            if mep_value:
                header_info["Maximum End User Price (MEP)"] = f"{mep_value:,.2f}"
                add_debug(f"[MEP] MEP set to: {mep_value:,.2f}")
        elif "IBM Opportunity Number:" in line:
            # Extract the opportunity number from the same or next line
            opp_match = re.search(r'[A-Z0-9]{10,}', line)
            if opp_match:
                header_info["IBM Opportunity Number"] = opp_match.group()
            elif i + 1 < len(lines):
                opp_match = re.search(r'[A-Z0-9]{10,}', lines[i + 1])
                if opp_match:
                    header_info["IBM Opportunity Number"] = opp_match.group()
    
    add_debug("\n" + "="*80)
    add_debug("HEADER INFORMATION EXTRACTED")
    add_debug("="*80)
    for key, value in header_info.items():
        add_debug(f"{key:30s}: {value}")
    add_debug("="*80 + "\n")
    
    # Fallback: If MEP still not found, search for "Maximum End User Price" pattern with value on next line
    if not header_info.get("Maximum End User Price (MEP)"):
        add_debug("[MEP FALLBACK] MEP not found in main loop, searching entire header...")
        for i, line in enumerate(lines):
            if ("Maximum End User Price" in line or "MEP" in line) and i + 1 < len(lines):
                next_line = lines[i + 1].strip()
                add_debug(f"[MEP FALLBACK] Found MEP-like line at index {i}: '{line}'")
                add_debug(f"[MEP FALLBACK] Next line {i+1}: '{next_line}'")
                # Check if next line contains USD amount or is a number
                if "USD" in next_line or "," in next_line or any(c.isdigit() for c in next_line):
                    mep_clean = re.sub(r'\s*(USD|AED|EUR).*$', '', next_line).strip()
                    add_debug(f"[MEP FALLBACK] Cleaned value: '{mep_clean}'")
                    mep_value = parse_number(mep_clean)
                    add_debug(f"[MEP FALLBACK] Parsed number: {mep_value}")
                    if mep_value:
                        header_info["Maximum End User Price (MEP)"] = f"{mep_value:,.2f}"
                        add_debug(f"[MEP FALLBACK SUCCESS] MEP set to: {mep_value:,.2f}")
                        break
                else:
                    add_debug(f"[MEP FALLBACK] Next line doesn't look like a price, skipping")
    
    # Extract line items (Subscription Parts)
    extracted_data = []
    global_channel_discount = 0.08  # Track the channel discount globally
    
    add_debug("\n" + "="*80)
    add_debug("STARTING LINE ITEM EXTRACTION")
    add_debug("="*80)
    
    # Pattern for subscription part numbers (like D1009ZX, D100AZX, D28B4LL)
    subscription_part_re = re.compile(r'\b[A-Z][A-Z0-9]{4,8}\b')  # More flexible: 1 letter + 4-8 alphanumeric chars
    date_pattern = re.compile(r'\b\d{2}-[A-Za-z]{3}-\d{4}\b')
    
    add_debug(f"Subscription Part Pattern: {subscription_part_re.pattern}")
    add_debug(f"Date Pattern: {date_pattern.pattern}")
    
    # PRE-SCAN: Detect if this is a multi-row case (e.g., rows 001-006)
    add_debug("\n[PRE-SCAN] Checking for multi-row table case...")
    table_row_pattern = re.compile(r'^00[1-9]$|^0[1-9][0-9]$')
    table_row_count = 0
    for line in lines:
        if table_row_pattern.match(line.strip()):
            table_row_count += 1
    
    is_multi_row_case = table_row_count >= 2
    add_debug(f"[PRE-SCAN] Table row markers found: {table_row_count}")
    add_debug(f"[PRE-SCAN] Multi-row case detected: {is_multi_row_case}")
    
    # Look for "Software as a Service" sections
    i = 0
    line_item_count = 0
    while i < len(lines):
        line = lines[i]
        
        # Check if we're in a service section
        # Include: "Subscription Part#:" and "Overage Part#:"
        # Exclude: "Corresponding Subscription Part#:"
        is_subscription_part = 'Subscription Part#:' in line and 'Corresponding' not in line
        is_overage_part = 'Overage Part#:' in line
        
        # STRATEGY 1: Skip if multi-row case detected
        if (is_subscription_part or is_overage_part) and not is_multi_row_case:
            try:
                line_item_count += 1
                add_debug(f"\n{'='*60}")
                add_debug(f"LINE ITEM #{line_item_count} - Found at line index {i}")
                add_debug(f"{'='*60}")
                add_debug(f"Trigger line: {line}")
                
                # Determine part type
                part_type = "Overage" if is_overage_part else "Subscription"
                add_debug(f"Part Type: {part_type}")
                
                # Extract SKU from current or next line
                sku = None
                sku_match = subscription_part_re.search(line)
                if sku_match:
                    sku = sku_match.group()
                    add_debug(f"✓ SKU found in trigger line: {sku}")
                elif i + 1 < len(lines):
                    sku_match = subscription_part_re.search(lines[i + 1])
                    if sku_match:
                        sku = sku_match.group()
                        add_debug(f"✓ SKU found in next line: {sku} (from: {lines[i+1]})")
                
                if not sku:
                    add_debug("✗ No SKU found - skipping this entry")
                    i += 1
                    continue
                
                # Extract service description (capture full block of information)
                desc_lines = []
                add_debug(f"\n[DESCRIPTION] Searching lines {max(0, i-15)} to {min(i+10, len(lines))} for full service block:")
                
                # First, find the main IBM service line (going backwards) - look for product/service description
                service_line_idx = None
                for j in range(i - 1, max(0, i - 15), -1):  # Search BACKWARDS from trigger line
                    line_text = lines[j].strip()
                    # Look for IBM service lines that describe the actual product (not distributor/company info)
                    # Should contain keywords like "Maximo", "Application", "Suite", "License", etc.
                    if line_text.startswith('IBM') and len(line_text) > 20:
                        # Exclude generic company/distributor lines
                        if not any(x in line_text for x in ['Building', 'Industrial Park', 'Campus', 'Dublin', 'Ireland']):
                            service_line_idx = j
                            add_debug(f"  Found service line at line {j}: {line_text}")
                            break
                
                if service_line_idx is not None:
                    # Collect the full service block from service line through additional billing details
                    # Extend search to include lines after the part line for billing info
                    end_range = min(i + 15, len(lines))  # Look 15 lines after part line
                    add_debug(f"  Collecting service block from line {service_line_idx} to {end_range}:")
                    
                    for j in range(service_line_idx, end_range):
                        line_text = lines[j].strip()
                        if line_text:  # Non-empty lines
                            # Include relevant lines that describe the service
                            if any(keyword in line_text for keyword in [
                                'IBM', 'Projected Service Start Date', 'Service Level Agreement',
                                'Current Transaction', 'Billing:', 'Subscription Length:',
                                'Renewal Type:', 'Renewal:', 'Resource Unit Overage',
                                'Corresponding Subscription Part#', 'Subscription Part#:', 'Overage Part#:'
                            ]):
                                desc_lines.append(line_text)
                                add_debug(f"    Line {j}: {line_text}")
                            # Stop if we hit another service or a new major section
                            elif line_text.startswith('IBM') and j > i + 5:
                                add_debug(f"    Stopping at line {j}: Next service section detected")
                                break
                
                # Join all lines with newlines to form complete description
                desc = '\n'.join(desc_lines) if desc_lines else ""
                
                if desc:
                    add_debug(f"✓ Full service block extracted ({len(desc_lines)} lines)")
                    add_debug(f"Complete description:")
                    add_debug(f"    {desc}")
                else:
                    # Fallback: Just get the IBM service name
                    for j in range(max(0, i - 10), i):
                        line_text = lines[j].strip()
                        if line_text.startswith('IBM') and len(line_text) > 15:
                            desc = line_text
                            add_debug(f"✓ Fallback single-line description: {desc}")
                            break
                    
                    if not desc:
                        desc = f"IBM Service - {part_type} Part"
                        add_debug(f"✗ No description found, using default: {desc}")
                
                # Extract start date (Projected Service Start Date)
                start_date = ""
                add_debug(f"\n[START DATE] Searching lines {max(0, i-5)} to {min(i+5, len(lines))}:")
                for j in range(max(0, i - 5), min(i + 5, len(lines))):
                    if 'Start Date:' in lines[j] or 'start date' in lines[j].lower():
                        add_debug(f"  Line {j} (Match!): {lines[j]}")
                        date_match = date_pattern.search(lines[j])
                        if date_match:
                            start_date = date_match.group()
                            add_debug(f"✓ Start date found in same line: {start_date}")
                        elif j + 1 < len(lines):
                            date_match = date_pattern.search(lines[j + 1])
                            if date_match:
                                start_date = date_match.group()
                                add_debug(f"✓ Start date found in next line: {start_date}")
                        break
                if not start_date:
                    add_debug("✗ No start date found")
                
                # Extract subscription length to calculate end date
                end_date = ""
                subscription_length = 12  # Default
                add_debug(f"\n[SUBSCRIPTION LENGTH] Searching lines {i} to {min(i+20, len(lines))}:")
                for j in range(i, min(i + 20, len(lines))):
                    if 'Subscription Length:' in lines[j] or 'subscription length' in lines[j].lower():
                        add_debug(f"  Line {j} (Match!): {lines[j]}")
                        length_match = re.search(r'(\d+)\s*Months?', lines[j], re.I)
                        if length_match:
                            subscription_length = int(length_match.group(1))
                            add_debug(f"✓ Subscription length found: {subscription_length} months")
                        break
                if subscription_length == 12:
                    add_debug("  Using default: 12 months")
                
                # Calculate end date if we have start date
                if start_date:
                    try:
                        from datetime import datetime
                        from dateutil.relativedelta import relativedelta
                        start_dt = datetime.strptime(start_date, '%d-%b-%Y')
                        end_dt = start_dt + relativedelta(months=subscription_length)
                        end_date = end_dt.strftime('%d-%b-%Y')
                        add_debug(f"✓ End date calculated: {start_date} + {subscription_length} months = {end_date}")
                    except Exception as e:
                        end_date = ""
                        add_debug(f"✗ Error calculating end date: {e}")
                else:
                    add_debug("✗ Cannot calculate end date - no start date")
                
                # Extract quantity (look for table data and line item mapping)
                qty = 1  # Fallback default
                add_debug(f"\n[QUANTITY] Searching for quantity data for SKU {sku}:")
                
                # Strategy 1: Look for large quantities first (for D100AZX type SKUs)
                found_qty = False
                if 'D100AZX' in sku:
                    add_debug(f"  Looking for large quantity for D100AZX:")
                    # Search a wider range for large quantities like 672
                    for j in range(max(0, i - 50), min(i + 100, len(lines))):
                        line_text = lines[j].strip()
                        # Look for 3-4 digit numbers that could be quantities
                        if re.match(r'^\d{2,4}$', line_text):
                            potential_qty = int(line_text)
                            if 50 <= potential_qty <= 1000:  # Reasonable range for bulk quantities
                                qty = potential_qty
                                add_debug(f"✓ Large quantity found for D100AZX at line {j}: {qty}")
                                found_qty = True
                                break
                        # Also check comma-formatted numbers
                        elif re.match(r'^\d{1,3}(,\d{3})*$', line_text):
                            potential_qty = int(line_text.replace(',', ''))
                            if 50 <= potential_qty <= 10000:
                                qty = potential_qty
                                add_debug(f"✓ Comma-formatted large quantity found: {qty}")
                                found_qty = True
                                break
                
                # Strategy 2: Find the line item number for this SKU
                if not found_qty:
                    line_item_number = None
                    for j in range(i, min(i + 50, len(lines))):  # Extend search range to i+50
                        line_text = lines[j].strip()
                        # Look for "Line Item" or standalone numbers near our SKU
                        if re.match(r'^00[1-9]$', line_text):  # 001, 002, 003, etc.
                            # Check if this line item is close to our current SKU
                            distance_to_sku = abs(j - i)
                            if distance_to_sku < 50:  # Within reasonable distance
                                line_item_number = line_text
                                add_debug(f"  Found line item number {line_item_number} at line {j} (distance: {distance_to_sku})")
                                break
                    
                    # Strategy 3: Extract quantity from table structure
                    if line_item_number:
                        add_debug(f"  Looking for table data for line item {line_item_number}:")
                        
                        # Search globally for our line item number and its table data
                        for j in range(len(lines)):
                            line_text = lines[j].strip()
                            
                            # If we find our line item number, the next few lines should contain table data
                            if line_text == line_item_number:
                                add_debug(f"    Found line item {line_item_number} at line {j}")
                                # Look at the next several lines for quantity
                                for k in range(j + 1, min(j + 15, len(lines))):
                                    qty_text = lines[k].strip()
                                    add_debug(f"    Line {k}: '{qty_text}'")
                                    
                                    # Look for quantity (numeric value, not decimal prices)
                                    if re.match(r'^\d+$', qty_text):
                                        potential_qty = int(qty_text)
                                        # Skip line item numbers and focus on reasonable quantities
                                        if potential_qty >= 1 and potential_qty <= 10000:
                                            # Avoid line item numbers like 001, 002, 003
                                            if not (potential_qty <= 3 and len(qty_text) == 1):
                                                qty = potential_qty
                                                add_debug(f"✓ Quantity found for line item {line_item_number}: {qty}")
                                                found_qty = True
                                                break
                                            elif potential_qty <= 10 and k <= j + 3:
                                                # Small quantities are valid if they appear early
                                                qty = potential_qty
                                                add_debug(f"✓ Small quantity found for line item {line_item_number}: {qty}")
                                                found_qty = True
                                                break
                                    
                                    # Also check for quantities with commas (e.g., "1,000")
                                    elif re.match(r'^\d{1,3}(,\d{3})*$', qty_text):
                                        potential_qty = int(qty_text.replace(',', ''))
                                        if potential_qty >= 1 and potential_qty <= 10000:
                                            qty = potential_qty
                                            add_debug(f"✓ Comma-formatted quantity found: {qty}")
                                            found_qty = True
                                            break
                                if found_qty:
                                    break
                    
                    # Strategy 4: If no table mapping, look for nearby quantities
                    if not found_qty:
                        add_debug(f"  No line item/table data found, searching for nearby quantities:")
                        for j in range(max(0, i - 10), min(i + 30, len(lines))):
                            line_text = lines[j].strip()
                            if re.match(r'^\d+$', line_text):
                                potential_qty = int(line_text)
                                # Avoid obvious line item numbers
                                if 1 <= potential_qty <= 10000 and line_text not in ['001', '002', '003']:
                                    qty = potential_qty
                                    add_debug(f"✓ Nearby quantity found at line {j}: {qty}")
                                    found_qty = True
                                    break
                
                if not found_qty:
                    add_debug(f"  No quantity found, using fallback: {qty}")
                
                add_debug(f"✓ Final quantity for {sku}: {qty}")
                
                # Extract duration (look for patterns like "1-12")
                duration = None
                add_debug(f"\n[DURATION] Searching for duration pattern for SKU {sku}:")
                
                # Look for duration patterns in nearby lines
                for j in range(max(0, i - 20), min(i + 50, len(lines))):
                    line_text = lines[j]
                    # Look for pattern like "1-12", "1 - 12", "1–12", etc.
                    duration_match = re.search(r'(\d+)\s*[\u2010\u2011\u2013-]\s*(\d+)', line_text)
                    if duration_match:
                        start_month = int(duration_match.group(1))
                        end_month = int(duration_match.group(2))
                        # Validate it looks like a duration (reasonable range)
                        if 1 <= start_month <= end_month <= 24:
                            duration = f"{start_month}-{end_month}"
                            add_debug(f"✓ Duration found at line {j}: '{duration}'")
                            add_debug(f"  Source line: {line_text.strip()}")
                            break
                
                if not duration:
                    add_debug("✗ No duration pattern found")
                else:
                    add_debug(f"✓ Final duration for {sku}: {duration}")
                
                # Extract pricing from table rows AND summary sections
                bid_unit_price = None
                bid_total_price = None
                
                add_debug(f"\n[PRICING] Searching for price data for SKU {sku}:")
                
                # Strategy 1: Look for line-item specific prices in table format
                add_debug(f"  Strategy 1: Looking for line-item table prices around lines {i}-{min(i+35, len(lines))}:")
                
                for j in range(max(0, i-10), min(i + 35, len(lines))):
                    line_text = lines[j]
                    
                    # Look for lines with decimal numbers (including 0,00 values)
                    price_matches = re.findall(r'\b\d{1,3}(?:[.,]\d{3})*[.,]\d{2}\b', line_text)
                    
                    # Also look for specific table patterns with line item numbers
                    if re.match(r'^\s*00[1-9]', line_text):  # Line starts with 001, 002, etc.
                        add_debug(f"    Line {j}: Found line item row: {line_text[:100]}")
                        
                        # Collect all price values from the next 10 lines after finding the row number
                        all_prices = []
                        for k in range(j+1, min(j+12, len(lines))):
                            price_line = lines[k]
                            # Look for European formatted numbers (including 0,00 and 215.712,00)
                            found_prices = re.findall(r'\b\d{1,3}(?:[.,]\d{3})*[.,]\d{2}(?:\s*USD)?\b', price_line)
                            if found_prices:
                                # Clean USD suffix and add to collection
                                clean_prices = [p.replace(' USD', '').strip() for p in found_prices]
                                all_prices.extend(clean_prices)
                                add_debug(f"             Line {k}: {price_line.strip()} -> {clean_prices}")
                        
                        if len(all_prices) >= 1:
                            add_debug(f"             All collected prices: {all_prices}")
                            try:
                                # For Template 2, we need to identify the "Bid Total Commit Value" column
                                # Based on your PDF table structure, this is typically the 7th-8th price value
                                # Look for patterns: often appears as the middle-to-later significant price
                                
                                total_price_candidates = [p for p in all_prices if not p.startswith('0,')]
                                
                                if total_price_candidates:
                                    # Use POSITIONAL logic instead of smallest price logic
                                    # Based on PDF table structure, "Bid Total Commit Value" is typically
                                    # around position 6-7 in the price sequence
                                    
                                    add_debug(f"             Price candidates: {total_price_candidates}")
                                    
                                    # Strategy: Look for the 4th position (index 3) for "Bid Total Commit Value"
                                    if len(total_price_candidates) >= 4:
                                        # Use 4th position (index 3) as it's typically "Bid Total Commit Value" 
                                        total_str = total_price_candidates[3]
                                        add_debug(f"             Selected position 4 (index 3): {total_str}")
                                    elif len(total_price_candidates) >= 2:
                                        # Use 2nd position for shorter sequences
                                        total_str = total_price_candidates[1]
                                        add_debug(f"             Selected position 2: {total_str}")
                                    else:
                                        # Only one price available
                                        total_str = total_price_candidates[0]
                                        add_debug(f"             Selected only available: {total_str}")
                                    
                                    def parse_european_price(price_str):
                                        """Convert European format like 107.856,00 to float"""
                                        if ',' in price_str and price_str.count(',') == 1:
                                            # European format: remove dots (thousands sep), replace comma with dot
                                            clean_str = price_str.replace('.', '').replace(',', '.')
                                        else:
                                            # Already standard format
                                            clean_str = price_str.replace(',', '')
                                        return float(clean_str)
                                    
                                    total_val = parse_european_price(total_str)
                                else:
                                    # All prices are 0,00 - this is valid pricing
                                    total_str = all_prices[0].replace(',', '.')
                                    total_val = float(total_str)
                                    add_debug(f"             All prices are zero - using: {total_str}")
                                
                                # Unit price calculation
                                unit_val = total_val / qty if qty > 0 else total_val
                                
                                bid_unit_price = unit_val
                                bid_total_price = total_val
                                add_debug(f"✓ LINE-ITEM PRICES FOUND at line {j}: Unit=${bid_unit_price}, Total=${bid_total_price}")
                                add_debug(f"  Selected from candidates: {total_price_candidates} -> {total_str}")
                                break
                                
                            except Exception as e:
                                add_debug(f"             ✗ Error parsing table row: {e}")
                                continue
                    
                    # Original logic for lines with multiple price matches
                    elif len(price_matches) >= 2:
                        add_debug(f"    Line {j}: Found {len(price_matches)} price values: {price_matches}")
                        add_debug(f"             Full line: {line_text[:100]}")
                        try:
                            # Usually: [...other values...] [unit_price] [total_price] USD
                            unit_str = price_matches[-2].replace(',', '.')
                            total_str = price_matches[-1].replace(',', '.')
                            
                            unit_val = float(unit_str)
                            total_val = float(total_str)
                            
                            add_debug(f"             Parsed: unit={unit_val}, total={total_val}")
                            add_debug(f"             Validation: total ({total_val}) vs unit*qty ({unit_val * qty}) = diff {abs(total_val - (unit_val * qty))}")
                            
                            # Validate that total ≈ unit * qty
                            if abs(total_val - (unit_val * qty)) < 1.0:
                                bid_unit_price = unit_val
                                bid_total_price = total_val
                                add_debug(f"✓ LINE-ITEM PRICES FOUND at line {j}: Unit=${bid_unit_price}, Total=${bid_total_price}")
                                break
                            else:
                                add_debug(f"             ✗ Validation failed - looking for better match...")
                        except Exception as e:
                            add_debug(f"             ✗ Error parsing: {e}")
                            continue
                
                # Strategy 2: Calculate unit price if we have total but not unit
                if bid_total_price and not bid_unit_price and qty > 0:
                    bid_unit_price = bid_total_price / qty
                    add_debug(f"✓ CALCULATED Unit Price: ${bid_unit_price} = ${bid_total_price} / {qty}")
                
                # No fallback - if table prices not found, leave blank
                if not bid_total_price:
                    add_debug("✗ No line-item pricing found - leaving prices blank")
                
                # Convert USD to AED - Total Price first, then Unit Price from Total
                bid_total_aed = round(bid_total_price * USD_TO_AED, 2) if bid_total_price else None
                bid_unit_aed = round(bid_total_aed / qty, 2) if bid_total_aed and qty > 0 else bid_total_aed
                
                # Calculate Partner Price in AED using Channel Discount
                # Formula: ROUNDUP(Unit Price AED * (1 - Channel Discount%), 2) * Qty
                partner_price_aed = None
                channel_discount_pct = 0.08  # Default 8%, will extract from PDF
                
                # Try to extract Channel Discount from nearby lines
                for j in range(max(0, i - 20), min(i + 20, len(lines))):
                    line_text = lines[j]
                    discount_match = re.search(r'Channel Discount:\s*(\d+)%', line_text, re.I)
                    if discount_match:
                        discount_value = int(discount_match.group(1))
                        channel_discount_pct = discount_value / 100.0
                        global_channel_discount = channel_discount_pct  # Update global value
                        add_debug(f"  Found Channel Discount: {discount_value}% = {channel_discount_pct}")
                        break
                
                if bid_unit_aed is not None and qty:
                    # Partner unit price with discount applied and rounded up
                    partner_unit_discounted = bid_unit_aed * (1 - channel_discount_pct)
                    import math
                    partner_unit_rounded = math.ceil(partner_unit_discounted * 100) / 100  # ROUNDUP to 2 decimals
                    partner_price_aed = round(partner_unit_rounded * qty, 2)
                    add_debug(f"  Partner Price: Unit AED {bid_unit_aed} × (1-{channel_discount_pct}) = {partner_unit_discounted}")
                    add_debug(f"  Partner Price: ROUNDUP({partner_unit_discounted}, 2) = {partner_unit_rounded}")
                    add_debug(f"  Partner Price: {partner_unit_rounded} × {qty} = AED {partner_price_aed}")
                
                add_debug(f"\n[CURRENCY CONVERSION] USD to AED (rate: {USD_TO_AED}):")
                add_debug(f"  Total: ${bid_total_price} → AED {bid_total_aed}")
                add_debug(f"  Unit: AED {bid_total_aed} ÷ {qty} → AED {bid_unit_aed}")
                add_debug(f"  Partner: AED {partner_price_aed} (with {channel_discount_pct*100}% discount)")
                
                # Add to extracted data
                # Format: [sku, desc, qty, duration, start_date, end_date, bid_unit_aed, bid_total_aed, partner_price_aed]
                row_data = [
                    sku,
                    desc,
                    qty,
                    duration,
                    start_date,
                    end_date,
                    bid_unit_aed,
                    bid_total_aed,
                    partner_price_aed
                ]
                extracted_data.append(row_data)
                
                add_debug(f"\n{'='*60}")
                add_debug(f"✓ LINE ITEM #{line_item_count} COMPLETE")
                add_debug(f"{'='*60}")
                add_debug(f"  SKU: {sku}")
                add_debug(f"  Description (full):")
                add_debug(f"    {desc}")
                add_debug(f"  Quantity: {qty}")
                add_debug(f"  Start Date: {start_date}")
                add_debug(f"  End Date: {end_date}")
                add_debug(f"  Unit Price (AED): {bid_unit_aed}")
                add_debug(f"  Total Price (AED): {bid_total_aed}")
                add_debug(f"{'='*60}\n")
                
            except Exception as e:
                add_debug(f"\n✗✗✗ ERROR in line item {line_item_count} ✗✗✗")
                add_debug(f"Exception: {str(e)}")
                import traceback
                add_debug(f"Traceback:\n{traceback.format_exc()}")
                add_debug(f"{'='*60}\n")
        
        # STRATEGY 2: Extract from table rows (001, 002, 003, etc.) if multi-row case
        if is_multi_row_case:
            line = lines[i]
            line_stripped = line.strip()
            
            # Check if this line is a table row marker
            if table_row_pattern.match(line_stripped):
                try:
                    add_debug(f"\n[STRATEGY 2] Processing table row: {line_stripped}")
                    
                    # Extract quantity from next line
                    qty = 1
                    if i + 1 < len(lines):
                        qty_line = lines[i + 1].strip()
                        add_debug(f"  Qty line: '{qty_line}'")
                        
                        # Try to parse quantity
                        # Handle European period format: 1.550 = 1550
                        if re.match(r'^\d{1,3}(\.\d{3})*$', qty_line):
                            qty = int(qty_line.replace('.', ''))
                            add_debug(f"  ✓ European period format quantity: {qty}")
                        # Handle comma format: 1,550 = 1550
                        elif re.match(r'^\d{1,3}(,\d{3})*$', qty_line):
                            qty = int(qty_line.replace(',', ''))
                            add_debug(f"  ✓ Comma format quantity: {qty}")
                        # Handle plain integer
                        elif re.match(r'^\d+$', qty_line):
                            qty = int(qty_line)
                            add_debug(f"  ✓ Plain integer quantity: {qty}")
                        # Handle mixed format with decimal
                        elif re.match(r'^\d+[.,]\d+$', qty_line):
                            if '.' in qty_line:
                                parts = qty_line.split('.')
                                # If 3 digits after period, it's thousands separator
                                if len(parts[1]) == 3:
                                    qty = int(qty_line.replace('.', ''))
                                else:
                                    qty = float(qty_line.replace(',', '.'))
                            else:
                                qty = float(qty_line.replace(',', '.'))
                            add_debug(f"  ✓ Mixed format quantity: {qty}")
                    
                    # Extract duration (e.g., "1-12" or "13-24")
                    duration = "1-12"
                    if i + 2 < len(lines):
                        duration_line = lines[i + 2].strip()
                        add_debug(f"  Duration line: '{duration_line}'")
                        duration_match = re.search(r'(\d+)-(\d+)', duration_line)
                        if duration_match:
                            duration = f"{duration_match.group(1)}-{duration_match.group(2)}"
                            add_debug(f"  ✓ Duration extracted: {duration}")
                    
                    # Extract SKU (search entire document backwards from current position)
                    sku_table = None
                    for j in range(i - 1, -1, -1):
                        search_text = lines[j]
                        sku_match = subscription_part_re.search(search_text)
                        if sku_match:
                            potential_sku = sku_match.group()
                            # Validate SKU (should contain both letters and digits)
                            if any(c.isalpha() for c in potential_sku) and any(c.isdigit() for c in potential_sku):
                                if 5 <= len(potential_sku) <= 20:
                                    sku_table = potential_sku
                                    add_debug(f"  ✓ SKU found at line {j}: {sku_table}")
                                    break
                    
                    if not sku_table:
                        add_debug(f"  ✗ No SKU found for table row {line_stripped}")
                        i += 1
                        continue
                    
                    # Extract description using SAME LOGIC as Strategy 1 (but only once per SKU with caching)
                    if not hasattr(extract_ibm_template2_from_pdf, '_desc_cache'):
                        extract_ibm_template2_from_pdf._desc_cache = {}
                    
                    desc_cache = extract_ibm_template2_from_pdf._desc_cache
                    
                    if sku_table not in desc_cache:
                        # Extract description using Strategy 1 logic
                        desc_lines = []
                        add_debug(f"  [DESC] Extracting description for {sku_table}...")
                        
                        # Find where the SKU was mentioned (subscription part line)
                        sku_line_idx = None
                        for j in range(i - 1, max(0, i - 50), -1):
                            if 'Subscription Part#:' in lines[j] or 'Overage Part#:' in lines[j]:
                                sku_line_idx = j
                                add_debug(f"    Found subscription part line at line {j}")
                                break
                        
                        if sku_line_idx is None:
                            sku_line_idx = i  # Fallback to current position
                        
                        # Find the main IBM service line (searching BEFORE the subscription part line)
                        service_line_idx = None
                        for j in range(sku_line_idx - 1, max(0, sku_line_idx - 30), -1):  # Search backwards
                            line_text = lines[j].strip()
                            if line_text.startswith('IBM') and len(line_text) > 20:
                                # Exclude generic company/distributor lines and opportunity numbers
                                if not any(x in line_text for x in ['Building', 'Industrial Park', 'Campus', 'Dublin', 'Ireland', 'Opportunity Number']):
                                    service_line_idx = j
                                    add_debug(f"    Found service line at line {j}: {line_text}")
                                    break
                        
                        if service_line_idx is not None:
                            # Collect the full service block from service line THROUGH billing/renewal details
                            # Go 15-20 lines after the subscription part line to capture all details
                            end_range = min(sku_line_idx + 20, len(lines))
                            add_debug(f"    Collecting service block from line {service_line_idx} to {end_range}:")
                            
                            for j in range(service_line_idx, end_range):
                                line_text = lines[j].strip()
                                if line_text:  # Non-empty lines
                                    # Include relevant lines that describe the service
                                    if any(keyword in line_text for keyword in [
                                        'IBM', 'Projected Service Start Date', 'Service Level Agreement',
                                        'Current Transaction', 'Billing:', 'Subscription Length:',
                                        'Renewal Type:', 'Renewal:', 'Resource Unit Overage',
                                        'Corresponding Subscription Part#', 'Subscription Part#:', 'Overage Part#:',
                                        'Channel Discount:', 'Customer Unit Price:', 'Quote Rate:', 'Committed Term:'
                                    ]):
                                        desc_lines.append(line_text)
                                        add_debug(f"      Line {j}: {line_text}")
                                    # Stop if we hit the table headers (Item, Quantity, etc.)
                                    elif re.match(r'^(Item|Line|Qty|Quantity|SI|Customer|Entitled|Months|Discount|Quote)\s*$', line_text, re.I):
                                        add_debug(f"      Stopping at line {j}: Table header detected")
                                        break
                        
                        # Join all lines with newlines to form complete description
                        full_desc = '\n'.join(desc_lines) if desc_lines else ""
                        
                        if full_desc:
                            add_debug(f"    ✓ Description extracted ({len(desc_lines)} lines)")
                        else:
                            add_debug(f"    ✗ No description found")
                        
                        desc_cache[sku_table] = full_desc
                    
                    desc_table = desc_cache[sku_table]
                    add_debug(f"  ✓ Using cached description for {sku_table}")
                    
                    # Extract pricing using SAME LOGIC as Strategy 1
                    unit_price_aed = 0
                    total_price_aed = 0
                    
                    add_debug(f"  [PRICING] Searching lines {i+1} to {min(i+15, len(lines))} for prices (same as Strategy 1):")
                    
                    # Collect all price values from lines after the row marker
                    all_prices = []
                    for j in range(i + 1, min(i + 15, len(lines))):
                        price_line = lines[j].strip()
                        # Look for European formatted numbers (including 0,00 and 864.960,00)
                        found_prices = re.findall(r'\b\d{1,3}(?:[.,]\d{3})*[.,]\d{2}(?:\s*USD)?\b', price_line)
                        if found_prices:
                            # Clean USD suffix and add to collection
                            clean_prices = [p.replace(' USD', '').strip() for p in found_prices]
                            all_prices.extend(clean_prices)
                            add_debug(f"    Line {j}: {price_line} -> {clean_prices}")
                    
                    if len(all_prices) >= 1:
                        add_debug(f"    All collected prices: {all_prices}")
                        try:
                            # Filter out prices that start with 0 (these are discounts or small values)
                            price_candidates = [p for p in all_prices if not p.startswith('0,')]
                            
                            if price_candidates:
                                add_debug(f"    Price candidates (non-zero): {price_candidates}")
                                
                                # Strategy: Use 4th position (index 3) for "Bid Total Commit Value" (same as Strategy 1)
                                if len(price_candidates) >= 4:
                                    total_str = price_candidates[3]
                                    add_debug(f"    Selected position 4 (index 3): {total_str}")
                                elif len(price_candidates) >= 2:
                                    total_str = price_candidates[1]
                                    add_debug(f"    Selected position 2: {total_str}")
                                else:
                                    total_str = price_candidates[0]
                                    add_debug(f"    Selected only available: {total_str}")
                                
                                # Parse European format
                                if ',' in total_str and total_str.count(',') == 1:
                                    clean_str = total_str.replace('.', '').replace(',', '.')
                                else:
                                    clean_str = total_str.replace(',', '')
                                
                                total_price_usd = float(clean_str)
                                unit_price_usd = total_price_usd / qty if qty > 0 else total_price_usd
                                
                                add_debug(f"    ✓ Extracted: Total USD {total_price_usd:,.2f}, Unit USD {unit_price_usd:,.2f}")
                                
                                # Convert USD to AED
                                unit_price_aed = unit_price_usd * USD_TO_AED
                                total_price_aed = total_price_usd * USD_TO_AED
                                
                                add_debug(f"  ✓ Prices found: USD {unit_price_usd:,.2f} → AED {unit_price_aed:,.2f}")
                                add_debug(f"  ✓ Total: USD {total_price_usd:,.2f} → AED {total_price_aed:,.2f}")
                            else:
                                add_debug(f"    ✗ All prices are zero or no valid candidates")
                        except Exception as e:
                            add_debug(f"    ✗ Error parsing prices: {e}")
                    else:
                        add_debug(f"  ✗ No prices found")
                    
                    # Extract dates
                    start_date = ""
                    end_date = ""
                    
                    for j in range(max(0, i - 30), min(i + 20, len(lines))):
                        date_matches = date_pattern.findall(lines[j])
                        if date_matches:
                            if not start_date:
                                start_date = date_matches[0]
                            if len(date_matches) > 1:
                                end_date = date_matches[1]
                            if start_date and end_date:
                                add_debug(f"  ✓ Dates found: {start_date} to {end_date}")
                                break
                    
                    # Add extracted row to results - use same format as Strategy 1
                    # Format: [sku, desc, qty, duration, start_date, end_date, unit_price_aed, total_price_aed, partner_price_aed]
                    partner_price_aed = round(total_price_aed * (1 - 0.08), 2)  # 8% discount
                    row_data = [
                        sku_table,
                        desc_table,
                        int(qty) if isinstance(qty, float) and qty.is_integer() else qty,
                        duration,
                        start_date,
                        end_date,
                        round(unit_price_aed, 2),
                        round(total_price_aed, 2),
                        partner_price_aed
                    ]
                    extracted_data.append(row_data)
                    add_debug(f"  ✓ Row added: {sku_table} x {qty}")
                    
                except Exception as e:
                    add_debug(f"  ✗ Error processing table row: {e}")
                    import traceback
                    add_debug(f"Traceback:\n{traceback.format_exc()}")
        
        i += 1
    
    add_debug("\n" + "="*80)
    add_debug("EXTRACTION COMPLETE")
    add_debug("="*80)
    add_debug(f"Total line items found: {line_item_count}")
    add_debug(f"Successfully extracted: {len(extracted_data)}")
    add_debug(f"Failed/Skipped: {line_item_count - len(extracted_data)}")
    
    if extracted_data:
        add_debug("\n" + "="*80)
        add_debug("FINAL EXTRACTED DATA SUMMARY")
        add_debug("="*80)
        for idx, row in enumerate(extracted_data, 1):
            add_debug(f"\nRow {idx}:")
            add_debug(f"  SKU: {row[0]}")
            add_debug(f"  Description (full):")
            add_debug(f"    {row[1]}")
            add_debug(f"  Qty: {row[2]}")
            add_debug(f"  Duration: {row[3]}")
            add_debug(f"  Dates: {row[4]} to {row[5]}")
            add_debug(f"  Prices: Unit={row[6]}, Total={row[7]}")
    
    add_debug("="*80 + "\n")
    logger.info(f"Template 2 extraction completed: {len(extracted_data)} items extracted")
    
    # Add channel discount to header_info for Excel generation
    header_info["Channel Discount"] = f"{global_channel_discount*100:.0f}%"
    
    # Save all debug info to file
    save_debug_to_file()
    
    return extracted_data, header_info


def create_template2_styled_excel(
    data: list,
    header_info: dict,
    logo_path: str,
    output: BytesIO,
    compliance_text: str,
    ibm_terms_text: str
):
    """
    Create Excel for Template 2 with custom headers:
    SI | SKU | Product Description | Quantity | Duration | Unit Price in AED | Cost | Total Price in AED | Partner Price in AED
    
    data rows: [sku, desc, qty, duration, bid_unit_aed, bid_total_aed] (start/end dates removed)
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, NamedStyle
    from openpyxl.drawing.image import Image
    from openpyxl.worksheet.page import PageMargins
    from openpyxl.utils import get_column_letter
    import os
    
    logger.info(f"[TEMPLATE2 EXCEL] Creating Excel with {len(data)} rows")
    add_debug(f"[TEMPLATE2 EXCEL] Creating Excel with {len(data)} rows")
    
    # Calculate total price for terms
    total_price_sum = sum(row[7] if len(row) > 7 else 0 for row in data)
    add_debug(f"[TEMPLATE2 EXCEL] Total price sum: AED {total_price_sum:,.2f}")
    add_debug(f"[TEMPLATE2 EXCEL] Header info before terms: MEP='{header_info.get('Maximum End User Price (MEP)', 'EMPTY')}'")
    
    # Get terms section from template
    try:
        terms = get_terms_section(header_info, total_price_sum)
        add_debug(f"[TEMPLATE2 EXCEL] Terms section generated with {len(terms)} cells")
        # Show first term cell content
        if terms and len(terms) > 1:
            add_debug(f"[TEMPLATE2 EXCEL] First term cell: {terms[1][0]} = {str(terms[1][1])[:150]}")
    except Exception as e:
        add_debug(f"[TEMPLATE2 EXCEL] ERROR generating terms: {e}")
        import traceback
        add_debug(traceback.format_exc())
        terms = []
    
    wb = Workbook()
    ws = wb.active
    ws.title = "IBM Software as a Service Quotation"
    ws.sheet_view.showGridLines = False
    
    # --- Header / Branding ---
    ws.merge_cells("B1:C2")  
    if logo_path and os.path.exists(logo_path):
        try:
            img = Image(logo_path)
            img.width = 1.87 * 96  # 1.87 inches * 96 dpi
            img.height = 0.56 * 96  
            ws.add_image(img, "B1")
            logger.info("Logo added successfully")
        except Exception as e:
            logger.error(f"Error adding logo: {e}")
    else:
        logger.warning(f"Logo file not found: {logo_path}")
    
    # Styling
    blue_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True, size=10)
    blue_font = Font(color="1F497D", bold=True, size=12)
    normal_font = Font(size=10)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Company info (moved to right side)
    ws["E1"] = "Mind Tool Information Technology"
    ws["E1"].font = blue_font
    ws["E2"] = "Trading License No: 1088516"
    ws["E2"].font = Font(size=9)
    ws["E3"] = "Tel: +971 4 320 5334"
    ws["E3"].font = Font(size=9)
    
    # Document title
    ws.merge_cells("B5:I5")
    ws["B5"] = "IBM Software as a Service Quotation"
    ws["B5"].font = Font(color="1F497D", bold=True, size=16)
    ws["B5"].alignment = Alignment(horizontal='center')
    
    # Customer info section
    row = 7
    customer_fields = [
        ("Customer Name:", header_info.get('customer_name', 'N/A')),
        ("Contact Person:", header_info.get('contact_person', 'N/A')), 
        ("Bid Number:", header_info.get('bid_number', 'N/A')),
        ("Territory:", header_info.get('territory', 'UAE')),
        ("Currency:", "AED"),
        ("Quotation Date:", header_info.get('quote_date', 'N/A'))
    ]
    
    for label, value in customer_fields:
        ws[f"B{row}"] = label
        ws[f"B{row}"].font = Font(bold=True, size=10)
        ws[f"C{row}"] = value
        ws[f"C{row}"].font = normal_font
        row += 1
    
    # Table headers for Template 2 (exact format requested)
    table_start_row = row + 2
    headers = [
        "SI",
        "SKU", 
        "Product Description",
        "Quantity",
        "Duration",
        "Unit Price in AED",
        "Cost",
        "Total Price in AED", 
        "Partner Price in AED"
    ]
    
    # Create header row
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=table_start_row, column=col_idx)
        cell.value = header
        cell.font = white_font
        cell.fill = blue_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data rows
    total_amount = 0
    for row_idx, row_data in enumerate(data, 1):
        excel_row = table_start_row + row_idx
        
        # Extract data: [sku, desc, qty, duration, start_date, end_date, unit_price_aed, total_price_aed, partner_price_aed]
        sku = row_data[0]
        description = row_data[1]
        quantity = row_data[2]
        duration = row_data[3] if row_data[3] else ""
        unit_price = row_data[6]  # bid_unit_aed
        total_price = row_data[7]  # bid_total_aed
        partner_price = row_data[8]  # partner_price_aed
        
        # Calculate USD cost (reverse conversion from AED)
        cost_usd = round(unit_price / USD_TO_AED, 2) if unit_price else ""
        
        values = [
            row_idx,  # SI (serial number)
            sku,      # SKU
            description,  # Product Description
            quantity,     # Quantity
            duration,     # Duration
            f"{unit_price:.2f}" if unit_price else "",  # Unit Price in AED
            f"{cost_usd:.2f}" if cost_usd else "",     # Cost
            f"{total_price:.2f}" if total_price else "", # Total Price in AED
            f"{partner_price:.2f}" if partner_price else ""  # Partner Price in AED
        ]
        
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=excel_row, column=col_idx)
            cell.value = value
            cell.font = normal_font
            cell.border = thin_border
            
            # Center align SI, SKU, Quantity, Duration
            if col_idx in [1, 2, 4, 5]:  # SI, SKU, Quantity, Duration
                cell.alignment = Alignment(horizontal='center', vertical='top')
            # Right align prices and cost
            elif col_idx in [6, 7, 8, 9]:  # Price columns
                cell.alignment = Alignment(horizontal='right', vertical='top')
            # Left align description
            else:
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        if total_price:
            total_amount += total_price
    
    # Total row
    total_row = table_start_row + len(data) + 1
    ws.merge_cells(f"A{total_row}:G{total_row}")
    ws[f"A{total_row}"] = "TOTAL AMOUNT"
    ws[f"A{total_row}"].font = Font(bold=True, size=11)
    ws[f"A{total_row}"].alignment = Alignment(horizontal='right')
    ws[f"A{total_row}"].border = thin_border
    
    ws[f"H{total_row}"] = f"{total_amount:.2f}"
    ws[f"H{total_row}"].font = Font(bold=True, size=11)
    ws[f"H{total_row}"].alignment = Alignment(horizontal='right')
    ws[f"H{total_row}"].border = thin_border
    
    # Partner total in column I
    ws[f"I{total_row}"] = f"{total_amount:.2f}"
    ws[f"I{total_row}"].font = Font(bold=True, size=11)
    ws[f"I{total_row}"].alignment = Alignment(horizontal='right')
    ws[f"I{total_row}"].border = thin_border
    
    # Adjust column widths
    column_widths = {
        'A': 8,   # SI
        'B': 15,  # SKU
        'C': 40,  # Description
        'D': 12,  # Quantity
        'E': 12,  # Duration
        'F': 18,  # Unit Price AED
        'G': 15,  # Cost
        'H': 18,  # Total Price AED
        'I': 18   # Partner Price AED
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # Terms and Conditions from template
    terms_start_row = total_row + 3
    add_debug(f"[TEMPLATE2 EXCEL] Adding terms at row {terms_start_row}")
    
    if terms:
        # Apply all terms cells from get_terms_section()
        try:
            for cell_ref, value, *style_args in terms:
                style_info = style_args[0] if style_args else {}
                add_debug(f"[TEMPLATE2 TERMS] Writing to {cell_ref}: {str(value)[:50]}...")
                
                ws[cell_ref] = value
                
                # Apply styling if provided
                if style_info:
                    if style_info.get("bold"):
                        ws[cell_ref].font = Font(bold=True, size=style_info.get("size", 11), color=style_info.get("color", "000000"))
                    if style_info.get("merge_rows"):
                        # Handle merge rows directive (not directly applicable to cell-by-cell approach)
                        ws[cell_ref].alignment = Alignment(wrap_text=True, vertical='top')
                else:
                    ws[cell_ref].font = Font(size=11)
                    ws[cell_ref].alignment = Alignment(wrap_text=True, vertical='top')
            
            add_debug(f"[TEMPLATE2 TERMS] Successfully added {len(terms)} term cells")
        except Exception as e:
            add_debug(f"[TEMPLATE2 TERMS] ERROR adding terms: {e}")
    else:
        add_debug("[TEMPLATE2 TERMS] No terms data available")
    
    # Page setup
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = False
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)
    
    # Create IBM Terms sheet
    terms_ws = wb.create_sheet("IBM Terms")
    if ibm_terms_text:
        terms_ws["A1"] = ibm_terms_text
        terms_ws["A1"].font = Font(size=10)
        terms_ws["A1"].alignment = Alignment(wrap_text=True, vertical='top')
        terms_ws.column_dimensions['A'].width = 100
    
    # Save workbook
    wb.save(output)
    logger.info("Template 2 Excel file generated successfully")
    add_debug("[TEMPLATE2 EXCEL] Workbook saved successfully")
    add_debug(f"[TEMPLATE2 EXCEL] Final workbook has sheet: {ws.title}")