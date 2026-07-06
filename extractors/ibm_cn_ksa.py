# extractors/ibm_cn_ksa.py
"""
IBM Credit Note Automation (KSA)
Upload multiple IBM credit note PDFs and produce one CNTS-UPLOAD workbook
with two sheets: CNTS_HEADER and CNTS_ITEM (one row per PDF in each).
"""

import io
import re
from datetime import datetime
from typing import Dict, List, Tuple

import fitz  # PyMuPDF
import pandas as pd
from openpyxl.utils import get_column_letter

CNTS_HEADER_COLS = [
    "S.No",
    "Date - (dd/MM/yyyy)",
    "Supp_Code",
    "Curr_Code",
    "Form_Code",
    "Doc_Src_Locn",
    "Location_Code",
    "Remarks",
    "Supplier Ref",
    "Supplier Ref Date",
]

CNTS_ITEM_COLS = [
    "S.No",
    "Ref. Key",
    "Item_Code",
    "Item_Name",
    "Grade1",
    "Grade2",
    "UOM",
    "Qty",
    "Qty_Ls",
    "Rate",
    "Main_Account",
    "Sub_Account",
    "Division",
    "Department",
    "Analysis-2",
]

IBM_SUPP_CODE = "SDII020"
IBM_FORM_CODE = 0
IBM_DOC_SRC_LOCN = "KA000"
IBM_LOCATION_CODE = "KA200"
IBM_ITEM_CODE = "NS"
IBM_ITEM_NAME = "Non Stock (Software type)"
IBM_GRADE1 = "NA"
IBM_GRADE2 = "NA"
IBM_UOM = "NOS"
IBM_QTY = 1
IBM_QTY_LS = 0
IBM_MAIN_ACCOUNT = 12741
IBM_SUB_ACCOUNT = "SDII020"
IBM_DIVISION = "ASAU"
IBM_DEPARTMENT = "FIN"
IBM_ANALYSIS2 = "ZZ-COMM"


def _read_pdf_text_all_pages(pdf_bytes: bytes) -> str:
    with fitz.open(stream=pdf_bytes, filetype="pdf") as doc:
        out = []
        for i in range(doc.page_count):
            out.append(doc.load_page(i).get_text("text") or "")
        return "\n".join(out)


def _extract_credit_fields(text: str) -> Dict[str, str]:
    """
    Pull document number, credit note date, currency and total amount from
    an IBM credit note.
    """
    fields = {
        "doc_number": "",
        "credit_date": "",
        "currency": "",
        "total_amount": "",
    }

    compact = " ".join(text.split())

    # PyMuPDF interleaves labels and values:
    # "Document number 6902321968 Reference number / 0101013060 Date 30/03/2026 ..."
    m = re.search(r"Document\s+number\s+(\d+)", compact, re.IGNORECASE)
    if m:
        fields["doc_number"] = m.group(1).strip()
    else:
        # Layout fallback: "... new/old 6902321968 / 0101013060 30/03/2026 ..."
        m = re.search(r"(\d{10})\s*/\s*\d{10}\s+\d{2}/\d{2}/\d{4}", compact)
        if m:
            fields["doc_number"] = m.group(1).strip()

    m = re.search(r"\bDate\s+(\d{2}/\d{2}/\d{4})", compact, re.IGNORECASE)
    if m:
        fields["credit_date"] = m.group(1).strip()

    # "Total amount USD -1,787.33" (not "Total Net amount")
    m = re.search(
        r"Total\s+amount\s+([A-Z]{3})\s+(-?[\d,]+\.\d{2})",
        compact,
        re.IGNORECASE,
    )
    if not m:
        m = re.search(
            r"Total\s+Net\s+amount\s+([A-Z]{3})\s+(-?[\d,]+\.\d{2})",
            compact,
            re.IGNORECASE,
        )
    if m:
        fields["currency"] = m.group(1).strip()
        fields["total_amount"] = m.group(2).replace(",", "")

    return fields


def _build_rows(doc_no: int, fields: Dict[str, str]) -> Tuple[Dict, Dict]:
    """
    Build the CNTS_HEADER and CNTS_ITEM rows for one PDF.
    Rate is the credit note total as a positive value.
    """
    amount = abs(float(fields["total_amount"])) if fields["total_amount"] else 0.0
    rate = int(amount) if amount.is_integer() else amount

    header_row = {
        "S.No": doc_no,
        "Date - (dd/MM/yyyy)": datetime.now().strftime("%d/%m/%Y"),
        "Supp_Code": IBM_SUPP_CODE,
        "Curr_Code": fields["currency"] or "USD",
        "Form_Code": IBM_FORM_CODE,
        "Doc_Src_Locn": IBM_DOC_SRC_LOCN,
        "Location_Code": IBM_LOCATION_CODE,
        "Remarks": f"Booking of IBM CN# {fields['doc_number']}".strip(),
        "Supplier Ref": fields["doc_number"],
        "Supplier Ref Date": fields["credit_date"],
    }

    item_row = {
        "S.No": doc_no,
        "Ref. Key": doc_no,
        "Item_Code": IBM_ITEM_CODE,
        "Item_Name": IBM_ITEM_NAME,
        "Grade1": IBM_GRADE1,
        "Grade2": IBM_GRADE2,
        "UOM": IBM_UOM,
        "Qty": IBM_QTY,
        "Qty_Ls": IBM_QTY_LS,
        "Rate": rate,
        "Main_Account": IBM_MAIN_ACCOUNT,
        "Sub_Account": IBM_SUB_ACCOUNT,
        "Division": IBM_DIVISION,
        "Department": IBM_DEPARTMENT,
        "Analysis-2": IBM_ANALYSIS2,
    }

    return header_row, item_row


def process_ibm_ksa_pdfs(files: List[Tuple[str, bytes]]) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    Input: list of (filename, pdf_bytes)
    Output: two DataFrames for CNTS_HEADER and CNTS_ITEM.
    One PDF produces one row in each sheet.
    """
    header_rows: List[Dict] = []
    item_rows: List[Dict] = []

    doc_no = 1
    for _, blob in files:
        text = _read_pdf_text_all_pages(blob)
        fields = _extract_credit_fields(text)

        header_row, item_row = _build_rows(doc_no, fields)
        header_rows.append(header_row)
        item_rows.append(item_row)
        doc_no += 1

    header_df = pd.DataFrame(header_rows, columns=CNTS_HEADER_COLS)
    item_df = pd.DataFrame(item_rows, columns=CNTS_ITEM_COLS)
    return header_df, item_df


def prepare_ibm_ksa_excel_bytes(header_df: pd.DataFrame, item_df: pd.DataFrame) -> bytes:
    """
    Build the CNTS-UPLOAD workbook with two sheets:
    - CNTS_HEADER
    - CNTS_ITEM
    """
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        header_df.to_excel(writer, index=False, sheet_name="CNTS_HEADER")
        item_df.to_excel(writer, index=False, sheet_name="CNTS_ITEM")

        header_ws = writer.sheets["CNTS_HEADER"]
        item_ws = writer.sheets["CNTS_ITEM"]

        # Keep everything as General, with date columns explicitly treated as text.
        for ws in (header_ws, item_ws):
            for row in ws.iter_rows():
                for cell in row:
                    cell.number_format = "General"

        for cell in header_ws["B"][1:]:
            cell.number_format = "@"
        for cell in header_ws["J"][1:]:
            cell.number_format = "@"

        # Make the workbook a little easier to read.
        for ws in (header_ws, item_ws):
            for col_idx, column in enumerate(ws.iter_cols(1, ws.max_column), start=1):
                max_len = 0
                for cell in column:
                    value = "" if cell.value is None else str(cell.value)
                    if len(value) > max_len:
                        max_len = len(value)
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)

    buf.seek(0)
    return buf.getvalue()


def build_ibm_ksa_output_filename():
    return "CNTS-UPLOAD.xlsx"
