# ibm.py
from decimal import Decimal
import os
import re
import logging
from datetime import datetime
from io import BytesIO
import pandas as pd
import fitz  # PyMuPDF
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from terms_template import get_terms_section

# ----------------------------------------------------------------------
# Debug info collector for live environment
# ----------------------------------------------------------------------
debug_info = []

def add_debug(message):
    """Add debug info that can be displayed in Streamlit"""
    debug_info.append(message)
    if len(debug_info) > 300:  # Keep more debug messages
        debug_info.pop(0)

def get_debug_info():
    """Get collected debug info"""
    return debug_info.copy()

def clear_debug():
    """Clear debug info"""
    debug_info.clear()

# ----------------------------------------------------------------------
# Minimal logging for Excel verification
# ----------------------------------------------------------------------
def setup_debug_logging():
    """Setup minimal debug logging to debug.log file"""
    debug_logger = logging.getLogger('ibm_debug')
    debug_logger.setLevel(logging.INFO)  # Only INFO and above
    
    # Remove existing handlers
    for handler in debug_logger.handlers[:]:
        debug_logger.removeHandler(handler)
    
    # Create file handler for debug.log
    file_handler = logging.FileHandler('debug.log', mode='w', encoding='utf-8')
    file_handler.setLevel(logging.INFO)
    
    # Create simple formatter
    formatter = logging.Formatter('%(message)s')  # Simple format
    file_handler.setFormatter(formatter)
    
    debug_logger.addHandler(file_handler)
    return debug_logger

# Initialize debug logger
debug_logger = setup_debug_logging()

# Also keep simplified logging for live environment
logging.basicConfig(level=logging.ERROR)  # Only log errors in live

# ----------------------------------------------------------------------
# Constants
# ----------------------------------------------------------------------
USD_TO_AED = 3.6725

# ----------------------------------------------------------------------
# Number parsers
# ----------------------------------------------------------------------
def parse_euro_number(value: str):
    """
    Parse EU-formatted numbers like:
    - '733,00'         -> 733.00
    - '114.030,00'     -> 114030.00
    - '60,770'         -> 60.770
    """
    try:
        if value is None:
            return None
        s = str(value).strip().replace(" ", "")
        if "." in s and "," in s:
            if s.rfind(",") > s.rfind("."):
                # thousands '.', decimal ','
                s = s.replace(".", "").replace(",", ".")
            else:
                # thousands ',', decimal '.'
                s = s.replace(",", "")
        else:
            s = s.replace(",", ".")
        return float(s)
    except Exception:
        return None

# ----------------------------------------------------------------------
# Debug function for data integrity
# ----------------------------------------------------------------------
def debug_extracted_data(extracted_data):
    """Debug function to check data integrity"""
    add_debug(f"[DATA DEBUG] Total extracted rows: {len(extracted_data)}")
    for i, row in enumerate(extracted_data):
        add_debug(f"[DATA ROW {i+1}] Length: {len(row)}, SKU: '{row[0]}', Desc: '{row[1][:30]}...'")
        if len(row) != 7:
            add_debug(f"[DATA ERROR] Row {i+1} has {len(row)} columns instead of 7!")
        if not row[0]:  # Check if SKU is empty
            add_debug(f"[DATA ERROR] Row {i+1} has empty SKU!")
        if i == 10:  # Row 11 specifically
            add_debug(f"[ROW 11 SPECIFIC] Full content: {row}")
    return extracted_data

# ----------------------------------------------------------------------
# Description correction
# ----------------------------------------------------------------------
def correct_descriptions(extracted_data, master_data=None):
    """
    Each row: [sku, desc, qty, start_date, end_date, bid_unit_svp_aed, bid_ext_svp_aed]
    Description policy:
    - If master_data uploaded: Use ONLY master CSV descriptions (blank if SKU not found)
    - If master_data NOT uploaded: Set ALL descriptions to blank
    - Never use PDF descriptions
    """
    debug_logger.info("=== DESCRIPTION CORRECTION ===")
    corrected = []
    
    add_debug(f"[DESC CORRECTION] Starting with {len(extracted_data)} rows")
    
    # First debug the data before correction
    debug_extracted_data(extracted_data)
    
    if master_data is not None:
        debug_logger.info(f"Using master CSV with {len(master_data)} records")
        try:
            master_map = dict(zip(master_data['SKU'], master_data['SKU DESCRIPTION']))
            add_debug(f"[MASTER DATA] Using master CSV with {len(master_map)} SKU mappings")
            
            corrections_made = 0
            corrections_blank = 0
            
            for i, row in enumerate(extracted_data):
                try:
                    sku = row[0]
                    
                    if sku in master_map:
                        row[1] = master_map[sku]
                        add_debug(f"[DESC FROM MASTER] Row {i+1} - SKU '{sku}': Found in master CSV")
                        corrections_made += 1
                    else:
                        row[1] = ""  # Blank if SKU not found in master
                        add_debug(f"[DESC BLANK] Row {i+1} - SKU '{sku}' not found in master CSV")
                        corrections_blank += 1
                        
                except Exception as e:
                    add_debug(f"[DESC ERROR] Row {i+1} - Error: {e}")
                    row[1] = ""
                    
                corrected.append(row)
            
            debug_logger.info(f"Corrections: {corrections_made} updated, {corrections_blank} blank")
                
        except Exception as e:
            add_debug(f"[MASTER DATA ERROR] Could not process master data: {e}")
            for row in extracted_data:
                row[1] = ""
            corrected = extracted_data
    else:
        add_debug(f"[NO MASTER DATA] Setting all descriptions to blank")
        debug_logger.info("No master data - setting descriptions to blank")
        for i, row in enumerate(extracted_data):
            row[1] = ""
        corrected = extracted_data
    
    add_debug(f"[DESC CORRECTION COMPLETE] Processed {len(corrected)} rows")
    debug_logger.info(f"Description correction complete: {len(corrected)} rows")
    
    # Debug after correction
    add_debug(f"[AFTER CORRECTION] Row 11: {corrected[10] if len(corrected) > 10 else 'N/A'}")
    
    return corrected

# ----------------------------------------------------------------------
# Enhanced regexes
# ----------------------------------------------------------------------
date_re = re.compile(r'\b\d{2}[‐‑–-][A-Za-z]{3}[‐‑–-]\d{4}\b')  # hyphen variants
sku_line_re = re.compile(r'^[A-Z0-9\-\._/]{5,20}$')
token_sku_re = re.compile(r'\b[A-Z0-9\-\._/]{5,20}\b')
int_re = re.compile(r'\b\d+\b')
money_with_sep_re = re.compile(r'\d[\d.,]*[.,]\d+')

header_blacklist_re = re.compile(
    r'\b('
    r'Bid Number|Bid Request|Opportunity Number|Distributor|Distributor Name|Supplier|'
    r'Page \d+ of|IBM Ireland Product Distribution Limited|IBM Terms|Customer Information|'
    r'PA Anniversary Date|MEP|Maximum End User Price|Parts Information|Current PA Price Level|'
    r'SVP|GV|Yes|No'
    r')\b', re.I
)

def looks_like_valid_sku(tok: str) -> bool:
    """Enhanced SKU validation for IBM part numbers"""
    if not tok:
        return False
    
    # CRITICAL: Explicitly reject serial numbers that start with IE and are long
    if tok.startswith('IE') and len(tok) > 8:
        return False
    
    # Skip pure numbers
    if tok.isdigit():
        return False
    
    # Basic pattern check
    if not sku_line_re.match(tok):
        return False
    
    # Must contain at least one letter and one digit
    if not re.search(r'[A-Z]', tok):
        return False
    if not re.search(r'\d', tok):
        return False
    
    # IBM SKUs are typically 7-8 characters, allow some flexibility
    if not (6 <= len(tok) <= 9):
        return False
    
    # Additional IBM-specific patterns (most start with D0, Y0, etc.)
    if re.match(r'^[A-Z]\d[A-Z0-9]{5,7}$', tok):
        return True
    
    return False

# ----------------------------------------------------------------------
# Qty inference (ANY qty, no small-number assumptions)
# ----------------------------------------------------------------------
def _pick_qty_from_candidates(candidates, unit, ext, abs_tol=0.02):
    """
    Choose qty q in candidates that minimizes |ext - unit*q|.
    Accept if min absolute error <= abs_tol (≈ 2 decimal rounding).
    If no candidates fit, try round(ext/unit).
    """
    if unit is None or ext is None or unit <= 0:
        return None
    best_q, best_err = None, float("inf")
    for q in candidates:
        err = abs(ext - unit * q)
        if err < best_err:
            best_q, best_err = q, err
    if best_q is not None and best_err <= abs_tol:
        return best_q
    # Fallback by division
    q_est = int(round(ext / unit))
    if q_est > 0 and abs(ext - unit * q_est) <= abs_tol:
        return q_est
    return None

def infer_qty_and_prorate(after_end: str, abs_tol=0.02):
    """
    Infer Qty using Entitled Ext ≈ Qty * Entitled Unit (to cent rounding).
    Split using STRICT money regex (must contain ',' or '.') so plain ints remain as ints.
    Returns: (qty:int|None, prorate:int|None, money_tokens:list[str])
    """
    # 1) First strict money token
    m_first = money_with_sep_re.search(after_end)
    if not m_first:
        return None, None, []
    # 2) Pre-money integers zone and money zone
    ints_zone = after_end[:m_first.start()]
    money_zone = after_end[m_first.start():]
    ints = [int(x) for x in int_re.findall(ints_zone)]
    tokens = money_with_sep_re.findall(money_zone)
    qty = None
    # Try Entitled pair first (tokens[0]=Entitled Unit, tokens[1]=Entitled Ext)
    if len(tokens) >= 2:
        m0 = parse_euro_number(tokens[0])
        m1 = parse_euro_number(tokens[1])
        qty = _pick_qty_from_candidates(ints, m0, m1, abs_tol=abs_tol)
    # Then try Bid pair (tokens[3], tokens[4]) if needed
    if qty is None and len(tokens) >= 5:
        bu = parse_euro_number(tokens[3])
        be = parse_euro_number(tokens[4])
        qty = _pick_qty_from_candidates(ints, bu, be, abs_tol=abs_tol)
    # Last resort: if we have Entitled pair, try division only
    if qty is None and len(tokens) >= 2:
        m0 = parse_euro_number(tokens[0])
        m1 = parse_euro_number(tokens[1])
        if m0 and m1:
            q_est = int(round(m1 / m0))
            if q_est > 0 and abs(m1 - m0 * q_est) <= abs_tol:
                qty = q_est
    # Optional: deduce prorate as the first different int in pre-money zone
    prorate = None
    if ints and qty is not None:
        for n in ints:
            if n != qty:
                prorate = n
                break
    return qty, prorate, tokens

# ----------------------------------------------------------------------
# Helpers to handle wrapped rows (extend after_end)
# ----------------------------------------------------------------------
def _extend_after_end_with_following_lines(lines, start_idx, window, max_extra_lines=8):
    """
    If the chunk cut off the amounts, extend the 'after_end' text with a few following
    lines to capture remaining tokens (Disc%, Bid Unit, Bid Ext). Stop early if we see
    strong signs of a new section/item; otherwise just append.
    """
    ext_parts = []
    for j in range(start_idx + window, min(start_idx + window + max_extra_lines, len(lines))):
        ln = lines[j].strip()
        # Heuristics to cautiously stop if a new section is likely
        if header_blacklist_re.search(ln):
            break
        ext_parts.append(" " + ln)
    return "".join(ext_parts)

# ----------------------------------------------------------------------
# Core PDF extraction
# ----------------------------------------------------------------------
def extract_ibm_data_from_pdf(file_like) -> tuple[list, dict]:
    """
    Extracts line items and header info from an IBM Quotation PDF.
    Args:
        file_like: PDF file stream
    Returns:
      - extracted_data: list of rows
          [sku, desc, qty, start_date, end_date, bid_unit_svp_aed, bid_ext_svp_aed]
      - header_info: dict of customer/bid metadata
    """
    debug_logger.info("=== IBM PDF EXTRACTION STARTED ===")
    clear_debug()  # Clear previous debug info
    
    # Open PDF
    doc = fitz.open(stream=file_like.read(), filetype="pdf")
    debug_logger.info(f"PDF: {len(doc)} pages, extracting data...")
    
    # Collect lines
    lines = []
    for page_num, page in enumerate(doc):
        page_text = page.get_text("text") or page.get_text()
        page_lines = []
        for l in page_text.splitlines():
            if l and l.strip():
                lines.append(l.rstrip())
                page_lines.append(l.rstrip())
    
    debug_logger.info(f"Total lines extracted: {len(lines)}")
    add_debug(f"[PDF INFO] Total lines extracted: {len(lines)}")
    
    # Header fields
    debug_logger.info("Extracting header information...")
    header_info = {
        "Customer Name": "",
        "Bid Number": "",
        "PA Agreement Number": "",
        "PA Site Number": "",
        "Select Territory": "",
        "Government Entity (GOE)": "",
        "IBM Opportunity Number": "",
        "Reseller Name": "",
        "City": "",
        "Country": "",
        "Maximum End User Price (MEP)": ""
    }
    
    # Parse header info (simple look-ahead by 1 line)
    header_fields_found = 0
    for i, line in enumerate(lines):
        if "Customer Name:" in line:
            header_info["Customer Name"] = lines[i + 1].strip() if i + 1 < len(lines) else ""
            header_fields_found += 1
        if "Reseller Name:" in line:
            header_info["Reseller Name"] = lines[i + 1].strip() if i + 1 < len(lines) else ""
        if "Bid Number:" in line:
            header_info["Bid Number"] = lines[i + 1].strip() if i + 1 < len(lines) else ""
        if "PA Agreement Number:" in line:
            header_info["PA Agreement Number"] = ""
        if "PA Site Number:" in line:
            header_info["PA Site Number"] = lines[i + 1].strip() if i + 1 < len(lines) else ""
            header_fields_found += 1
        if "Select Territory:" in line:
            header_info["Select Territory"] = lines[i + 1].strip() if i + 1 < len(lines) else ""
            header_fields_found += 1
        if "Government Entity" in line:
            header_info["Government Entity (GOE)"] = lines[i + 1].strip() if i + 1 < len(lines) else ""
            header_fields_found += 1
        if "IBM Opportunity Number:" in line:
            header_info["IBM Opportunity Number"] = lines[i + 1].strip() if i + 1 < len(lines) else ""
            header_fields_found += 1
        if "City:" in line:
            header_info["City"] = lines[i + 1].strip() if i + 1 < len(lines) else ""
            header_fields_found += 1
        if "Country:" in line:
            header_info["Country"] = lines[i + 1].strip() if i + 1 < len(lines) else ""
            header_fields_found += 1
        if "Maximum End User Price" in line or "MEP" in line:
            # Look for MEP value in same line or next line
            if ":" in line:
                mep_part = line.split(":", 1)[1].strip()
                if mep_part:
                    # Remove currency suffixes like "USD", "AED", etc.
                    mep_clean = re.sub(r'\s*(USD).*$', '', mep_part).strip()
                    # Parse European number format and convert to proper value
                    mep_value = parse_euro_number(mep_clean)
                    if mep_value:
                        header_info["Maximum End User Price (MEP)"] = f"{mep_value:,.2f}"
                        debug_logger.info(f"MEP found in same line: '{mep_part}' -> cleaned: '{mep_clean}' -> {mep_value}")
                        header_fields_found += 1
                elif i + 1 < len(lines):
                    next_line = lines[i + 1].strip()
                    # Remove currency suffixes like "USD", "AED", etc.
                    next_clean = re.sub(r'\s*(USD|AED|EUR).*$', '', next_line).strip()
                    mep_value = parse_euro_number(next_clean)
                    if mep_value:
                        header_info["Maximum End User Price (MEP)"] = f"{mep_value:,.2f}"
                        debug_logger.info(f"MEP found in next line: '{next_line}' -> cleaned: '{next_clean}' -> {mep_value}")
                        header_fields_found += 1
    
    debug_logger.info(f"Header fields found: {header_fields_found}")
    debug_logger.info(f"MEP extracted: '{header_info.get('Maximum End User Price (MEP)', 'Not found')}')")
    
    # === Line Item Extraction ===
    debug_logger.info("Extracting line items...")
    extracted_data = []
    i = 0
    max_window = 12  # Try wider chunks first to capture wrapped rows
    processed_positions = set()  # Track processed line positions to avoid duplicates
    
    add_debug(f"[EXTRACTION START] Beginning extraction from {len(lines)} lines")
    
    while i < len(lines):
        matched = False
        
        # Prefer larger chunks first (helps capture qty + amounts in one chunk)
        for window in range(max_window, 0, -1):
            if i + window > len(lines):
                continue
            chunk_lines = lines[i:i + window]
            chunk = " | ".join(chunk_lines)
            
            # Skip header-ish chunks
            if header_blacklist_re.search(chunk):
                continue
                
            # Must have at least two date tokens
            dates = date_re.findall(chunk)
            if len(dates) < 2:
                continue
            start_date, end_date = dates[0], dates[1]
            
            # ENHANCED SKU identification - look for the ACTUAL SKU, not serial numbers
            sku = None
            desc_start_index = None
            
            add_debug(f"[CHUNK ANALYSIS] Lines {i}-{i+window}: {[line.strip() for line in chunk_lines]}")
            
            # Strategy: Find all valid SKUs, then pick the one that's NOT a serial number
            valid_skus_found = []
            for line_idx, line in enumerate(chunk_lines):
                line = line.strip()
                
                # Skip obvious serial/row numbers  
                if line.isdigit():
                    continue
                
                # Find SKU patterns in this line
                sku_candidates = token_sku_re.findall(line)
                for candidate in sku_candidates:
                    if looks_like_valid_sku(candidate):
                        # Additional check: reject obvious serial numbers
                        if candidate.startswith('IE') and len(candidate) > 8:
                            add_debug(f"[SKU REJECT] Serial number: '{candidate}' in line {line_idx}")
                            continue
                        
                        valid_skus_found.append((candidate, line_idx, line))
                        add_debug(f"[SKU CANDIDATE] Found '{candidate}' in line {line_idx}: '{line}'")

            # Pick the BEST SKU (prefer shorter, IBM-style part numbers)
            if valid_skus_found:
                # Sort by: 1) Not starting with 'IE', 2) Length (shorter preferred), 3) Position
                def sku_priority(sku_info):
                    candidate, line_idx, line = sku_info
                    starts_with_ie = candidate.startswith('IE')
                    length = len(candidate)
                    return (starts_with_ie, length, line_idx)
                
                best_sku_info = sorted(valid_skus_found, key=sku_priority)[0]
                sku, sku_line_idx, _ = best_sku_info
                desc_start_index = sku_line_idx + 1  # Description starts after SKU line
                add_debug(f"[SKU SELECTED] Best SKU: '{sku}' from {len(valid_skus_found)} candidates")
                    
            else:
                add_debug(f"[SKU NOT FOUND] No valid SKUs in chunk lines {i}-{i+window}")
                continue

            # Additional validation: Avoid processing same SKU position multiple times 
            sku_position_key = f"{i + sku_line_idx}_{sku}"  # Use SKU line position + SKU name
            if sku_position_key in processed_positions:
                add_debug(f"[POSITION SKIP] Already processed SKU '{sku}' at position {i + sku_line_idx}")
                continue
            processed_positions.add(sku_position_key)
            
            # Keep only chunks where a money token is "near" the start date (reduces false positives)
            pos_date = chunk.find(start_date)
            near_money = False
            if pos_date >= 0:
                window_text = chunk[max(0, pos_date - 60): pos_date + len(start_date) + 60]
                if money_with_sep_re.search(window_text):
                    near_money = True
            if not near_money:
                continue
            
            # Enhanced description extraction with cleaning
            if desc_start_index is not None and desc_start_index < len(chunk_lines):
                desc_parts = []
                for ln in chunk_lines[desc_start_index:]:
                    # Stop at date patterns
                    if date_re.search(ln):
                        break
                    # Clean the line
                    clean_line = ln.strip()
                    # Remove leading/trailing pipes and extra characters
                    clean_line = re.sub(r'^\|?\s*', '', clean_line)
                    clean_line = re.sub(r'\s*\|?\s*$', '', clean_line)
                    if clean_line and not clean_line.isdigit():  # Skip digit-only lines
                        desc_parts.append(clean_line)
                
                desc = " ".join(desc_parts).strip()
                # Remove any remaining pipe characters and clean up
                desc = re.sub(r'\s*\|\s*', ' ', desc)
                desc = re.sub(r'\s+', ' ', desc).strip()
                add_debug(f"[DESC CLEANED] SKU '{sku}': '{desc[:50]}...'")
            else:
                # Fallback description extraction
                pos_sku = chunk.find(sku)
                pos_date0 = chunk.find(start_date)
                desc = chunk[pos_sku + len(sku):pos_date0].strip() if pos_sku >= 0 and pos_date0 > pos_sku else ""
                desc = re.sub(r'\s*\|\s*', ' ', desc)
                desc = re.sub(r'\s+', ' ', desc).strip()
                add_debug(f"[DESC FALLBACK] SKU '{sku}': '{desc[:50]}...'")
            
            # ---- Robust Qty inference (ANY value) ----
            chunk_flat = " ".join(chunk_lines)
            all_date_matches = list(date_re.finditer(chunk_flat))
            if len(all_date_matches) < 2:
                continue
            end_date_match = all_date_matches[1]
            after_end = chunk_flat[end_date_match.end():].strip()
            
            # 1) First pass qty + tokens
            qty, prorate, money_tokens = infer_qty_and_prorate(after_end, abs_tol=0.02)
            
            # Show raw PDF content for debugging
            debug_logger.info(f"=== RAW PDF CONTENT FOR SKU {sku} ===")
            debug_logger.info(f"Chunk lines from PDF:")
            for idx, line in enumerate(chunk_lines):
                debug_logger.info(f"  Line {idx}: '{line.strip()}'")
            debug_logger.info(f"Money tokens found: {money_tokens}")
            debug_logger.info(f"Date range: {start_date} to {end_date}")
            debug_logger.info("=" * 50)
            add_debug(f"[QTY] sku={sku} qty={qty}, money_tokens={len(money_tokens)}")
            
            # 2) If we didn't get all money tokens, extend with a few following lines
            if len(money_tokens) < 5:
                extra = _extend_after_end_with_following_lines(lines, i, window, max_extra_lines=8)
                if extra:
                    qty2, prorate2, money_tokens2 = infer_qty_and_prorate(after_end + " " + extra, abs_tol=0.02)
                    # keep first valid qty but prefer longer token list
                    if qty is None and qty2 is not None:
                        qty, prorate = qty2, prorate2
                    if len(money_tokens2) > len(money_tokens):
                        money_tokens = money_tokens2
                        add_debug(f"[EXTENDED] Extended tokens for sku={sku}: {len(money_tokens)} tokens")
            
            
            
                        # Replace the fallback quantity detection section around line 395-405:
            
                        # Replace the fallback quantity detection:
            
            if qty is None:
                # Strategy 1: Look for decimal quantities FIRST (like 1.780)
                for line_idx, line in enumerate(chunk_lines[:8]):  # Check first 8 lines
                    line = line.strip()
                    # Check for decimal numbers that could be quantities
                    if re.match(r'^\d+\.\d{3}$', line):  # Pattern like 1.780
                        # Convert decimal to integer by multiplying by 1000
                        try:
                            decimal_qty = float(line)
                            if 0.1 <= decimal_qty <= 100:  # Allow up to 100.999 (becomes 100,999)
                                qty = int(decimal_qty * 1000)  # 1.780 * 1000 = 1780
                                add_debug(f"[DECIMAL QTY] sku={sku} converted {line} to {qty} (x1000) at position {line_idx}")
                                break
                        except ValueError:
                            continue
                    # Check for comma-separated thousands (like 1,780)
                    elif re.match(r'^\d{1,3}(,\d{3})+$', line):
                        comma_qty = int(line.replace(',', ''))
                        if 1 <= comma_qty <= 100000:
                            qty = comma_qty
                            add_debug(f"[COMMA QTY] sku={sku} converted {line} to {qty} at position {line_idx}")
                            break
                
                # Strategy 2: Only use first line if no decimal found
                if qty is None:
                    first_line = chunk_lines[0].strip()
                    if first_line.isdigit() and 1 <= int(first_line) <= 100000:
                        qty = int(first_line)
                        add_debug(f"[FALLBACK QTY] sku={sku} using first line qty={qty}")
            
            if qty is None or not (1 <= qty <= 999999):
                add_debug(f"[QTY INVALID] sku={sku} invalid qty={qty}")
                continue
            
            # ---- Extract Standard/List Price instead of Bid Price ----
            bid_unit_svp = None
            bid_ext_svp = None
            try:
                # Strategy: Look for the highest value in money_tokens as it's likely the Standard Price
                if len(money_tokens) >= 1:
                    debug_logger.info(f"=== PRICE ANALYSIS FOR SKU {sku} ===")
                    debug_logger.info(f"All money tokens from PDF: {money_tokens}")
                    
                    # Parse all money values and find the Standard Price (usually the highest unit price)
                    parsed_values = []
                    for i, token in enumerate(money_tokens):
                        try:
                            value = parse_euro_number(token)
                            if value and value > 0:
                                parsed_values.append((value, i, token))
                                debug_logger.info(f"  Token {i}: '{token}' = {value}")
                        except:
                            debug_logger.info(f"  Token {i}: '{token}' = PARSE ERROR")
                            continue
                    
                    if parsed_values:
                        debug_logger.info(f"All parsed values:")
                        for val, idx, token in parsed_values:
                            debug_logger.info(f"  Position {idx}: '{token}' = {val}")
                        
                        # Strategy: Extract both unit cost and extended cost from positions 4 & 5
                        cost_value = None
                        cost_token = None
                        ext_cost_value = None
                        ext_cost_token = None
                        
                        # Look for position 4 (extended cost) and position 5 (unit cost) 
                        for val, idx, token in parsed_values:
                            if idx == 4:  # Extended cost is typically at position 4
                                ext_cost_value = val
                                ext_cost_token = token
                                debug_logger.info(f"FOUND EXTENDED COST: Position 4 '{token}' = {val}")
                            elif idx == 5:  # Unit cost at position 5
                                cost_value = val
                                cost_token = token
                                debug_logger.info(f"FOUND UNIT COST: Position 5 '{token}' = {val}")
                        
                        # Use extended cost if found, otherwise fallback logic
                        if ext_cost_value is not None and ext_cost_value > 10:  # Allow smaller extended costs
                            bid_unit_svp = cost_value if cost_value and cost_value > 100 else ext_cost_value / qty if qty > 0 else ext_cost_value
                            bid_ext_svp = ext_cost_value
                            debug_logger.info(f"USING EXTENDED COST: Unit={bid_unit_svp}, Extended={bid_ext_svp}")
                        elif cost_value is not None and cost_value > 100:
                            bid_unit_svp = cost_value
                            bid_ext_svp = cost_value * qty if qty else cost_value
                            debug_logger.info(f"USING UNIT COST: Unit={bid_unit_svp}, Extended={bid_ext_svp}")
                        else:
                            # Fallback to highest reasonable value
                            reasonable_values = [x for x in parsed_values if x[0] > 1000]
                            if reasonable_values:
                                reasonable_values.sort(key=lambda x: x[0], reverse=True)
                                fallback_value = reasonable_values[0][0]
                                fallback_token = reasonable_values[0][2]
                                debug_logger.info(f"FALLBACK TO HIGHEST REASONABLE: '{fallback_token}' = {fallback_value}")
                                bid_unit_svp = fallback_value
                                bid_ext_svp = fallback_value * qty if qty else fallback_value
                            else:
                                # Last resort: use highest value regardless
                                parsed_values.sort(key=lambda x: x[0], reverse=True)
                                fallback_value = parsed_values[0][0]
                                fallback_token = parsed_values[0][2]
                                debug_logger.info(f"FALLBACK TO HIGHEST: '{fallback_token}' = {fallback_value}")
                                bid_unit_svp = fallback_value
                                bid_ext_svp = fallback_value * qty if qty else fallback_value
                        
                        debug_logger.info(f"SELECTED: Using Extended={bid_ext_svp}, Unit={bid_unit_svp}")
                        debug_logger.info(f"FINAL: Unit={bid_unit_svp}, Total={bid_ext_svp}")
                        debug_logger.info("=" * 50)
                        
                        add_debug(f"[COST PRICE] SKU '{sku}' - Unit={bid_unit_svp}, Extended={bid_ext_svp}")
                    
                if bid_unit_svp is None and len(money_tokens) >= 5:
                    # Fallback to original logic if Standard Price detection fails
                    bid_unit_svp = parse_euro_number(money_tokens[3])
                    bid_ext_svp  = parse_euro_number(money_tokens[4])
                    debug_logger.info(f"FALLBACK: Using tokens[3]={money_tokens[3]} -> {bid_unit_svp}")
                    add_debug(f"[FALLBACK PRICE] SKU '{sku}' - BidUnit={bid_unit_svp}, BidExt={bid_ext_svp}")
                    
            except Exception as e:
                add_debug(f"[PRICE ERROR] sku={sku} err={e}")
            
            # Convert to AED
            bid_unit_svp_aed = round(bid_unit_svp * USD_TO_AED, 2) if bid_unit_svp is not None else None
            bid_ext_svp_aed  = round(bid_ext_svp  * USD_TO_AED, 2) if bid_ext_svp  is not None else None
            
            # Final description cleanup
            desc = re.sub(r'\s{2,}', ' ', desc).strip()
            
            extracted_data.append([sku, desc, qty, start_date, end_date, bid_unit_svp_aed, bid_ext_svp_aed])
            i += window
            matched = True
            add_debug(f"[ROW EXTRACTED] Row {len(extracted_data)}: SKU='{sku}', Qty={qty}")
            break  # break window loop
        if not matched:
            i += 1
    
    add_debug(f"[EXTRACTION COMPLETE] Total rows extracted: {len(extracted_data)}")
    debug_logger.info(f"=== EXTRACTION COMPLETE ===")
    debug_logger.info(f"Total line items: {len(extracted_data)}")
    
    # Log summary for Excel verification
    if extracted_data:
        total_value = sum(row[6] for row in extracted_data if len(row) > 6 and row[6])
        debug_logger.info(f"Total quotation value: AED {total_value:,.2f}")
        debug_logger.info("=== FINAL EXCEL DATA ===")
        for i, row in enumerate(extracted_data, 1):
            if len(row) >= 7:
                debug_logger.info(f"Row {i}: {row[0]} | Qty: {row[2]} | Unit: AED {row[5]:.2f} | Total: AED {row[6]:.2f}")
    
    return extracted_data, header_info

# ----------------------------------------------------------------------
# Extract last page text (for "IBM Terms" sheet)
# ----------------------------------------------------------------------
def extract_last_page_text(file_like) -> str:
    doc = fitz.open(stream=file_like.read(), filetype="pdf")
    last_page = doc[-1]
    full_text = last_page.get_text("text") or last_page.get_text()
    
    # Filter to extract IBM terms content
    lines = full_text.splitlines()
    
    # First, collect the "Useful/Important web resources" section if it appears before IBM Terms
    useful_resources_section = []
    ibm_terms_section = []
    
    # Split into sections
    capture_useful = False
    capture_ibm_terms = False
    
    for line in lines:
        line = line.strip()
        
        # Start capturing useful resources section
        if "Useful/Important web resources:" in line:
            capture_useful = True
            useful_resources_section.append(line)
            continue
            
        # Start capturing IBM terms section
        if "IBM Terms and Conditions" in line:
            capture_ibm_terms = True
            capture_useful = False  # Stop capturing useful resources
            continue  # Skip the header itself
            
        # Skip company header info at the top
        if not capture_useful and not capture_ibm_terms:
            if any(skip_pattern in line for skip_pattern in ["IBM Ireland Product", "VAT: Reg", "Building", "Mulhuddart", "Dublin"]):
                continue
                
        # Capture useful resources content
        if capture_useful and line:
            useful_resources_section.append(line)
            
        # Capture IBM terms content  
        if capture_ibm_terms and line:
            # Stop at page numbers
            if line.lower().startswith("page ") and line.count(" ") <= 3:
                break
            ibm_terms_section.append(line)
    
    # Now reconstruct the IBM terms section into proper paragraphs
    reconstructed_ibm_terms = []
    current_paragraph = []
    
    for line in ibm_terms_section:
        line = line.strip()
        if not line:
            continue
            
        # Check if this line starts a new section/paragraph
        if (line.startswith("IBM International") or 
            line.startswith("The quote or order") or
            line.startswith("Unless specifically") or
            line.startswith("The terms of the IBM") or
            line.startswith("If you have any trouble")):
            
            # Save previous paragraph if exists
            if current_paragraph:
                reconstructed_ibm_terms.append(" ".join(current_paragraph))
                current_paragraph = []
            
            # Start new paragraph
            current_paragraph = [line]
        else:
            # Continue current paragraph
            if current_paragraph:
                current_paragraph.append(line)
            else:
                current_paragraph = [line]
    
    # Add the last paragraph
    if current_paragraph:
        reconstructed_ibm_terms.append(" ".join(current_paragraph))
    
    # Combine sections: IBM Terms first, then Useful Resources at the end
    all_content = reconstructed_ibm_terms
    if useful_resources_section:
        all_content.append("")  # Add spacing
        all_content.extend(useful_resources_section)
    
    result = "\n\n".join(all_content)  # Use double newlines for paragraph separation
    return result

# ----------------------------------------------------------------------
# Excel creation with enhanced debugging
# ----------------------------------------------------------------------
def create_styled_excel(
    data: list,
    header_info: dict,
    logo_path: str,
    output: BytesIO,
    compliance_text: str,
    ibm_terms_text: str
):
    """
    data rows: [SKU, Product Description, Quantity, Start Date, End Date, Unit Price AED, Total Price AED]
    """
    add_debug(f"[EXCEL START] Creating Excel with {len(data)} rows")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Quotation"
    ws.sheet_view.showGridLines = False
    
    # --- Header / Branding ---
    ws.merge_cells("B1:C2")  # Move logo to row 1-2
    if logo_path and os.path.exists(logo_path):
        img = Image(logo_path)
        img.width = 1.87 * 96  # 1.87 inches * 96 dpi
        img.height = 0.56 * 96
        ws.add_image(img, "B1")  # Logo starts at B1
    ws.merge_cells("D3:G3")  # Move title to row 3
    ws["D3"] = "Quotation"
    ws["D3"].font = Font(size=20, color="1F497D")
    ws["D3"].alignment = Alignment(horizontal="center", vertical="center")
    
    ws.column_dimensions[get_column_letter(2)].width  = 8   # B (Sl) 
    ws.column_dimensions[get_column_letter(3)].width  = 15  # C (SKU)
    ws.column_dimensions[get_column_letter(4)].width  = 50  # D (Description) - balanced width
    ws.column_dimensions[get_column_letter(5)].width  = 10  # E (Qty)
    ws.column_dimensions[get_column_letter(6)].width  = 14  # F (Start Date)
    ws.column_dimensions[get_column_letter(7)].width  = 14  # G (End Date) 
    ws.column_dimensions[get_column_letter(8)].width  = 15  # H (Unit Price in AED)
    ws.column_dimensions[get_column_letter(9)].width  = 15  # I (Cost)
    ws.column_dimensions[get_column_letter(10)].width = 18  # J (Total Price in AED)
    ws.column_dimensions[get_column_letter(11)].width = 15  # K (Partner Discount)
    ws.column_dimensions[get_column_letter(12)].width = 18  # L (Partner Price in AED)
    # Left block
    left_labels = ["Date:", "From:", "Email:", "Contact:", "", "Company:", "Attn:", "Email:"]
    left_values = [
        datetime.today().strftime('%d/%m/%Y'),
        "Sneha Lokhandwala",
        "s.lokhandwala@mindware.net",
        "+971 55 456 6650",
        "",
        header_info.get('Reseller Name', 'empty'),
        "empty",
        "empty"
    ]
    row_positions = [5, 6, 7, 8, 9, 10, 11, 12]  # Move up by 1 row
    for row, label, value in zip(row_positions, left_labels, left_values):
        if label:
            ws[f"C{row}"] = label
            ws[f"C{row}"].font = Font(bold=True, color="1F497D")
        if value:
            ws[f"D{row}"] = value
            ws[f"D{row}"].font = Font(color="1F497D")
    
    # IBM Opp no.
    ws["C14"] = "IBM Opportunity Number: "  # Move up by 1 row
    ws["C14"].font = Font(bold=True, underline="single", color="000000")
    ws["D14"] = header_info.get('IBM Opportunity Number', '')
    ws["D14"].font = Font(bold=True, italic=True, underline="single", color="000000")
    
    # Right block
    right_labels = [
        "End User:", "Bid Number:", "Agreement Number:", "PA Site Number:", "",
        "Select Territory:", "Government Entity (GOE):", "Payment Terms:"
    ]
    right_values = [
        header_info.get('Customer Name', ''),
        header_info.get('Bid Number', ''),
        header_info.get('PA Agreement Number', ''),
        header_info.get('PA Site Number', ''),
        "",
        header_info.get('Select Territory', ''),
        header_info.get('Government Entity (GOE)', ''),
        "As aligned with Mindware"
    ]
    for row, label, value in zip(row_positions, right_labels, right_values):
        ws.merge_cells(f"H{row}:I{row}")  # Only merge to column I to stay within table bounds
        ws[f"H{row}"] = f"{label} {value}"
        ws[f"H{row}"].font = Font(bold=True, color="1F497D")
        ws[f"H{row}"].alignment = Alignment(horizontal="left", vertical="center")
    
    # --- Table header (8 columns + serial) ---
    headers = [
    "Sl",                      # Column A - Serial number
    "SKU",                     # Column B - Product code  
    "Product Description",     # Column C - Description
    "Quantity",                # Column D - Number of units
    "Start Date",              # Column E - Coverage start
    "End Date",                # Column F - Coverage end
    "Unit Price in AED",       # Column G - Price per unit
    "Cost (USD)",              # Column H - Base cost in USD
    "Total Price in AED",      # Column I - Final amount
    "Partner Discount",        # Column J - Discount percentage
    "Partner Price in AED"     # Column K - Discounted price
]
    header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    for col, header in enumerate(headers, start=2):
        ws.merge_cells(start_row=16, start_column=col, end_row=17, end_column=col)  # Move up by 1 row
        cell = ws.cell(row=16, column=col, value=header)  # Move up by 1 row
        cell.font = Font(bold=True, size=13, color="1F497D")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = header_fill
    
    # --- Data rows with enhanced debugging ---
    row_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    start_row = 18  # Move up by 1 row
    
    # Build a name->column index map from the header row for robust formatting
    headers_map = {ws.cell(row=16, column=c).value: c for c in range(2, 2 + len(headers))}  # Update header row reference
    col_unit = headers_map.get("Unit Price in AED")
    col_total = headers_map.get("Total Price in AED")
    
    for idx, row in enumerate(data, start=1):
        excel_row = start_row + idx - 1
        
        # Debug: Show what we're processing
        add_debug(f"[EXCEL WRITE] Processing row {idx}: SKU={row[0]}, Desc={row[1][:30]}...")
        
        # Serial number in column B (2)
        cell_sl = ws.cell(row=excel_row, column=2, value=idx)
        cell_sl.font = Font(size=11, color="1F497D")
        cell_sl.alignment = Alignment(horizontal="center", vertical="center")
        
        # Data values in columns C-I (3-9)
                # Extract and calculate data for 11-column structure
        sku = row[0] if len(row) > 0 else ""
        desc = row[1] if len(row) > 1 else ""
        qty = row[2] if len(row) > 2 else 0
        start_date = row[3] if len(row) > 3 else ""
        end_date = row[4] if len(row) > 4 else ""
        bid_unit_svp_aed = row[5] if len(row) > 5 else 0
        bid_ext_svp_aed = row[6] if len(row) > 6 else 0
        
        # Calculate new columns - Cost shows Extended Cost in USD, Unit Price and Total Price in AED
        # Extract USD values from the AED values (reverse conversion)
        bid_unit_svp = bid_unit_svp_aed / USD_TO_AED if bid_unit_svp_aed else 0  # Unit cost in USD 
        bid_ext_svp = bid_ext_svp_aed / USD_TO_AED if bid_ext_svp_aed else 0  # Extended cost in USD 
        cost_usd = bid_ext_svp  # Cost column shows EXTENDED cost in USD (e.g., 90,571.95)
        total_price_aed = cost_usd * USD_TO_AED if cost_usd else 0  # Total Price = Extended Cost × conversion
        unit_price_aed = total_price_aed / qty if qty and qty > 0 else 0  # Unit Price = Total / Quantity
        
        # LIVE DEBUG: Cost column calculation
        print(f"🔍 ROW {idx} DEBUG:")
        print(f"   SKU: {sku}")
        print(f"   Quantity: {qty}")
        print(f"   bid_unit_svp_aed (Unit AED): {bid_unit_svp_aed}")
        print(f"   bid_ext_svp_aed (Ext AED): {bid_ext_svp_aed}")
        print(f"   bid_ext_svp (Ext USD): {bid_ext_svp}")
        print(f"   cost_usd (Extended Cost): {cost_usd}")
        print(f"   total_price_aed (ext cost × 3.6725): {total_price_aed}")
        print(f"   unit_price_aed (total / qty): {unit_price_aed}")
        print("---")
        
        # Also add to debug log
        add_debug(f"[COST DEBUG] Row {idx}: qty={qty}, cost_usd={cost_usd}, total_price_aed={total_price_aed}, unit_price_aed={unit_price_aed}")
        partner_discount = round(unit_price_aed * 0.99, 2) if unit_price_aed else 0
        partner_price_aed = round(partner_discount * qty, 2) if partner_discount and qty else 0
        # Write 10 columns (C-L) - Cost in USD, others in AED
        excel_data = [sku, desc, qty, start_date, end_date, unit_price_aed, cost_usd, total_price_aed, partner_discount, partner_price_aed]
        for j, value in enumerate(excel_data):
            excel_col = j + 3  # C=3, D=4, E=5, F=6, G=7, H=8, I=9, J=10, K=11, L=12
            cell = ws.cell(row=excel_row, column=excel_col, value=value)
            cell.font = Font(size=11, color="1F497D")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            
        # Currency formatting for price columns - Cost (I=9) in USD, others in AED
        for price_col in [8, 10, 11, 12]:  # AED columns: Unit Price, Total, Partner Discount, Partner Price
            ws.cell(row=excel_row, column=price_col).number_format = '"AED"#,##0.00'
        # USD formatting for Cost column (I=9)
        ws.cell(row=excel_row, column=9).number_format = '"USD"#,##0.00'
        # Row fill
        for col in range(2, 2 + len(headers)):
            ws.cell(row=excel_row, column=col).fill = row_fill
        
        # Description wrap & left align (column D = 4)
        ws.cell(row=excel_row, column=4).alignment = Alignment(wrap_text=True, horizontal="left", vertical="center")
        
        # Currency formats
        if col_unit:
            ws.cell(row=excel_row, column=col_unit).number_format = '"AED"#,##0.00'
        if col_total:
            ws.cell(row=excel_row, column=col_total).number_format = '"AED"#,##0.00'
    
    # --- Summary rows ---
    summary_row = start_row + len(data) + 2
    ws.merge_cells(f"C{summary_row}:G{summary_row}")
    ws[f"C{summary_row}"] = "TOTAL Bid Discounted Price"
    ws[f"C{summary_row}"].font = Font(bold=True, color="1F497D")
    ws[f"C{summary_row}"].alignment = Alignment(horizontal="right")
    
    # Sum Total Price (index 6 in data list)
        # Replace around line 825:
    
    # Sum Total Price in AED (sum of cost × USD_TO_AED for each row)
    total_bid_aed = sum(((row[5] / USD_TO_AED * USD_TO_AED) if len(row) >= 6 and row[5] else 0) for row in data)
    ws[f"J{summary_row}"] = total_bid_aed  # Put total in Total Price AED column (J)
    ws[f"J{summary_row}"].number_format = '"AED"#,##0.00'
    ws[f"J{summary_row}"].font = Font(bold=True, color="1F497D")
    ws[f"J{summary_row}"].fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    # Second summary row - TOTAL BP Special Discounted Price
    bp_summary_row = summary_row + 1
    ws.merge_cells(f"C{bp_summary_row}:G{bp_summary_row}")
    ws[f"C{bp_summary_row}"] = "TOTAL BP Special Discounted Price excluding VAT:"
    ws[f"C{bp_summary_row}"].font = Font(bold=True, color="1F497D")
    ws[f"C{bp_summary_row}"].alignment = Alignment(horizontal="right")
    
    # Calculate sum of Partner Price in AED (column L values)
    # Partner price is calculated as: partner_discount * qty for each row
    total_bp_special = 0
    for row in data:
        if len(row) >= 3:  # Need at least sku, desc, qty
            qty = row[2] if row[2] else 0
            bid_unit_svp_aed = row[5] if len(row) > 5 and row[5] else 0
            cost_usd = bid_unit_svp_aed / USD_TO_AED if bid_unit_svp_aed else 0
            total_price_aed = cost_usd * USD_TO_AED if cost_usd else 0
            unit_price_aed = total_price_aed / qty if qty and qty > 0 else 0
            partner_discount = unit_price_aed * 0.99 if unit_price_aed else 0
            partner_price_aed = partner_discount * qty if partner_discount and qty else 0
            total_bp_special += partner_price_aed
    
    ws[f"L{bp_summary_row}"] = total_bp_special  # Put total in Partner Price AED column (L)
    ws[f"L{bp_summary_row}"].number_format = '"AED"#,##0.00'
    ws[f"L{bp_summary_row}"].font = Font(bold=True, color="1F497D")
    ws[f"L{bp_summary_row}"].fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    # --- Dynamic Terms block (main sheet) ---
    total_price_sum = total_bid_aed
    terms = get_terms_section(header_info, total_price_sum)
    
    def estimate_line_count(text, max_chars_per_line=80):
        lines = text.split('\n')
        total_lines = 0
        for line in lines:
            if not line:
                total_lines += 1
            else:
                wrapped = len(line) // max_chars_per_line + (1 if (len(line) % max_chars_per_line) else 0)
                total_lines += max(1, wrapped)
        return total_lines
    
    # Calculate the actual end row of the table content dynamically (now includes BP Special row)
    table_end_row = start_row + len(data) + 5  # data rows + 2 summary rows + spacing
    terms_start_row = max(29, table_end_row + 2)  # Ensure terms start after table
    
    # Adjust terms positioning dynamically
    adjusted_terms = []
    row_offset = terms_start_row - 29  # Calculate offset from original row 29
    
    for cell_addr, text, *style in terms:
        try:
            # Extract row number and adjust it
            if len(cell_addr) >= 2 and cell_addr[1:].isdigit():
                col_letter = cell_addr[0]
                original_row = int(cell_addr[1:])
                new_row = original_row + row_offset
                new_cell_addr = f"{col_letter}{new_row}"
                adjusted_terms.append((new_cell_addr, text, *style))
            else:
                # Keep original if parsing fails
                adjusted_terms.append((cell_addr, text, *style))
        except Exception as e:
            adjusted_terms.append((cell_addr, text, *style))
    
    # Render the terms blocks
    for cell_addr, text, *style in adjusted_terms:
        try:
            if len(cell_addr) >= 2 and cell_addr[1:].isdigit():
                row_num = int(cell_addr[1:])
                col_letter = cell_addr[0]
                merge_rows = style[0].get("merge_rows") if style else None
                end_row = row_num + (merge_rows - 1 if merge_rows else 0)
                
                ws.merge_cells(f"{col_letter}{row_num}:H{end_row}")
                ws[cell_addr] = text
                ws[cell_addr].alignment = Alignment(wrap_text=True, vertical="top")
                # Height by estimated wrap - balanced for content visibility
                line_count = estimate_line_count(text, max_chars_per_line=80)
                total_height = max(18, line_count * 16)
                if merge_rows:
                   per_row = total_height / merge_rows
                   for r in range(row_num, end_row + 1):
                        ws.row_dimensions[r].height = per_row
                else:
                    ws.row_dimensions[row_num].height = total_height
                if style and "bold" in style[0]:
                    ws[cell_addr].font = Font(**style[0])
        except Exception as e:
            pass

    # Divider line across current header row
    border_row = 4
    bottom_border = Border(bottom=Side(style="thin", color="000000"))
    table_last_col = 9  # Only go to column I
    for col in range(1, table_last_col + 1):
        ws.cell(row=border_row, column=col).border = bottom_border
    
    # Calculate IBM Terms start position
    try:
        last_terms_row = max([int(addr[1:]) + (style[0].get("merge_rows", 1) - 1) 
                             for addr, text, *style in adjusted_terms 
                             if style and len(addr) >= 2 and addr[1:].isdigit()], 
                             default=terms_start_row + 10)
    except Exception:
        last_terms_row = terms_start_row + 10
        
    current_row = last_terms_row + 3
    
    # Force a page break before IBM Terms section to prevent awkward splits
    ws.row_breaks.append(current_row - 1)
    
    # IBM Terms header
    ibm_header_cell = ws[f"C{current_row}"]
    ibm_header_cell.value = "IBM Terms and Conditions"
    ibm_header_cell.font = Font(bold=True, size=12, color="1F497D")
    current_row += 2
    
    # Add IBM Terms content - use complete paragraphs with proper text wrapping
    # Split by double newlines to get complete paragraphs
    paragraphs = [p.strip() for p in ibm_terms_text.split('\n\n') if p.strip()]
    
    for paragraph in paragraphs:
        if paragraph:
            # Use one cell per complete paragraph - let Excel handle text wrapping
            ws.merge_cells(f"C{current_row}:H{current_row}")
            cell = ws[f"C{current_row}"]
            
            # Check if paragraph contains a URL
            url_pattern = r'https?://[^\s]+'
            urls = re.findall(url_pattern, paragraph)
            
            if urls:
                for url in urls:
                    cell.hyperlink = url
                    cell.value = paragraph
                    cell.font = Font(size=10, color="0563C1", underline="single")
            else:
                cell.value = paragraph
                cell.font = Font(size=10, color="000000")
            
            # Enable text wrapping for proper paragraph display
            cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
            
            # Calculate row height based on paragraph length
            estimated_lines = max(2, len(paragraph) // 100 + 1)  # More generous estimation
            row_height = max(25, estimated_lines * 15)  # 15 points per line
            ws.row_dimensions[current_row].height = row_height
            
            current_row += 1
            
            # Add extra spacing between major sections
            if "Useful/Important web resources" in paragraph:
                current_row += 1
                current_row += 1
    
    # Page setup - optimized for 11 columns (B through L)
    first_col = 2
    last_col = 12  # Column L
    last_row = ws.max_row
    
    ws.print_area = f"B1:L{last_row}"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1  # Fit all columns to page width
    ws.page_setup.fitToHeight = 0  # Allow multiple pages vertically
    
    # Reduced margins to maximize space for 11 columns
    ws.page_margins.left = 0.2
    ws.page_margins.right = 0.2
    ws.page_margins.top = 0.3
    ws.page_margins.bottom = 0.3
    ws.page_margins.header = 0.2
    ws.page_margins.footer = 0.2
    
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.draft = False
    ws.page_setup.blackAndWhite = False
    # Remove fixed scale - let fitToWidth handle scaling automatically
    ws.sheet_properties.pageSetUpPr.fitToPage = True  # Enable fit-to-page
    
    add_debug(f"[EXCEL COMPLETE] Saved Excel with {len(data)} data rows")
    wb.save(output)

# Function to get debug info for Streamlit display
def get_extraction_debug():
    """Get debug info for display in Streamlit"""
    return get_debug_info()