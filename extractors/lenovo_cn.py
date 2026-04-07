# extractors/lenovo_credit_note.py
"""
Lenovo Credit Note Extractor
- UAE: existing single-sheet CN upload format
- KSA: two-sheet CNTS workbook format
"""

import io
import re
from datetime import datetime
from typing import Dict, List, Tuple

import fitz  # PyMuPDF
import pandas as pd
from openpyxl.utils import get_column_letter

# === Output header with Doc No first ===
CN_HEADERS = [
    "Doc No", "Doc Dt", "Seq No", "Ref Seq No", "Manual Entry Y/N",
    "Main A/C", "Sub A/C", "Div", "Dept", "Anly1", "Anly2", "Acty1", "Acty2",
    "Currency", "FC Amt", "LC Amt", "Dr/Cr", "Detail Narration", "Header Narration",
    "Paym Mode", "Chq Book Id", "Chq No", "Chq Dt", "Payee Name", "Val Date",
    "Doc Ref", "TH Doc ref", "Due Dt",
] + [f"FLEX_{i:02d}" for i in range(1, 51)] + [
    "Party Code", "NOP/NOR", "Tax Code", "Expense Code", "DISC Code",
] + [f"TH_FLEX_{i:02d}" for i in range(1, 51)]

# === Constants / defaults ===
SUB_AC = "SDIL006"
DIV = "PUHO"
DEPT = "GEN"
NARR_PREFIX = "LENOVO(PCG) SELLOUT REBATE RECEIPT / AGREEMENT # "
MAIN_AC_D = "14301"   # when Dr/Cr = D
MAIN_AC_C = "12741"   # when Dr/Cr = C

def _read_pdf_text_all_pages(pdf_bytes: bytes) -> str:
    """
    Lenovo CN totals can appear on later pages (e.g., page 3).
    Read ALL pages to avoid missing 'Total of Products/Services'.
    """
    with fitz.open(stream=pdf_bytes, filetype="pdf") as doc:
        out = []
        for i in range(doc.page_count):
            out.append(doc.load_page(i).get_text("text") or "")
        return "\n".join(out)

def _normalize_date(d: str) -> str:
    """
    Convert dates like 12-FEB-2026 to dd/mm/yyyy.
    """
    for fmt in ("%d-%b-%Y", "%d-%b-%Y"):
        try:
            return datetime.strptime(d.strip().upper(), fmt).strftime("%d/%m/%Y")
        except Exception:
            pass
    return d  # return as-is if parse fails

def _extract_fields(text: str) -> Dict[str, str]:
    """
    Pull: Credit No, Credit Date, Currency, Total Amount, Program/Agreement ID.
    Be robust to '**', spacing, and page breaks.
    """
    fields = {"credit_no": "", "credit_date": "", "currency": "", "total_amount": "", "program_id": ""}

    # Compact whitespace but keep order
    compact = " ".join(text.split())

    # --- Credit No / Credit Date ---
    m = re.search(
        r"Credit\s*No\.?\s*:\s*([A-Z0-9/-]+)\s+Credit\s*Date\s*:\s*([0-9]{2}-[A-Za-z]{3}-[0-9]{4})",
        compact, re.IGNORECASE
    )
    if m:
        fields["credit_no"] = m.group(1).strip()
        fields["credit_date"] = _normalize_date(m.group(2))

    # --- Program/Agreement (Program ID or Claim ref ID) ---
    # Program ID (unchanged)
    m = re.search(r"\bProgram\s*ID\s*:\s*(SM-[0-9-]+)", compact, re.IGNORECASE)
    if m:
        fields["program_id"] = m.group(1).strip()
    else:
        # Claim ref ID (NEW - accept letters, numbers, underscore, hyphen)
        m = re.search(r"\bClaim\s*ref\s*ID\s*:\s*([A-Za-z0-9_-]+)", compact, re.IGNORECASE)
        if m:
            fields["program_id"] = m.group(1).strip()

    # --- Currency + Total ---
    # Primary (handles: Total of Products/Services** USD 200.02, extra symbols/spaces)
    m = re.search(
        r"Total\s*of\s*Products/Services.*?([A-Z]{3})\s*([0-9,]+\.\d{2})",
        compact, re.IGNORECASE
    )
    if m:
        fields["currency"] = m.group(1).strip()
        fields["total_amount"] = m.group(2).replace(",", "")
    else:
        # Secondary fallback: any 'Total ... CUR AMT'
        m = re.search(
            r"Total[^A-Za-z0-9]+.*?([A-Z]{3})\s*([0-9,]+\.\d{2})",
            compact, re.IGNORECASE
        )
        if m:
            fields["currency"] = m.group(1).strip()
            fields["total_amount"] = m.group(2).replace(",", "")
        else:
            # Tertiary: 'Sub total 200.02' (no currency). Pair with any detected currency token if present.
            m_sub = re.search(r"Sub\s*total\s*([0-9,]+\.\d{2})", compact, re.IGNORECASE)
            if m_sub:
                fields["total_amount"] = m_sub.group(1).replace(",", "")
                # Try to find a currency token somewhere nearby/global (e.g., USD)
                m_cur = re.search(r"\b([A-Z]{3})\b\s*[0-9,]+\.\d{2}", compact)
                fields["currency"] = (m_cur.group(1) if m_cur else "USD")

    return fields

def _build_two_rows(fields: Dict[str, str], doc_no: int, seq_no: int) -> List[List]:
    """
    Build the two ledger rows (D & C) in the exact header order.
    - LC Amt mirrors FC Amt.
    - All FLEX and TH_FLEX left blank; Party/Tax/Expense/DISC blank.
    - 'Doc No' repeats for both D and C rows.
    - 'Seq No' and 'Ref Seq No' use seq_no.
    """
    doc_dt = fields["credit_date"]
    doc_ref = fields["credit_no"]
    currency = fields["currency"] or "USD"
    amt = fields["total_amount"] or "0.00"
    narration = f"{NARR_PREFIX}{fields['program_id']}".strip()

    def base_row(main_ac: str, drcr: str) -> List:
        row = [
            doc_no,          # Doc No (repeats for D & C)
            doc_dt,          # Doc Dt
            seq_no,          # Seq No
            "",          # Ref Seq No
            "",              # Manual Entry Y/N
            main_ac,         # Main A/C
            SUB_AC,          # Sub A/C
            DIV,             # Div
            DEPT,            # Dept
            "", "", "", "",  # Anly1, Anly2, Acty1, Acty2
            currency,        # Currency
            amt,             # FC Amt
            "",             # LC Amt (mirror FC)
            drcr,            # Dr/Cr
            narration,       # Detail Narration
            narration,       # Header Narration
            "", "", "",      # Paym Mode, Chq Book Id, Chq No
            "",          # Chq Dt  (align with your sample)
            "",              # Payee Name
            doc_dt,          # Val Date
            doc_ref,         # Doc Ref
            doc_ref,         # TH Doc ref
            doc_dt,          # Due Dt
        ]
        row += [""] * 50                        # FLEX_01..50
        row += ["", "", "", "", ""]             # Party, NOP/NOR, Tax, Expense, DISC
        row += [""] * 50                        # TH_FLEX_01..50
        return row

    return [
        base_row(MAIN_AC_D, "D"),
        base_row(MAIN_AC_C, "C"),
    ]

def process_lenovo_credit_pdfs(files: List[Tuple[str, bytes]]) -> pd.DataFrame:
    """
    Input: list of (filename, pdf_bytes)
    Output: DataFrame with CN_HEADERS and rows = 2 per PDF (D & C), Doc No = 1,1 then 2,2 etc.
    """
    all_rows: List[List] = []
    doc_no = 1     # increments per PDF
    seq_no = 1     # increments per PDF (mirrors Doc No)

    for name, blob in files:
        text = _read_pdf_text_all_pages(blob)  # <-- read ALL pages
        fields = _extract_fields(text)

        # Minimal guardrail: if amount still missing, try to salvage from any "<CUR> <amount>" pattern
        if not fields["total_amount"]:
            m_any = re.search(r"\b([A-Z]{3})\s*([0-9,]+\.\d{2})\b", " ".join(text.split()))
            if m_any:
                fields["currency"] = fields["currency"] or m_any.group(1)
                fields["total_amount"] = m_any.group(2).replace(",", "")

        rows = _build_two_rows(fields, doc_no=doc_no, seq_no=seq_no)
        all_rows.extend(rows)
        doc_no += 1
        seq_no += 1

    df = pd.DataFrame(all_rows, columns=CN_HEADERS)
    return df

def prepare_excel_bytes(df: pd.DataFrame) -> bytes:
    # Convert MAIN A/C (Column F) & FC Amt (Column O) to numeric
    # Column names: "Main A/C" and "FC Amt"
    df["Main A/C"] = pd.to_numeric(df["Main A/C"], errors="coerce")
    df["FC Amt"] = pd.to_numeric(df["FC Amt"], errors="coerce")

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False)

        # Optional: Apply Excel number formatting
        ws = writer.sheets["Sheet1"]

        # Main A/C column (F) - integer format
        for cell in ws["F"][1:]:      # skip header row
            cell.number_format = "0"

        # FC Amt column (O) - numeric with 2 decimals
        for cell in ws["O"][1:]:      # skip header row
            cell.number_format = "#,##0.00"

    buf.seek(0)
    return buf.getvalue()


def build_output_filename():
    """
    Returns the filename with today's date in DD-MM-YYYY format.
    Example: lenovo_credit_notes - 10-03-2026.xlsx
    """
    today = datetime.now().strftime("%d-%m-%Y")
    return f"lenovo_credit_notes - {today}.xlsx"


# ================= KSA output =================

KSA_HEADER_HEADERS = [
    "S.No",
    "Date - (dd/MM/yyyy)",
    "Supp_Code",
    "Curr_Code",
    "Form_Code",
    "Doc_Src_Locn",
    "Location_Code",
    "Remarks",
    "Supplier Ref",
    "Supplier Ref Date",
]

KSA_ITEM_HEADERS = [
    "S.No",
    "Ref. Key",
    "Item_Code",
    "Item_Name",
    "Grade1",
    "Grade2",
    "UOM",
    "Qty",
    "Qty_Ls",
    "Rate",
    "Main_Account",
    "Sub_Account",
    "Division",
    "Department",
    "Analysis-2",
]

KSA_SUPP_CODE = "SKIL009"
KSA_CURR_CODE = "USD"
KSA_FORM_CODE = 0
KSA_DOC_SRC_LOCN = "KA000"
KSA_LOCATION_CODE = "KA200"
KSA_ITEM_CODE = "NS"
KSA_ITEM_NAME = "Non Stock (Software type)"
KSA_GRADE1 = "NA"
KSA_GRADE2 = "NA"
KSA_UOM = "NOS"
KSA_QTY = 1
KSA_QTY_LS = 0
KSA_MAIN_ACCOUNT = 12741
KSA_SUB_ACCOUNT = "SDIL003"
KSA_DIVISION = "ASAU"
KSA_DEPARTMENT = "FIN"
KSA_ANALYSIS2 = "ZZ-COMM"


def _extract_credit_fields(text: str) -> Dict[str, str]:
    """
    Shared credit note extraction for Lenovo PDFs.
    Returns credit_no, credit_date, currency, total_amount, program_id.
    """
    fields = {
        "credit_no": "",
        "credit_date": "",
        "currency": "",
        "total_amount": "",
        "program_id": "",
    }

    compact = " ".join(text.split())

    m = re.search(
        r"Credit\s*No\.?\s*:\s*([A-Z0-9/-]+)\s+Credit\s*Date\s*:\s*([0-9]{2}-[A-Za-z]{3}-[0-9]{4})",
        compact,
        re.IGNORECASE,
    )
    if m:
        fields["credit_no"] = m.group(1).strip()
        fields["credit_date"] = _normalize_date(m.group(2))

    m = re.search(r"\bProgram\s*ID\s*:\s*(SM-[0-9-]+)", compact, re.IGNORECASE)
    if m:
        fields["program_id"] = m.group(1).strip()
    else:
        m = re.search(r"\bClaim\s*ref\s*ID\s*:\s*([A-Za-z0-9_-]+)", compact, re.IGNORECASE)
        if m:
            fields["program_id"] = m.group(1).strip()

    m = re.search(
        r"Total\s*of\s*Products/Services.*?([A-Z]{3})\s*([0-9,]+\.\d{2})",
        compact,
        re.IGNORECASE,
    )
    if m:
        fields["currency"] = m.group(1).strip()
        fields["total_amount"] = m.group(2).replace(",", "")
    else:
        m = re.search(
            r"Total[^A-Za-z0-9]+.*?([A-Z]{3})\s*([0-9,]+\.\d{2})",
            compact,
            re.IGNORECASE,
        )
        if m:
            fields["currency"] = m.group(1).strip()
            fields["total_amount"] = m.group(2).replace(",", "")
        else:
            m_sub = re.search(r"Sub\s*total\s*([0-9,]+\.\d{2})", compact, re.IGNORECASE)
            if m_sub:
                fields["total_amount"] = m_sub.group(1).replace(",", "")
                m_cur = re.search(r"\b([A-Z]{3})\b\s*[0-9,]+\.\d{2}", compact)
                fields["currency"] = m_cur.group(1) if m_cur else "USD"

    return fields


def _build_ksa_rows(doc_no: int, total_amount: str, credit_no: str) -> Tuple[Dict[str, str], Dict[str, str]]:
    """
    Build the KSA header and item rows for one PDF.
    """
    today = datetime.now().strftime("%d/%m/%Y")
    rate_value = float(total_amount) if total_amount else 0.0
    rate = int(rate_value) if rate_value.is_integer() else rate_value

    header_row = {
        "S.No": doc_no,
        "Date - (dd/MM/yyyy)": today,
        "Supp_Code": KSA_SUPP_CODE,
        "Curr_Code": KSA_CURR_CODE,
        "Form_Code": KSA_FORM_CODE,
        "Doc_Src_Locn": KSA_DOC_SRC_LOCN,
        "Location_Code": KSA_LOCATION_CODE,
        "Remarks": f"Booking of Lenovo CN# {credit_no}".strip(),
        "Supplier Ref": credit_no,
        "Supplier Ref Date": today,
    }

    item_row = {
        "S.No": doc_no,
        "Ref. Key": doc_no,
        "Item_Code": KSA_ITEM_CODE,
        "Item_Name": KSA_ITEM_NAME,
        "Grade1": KSA_GRADE1,
        "Grade2": KSA_GRADE2,
        "UOM": KSA_UOM,
        "Qty": KSA_QTY,
        "Qty_Ls": KSA_QTY_LS,
        "Rate": rate,
        "Main_Account": KSA_MAIN_ACCOUNT,
        "Sub_Account": KSA_SUB_ACCOUNT,
        "Division": KSA_DIVISION,
        "Department": KSA_DEPARTMENT,
        "Analysis-2": KSA_ANALYSIS2,
    }

    return header_row, item_row


def process_lenovo_ksa_pdfs(files: List[Tuple[str, bytes]]) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    Input: list of (filename, pdf_bytes)
    Output: two DataFrames for CNTS_HEADER and CNTS_ITEM.
    One PDF produces one row in each sheet.
    """
    header_rows: List[Dict[str, str]] = []
    item_rows: List[Dict[str, str]] = []

    doc_no = 1
    for _, blob in files:
        text = _read_pdf_text_all_pages(blob)
        fields = _extract_credit_fields(text)

        header_row, item_row = _build_ksa_rows(
            doc_no=doc_no,
            total_amount=fields["total_amount"],
            credit_no=fields["credit_no"],
        )
        header_rows.append(header_row)
        item_rows.append(item_row)
        doc_no += 1

    header_df = pd.DataFrame(header_rows, columns=KSA_HEADER_HEADERS)
    item_df = pd.DataFrame(item_rows, columns=KSA_ITEM_HEADERS)
    return header_df, item_df


def prepare_ksa_excel_bytes(header_df: pd.DataFrame, item_df: pd.DataFrame) -> bytes:
    """
    Build the KSA workbook with two sheets:
    - CNTS_HEADER
    - CNTS_ITEM
    """
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        header_df.to_excel(writer, index=False, sheet_name="CNTS_HEADER")
        item_df.to_excel(writer, index=False, sheet_name="CNTS_ITEM")

        header_ws = writer.sheets["CNTS_HEADER"]
        item_ws = writer.sheets["CNTS_ITEM"]

        # Keep everything as General, with date columns explicitly treated as text.
        for ws in (header_ws, item_ws):
            for row in ws.iter_rows():
                for cell in row:
                    cell.number_format = "General"

        for cell in header_ws["B"][1:]:
            cell.number_format = "@"
        for cell in header_ws["J"][1:]:
            cell.number_format = "@"

        # Make the workbook a little easier to read.
        for ws in (header_ws, item_ws):
            for col_idx, column in enumerate(ws.iter_cols(1, ws.max_column), start=1):
                max_len = 0
                for cell in column:
                    value = "" if cell.value is None else str(cell.value)
                    if len(value) > max_len:
                        max_len = len(value)
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)

    buf.seek(0)
    return buf.getvalue()


def build_ksa_output_filename():
    """
    Returns the KSA filename with today's date in DD-MM-YYYY format.
    """
    today = datetime.now().strftime("%d-%m-%Y")
    return f"lenovo_ksa_credit_notes - {today}.xlsx"
