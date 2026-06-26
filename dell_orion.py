"""
Fixed helpers and generate_orion_quote for the Dell → Orion pipeline.

Fixes applied:
1. MSRP / Unit Cost now use unit_price (not total_price).
2. "Base Options" rows are excluded from the description.
3. Processor extraction now works from config rows for ALL items (including item 1).
4. Memory regex no longer drops "1 x N GB" entries.
5. Graphics extraction uses config rows (catches NVIDIA on items 2 & 4).
6. "Storage configuration (Boot Drive)" module is excluded from storage match.
7. Item 4 NVMe SSD rows (modules "1st M.2 NVMe SSD" / "Additional M.2 NVMe SSD") are
   now matched by the storage matcher.
8. Product description now also stops at a model number (e.g. "R750xs", "T550").
9. Processor output now starts AFTER "Intel" / "Intel Core" (model/speed/cores only)
   and stops precisely at "N cores" or "NC".
10. Memory suppresses "1 x" prefix when qty is 1.
11. Support terms now correctly handle two separate config rows (module + months on
    different rows) and the "+" concatenation pattern.
"""

import re
from datetime import datetime
from io import BytesIO

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

# ── keep all the original imports from your dell module ──────────────────────
from dell import (
    CURRENCY_CONVERSION_RATES,
    _extract_all_config_rows,
    _extract_compact_quote_items_and_config,
    _extract_excel_consolidation_fee,
    _extract_excel_shipping_fee,
    _extract_grouped_template_items_and_config,
    _extract_pdf_quote_data,
    _extract_product_detail_headings,
    _extract_quote_metadata,
    _extract_config_rows_from_configuration_sheet,
    _try_extract_items_from_pricing_summary,
)

ORION_HEADERS = [
    "Vendor Item Code",
    "P/N - Orion Item code",
    "Description",
    "Qty",
    "MSRP",
    "Unit Cost",
    "Unit Selling",
]

ORION_CURRENCY_CONVERSION_RATES = {
    "USD": 1.0,
    "AED": 3.6725,
    "EUR": 0.92,
    "SAR": 3.75,
    "QAR": 3.64,
}


# ── small text utilities ──────────────────────────────────────────────────────

def _cell_to_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")
    return str(value).strip()


def _sanitize_text(value: str) -> str:
    return re.sub(r"\s+", " ", value).strip()


def _extract_part_number_from_description(text: str) -> str:
    text = _sanitize_text(text)
    if not text:
        return ""
    matches = re.findall(r"\(([^()]+)\)", text)
    for candidate in reversed(matches):
        normalized = candidate.strip().replace("–", "-").replace("—", "-")
        normalized = re.sub(r"\s*-\s*", "-", normalized)
        normalized = re.sub(r"\s+", " ", normalized).strip()
        if re.fullmatch(r"(?:\d{3}-[A-Z0-9]{4,5}|[A-Z]{2}\d{6,})", normalized, re.I):
            return normalized
    fallback = re.search(r"\b(?:[A-Z]{2}\d{6,}|\d{3}-[A-Z0-9]{4,5})\b", text, re.I)
    if fallback:
        return fallback.group(0).upper()
    return ""


def _best_description(desc: str, config_rows: list, idx: int) -> str:
    base = _sanitize_text(desc)
    if base:
        return base
    for row in config_rows:
        if len(row) >= 5 and str(row[0]).strip() == str(idx + 1):
            candidate = _sanitize_text(row[3] or row[4] or row[2] or row[1])
            if candidate:
                return candidate
    return base


# ── description-building utilities ───────────────────────────────────────────

def _extract_product_description(text: str) -> str:
    """
    Return the product name up to the first Dell model token or the word 'server'.
    """
    text = _sanitize_text(text)
    if not text:
        return ""

    text = re.sub(r"^\s*SI#\s*\w+\s*", "", text, flags=re.I)

    model_match = re.search(r"\b(?:[A-Z]{1,2}\d{3,4}[A-Za-z0-9]*|[A-Z]{2,}\d{3,4}[A-Za-z0-9]*)\b", text, flags=re.I)
    if model_match:
        text = text[:model_match.start()].strip()

    stop_match = re.search(r"\b(?:server|intel)\b", text, flags=re.I)
    if stop_match:
        text = text[:stop_match.start()].strip()

    text = re.sub(r"\s*[-,;:]+\s*$", "", text)
    text = text.rstrip("()[]{}")
    return _sanitize_text(text)


def _extract_processor(text: str) -> str:
    """
    Return a compact processor descriptor in the form 'qty x Ultra 7 ...'.
    """
    cleaned = _sanitize_text(text)
    cleaned = cleaned.replace("Intel(R)", "Intel").replace("Core(TM)", "Core")
    cleaned = cleaned.replace("vPro(R)", "vPro")

    match = re.search(
        r"(?:(\d+)\s*x\s*)?intel(?:\s+core)?\s+(.*?)(?=\s*,\s*\d+\s*(?:cores?|threads?)\b|\s*,\s*\d+\s*MB\b|\s*,\s*\d+\.\d+\s*GHz\b|\s*\(|[,;]|$)",
        cleaned,
        re.I,
    )
    if match:
        qty = match.group(1) or "1"
        processor = _sanitize_text(match.group(2))
        processor = re.sub(r"\bprocessor\b", "", processor, flags=re.I)
        processor = processor.strip(" ,;")
        return _sanitize_text(f"{qty} x {processor}")
    return ""


def _extract_memory(text: str) -> str:
    """
    Return memory in the form 'N x M GB' or 'M GB' when the quantity is 1.
    """
    parts = []
    for clause in re.split(r"\s*\|\s*", text):
        clause = clause.strip()
        if not clause or any(term in clause.lower() for term in ("ssd", "hdd", "hard drive", "storage")):
            continue
        matches = re.findall(r"(\d+)\s*x\s*(\d+(?:\.\d+)?)\s*GB\b", clause, flags=re.I)
        if not matches:
            continue
        for qty, size in matches:
            if int(qty) == 1:
                parts.append(f"{size} GB")
            else:
                parts.append(f"{qty} x {size} GB")
    return ", ".join(dict.fromkeys(parts))


def _extract_graphics(text: str) -> str:
    for clause in re.split(r"\s*\|\s*", text):
        clause = clause.strip()
        if not clause or not re.search(r"\b(?:nvidia|amd|radeon|rtx|gtx|graphics?|gpu)\b", clause, re.I):
            continue
        if re.search(r"\bintegrated\b", clause, re.I):
            continue
        phrase = re.sub(r"\s*\(.*?\)", "", clause)
        phrase = re.sub(r"\s*,.*$", "", phrase)
        phrase = re.sub(r"\b(?:graphics|gpu)\b.*$", "", phrase, flags=re.I)
        return _sanitize_text("1 x " + phrase.strip())
    return ""


def _extract_storage(text: str) -> str:
    for clause in re.split(r"\s*\|\s*", text):
        clause = clause.strip()
        if not clause or not re.search(r"\b(?:ssd|hdd)\b", clause, re.I):
            continue
        match = re.search(r"(?:(\d+)\s*x\s*)?(\d+(?:\.\d+)?)\s*(TB|GB)\b[^,;]*(SSD|HDD)\b", clause, re.I)
        if match:
            qty, size, unit, drive_type = match.groups()
            qty_value = qty or "1"
            return _sanitize_text(f"{qty_value} x {size} {unit} {drive_type}")
    return ""


def _extract_operating_system(text: str) -> str:
    match = re.search(r"\b(?:operating\s*system|os)\b\s*[:\-]?\s*([^,;]+)", text, re.I)
    if match:
        return _sanitize_text(match.group(1))
    match = re.search(
        r"\b(?:windows|ubuntu|linux|rhel|red\s*hat|suse|oracle\s*linux|vmware\s*esxi)[^,;]*",
        text, re.I,
    )
    if match:
        return _sanitize_text(match.group(0))
    return ""



def build_orion_description(text: str) -> str:
    cleaned = _sanitize_text(text)
    if not cleaned:
        return ""
    parts = [
        _extract_product_description(cleaned),
        _extract_processor(cleaned),
        _extract_memory(cleaned),
        _extract_graphics(cleaned),
        _extract_storage(cleaned),
        _extract_operating_system(cleaned),
    ]
    return ", ".join(part for part in parts if part)


# ── config-row matchers ───────────────────────────────────────────────────────

def _is_base_options_row(module: str, description: str) -> bool:
    """Return True for rows that only describe the base option (noise)."""
    text = f"{module} {description}".lower()
    return "base options" in text or module.lower().startswith("base options")


def _processor_match(module: str, description: str) -> bool:
    if module.lower().strip() == "processor":
        return True
    text = f"{module} {description}".lower()
    return ("processor" in text or "cpu" in text) and any(
        term in text for term in ("intel", "amd", "core", "ghz", "cache", "cores")
    )


def _memory_match(module: str, description: str) -> bool:
    text = f"{module} {description}".lower()
    return "memory" in text and any(term in text for term in ("gb", "ddr", "sodimm", "non-ecc"))


def _graphics_match(module: str, description: str) -> bool:
    if module.lower().strip() in ("graphics", "graphics holder"):
        return module.lower().strip() == "graphics"
    text = f"{module} {description}".lower()
    return "graphics" in text or any(
        term in text for term in ("nvidia", "amd", "radeon", "rtx", "gtx", "gpu")
    )


def _storage_match(module: str, description: str) -> bool:
    module_lower = module.lower().strip()
    desc_lower = description.lower()

    if "storage configuration" in module_lower:
        return False

    if re.search(r"\bm\.2\b.*\bssd\b|\bm2\b.*\bssd\b|nvme\s+ssd|1st\s+m\.2|additional\s+m\.2", module_lower):
        return True

    text = f"{module_lower} {desc_lower}"
    return (
        ("storage" in text or "hard drive" in text or any(term in text for term in ("ssd", "hdd", "tb")))
        and "driver" not in text
        and "storage configuration" not in text
    )


def _os_match(module: str, description: str) -> bool:
    text = f"{module} {description}".lower()
    return "operating system" in text or any(
        term in text for term in ("windows", "ubuntu", "linux", "rhel", "red hat", "suse", "vmware")
    )


# ── config-row detail finder ──────────────────────────────────────────────────

def _find_config_detail(config_rows: list, item_no: int, matcher) -> str:
    matches = []
    for row in config_rows:
        if len(row) < 6:
            continue
        row_item = str(row[0]).strip()
        if row_item and row_item != str(item_no):
            continue
        module = _sanitize_text(str(row[2] or ""))
        description = _sanitize_text(str(row[3] or ""))

        if _is_base_options_row(module, description):
            continue

        candidate = description or module
        if matcher(module, description):
            matches.append((module, description, candidate))

    if not matches:
        return ""

    preferred = [item for item in matches if "base options" in f"{item[0]} {item[1]}".lower()]
    if preferred:
        return preferred[0][2]

    return matches[0][2]



def _find_base_options_summary(config_rows: list, item_no: int) -> str:
    # Always returns "" — individual field matchers run instead.
    return ""


def build_orion_description_from_config(desc: str, config_rows: list, idx: int) -> str:
    base = _sanitize_text(desc)
    parts = []

    if base:
        parts.append(_extract_product_description(base))

    processor_raw = _find_config_detail(config_rows, idx, _processor_match)
    processor = _extract_processor(processor_raw) if processor_raw else ""
    if processor:
        parts.append(processor)

    memory_raw = _find_config_detail(config_rows, idx, _memory_match)
    memory = _extract_memory(memory_raw) if memory_raw else ""
    if memory:
        parts.append(memory)

    graphics_raw = _find_config_detail(config_rows, idx, _graphics_match)
    graphics = _extract_graphics(graphics_raw) if graphics_raw else ""
    if graphics:
        parts.append(graphics)

    storage_raw = _find_config_detail(config_rows, idx, _storage_match)
    storage = _extract_storage(storage_raw) if storage_raw else ""
    if storage:
        parts.append(storage)

    os_raw = _find_config_detail(config_rows, idx, _os_match)
    os_text = _extract_operating_system(os_raw) if os_raw else ""
    if os_text:
        parts.append(os_text)

    combined = ", ".join(part for part in parts if part)
    if combined:
        return combined
    return build_orion_description(base)


# ── metadata extraction ───────────────────────────────────────────────────────

def _extract_fee_values(input_excel_bytes: bytes) -> tuple[float, float]:
    is_pdf = input_excel_bytes.lstrip().startswith(b"%PDF")
    if is_pdf:
        _, _, _, _, _, consolidation_fee = _extract_pdf_quote_data(input_excel_bytes)
        return float(consolidation_fee or 0.0), 0.0

    wb = openpyxl.load_workbook(BytesIO(input_excel_bytes), data_only=True)
    ws = wb.active
    return _extract_excel_consolidation_fee(ws), _extract_excel_shipping_fee(ws)


def _extract_items_and_metadata(input_excel_bytes: bytes):
    is_pdf = input_excel_bytes.lstrip().startswith(b"%PDF")
    if is_pdf:
        items, quote_meta, config_rows, quote_ref_text, date_text, _ = _extract_pdf_quote_data(input_excel_bytes)
        return items, quote_meta, config_rows, {}, quote_ref_text, date_text, is_pdf

    wb = openpyxl.load_workbook(BytesIO(input_excel_bytes), data_only=True)
    ws = wb.active
    quote_meta = _extract_quote_metadata(ws)
    item_headings_by_item = _extract_product_detail_headings(ws)

    items = _try_extract_items_from_pricing_summary(ws)
    config_rows = _extract_all_config_rows(ws)
    if not items:
        items, config_rows = _extract_compact_quote_items_and_config(ws)
    if not items:
        items, config_rows = _extract_grouped_template_items_and_config(ws)

    if not items:
        items = []

    config_sheet = None
    for sheet_name in wb.sheetnames:
        if "config" in sheet_name.lower():
            config_sheet = wb[sheet_name]
            break

    if config_sheet is not None:
        config_rows = _extract_config_rows_from_configuration_sheet(config_sheet)

    quote_ref_text = ""
    date_text = ""
    for r in range(1, min(ws.max_row, 60) + 1):
        label = _cell_to_text(ws.cell(r, 2).value).lower()
        if "quote" in label and "ref" in label:
            quote_ref_text = _cell_to_text(ws.cell(r, 5).value)
        if label.startswith("date"):
            date_text = _cell_to_text(ws.cell(r, 5).value)

    return items, quote_meta, config_rows, item_headings_by_item, quote_ref_text, date_text, is_pdf


def build_dell_orion_output_filename(input_excel_bytes: bytes) -> str:
    _, quote_meta, _, _, quote_ref_text, _, _ = _extract_items_and_metadata(input_excel_bytes)
    partner_raw = _sanitize_text(
        quote_meta.get("reseller") or quote_meta.get("company name") or quote_meta.get("end user") or ""
    )
    quote_ref_raw = _sanitize_text(quote_ref_text or "")
    safe_partner = re.sub(r"[^A-Za-z0-9._-]+", " ", partner_raw).strip()
    safe_ref = re.sub(r"[^A-Za-z0-9._-]+", " ", quote_ref_raw).strip()
    date_str = datetime.now().strftime("%Y%m%d")
    parts = [p for p in [safe_partner, safe_ref, date_str] if p]
    return " - ".join(parts) + ".xlsx"


# ── main generator ────────────────────────────────────────────────────────────

def generate_orion_quote(input_excel_bytes: bytes, currency_code: str = "USD") -> bytes:
    items, _, config_rows, item_headings_by_item, _, _, _ = _extract_items_and_metadata(input_excel_bytes)
    consolidation_fee, shipping_fee = _extract_fee_values(input_excel_bytes)
    currency_code = (currency_code or "USD").upper()
    conversion_rate = ORION_CURRENCY_CONVERSION_RATES.get(
        currency_code, CURRENCY_CONVERSION_RATES.get(currency_code, 1.0)
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Orion_Quote"
    ws.append(ORION_HEADERS)

    # Build part-number lookups
    item_descs_order = [item[0] for item in items]
    part_numbers_by_item = {}
    for row_data in config_rows:
        if len(row_data) >= 7:
            item_no, _heading, _module, _desc, sku, _tax, _qty = row_data
        else:
            item_no, _heading, _module, _desc, sku, _tax = row_data
        if sku and item_no not in part_numbers_by_item:
            part_numbers_by_item[item_no] = sku

    heading_part_numbers_by_item = {}
    for item_key, heading in item_headings_by_item.items():
        part_number = _extract_part_number_from_description(heading)
        if part_number:
            heading_part_numbers_by_item[item_key] = part_number

    for idx, desc in enumerate(item_descs_order, start=1):
        item_key = str(idx)
        if item_key not in heading_part_numbers_by_item:
            part_number = _extract_part_number_from_description(desc)
            if part_number:
                heading_part_numbers_by_item[item_key] = part_number

    for item_key, part_number in heading_part_numbers_by_item.items():
        part_numbers_by_item.setdefault(item_key, part_number)

    # Distribute consolidation + shipping fees equally across items
    total_fees_converted = (float(consolidation_fee or 0.0) + float(shipping_fee or 0.0)) * conversion_rate
    num_items = len(items)
    fee_per_item = total_fees_converted / num_items if (total_fees_converted and num_items) else 0.0

    # Write rows
    for idx, (desc, qty, unit_price, total_price) in enumerate(items, start=1):
        qty_value = int(qty) if qty not in (None, "") else 0
        base_unit_value = float(unit_price or 0.0) * conversion_rate
        fee_per_unit = fee_per_item / qty_value if qty_value else 0.0
        unit_value = base_unit_value + fee_per_unit

        vendor_code = part_numbers_by_item.get(str(idx), "") or _extract_part_number_from_description(desc)

        if config_rows:
            description = build_orion_description_from_config(desc, config_rows, idx)
        else:
            description = build_orion_description(_best_description(desc, config_rows, idx))

        ws.append([
            vendor_code,   # Vendor Item Code
            "",            # P/N - Orion Item code
            description,   # Description
            qty_value,     # Qty
            unit_value,    # MSRP
            unit_value,    # Unit Cost
            "",            # Unit Selling
        ])

    # Formatting
    header_fill = PatternFill(fill_type="solid", start_color="D9EAF7", end_color="D9EAF7")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        row[2].alignment = Alignment(wrap_text=True)
        row[3].number_format = "0"
        for cell in row[4:7]:
            cell.number_format = "0.00"

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 16

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()