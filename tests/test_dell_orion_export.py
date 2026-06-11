from io import BytesIO

import openpyxl

from dell_orion import generate_orion_quote


def test_generate_orion_quote_uses_unit_price_in_price_columns():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Quote"
    ws["B30"] = "Pricing Summary"
    ws["A33"] = "1"
    ws["B33"] = "Dell Pro Max 16 Premium"
    ws["K33"] = 10
    ws["L33"] = 25
    ws["N33"] = 250

    out = generate_orion_quote(wb.save(BytesIO()) or b"", currency_code="USD")
    out_wb = openpyxl.load_workbook(BytesIO(out), data_only=True)
    out_ws = out_wb.active

    assert out_ws["E2"].value == 25
    assert out_ws["F2"].value == 25
