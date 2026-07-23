"""Builds the rebate calculation workbook (backend-only output, UAE deals).

Layout matches the sample the business owner provided:
S.no | Part number | <incentive columns> ($) | Total ($)
... rows ...
                     Total ($) | <grand total, green>
Note: All rebates are subject to IBM approval ...
"""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from rebate.rate_card import REBATE_NOTE

HEADER_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
TOTAL_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
HEADER_FONT = Font(bold=True)
TOTAL_FONT = Font(bold=True)


def build_rebate_workbook(rows, columns):
    """rows/columns as returned by rebate.calculator.compute_rebate_rows."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Rebate"

    header_labels = ["S.no", "Part number"] + [f"{name} ($)" for name in columns] + ["Total ($)"]
    last_col = len(header_labels)

    for col_idx, label in enumerate(header_labels, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    row_idx = 2
    for i, row in enumerate(rows, start=1):
        ws.cell(row=row_idx, column=1, value=i).alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=2, value=row["part_number"]).alignment = Alignment(horizontal="center")
        for col_offset, name in enumerate(columns, start=3):
            cell = ws.cell(row=row_idx, column=col_offset, value=row["amounts"].get(name, 0.0))
            cell.number_format = "#,##0.00"
        ws.cell(row=row_idx, column=last_col, value=row["total"])
        row_idx += 1

    total_row = row_idx
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=last_col - 1)
    total_label_cell = ws.cell(row=total_row, column=1, value="Total ($)")
    total_label_cell.font = TOTAL_FONT
    total_label_cell.alignment = Alignment(horizontal="right", vertical="center")

    grand_total = round(sum(r["total"] for r in rows), 4)
    grand_total_cell = ws.cell(row=total_row, column=last_col, value=grand_total)
    grand_total_cell.font = TOTAL_FONT
    grand_total_cell.fill = TOTAL_FILL

    note_row = total_row + 2
    ws.cell(row=note_row, column=1, value=REBATE_NOTE).font = Font(italic=True, bold=True)

    ws.column_dimensions[get_column_letter(1)].width = 8
    ws.column_dimensions[get_column_letter(2)].width = 16
    for col_idx in range(3, last_col + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 22

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
